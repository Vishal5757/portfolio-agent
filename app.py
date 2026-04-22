import argparse
import ast
import base64
import datetime as dt
import errno
import html
import http.cookiejar
import io
import json
import math
import os
import random
import re
import sqlite3
import subprocess
import threading
import time
import urllib.parse
import urllib.request
from collections import defaultdict, deque
from contextlib import contextmanager
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl

from portfolio_agent.software_performance import (
    get_software_perf_agent_config as _mod_get_software_perf_agent_config,
    list_software_perf_actions as _mod_list_software_perf_actions,
    list_software_perf_snapshots as _mod_list_software_perf_snapshots,
    maybe_run_software_perf_agent_once as _mod_maybe_run_software_perf_agent_once,
    run_software_perf_agent_once as _mod_run_software_perf_agent_once,
    set_software_perf_agent_config as _mod_set_software_perf_agent_config,
    software_performance_worker as _mod_software_performance_worker,
)
from portfolio_agent.risk_analysis import (
    get_risk_agent_config as _mod_get_risk_agent_config,
    list_risk_analysis_snapshots as _mod_list_risk_analysis_snapshots,
    maybe_run_risk_analysis_agent_once as _mod_maybe_run_risk_analysis_agent_once,
    risk_analysis_worker as _mod_risk_analysis_worker,
    run_risk_analysis_agent_once as _mod_run_risk_analysis_agent_once,
    set_risk_agent_config as _mod_set_risk_agent_config,
)


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
WEB_DIR = ROOT / "web"
SCHEMA_PATH = ROOT / "schema.sql"
DB_PATH = DATA_DIR / "portfolio.db"
MARKET_HISTORY_DB_PATH = DATA_DIR / "market_history.db"
UPLOAD_DIR = DATA_DIR / "uploads"
BACKUP_DIR = DATA_DIR / "backups"
REPO_DATA_DIR = ROOT / "repo_data"
TENANTS_ROOT = DATA_DIR / "tenants"
TENANTS_META_PATH = DATA_DIR / "tenants.json"
TENANT_MAX_COUNT = 5
DEFAULT_TENANT_KEY = "default"
TENANT_KEY_RE = re.compile(r"^[a-z0-9][a-z0-9_-]{0,31}$")
TENANT_NAME_MAX_LEN = 48
DEFAULT_XLSX_CANDIDATE_PATHS = (
    DATA_DIR / "Portfolio.xlsx",
    DATA_DIR / "seed" / "Portfolio.xlsx",
    ROOT / "Portfolio.xlsx",
)
LIVE_REFRESH_MIN_SEC = 5
DB_BACKUP_INTERVAL_SEC = 2 * 24 * 60 * 60
STRATEGY_REFRESH_MIN_SEC = 24 * 60 * 60
HISTORY_REFRESH_DEFAULT_SEC = 24 * 60 * 60
HISTORY_REFRESH_MIN_SEC = 60 * 60
CHART_AGENT_INTERVAL_DEFAULT_SEC = 6 * 60 * 60
CHART_AGENT_MIN_INTERVAL_SEC = 60 * 15
SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC = 15 * 60
SOFTWARE_PERF_AGENT_MIN_INTERVAL_SEC = 60
TAX_MONITOR_INTERVAL_DEFAULT_SEC = 24 * 60 * 60
TAX_MONITOR_MIN_INTERVAL_SEC = 60 * 60
LLM_DEFAULT_MODEL = "gpt-4.1-mini"
LLM_DEFAULT_API_URL = "https://api.openai.com/v1/responses"
RISK_AGENT_INTERVAL_DEFAULT_SEC = 6 * 60 * 60
RISK_AGENT_MIN_INTERVAL_SEC = 15 * 60
REPO_SYNC_INTERVAL_DEFAULT_SEC = 60 * 60
REPO_SYNC_MIN_INTERVAL_SEC = 5 * 60
GIT_EXE_CANDIDATES = (
    Path(r"C:\Program Files\Git\cmd\git.exe"),
    Path(r"C:\Program Files\Git\bin\git.exe"),
)
INTEL_FIN_BACKFILL_MIN_INTERVAL_SEC = 6 * 60 * 60
INTEL_FIN_BACKFILL_MAX_SYMBOLS = 60
INTEL_FIN_BACKFILL_MAX_RUNTIME_SEC = 18
DEFAULT_CHART_AGENT_SOURCES = "market_history,tradingview_scan,quote_samples"
BUILTIN_LIVE_QUOTE_SOURCES = (
    "nse_api",
    "bse_api",
    "yahoo_finance",
    "google_scrape",
    "screener_scrape",
    "trendlyne_scrape",
    "cnbc_scrape",
)
DEFAULT_LIVE_QUOTE_SOURCES = ",".join(BUILTIN_LIVE_QUOTE_SOURCES)
DEFAULT_TAX_MONITOR_TAX_URL = "https://zerodha.com/z-connect/business-updates/what-changes-for-investors-after-budget-2024"
DEFAULT_TAX_MONITOR_CHARGES_URL = "https://zerodha.com/charges/"
LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT = 8.0
LIVE_QUOTE_TOP_K_DEFAULT = 4
LIVE_QUOTE_EXPLORE_RATIO_DEFAULT = 0.2
QUOTE_SOURCE_HARD_FAIL_COOLDOWN_SEC = 6 * 60 * 60
STOCK_NSE_INDIA_API_BASE_DEFAULT = "http://127.0.0.1:3000"
IST_TZ = dt.timezone(dt.timedelta(hours=5, minutes=30))
ZERO_QTY_EOD_IST_HOUR = 15
ZERO_QTY_EOD_IST_MINUTE = 35
LAST_TRADE_FALLBACK_MAX_AGE_DAYS = 15
GOLD_QUOTE_STALE_MAX_AGE_SEC = 2 * 60 * 60
CHART_INDEX_PROXY_SYMBOLS = (
    "NIFTYBEES",
    "RELIANCE",
    "HDFCBANK",
    "ICICIBANK",
    "SBIN",
    "TCS",
    "INFY",
    "ITC",
    "LT",
)
ASSET_CLASS_EQUITY = "EQUITY"
ASSET_CLASS_GOLD = "GOLD"
GOLD_HINT_PATTERNS = (
    re.compile(r"\bGOLD\b", re.IGNORECASE),
    re.compile(r"\bSGB[A-Z0-9]*\b", re.IGNORECASE),
    re.compile(r"\bGOLDBEES\b", re.IGNORECASE),
    re.compile(r"\bSILVER\b", re.IGNORECASE),
)

INTEL_POSITIVE_WORDS = {
    "beat", "beats", "growth", "upside", "upgrade", "strong", "improve", "improved", "expansion",
    "capex", "incentive", "support", "reform", "tailwind", "order", "orders", "margin", "profit",
    "record", "guidance", "bullish", "healthy", "resilient", "surplus", "ratecut", "taxcut", "relief",
}
INTEL_NEGATIVE_WORDS = {
    "miss", "downside", "downgrade", "weak", "decline", "pressure", "cuts", "cut", "slowdown",
    "deficit", "inflation", "headwind", "risk", "lawsuit", "penalty", "loss", "debt", "stressed",
    "volatile", "uncertain", "bearish", "selloff", "contraction", "tightening", "hike", "taxhike",
}
INTEL_POLICY_THEMES = {
    "infrastructure": ["budget", "capex", "roads", "rail", "infra", "construction", "housing", "cement"],
    "banking": ["liquidity", "credit", "banking", "nbfc", "rate cut", "rate hike", "repo", "deposit"],
    "manufacturing": ["pli", "manufacturing", "export", "duty", "import", "tariff", "factory"],
    "technology": ["it services", "digital", "ai", "software", "cloud", "tech spend"],
    "pharma": ["healthcare", "pharma", "drug", "api", "hospital"],
    "energy": ["oil", "gas", "power", "renewable", "solar", "wind", "coal"],
}
INTEL_THEME_SYMBOL_HINTS = {
    "infrastructure": ["L&T", "LT", "IRB", "IRCON", "NBCC", "KNR", "ADANI", "ULTRA", "CEMENT"],
    "banking": ["BANK", "FIN", "HDFC", "ICICI", "AXIS", "KOTAK", "SBI", "FEDERAL", "INDUSIND"],
    "manufacturing": ["TATA", "MOTORS", "AUTO", "MARUTI", "EICHER", "M&M", "SIEMENS", "ABB"],
    "technology": ["TCS", "INFY", "WIPRO", "TECHM", "LTIM", "PERSISTENT", "COFORGE"],
    "pharma": ["PHARMA", "DRREDDY", "SUN", "CIPLA", "DIVI", "LUPIN"],
    "energy": ["POWER", "NTPC", "ONGC", "COAL", "IOC", "BPCL", "GAIL", "ADANI"],
}

_tenant_local = threading.local()
_startup_market_data_repair_done = False
_startup_market_data_repair_lock = threading.Lock()
_active_tenant_lock = threading.Lock()
_active_tenant_key = DEFAULT_TENANT_KEY


def sanitize_tenant_key(value):
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    raw = re.sub(r"[^a-z0-9_-]+", "-", raw)
    raw = raw.strip("-_")
    if not raw:
        return ""
    if not TENANT_KEY_RE.match(raw):
        return ""
    return raw


def sanitize_tenant_name(value, fallback="Tenant"):
    name = str(value or "").strip()
    if not name:
        name = str(fallback or "Tenant").strip() or "Tenant"
    name = re.sub(r"[\r\n\t]+", " ", name)
    if len(name) > TENANT_NAME_MAX_LEN:
        name = name[:TENANT_NAME_MAX_LEN].rstrip()
    return name or "Tenant"


def get_active_tenant_key():
    with _active_tenant_lock:
        return str(_active_tenant_key or DEFAULT_TENANT_KEY)


def get_request_tenant_key():
    return str(getattr(_tenant_local, "tenant_key", "") or "").strip()


def get_current_tenant_key():
    req = get_request_tenant_key()
    if req:
        return req
    return get_active_tenant_key()


def tenant_storage_dir(tenant_key):
    key = sanitize_tenant_key(tenant_key) or DEFAULT_TENANT_KEY
    if key == DEFAULT_TENANT_KEY:
        return DATA_DIR
    return TENANTS_ROOT / key


def tenant_paths(tenant_key=None):
    key = sanitize_tenant_key(tenant_key or get_current_tenant_key()) or DEFAULT_TENANT_KEY
    data_dir = tenant_storage_dir(key)
    return {
        "tenant_key": key,
        "data_dir": data_dir,
        "db_path": data_dir / "portfolio.db",
        "market_db_path": data_dir / "market_history.db",
        "upload_dir": data_dir / "uploads",
        "backup_dir": data_dir / "backups",
    }


def get_current_tenant_data_dir():
    return tenant_paths(get_current_tenant_key())["data_dir"]


def _default_tenant_meta():
    created = now_iso()
    return {
        "active": DEFAULT_TENANT_KEY,
        "tenants": [
            {
                "key": DEFAULT_TENANT_KEY,
                "name": "Default",
                "created_at": created,
            }
        ],
    }


def load_tenant_meta():
    DATA_DIR.mkdir(exist_ok=True)
    meta = {}
    if TENANTS_META_PATH.exists():
        try:
            meta = json.loads(TENANTS_META_PATH.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    if not isinstance(meta, dict):
        meta = {}

    tenants_raw = meta.get("tenants")
    if not isinstance(tenants_raw, list):
        tenants_raw = []

    cleaned = []
    seen = set()
    for t in tenants_raw:
        if not isinstance(t, dict):
            continue
        key = sanitize_tenant_key(t.get("key"))
        if not key or key in seen:
            continue
        seen.add(key)
        cleaned.append(
            {
                "key": key,
                "name": sanitize_tenant_name(t.get("name"), fallback=key.upper()),
                "created_at": str(t.get("created_at") or now_iso()),
            }
        )

    if DEFAULT_TENANT_KEY not in seen:
        cleaned.insert(
            0,
            {
                "key": DEFAULT_TENANT_KEY,
                "name": "Default",
                "created_at": now_iso(),
            },
        )
    if len(cleaned) > TENANT_MAX_COUNT:
        if cleaned[0]["key"] != DEFAULT_TENANT_KEY:
            cleaned = [x for x in cleaned if x["key"] == DEFAULT_TENANT_KEY] + [
                x for x in cleaned if x["key"] != DEFAULT_TENANT_KEY
            ]
        cleaned = cleaned[:TENANT_MAX_COUNT]

    active = sanitize_tenant_key(meta.get("active")) or DEFAULT_TENANT_KEY
    valid = {t["key"] for t in cleaned}
    if active not in valid:
        active = DEFAULT_TENANT_KEY

    return {"active": active, "tenants": cleaned}


def save_tenant_meta(meta):
    DATA_DIR.mkdir(exist_ok=True)
    TENANTS_META_PATH.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def ensure_tenant_bootstrap():
    meta = load_tenant_meta()
    save_tenant_meta(meta)
    with _active_tenant_lock:
        global _active_tenant_key
        _active_tenant_key = sanitize_tenant_key(meta.get("active")) or DEFAULT_TENANT_KEY
    active_paths = tenant_paths(_active_tenant_key)
    active_paths["data_dir"].mkdir(parents=True, exist_ok=True)
    active_paths["upload_dir"].mkdir(parents=True, exist_ok=True)
    active_paths["backup_dir"].mkdir(parents=True, exist_ok=True)
    return meta


def set_active_tenant_key(tenant_key, persist=True):
    key = sanitize_tenant_key(tenant_key)
    if not key:
        raise ValueError("invalid_tenant")
    meta = load_tenant_meta()
    keys = {t["key"] for t in meta["tenants"]}
    if key not in keys:
        raise ValueError("tenant_not_found")
    if persist:
        meta["active"] = key
        save_tenant_meta(meta)
    with _active_tenant_lock:
        global _active_tenant_key
        _active_tenant_key = key
    return key


def set_request_tenant_key(tenant_key):
    key = sanitize_tenant_key(tenant_key)
    if not key:
        raise ValueError("invalid_tenant")
    _tenant_local.tenant_key = key
    return key


def clear_request_tenant_key():
    if hasattr(_tenant_local, "tenant_key"):
        delattr(_tenant_local, "tenant_key")


@contextmanager
def tenant_context(tenant_key):
    prev = get_request_tenant_key()
    set_request_tenant_key(tenant_key)
    try:
        yield
    finally:
        if prev:
            _tenant_local.tenant_key = prev
        else:
            clear_request_tenant_key()


def list_tenants():
    meta = load_tenant_meta()
    active = sanitize_tenant_key(meta.get("active")) or DEFAULT_TENANT_KEY
    out = []
    for t in meta.get("tenants", []):
        key = sanitize_tenant_key(t.get("key"))
        if not key:
            continue
        p = tenant_paths(key)
        out.append(
            {
                "key": key,
                "name": sanitize_tenant_name(t.get("name"), fallback=key.upper()),
                "created_at": str(t.get("created_at") or ""),
                "active": key == active,
                "db_path": str(p["db_path"]),
                "market_db_path": str(p["market_db_path"]),
            }
        )
    return out


def create_tenant(key=None, name=None, activate=False):
    meta = load_tenant_meta()
    current = list(meta.get("tenants", []))
    if len(current) >= TENANT_MAX_COUNT:
        raise ValueError("tenant_limit_reached_max_5")
    key_s = sanitize_tenant_key(key or name)
    if not key_s or key_s == DEFAULT_TENANT_KEY:
        raise ValueError("invalid_tenant")
    if any(sanitize_tenant_key(t.get("key")) == key_s for t in current):
        raise ValueError("tenant_already_exists")
    row = {
        "key": key_s,
        "name": sanitize_tenant_name(name, fallback=key_s.upper()),
        "created_at": now_iso(),
    }
    current.append(row)
    meta["tenants"] = current
    if activate:
        meta["active"] = key_s
    save_tenant_meta(meta)
    with tenant_context(key_s):
        init_db()
    if activate:
        set_active_tenant_key(key_s, persist=False)
    p = tenant_paths(key_s)
    return {
        "key": key_s,
        "name": row["name"],
        "created_at": row["created_at"],
        "active": bool(activate),
        "db_path": str(p["db_path"]),
        "market_db_path": str(p["market_db_path"]),
    }


def now_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat()


def ist_now():
    return dt.datetime.now(IST_TZ)


def is_zero_qty_eod_window(now_ist=None):
    now_ist = now_ist or ist_now()
    h = int(now_ist.hour)
    m = int(now_ist.minute)
    return (h > ZERO_QTY_EOD_IST_HOUR) or (h == ZERO_QTY_EOD_IST_HOUR and m >= ZERO_QTY_EOD_IST_MINUTE)


def parse_float(value, default=0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        raw = str(value or "").strip()
        if not raw:
            return default
        cleaned = (
            raw.replace(",", "")
            .replace("â‚¹", "")
            .replace("₹", "")
            .replace("$", "")
            .replace("rs.", "")
            .replace("rs", "")
            .replace("inr", "")
            .replace("%", "")
            .replace("+", "")
            .replace("(", "-")
            .replace(")", "")
            .replace("\u2212", "-")
            .strip()
        )
        if cleaned == "":
            return default
        try:
            return float(cleaned)
        except ValueError:
            low = cleaned.lower()
            # Common compact/financial suffixes used in market data pages.
            multipliers = {
                "k": 1e3,
                "thousand": 1e3,
                "l": 1e5,
                "lac": 1e5,
                "lakh": 1e5,
                "lakhs": 1e5,
                "cr": 1e7,
                "crore": 1e7,
                "crores": 1e7,
                "m": 1e6,
                "mn": 1e6,
                "million": 1e6,
                "b": 1e9,
                "bn": 1e9,
                "billion": 1e9,
            }
            m = re.search(r"([+\-]?\d+(?:\.\d+)?)\s*([a-zA-Z]+)?", low)
            if not m:
                return default
            try:
                base = float(m.group(1))
            except Exception:
                return default
            unit = str(m.group(2) or "").strip().lower()
            mul = multipliers.get(unit, 1.0)
            return base * mul
    return default


def money(value):
    return f"₹{parse_float(value, 0.0):,.2f}"


def clamp(value, low, high):
    return max(low, min(high, value))


def parse_bool(value, default=None):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))
    s = str(value).strip().lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    if default is not None:
        return default
    raise ValueError("invalid_boolean")


def parse_source_list(value, default=None):
    raw = str(value or "").strip()
    if not raw:
        raw = default or DEFAULT_LIVE_QUOTE_SOURCES
    parts = [p.strip().lower() for p in raw.split(",") if str(p).strip()]
    out = []
    for p in parts:
        if not re.match(r"^[a-z0-9_:-]{2,64}$", p):
            continue
        if p not in out:
            out.append(p)
    if not out:
        out = [p for p in BUILTIN_LIVE_QUOTE_SOURCES]
    return out


def parse_token_list(value, default=None):
    raw = str(value or "").strip()
    if not raw:
        raw = str(default or "")
    parts = [p.strip().lower() for p in raw.split(",") if str(p).strip()]
    out = []
    for p in parts:
        if not re.match(r"^[a-z0-9_:-]{2,64}$", p):
            continue
        if p not in out:
            out.append(p)
    return out


def parse_chart_source_list(value, default=None):
    raw = str(value or "").strip()
    if not raw:
        raw = str(default or DEFAULT_CHART_AGENT_SOURCES)
    parts = [p.strip().lower() for p in raw.split(",") if str(p).strip()]
    out = []
    for p in parts:
        if not re.match(r"^[a-z0-9_:-]{2,64}$", p):
            continue
        if p not in out:
            out.append(p)
    if not out:
        out = [p for p in str(DEFAULT_CHART_AGENT_SOURCES).split(",") if p]
    return out


def median_value(values):
    vals = sorted(float(v) for v in values if parse_float(v, 0.0) > 0)
    if not vals:
        return 0.0
    n = len(vals)
    m = n // 2
    if n % 2 == 1:
        return float(vals[m])
    return float((vals[m - 1] + vals[m]) / 2.0)


def parse_excel_date(value):
    def valid_year(d):
        return 1990 <= d.year <= (dt.date.today().year + 2)

    if value is None:
        return None
    if isinstance(value, dt.datetime):
        d = value.date()
        return d if valid_year(d) else None
    if isinstance(value, dt.date):
        return value if valid_year(value) else None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                d = dt.datetime.strptime(raw[:10], fmt).date()
                if valid_year(d):
                    return d
            except ValueError:
                continue
    return None


def db_connect():
    p = tenant_paths(get_current_tenant_key())
    p["data_dir"].mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(p["db_path"])
    conn.row_factory = sqlite3.Row
    return conn


def market_db_connect():
    init_market_history_db()
    p = tenant_paths(get_current_tenant_key())
    conn = sqlite3.connect(p["market_db_path"])
    conn.row_factory = sqlite3.Row
    return conn


def init_market_history_db(tenant_key=None):
    p = tenant_paths(tenant_key or get_current_tenant_key())
    p["data_dir"].mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(p["market_db_path"])
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_prices (
              symbol TEXT NOT NULL,
              price_date TEXT NOT NULL,
              close REAL NOT NULL,
              source TEXT NOT NULL,
              fetched_at TEXT NOT NULL,
              PRIMARY KEY(symbol, price_date)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_prices_symbol_date ON daily_prices(symbol, price_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_prices_date ON daily_prices(price_date)")
        conn.commit()
    finally:
        conn.close()


def init_db():
    p = tenant_paths(get_current_tenant_key())
    p["data_dir"].mkdir(parents=True, exist_ok=True)
    p["upload_dir"].mkdir(parents=True, exist_ok=True)
    p["backup_dir"].mkdir(parents=True, exist_ok=True)
    init_market_history_db()
    with db_connect() as conn:
        conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
        conn.commit()
    ensure_schema_migrations()
    ensure_default_strategy()
    ensure_default_config()


def ensure_schema_migrations():
    with db_connect() as conn:
        cols = [r["name"] for r in conn.execute("PRAGMA table_info(instruments)").fetchall()]
        if "feed_code" not in cols:
            conn.execute("ALTER TABLE instruments ADD COLUMN feed_code TEXT")
        if "price_source" not in cols:
            conn.execute("ALTER TABLE instruments ADD COLUMN price_source TEXT NOT NULL DEFAULT 'exchange_api'")
        if "asset_class" not in cols:
            conn.execute("ALTER TABLE instruments ADD COLUMN asset_class TEXT NOT NULL DEFAULT 'EQUITY'")
        conn.execute(
            """
            UPDATE instruments
            SET asset_class = 'EQUITY'
            WHERE asset_class IS NULL OR TRIM(asset_class) = ''
            """
        )
        inst_rows = conn.execute("SELECT symbol, name, asset_class FROM instruments").fetchall()
        for r in inst_rows:
            desired = infer_asset_class(
                symbol=r["symbol"],
                name=r["name"],
                fallback=normalize_asset_class(r["asset_class"], fallback=ASSET_CLASS_EQUITY),
            )
            if str(r["asset_class"] or "").strip().upper() != desired:
                conn.execute("UPDATE instruments SET asset_class=? WHERE UPPER(symbol)=?", (desired, symbol_upper(r["symbol"])))
        tcols = [r["name"] for r in conn.execute("PRAGMA table_info(trades)").fetchall()]
        if "external_trade_id" not in tcols:
            conn.execute("ALTER TABLE trades ADD COLUMN external_trade_id TEXT")
        ccols = [r["name"] for r in conn.execute("PRAGMA table_info(cash_ledger)").fetchall()]
        if "external_entry_id" not in ccols:
            conn.execute("ALTER TABLE cash_ledger ADD COLUMN external_entry_id TEXT")
        if "source" not in ccols:
            conn.execute("ALTER TABLE cash_ledger ADD COLUMN source TEXT NOT NULL DEFAULT 'cashflow_upload'")
        src_rows = conn.execute(
            "SELECT COUNT(*) AS c FROM cash_ledger WHERE source IS NULL OR TRIM(source) = ''"
        ).fetchone()
        if src_rows and int(src_rows["c"]) > 0:
            conn.execute("UPDATE cash_ledger SET source = 'cashflow_upload' WHERE source IS NULL OR TRIM(source) = ''")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_trades_external_trade_id ON trades(external_trade_id)")
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_trades_upload_trade_id
            ON trades(source, external_trade_id)
            WHERE source = 'tradebook_upload' AND external_trade_id IS NOT NULL AND TRIM(external_trade_id) <> ''
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_cash_ledger_entry_date ON cash_ledger(entry_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_cash_ledger_external_entry_id ON cash_ledger(external_entry_id)")
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_cash_ledger_upload_external_id
            ON cash_ledger(source, external_entry_id)
            WHERE source = 'cashflow_upload' AND external_entry_id IS NOT NULL AND TRIM(external_entry_id) <> ''
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dividends (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              symbol TEXT NOT NULL,
              entry_date TEXT NOT NULL,
              amount REAL NOT NULL,
              reference_text TEXT,
              external_entry_id TEXT,
              source TEXT NOT NULL DEFAULT 'dividend_upload',
              created_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_dividends_symbol_date ON dividends(symbol, entry_date DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_dividends_date ON dividends(entry_date DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_dividends_external_entry_id ON dividends(external_entry_id)")
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_dividends_upload_external_id
            ON dividends(source, external_entry_id)
            WHERE source = 'dividend_upload' AND external_entry_id IS NOT NULL AND TRIM(external_entry_id) <> ''
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS scrip_position_guards (
              symbol TEXT PRIMARY KEY,
              min_value REAL NOT NULL DEFAULT 0,
              max_value REAL,
              updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_scrip_position_guards_updated ON scrip_position_guards(updated_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS rebalance_lots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              side TEXT NOT NULL CHECK (side IN ('BUY','SELL')),
              percent REAL NOT NULL,
              allocation_basis TEXT NOT NULL DEFAULT 'portfolio_market_value',
              target_trade_value REAL NOT NULL DEFAULT 0,
              total_current_market_value REAL NOT NULL DEFAULT 0,
              status TEXT NOT NULL CHECK (status IN ('active','completed','reset')) DEFAULT 'active',
              created_at TEXT NOT NULL,
              completed_at TEXT,
              reset_at TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_rebalance_lots_status_created ON rebalance_lots(status, created_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS rebalance_lot_items (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              lot_id INTEGER NOT NULL REFERENCES rebalance_lots(id) ON DELETE CASCADE,
              symbol TEXT NOT NULL,
              planned_qty REAL NOT NULL DEFAULT 0,
              planned_trade_value REAL NOT NULL DEFAULT 0,
              ltp_at_lock REAL NOT NULL DEFAULT 0,
              market_value_at_lock REAL NOT NULL DEFAULT 0,
              min_value_at_lock REAL NOT NULL DEFAULT 0,
              max_value_at_lock REAL,
              note TEXT,
              completed INTEGER NOT NULL DEFAULT 0,
              completed_at TEXT,
              completion_note TEXT,
              execution_state TEXT NOT NULL DEFAULT 'pending',
              executed_price REAL,
              executed_at TEXT,
              buyback_completed INTEGER NOT NULL DEFAULT 0,
              buyback_completed_at TEXT,
              buyback_price REAL,
              buyback_note TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_rebalance_lot_items_lot ON rebalance_lot_items(lot_id, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_rebalance_lot_items_symbol ON rebalance_lot_items(symbol)")
        rcols = [r["name"] for r in conn.execute("PRAGMA table_info(rebalance_lot_items)").fetchall()]
        if "execution_state" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN execution_state TEXT NOT NULL DEFAULT 'pending'")
        if "executed_price" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN executed_price REAL")
        if "executed_at" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN executed_at TEXT")
        if "buyback_completed" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN buyback_completed INTEGER NOT NULL DEFAULT 0")
        if "buyback_completed_at" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN buyback_completed_at TEXT")
        if "buyback_price" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN buyback_price REAL")
        if "buyback_note" not in rcols:
            conn.execute("ALTER TABLE rebalance_lot_items ADD COLUMN buyback_note TEXT")
        conn.execute(
            """
            UPDATE rebalance_lot_items
            SET execution_state = CASE WHEN COALESCE(completed,0)=1 THEN 'closed' ELSE 'pending' END
            WHERE execution_state IS NULL OR TRIM(execution_state) = ''
            """
        )
        conn.execute(
            """
            UPDATE rebalance_lot_items
            SET execution_state = 'pending'
            WHERE LOWER(COALESCE(execution_state,'')) NOT IN ('pending','closed','skipped')
            """
        )
        conn.execute("UPDATE rebalance_lot_items SET buyback_completed = 0 WHERE buyback_completed IS NULL")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_target_plans (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              seed_capital REAL NOT NULL DEFAULT 10000,
              target_profit_pct REAL NOT NULL DEFAULT 1,
              target_profit_value REAL NOT NULL DEFAULT 0,
              top_n INTEGER NOT NULL DEFAULT 5,
              status TEXT NOT NULL CHECK (status IN ('active','completed','reset')) DEFAULT 'active',
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              last_recalibrated_at TEXT,
              closed_at TEXT,
              notes TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_target_plans_status_created ON daily_target_plans(status, created_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_target_plan_pairs (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              plan_id INTEGER NOT NULL REFERENCES daily_target_plans(id) ON DELETE CASCADE,
              priority_rank INTEGER NOT NULL DEFAULT 1,
              state TEXT NOT NULL DEFAULT 'pending',
              sell_symbol TEXT NOT NULL,
              sell_qty REAL NOT NULL DEFAULT 0,
              sell_ref_price REAL NOT NULL DEFAULT 0,
              sell_trade_value REAL NOT NULL DEFAULT 0,
              sell_target_price REAL NOT NULL DEFAULT 0,
              sell_score REAL NOT NULL DEFAULT 0,
              sell_reason TEXT,
              buy_symbol TEXT NOT NULL,
              buy_qty REAL NOT NULL DEFAULT 0,
              buy_ref_price REAL NOT NULL DEFAULT 0,
              buy_trade_value REAL NOT NULL DEFAULT 0,
              buy_target_exit_price REAL NOT NULL DEFAULT 0,
              buy_score REAL NOT NULL DEFAULT 0,
              buy_reason TEXT,
              expected_profit_value REAL NOT NULL DEFAULT 0,
              rotation_score REAL NOT NULL DEFAULT 0,
              current_sell_ref_price REAL NOT NULL DEFAULT 0,
              current_buy_ref_price REAL NOT NULL DEFAULT 0,
              target_progress_pct REAL NOT NULL DEFAULT 0,
              matched_sell_trade_id INTEGER,
              matched_buy_trade_id INTEGER,
              reconciliation_status TEXT NOT NULL DEFAULT 'unmatched',
              executed_sell_price REAL,
              executed_sell_at TEXT,
              executed_buy_price REAL,
              executed_buy_at TEXT,
              completion_note TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              last_recalibrated_at TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_target_pairs_plan_rank ON daily_target_plan_pairs(plan_id, priority_rank, id)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_target_pairs_state ON daily_target_plan_pairs(state, updated_at DESC)")
        pcols = [r["name"] for r in conn.execute("PRAGMA table_info(daily_target_plan_pairs)").fetchall()]
        if "state" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN state TEXT NOT NULL DEFAULT 'pending'")
        if "current_sell_ref_price" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN current_sell_ref_price REAL NOT NULL DEFAULT 0")
        if "current_buy_ref_price" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN current_buy_ref_price REAL NOT NULL DEFAULT 0")
        if "target_progress_pct" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN target_progress_pct REAL NOT NULL DEFAULT 0")
        if "matched_sell_trade_id" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN matched_sell_trade_id INTEGER")
        if "matched_buy_trade_id" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN matched_buy_trade_id INTEGER")
        if "reconciliation_status" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN reconciliation_status TEXT NOT NULL DEFAULT 'unmatched'")
        if "executed_sell_price" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN executed_sell_price REAL")
        if "executed_sell_at" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN executed_sell_at TEXT")
        if "executed_buy_price" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN executed_buy_price REAL")
        if "executed_buy_at" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN executed_buy_at TEXT")
        if "completion_note" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN completion_note TEXT")
        if "updated_at" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN updated_at TEXT")
        if "last_recalibrated_at" not in pcols:
            conn.execute("ALTER TABLE daily_target_plan_pairs ADD COLUMN last_recalibrated_at TEXT")
        conn.execute(
            """
            UPDATE daily_target_plan_pairs
            SET state = 'pending'
            WHERE LOWER(COALESCE(state,'')) NOT IN ('pending','sell_done','buy_done','executed','skipped','replaced')
            """
        )
        conn.execute(
            """
            UPDATE daily_target_plan_pairs
            SET reconciliation_status = 'unmatched'
            WHERE LOWER(COALESCE(reconciliation_status,'')) NOT IN ('unmatched','partial','matched')
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_target_pair_snapshots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              plan_id INTEGER NOT NULL REFERENCES daily_target_plans(id) ON DELETE CASCADE,
              pair_id INTEGER NOT NULL REFERENCES daily_target_plan_pairs(id) ON DELETE CASCADE,
              captured_at TEXT NOT NULL,
              sell_ref_price REAL NOT NULL DEFAULT 0,
              buy_ref_price REAL NOT NULL DEFAULT 0,
              expected_profit_value REAL NOT NULL DEFAULT 0,
              rotation_score REAL NOT NULL DEFAULT 0,
              buy_target_exit_price REAL NOT NULL DEFAULT 0,
              snapshot_note TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_target_pair_snapshots_plan_time ON daily_target_pair_snapshots(plan_id, captured_at DESC, id DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_target_positions (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              source_pair_id INTEGER NOT NULL REFERENCES daily_target_plan_pairs(id) ON DELETE CASCADE,
              symbol TEXT NOT NULL,
              qty REAL NOT NULL DEFAULT 0,
              initial_qty REAL NOT NULL DEFAULT 0,
              closed_qty REAL NOT NULL DEFAULT 0,
              entry_price REAL NOT NULL DEFAULT 0,
              entry_value REAL NOT NULL DEFAULT 0,
              realized_profit REAL NOT NULL DEFAULT 0,
              entry_at TEXT NOT NULL,
              exit_pair_id INTEGER,
              exit_price REAL,
              exit_value REAL,
              exit_at TEXT,
              status TEXT NOT NULL DEFAULT 'open',
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_target_positions_symbol_status ON daily_target_positions(symbol, status, entry_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_daily_target_positions_pair ON daily_target_positions(source_pair_id, id)")
        dtp_cols = [r["name"] for r in conn.execute("PRAGMA table_info(daily_target_positions)").fetchall()]
        if "initial_qty" not in dtp_cols:
            conn.execute("ALTER TABLE daily_target_positions ADD COLUMN initial_qty REAL NOT NULL DEFAULT 0")
        if "closed_qty" not in dtp_cols:
            conn.execute("ALTER TABLE daily_target_positions ADD COLUMN closed_qty REAL NOT NULL DEFAULT 0")
        if "realized_profit" not in dtp_cols:
            conn.execute("ALTER TABLE daily_target_positions ADD COLUMN realized_profit REAL NOT NULL DEFAULT 0")
        conn.execute("UPDATE daily_target_positions SET initial_qty = qty WHERE COALESCE(initial_qty,0) <= 0 AND COALESCE(qty,0) > 0")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS price_ticks (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              symbol TEXT NOT NULL,
              ltp REAL NOT NULL,
              change_abs REAL,
              fetched_at TEXT NOT NULL,
              source TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_price_ticks_symbol_time ON price_ticks(symbol, fetched_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS quote_samples (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              symbol TEXT NOT NULL,
              source TEXT NOT NULL,
              ltp REAL NOT NULL,
              change_abs REAL,
              latency_ms REAL,
              accuracy_error_pct REAL,
              fetched_at TEXT NOT NULL,
              selected INTEGER NOT NULL DEFAULT 0,
              consensus_ltp REAL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_quote_samples_symbol_time ON quote_samples(symbol, fetched_at DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_quote_samples_source_time ON quote_samples(source, fetched_at DESC)")
        qcols = [r["name"] for r in conn.execute("PRAGMA table_info(quote_samples)").fetchall()]
        if "latency_ms" not in qcols:
            conn.execute("ALTER TABLE quote_samples ADD COLUMN latency_ms REAL")
        if "accuracy_error_pct" not in qcols:
            conn.execute("ALTER TABLE quote_samples ADD COLUMN accuracy_error_pct REAL")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS quote_source_registry (
              source TEXT PRIMARY KEY,
              adapter TEXT NOT NULL,
              enabled INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              notes TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS quote_source_stats (
              source TEXT PRIMARY KEY,
              attempts INTEGER NOT NULL DEFAULT 0,
              successes INTEGER NOT NULL DEFAULT 0,
              failures INTEGER NOT NULL DEFAULT 0,
              total_latency_ms REAL NOT NULL DEFAULT 0,
              total_accuracy_error_pct REAL NOT NULL DEFAULT 0,
              accuracy_samples INTEGER NOT NULL DEFAULT 0,
              score REAL NOT NULL DEFAULT 0,
              last_success_at TEXT,
              last_error_at TEXT,
              updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_quote_source_stats_score ON quote_source_stats(score DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS strategy_runs (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              run_date TEXT NOT NULL UNIQUE,
              created_at TEXT NOT NULL,
              market_value REAL NOT NULL,
              invested_value REAL NOT NULL,
              cash_balance REAL NOT NULL,
              projected_start_value REAL NOT NULL,
              add_count INTEGER NOT NULL DEFAULT 0,
              trim_count INTEGER NOT NULL DEFAULT 0,
              hold_count INTEGER NOT NULL DEFAULT 0,
              review_count INTEGER NOT NULL DEFAULT 0,
              watch_add_count INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS strategy_recommendations (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              run_date TEXT NOT NULL,
              symbol TEXT NOT NULL,
              action TEXT NOT NULL,
              priority INTEGER NOT NULL DEFAULT 0,
              weight_current REAL NOT NULL DEFAULT 0,
              weight_target REAL NOT NULL DEFAULT 0,
              delta_weight REAL NOT NULL DEFAULT 0,
              confidence REAL NOT NULL DEFAULT 0,
              price_now REAL NOT NULL DEFAULT 0,
              buy_price_1 REAL,
              buy_price_2 REAL,
              sell_price_1 REAL,
              sell_price_2 REAL,
              expected_annual_return REAL NOT NULL DEFAULT 0,
              reason TEXT,
              source TEXT NOT NULL DEFAULT 'rotation_engine'
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_strategy_reco_run ON strategy_recommendations(run_date, priority DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_strategy_reco_symbol ON strategy_recommendations(symbol)")
        reco_cols = [r["name"] for r in conn.execute("PRAGMA table_info(strategy_recommendations)").fetchall()]
        if "intel_score" not in reco_cols:
            conn.execute("ALTER TABLE strategy_recommendations ADD COLUMN intel_score REAL NOT NULL DEFAULT 0")
        if "intel_confidence" not in reco_cols:
            conn.execute("ALTER TABLE strategy_recommendations ADD COLUMN intel_confidence REAL NOT NULL DEFAULT 0")
        if "intel_summary" not in reco_cols:
            conn.execute("ALTER TABLE strategy_recommendations ADD COLUMN intel_summary TEXT")
        sr_cols = [r["name"] for r in conn.execute("PRAGMA table_info(strategy_runs)").fetchall()]
        if "macro_regime" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN macro_regime TEXT NOT NULL DEFAULT 'neutral'")
        if "macro_score" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN macro_score REAL NOT NULL DEFAULT 0")
        if "macro_confidence" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN macro_confidence REAL NOT NULL DEFAULT 0")
        if "macro_thought" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN macro_thought TEXT")
        if "intel_score" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN intel_score REAL NOT NULL DEFAULT 0")
        if "intel_confidence" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN intel_confidence REAL NOT NULL DEFAULT 0")
        if "intel_thought" not in sr_cols:
            conn.execute("ALTER TABLE strategy_runs ADD COLUMN intel_thought TEXT")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS strategy_projection_points (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              run_date TEXT NOT NULL,
              scenario TEXT NOT NULL,
              year_offset INTEGER NOT NULL,
              annual_return REAL NOT NULL,
              projected_value REAL NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_strategy_projection_run ON strategy_projection_points(run_date, scenario, year_offset)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS strategy_audit_runs (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              strategy_run_date TEXT,
              audit_mode TEXT NOT NULL DEFAULT 'heuristic',
              overall_status TEXT NOT NULL DEFAULT 'ok',
              overall_score REAL NOT NULL DEFAULT 0,
              summary TEXT,
              recommendation TEXT,
              findings_count INTEGER NOT NULL DEFAULT 0,
              stats_json TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_strategy_audit_runs_created ON strategy_audit_runs(created_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS strategy_audit_findings (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              audit_id INTEGER NOT NULL REFERENCES strategy_audit_runs(id) ON DELETE CASCADE,
              severity TEXT NOT NULL,
              code TEXT NOT NULL,
              title TEXT NOT NULL,
              detail TEXT,
              symbol TEXT,
              metric_value REAL,
              expected_range TEXT,
              created_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_strategy_audit_findings_audit ON strategy_audit_findings(audit_id, severity, id)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS tax_rate_sync_runs (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              status TEXT NOT NULL CHECK (status IN ('success','error')),
              source_label TEXT NOT NULL,
              source_url TEXT,
              stcg_rate_pct REAL,
              ltcg_rate_pct REAL,
              ltcg_exemption_limit REAL,
              stt_delivery_rate REAL,
              stamp_buy_rate REAL,
              gst_rate REAL,
              dp_charge_sell REAL,
              detail TEXT,
              error TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_tax_rate_sync_runs_created ON tax_rate_sync_runs(created_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS attention_alerts (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              code TEXT NOT NULL UNIQUE,
              category TEXT NOT NULL,
              severity_rank INTEGER NOT NULL DEFAULT 0,
              severity_label TEXT NOT NULL DEFAULT 'info',
              status TEXT NOT NULL DEFAULT 'open' CHECK (status IN ('open','resolved')),
              title TEXT NOT NULL,
              detail TEXT,
              source_ref TEXT,
              detected_at TEXT NOT NULL,
              last_seen_at TEXT NOT NULL,
              resolved_at TEXT,
              meta_json TEXT,
              occurrence_count INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_attention_alerts_status_rank ON attention_alerts(status, severity_rank DESC, last_seen_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS agent_backtest_runs (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              from_date TEXT NOT NULL,
              to_date TEXT NOT NULL,
              horizon_days INTEGER NOT NULL,
              sample_count INTEGER NOT NULL DEFAULT 0,
              hit_rate REAL NOT NULL DEFAULT 0,
              avg_future_return REAL NOT NULL DEFAULT 0,
              momentum_hit_rate REAL NOT NULL DEFAULT 0,
              intel_hit_rate REAL NOT NULL DEFAULT 0,
              applied_tuning INTEGER NOT NULL DEFAULT 0,
              params_before_json TEXT,
              params_after_json TEXT,
              suggestions_json TEXT,
              diagnostics_json TEXT,
              errors_json TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_agent_backtest_runs_created ON agent_backtest_runs(created_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS intelligence_documents (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              doc_type TEXT NOT NULL,
              source TEXT,
              source_ref TEXT,
              doc_date TEXT NOT NULL,
              title TEXT,
              content TEXT NOT NULL,
              sentiment_score REAL NOT NULL DEFAULT 0,
              created_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_intel_docs_date ON intelligence_documents(doc_date DESC, id DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_intel_docs_type ON intelligence_documents(doc_type, doc_date DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS intelligence_impacts (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              doc_id INTEGER NOT NULL REFERENCES intelligence_documents(id) ON DELETE CASCADE,
              symbol TEXT NOT NULL,
              impact_score REAL NOT NULL,
              confidence REAL NOT NULL DEFAULT 0.5,
              reason TEXT,
              created_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_intel_impacts_symbol_time ON intelligence_impacts(symbol, created_at DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_intel_impacts_doc ON intelligence_impacts(doc_id)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS company_financials (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              symbol TEXT NOT NULL,
              fiscal_period TEXT NOT NULL,
              report_date TEXT NOT NULL,
              revenue REAL,
              pat REAL,
              operating_cash_flow REAL,
              investing_cash_flow REAL,
              financing_cash_flow REAL,
              debt REAL,
              fii_holding_pct REAL,
              dii_holding_pct REAL,
              promoter_holding_pct REAL,
              source TEXT NOT NULL DEFAULT 'manual',
              notes TEXT,
              created_at TEXT NOT NULL,
              UNIQUE(symbol, fiscal_period, source)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_company_fin_symbol_date ON company_financials(symbol, report_date DESC, id DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS agent_approvals (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              status TEXT NOT NULL CHECK (status IN ('pending','approved','rejected','executed','expired')),
              action_type TEXT NOT NULL,
              query_text TEXT,
              payload_json TEXT NOT NULL,
              summary TEXT,
              decided_at TEXT,
              executed_at TEXT,
              decision_note TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_agent_approvals_status_created ON agent_approvals(status, created_at DESC)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS peak_split_reviews (
              corporate_action_id INTEGER PRIMARY KEY,
              decision TEXT NOT NULL CHECK (decision IN ('apply','ignore')),
              decided_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS ix_peak_split_reviews_decision ON peak_split_reviews(decision)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS chart_analysis_snapshots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              symbol TEXT NOT NULL,
              as_of_date TEXT NOT NULL,
              score REAL NOT NULL DEFAULT 0,
              confidence REAL NOT NULL DEFAULT 0,
              signal TEXT NOT NULL DEFAULT 'NEUTRAL',
              trend_score REAL NOT NULL DEFAULT 0,
              momentum_score REAL NOT NULL DEFAULT 0,
              mean_reversion_score REAL NOT NULL DEFAULT 0,
              breakout_score REAL NOT NULL DEFAULT 0,
              relative_strength_score REAL NOT NULL DEFAULT 0,
              index_correlation REAL NOT NULL DEFAULT 0,
              source_summary TEXT,
              pattern_flags_json TEXT,
              details_json TEXT,
              created_at TEXT NOT NULL,
              UNIQUE(symbol, as_of_date)
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS ix_chart_snapshots_symbol_time ON chart_analysis_snapshots(symbol, created_at DESC)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS ix_chart_snapshots_date ON chart_analysis_snapshots(as_of_date DESC, created_at DESC)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS software_perf_snapshots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              live_stale_symbols INTEGER NOT NULL DEFAULT 0,
              live_zero_ltp_symbols INTEGER NOT NULL DEFAULT 0,
              live_missing_price_symbols INTEGER NOT NULL DEFAULT 0,
              weak_sources_count INTEGER NOT NULL DEFAULT 0,
              avg_quote_latency_ms REAL NOT NULL DEFAULT 0,
              quote_success_rate REAL NOT NULL DEFAULT 0,
              last_price_age_sec REAL NOT NULL DEFAULT 0,
              history_coverage_ratio REAL NOT NULL DEFAULT 0,
              issue_count INTEGER NOT NULL DEFAULT 0,
              notes_json TEXT
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS ix_software_perf_snapshots_created ON software_perf_snapshots(created_at DESC)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS software_perf_actions (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              action_type TEXT NOT NULL,
              status TEXT NOT NULL,
              summary TEXT,
              details_json TEXT
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS ix_software_perf_actions_created ON software_perf_actions(created_at DESC)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_analysis_snapshots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              created_at TEXT NOT NULL,
              lookback_days INTEGER NOT NULL DEFAULT 252,
              winsorize_pct REAL NOT NULL DEFAULT 0.05,
              symbols_in_portfolio INTEGER NOT NULL DEFAULT 0,
              symbols_with_history INTEGER NOT NULL DEFAULT 0,
              symbols_analyzed INTEGER NOT NULL DEFAULT 0,
              observation_count INTEGER NOT NULL DEFAULT 0,
              portfolio_volatility REAL NOT NULL DEFAULT 0,
              downside_volatility REAL NOT NULL DEFAULT 0,
              max_drawdown REAL NOT NULL DEFAULT 0,
              var_95 REAL NOT NULL DEFAULT 0,
              cvar_95 REAL NOT NULL DEFAULT 0,
              avg_pair_correlation REAL NOT NULL DEFAULT 0,
              pair_count INTEGER NOT NULL DEFAULT 0,
              concentration_hhi REAL NOT NULL DEFAULT 0,
              risk_score REAL NOT NULL DEFAULT 0,
              risk_level TEXT NOT NULL DEFAULT 'low',
              notes_json TEXT
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS ix_risk_analysis_snapshots_created ON risk_analysis_snapshots(created_at DESC)"
        )
        conn.commit()


def ensure_default_strategy():
    default_params = {
        "buy_l1_discount": 0.032,
        "buy_l2_discount": 0.062,
        "sell_s1_markup": 0.12,
        "sell_s2_markup": 0.50,
        "sell_s3_markup": 1.25,
        "mux_factor_default": 1.0,
        "brokerage_rate": 0.0027,
        "allocation_limit": 0.25,
        "exposure_cap": 0.07,
        "trim_trigger_overweight": 0.02,
        "add_trigger_underweight": 0.02,
        "max_position_weight": 0.12,
        "max_new_ideas": 2,
        "momentum_lookback_days": 30,
        "projection_years": 5,
        "projection_base_return": 0.12,
        "projection_conservative_delta": 0.04,
        "projection_aggressive_delta": 0.04,
        "strategy_confidence_floor": 0.45,
        "strategy_confidence_ceiling": 0.92,
        "intel_weight_commentary": 0.25,
        "intel_weight_policy": 0.25,
        "intel_weight_financials": 0.50,
        "intel_weight_chart": 0.20,
        "intel_decay_days": 45,
    }
    with db_connect() as conn:
        row = conn.execute(
            "SELECT id FROM strategy_sets WHERE name = 'Default Workbook Logic'"
        ).fetchone()
        if row is None:
            conn.execute(
                "INSERT INTO strategy_sets(name, is_active, created_at) VALUES (?, 1, ?)",
                ("Default Workbook Logic", now_iso()),
            )
            set_id = conn.execute(
                "SELECT id FROM strategy_sets WHERE name = 'Default Workbook Logic'"
            ).fetchone()["id"]
            for key, value in default_params.items():
                conn.execute(
                    "INSERT INTO strategy_parameters(set_id, key, value) VALUES (?, ?, ?)",
                    (set_id, key, value),
                )
            conn.commit()
        else:
            set_id = row["id"]
            existing = {
                r["key"]: r["value"]
                for r in conn.execute(
                    "SELECT key, value FROM strategy_parameters WHERE set_id = ?",
                    (set_id,),
                ).fetchall()
            }
            for key, value in default_params.items():
                if key not in existing:
                    conn.execute(
                        "INSERT INTO strategy_parameters(set_id, key, value) VALUES (?, ?, ?)",
                        (set_id, key, value),
                    )
            active_exists = conn.execute(
                "SELECT COUNT(*) AS c FROM strategy_sets WHERE is_active = 1"
            ).fetchone()["c"]
            if active_exists == 0:
                conn.execute("UPDATE strategy_sets SET is_active = 0")
                conn.execute("UPDATE strategy_sets SET is_active = 1 WHERE id = ?", (set_id,))
            conn.commit()


def ensure_default_config():
    defaults = {
        "live_refresh_enabled": "1",
        "live_refresh_interval_sec": "10",
        "live_refresh_cursor": "0",
        "live_quote_sources": DEFAULT_LIVE_QUOTE_SOURCES,
        "live_quote_max_deviation_pct": str(LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT),
        "live_quote_top_k": str(LIVE_QUOTE_TOP_K_DEFAULT),
        "live_quote_explore_ratio": str(LIVE_QUOTE_EXPLORE_RATIO_DEFAULT),
        "strategy_refresh_enabled": "1",
        "strategy_refresh_interval_sec": str(STRATEGY_REFRESH_MIN_SEC),
        "history_refresh_enabled": "1",
        "history_refresh_interval_sec": str(HISTORY_REFRESH_DEFAULT_SEC),
        "history_last_sync_at": "",
        "backup_agent_enabled": "1",
        "backup_last_run_at": "",
        "backup_last_file": "",
        "repo_sync_enabled": "1",
        "repo_sync_interval_sec": str(REPO_SYNC_INTERVAL_DEFAULT_SEC),
        "repo_sync_auto_push": "1",
        "repo_sync_last_run_at": "",
        "repo_sync_last_error": "",
        "self_learning_enabled": "1",
        "self_learning_interval_days": "7",
        "self_learning_last_run_at": "",
        "self_learning_min_samples": "30",
        "intel_autopilot_enabled": "1",
        "intel_autopilot_interval_sec": str(12 * 60 * 60),
        "intel_autopilot_last_run_at": "",
        "intel_autopilot_max_docs": "24",
        "intel_autopilot_symbols_limit": "20",
        "intel_financial_backfill_last_run_at": "",
        "intel_autopilot_sources": "google_news_rss,screener_financials,nse_announcements,company_site_ir",
        "intel_autopilot_query_seed": (
            "India union budget policy stocks\n"
            "RBI policy impact equities India\n"
            "SEBI circular equity market impact\n"
            "India sector rotation stocks"
        ),
        "chart_agent_enabled": "1",
        "chart_agent_interval_sec": str(CHART_AGENT_INTERVAL_DEFAULT_SEC),
        "chart_agent_last_run_at": "",
        "chart_agent_sources": DEFAULT_CHART_AGENT_SOURCES,
        "software_perf_agent_enabled": "1",
        "software_perf_agent_interval_sec": str(SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC),
        "software_perf_agent_last_run_at": "",
        "software_perf_agent_last_heal_at": "",
        "software_perf_agent_last_improvement_at": "",
        "software_perf_agent_last_cleanup_at": "",
        "software_perf_agent_auto_tune": "1",
        "software_perf_agent_write_changes": "1",
        "software_perf_agent_retention_days": "90",
        "software_perf_agent_core_objective": "Preserve portfolio data integrity and strategy objective.",
        "tax_monitor_enabled": "1",
        "tax_monitor_interval_sec": str(TAX_MONITOR_INTERVAL_DEFAULT_SEC),
        "tax_monitor_last_run_at": "",
        "tax_monitor_last_success_at": "",
        "tax_monitor_last_error": "",
        "tax_monitor_last_change_at": "",
        "tax_monitor_tax_source_url": DEFAULT_TAX_MONITOR_TAX_URL,
        "tax_monitor_charges_source_url": DEFAULT_TAX_MONITOR_CHARGES_URL,
        "tax_rate_effective_from": "2024-07-23",
        "tax_rate_stcg_pct": "20",
        "tax_rate_ltcg_pct": "12.5",
        "tax_rate_ltcg_exemption_limit": "125000",
        "zerodha_eq_delivery_txn_rate_nse": "0.0000307",
        "zerodha_eq_delivery_txn_rate_bse": "0.0000375",
        "zerodha_sebi_rate": "0.000001",
        "zerodha_stt_delivery_rate": "0.001",
        "zerodha_stamp_buy_rate": "0.00015",
        "zerodha_gst_rate": "0.18",
        "zerodha_dp_charge_sell_incl_gst": "15.34",
        "llm_api_key": "",
        "llm_model": LLM_DEFAULT_MODEL,
        "llm_api_url": LLM_DEFAULT_API_URL,
        "llm_last_status": "not_configured",
        "llm_last_error": "",
        "llm_last_checked_at": "",
        "risk_agent_enabled": "1",
        "risk_agent_interval_sec": str(RISK_AGENT_INTERVAL_DEFAULT_SEC),
        "risk_agent_last_run_at": "",
        "risk_agent_lookback_days": "252",
        "risk_agent_winsorize_pct": "0.05",
    }
    with db_connect() as conn:
        for key, value in defaults.items():
            row = conn.execute("SELECT key FROM app_config WHERE key = ?", (key,)).fetchone()
            if row is None:
                conn.execute(
                    "INSERT INTO app_config(key, value, updated_at) VALUES (?, ?, ?)",
                    (key, value, now_iso()),
                )
        ensure_quote_source_registry(conn, parse_source_list(DEFAULT_LIVE_QUOTE_SOURCES))
        conn.commit()


def _app_config_get_many(conn, keys):
    key_list = [str(k) for k in (keys or []) if str(k)]
    if not key_list:
        return {}
    placeholders = ",".join(["?"] * len(key_list))
    rows = conn.execute(
        f"SELECT key, value FROM app_config WHERE key IN ({placeholders})",
        key_list,
    ).fetchall()
    return {str(r["key"]): str(r["value"] or "") for r in rows}


def _app_config_upsert_many(conn, mapping):
    stamp = now_iso()
    for key, value in dict(mapping or {}).items():
        conn.execute(
            """
            INSERT INTO app_config(key, value, updated_at) VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
            """,
            (str(key), str(value), stamp),
        )


def get_tax_profile_config(conn):
    cfg = _app_config_get_many(
        conn,
        [
            "tax_rate_effective_from",
            "tax_rate_stcg_pct",
            "tax_rate_ltcg_pct",
            "tax_rate_ltcg_exemption_limit",
            "zerodha_eq_delivery_txn_rate_nse",
            "zerodha_eq_delivery_txn_rate_bse",
            "zerodha_sebi_rate",
            "zerodha_stt_delivery_rate",
            "zerodha_stamp_buy_rate",
            "zerodha_gst_rate",
            "zerodha_dp_charge_sell_incl_gst",
        ],
    )
    return {
        "effective_from": str(cfg.get("tax_rate_effective_from") or "2024-07-23"),
        "stcg_rate_pct": round(parse_float(cfg.get("tax_rate_stcg_pct"), 20.0), 4),
        "ltcg_rate_pct": round(parse_float(cfg.get("tax_rate_ltcg_pct"), 12.5), 4),
        "ltcg_exemption_limit": round(parse_float(cfg.get("tax_rate_ltcg_exemption_limit"), 125000.0), 2),
        "txn_rate_nse": parse_float(cfg.get("zerodha_eq_delivery_txn_rate_nse"), 0.0000307),
        "txn_rate_bse": parse_float(cfg.get("zerodha_eq_delivery_txn_rate_bse"), 0.0000375),
        "sebi_rate": parse_float(cfg.get("zerodha_sebi_rate"), 0.000001),
        "stt_delivery_rate": parse_float(cfg.get("zerodha_stt_delivery_rate"), 0.001),
        "stamp_buy_rate": parse_float(cfg.get("zerodha_stamp_buy_rate"), 0.00015),
        "gst_rate": parse_float(cfg.get("zerodha_gst_rate"), 0.18),
        "dp_charge_sell_incl_gst": round(parse_float(cfg.get("zerodha_dp_charge_sell_incl_gst"), 15.34), 2),
    }


def set_tax_profile_config(
    conn,
    stcg_rate_pct=None,
    ltcg_rate_pct=None,
    ltcg_exemption_limit=None,
    txn_rate_nse=None,
    txn_rate_bse=None,
    sebi_rate=None,
    stt_delivery_rate=None,
    stamp_buy_rate=None,
    gst_rate=None,
    dp_charge_sell_incl_gst=None,
    effective_from=None,
):
    updates = {}
    if stcg_rate_pct is not None:
        updates["tax_rate_stcg_pct"] = round(parse_float(stcg_rate_pct, 20.0), 4)
    if ltcg_rate_pct is not None:
        updates["tax_rate_ltcg_pct"] = round(parse_float(ltcg_rate_pct, 12.5), 4)
    if ltcg_exemption_limit is not None:
        updates["tax_rate_ltcg_exemption_limit"] = round(parse_float(ltcg_exemption_limit, 125000.0), 2)
    if txn_rate_nse is not None:
        updates["zerodha_eq_delivery_txn_rate_nse"] = parse_float(txn_rate_nse, 0.0000307)
    if txn_rate_bse is not None:
        updates["zerodha_eq_delivery_txn_rate_bse"] = parse_float(txn_rate_bse, 0.0000375)
    if sebi_rate is not None:
        updates["zerodha_sebi_rate"] = parse_float(sebi_rate, 0.000001)
    if stt_delivery_rate is not None:
        updates["zerodha_stt_delivery_rate"] = parse_float(stt_delivery_rate, 0.001)
    if stamp_buy_rate is not None:
        updates["zerodha_stamp_buy_rate"] = parse_float(stamp_buy_rate, 0.00015)
    if gst_rate is not None:
        updates["zerodha_gst_rate"] = parse_float(gst_rate, 0.18)
    if dp_charge_sell_incl_gst is not None:
        updates["zerodha_dp_charge_sell_incl_gst"] = round(parse_float(dp_charge_sell_incl_gst, 15.34), 2)
    if effective_from is not None:
        updates["tax_rate_effective_from"] = str(effective_from or "").strip() or "2024-07-23"
    if updates:
        _app_config_upsert_many(conn, updates)


def get_tax_monitor_config(conn):
    cfg = _app_config_get_many(
        conn,
        [
            "tax_monitor_enabled",
            "tax_monitor_interval_sec",
            "tax_monitor_last_run_at",
            "tax_monitor_last_success_at",
            "tax_monitor_last_error",
            "tax_monitor_last_change_at",
            "tax_monitor_tax_source_url",
            "tax_monitor_charges_source_url",
        ],
    )
    try:
        interval_sec = int(float(cfg.get("tax_monitor_interval_sec", str(TAX_MONITOR_INTERVAL_DEFAULT_SEC))))
    except Exception:
        interval_sec = TAX_MONITOR_INTERVAL_DEFAULT_SEC
    return {
        "enabled": str(cfg.get("tax_monitor_enabled", "1")) == "1",
        "interval_seconds": max(TAX_MONITOR_MIN_INTERVAL_SEC, interval_sec),
        "last_run_at": str(cfg.get("tax_monitor_last_run_at", "") or ""),
        "last_success_at": str(cfg.get("tax_monitor_last_success_at", "") or ""),
        "last_error": str(cfg.get("tax_monitor_last_error", "") or ""),
        "last_change_at": str(cfg.get("tax_monitor_last_change_at", "") or ""),
        "tax_source_url": str(cfg.get("tax_monitor_tax_source_url") or DEFAULT_TAX_MONITOR_TAX_URL),
        "charges_source_url": str(cfg.get("tax_monitor_charges_source_url") or DEFAULT_TAX_MONITOR_CHARGES_URL),
    }


def set_tax_monitor_config(
    conn,
    enabled=None,
    interval_seconds=None,
    last_run_at=None,
    last_success_at=None,
    last_error=None,
    last_change_at=None,
    tax_source_url=None,
    charges_source_url=None,
):
    updates = {}
    if enabled is not None:
        updates["tax_monitor_enabled"] = "1" if bool(enabled) else "0"
    if interval_seconds is not None:
        updates["tax_monitor_interval_sec"] = max(TAX_MONITOR_MIN_INTERVAL_SEC, int(interval_seconds))
    if last_run_at is not None:
        updates["tax_monitor_last_run_at"] = str(last_run_at or "")
    if last_success_at is not None:
        updates["tax_monitor_last_success_at"] = str(last_success_at or "")
    if last_error is not None:
        updates["tax_monitor_last_error"] = str(last_error or "")
    if last_change_at is not None:
        updates["tax_monitor_last_change_at"] = str(last_change_at or "")
    if tax_source_url is not None:
        updates["tax_monitor_tax_source_url"] = str(tax_source_url or "").strip() or DEFAULT_TAX_MONITOR_TAX_URL
    if charges_source_url is not None:
        updates["tax_monitor_charges_source_url"] = str(charges_source_url or "").strip() or DEFAULT_TAX_MONITOR_CHARGES_URL
    if updates:
        _app_config_upsert_many(conn, updates)


def backup_database():
    p = tenant_paths(get_current_tenant_key())
    p["backup_dir"].mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    target = p["backup_dir"] / f"portfolio_{stamp}.db"
    src = sqlite3.connect(p["db_path"])
    dst = sqlite3.connect(target)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()
    with db_connect() as conn:
        run_at = now_iso()
        conn.execute(
            """
            INSERT INTO app_config(key, value, updated_at) VALUES ('backup_last_run_at', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
            """,
            (run_at, run_at),
        )
        conn.execute(
            """
            INSERT INTO app_config(key, value, updated_at) VALUES ('backup_last_file', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
            """,
            (str(target), run_at),
        )
        conn.commit()
    return str(target)


def find_git_executable():
    for p in GIT_EXE_CANDIDATES:
        try:
            if p.exists():
                return str(p)
        except Exception:
            continue
    return None


def git_run(args, cwd=None, check=True):
    git_exe = find_git_executable()
    if not git_exe:
        raise RuntimeError("git_executable_not_found")
    proc = subprocess.run(
        [git_exe] + list(args),
        cwd=str(cwd or ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if check and proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or f"git_failed_{proc.returncode}").strip())
    return proc


def _sql_dump_for_db(db_path):
    dbp = Path(db_path)
    if not dbp.exists():
        return ""
    conn = sqlite3.connect(str(dbp))
    try:
        lines = list(conn.iterdump())
    finally:
        conn.close()
    body = "\n".join(lines).strip()
    if body:
        body += "\n"
    header = (
        f"-- Repository snapshot generated at {now_iso()}\n"
        f"-- Source database: {dbp.name}\n"
    )
    return header + body


def export_repo_data_snapshots():
    ensure_tenant_bootstrap()
    REPO_DATA_DIR.mkdir(parents=True, exist_ok=True)
    tenants_dir = REPO_DATA_DIR / "tenants"
    tenants_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        "generated_at": now_iso(),
        "tenants": [],
    }
    for tenant in list_tenants():
        key = sanitize_tenant_key(tenant.get("key")) or DEFAULT_TENANT_KEY
        paths = tenant_paths(key)
        out_dir = tenants_dir / key
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "portfolio.sql").write_text(_sql_dump_for_db(paths["db_path"]), encoding="utf-8")
        (out_dir / "market_history.sql").write_text(_sql_dump_for_db(paths["market_db_path"]), encoding="utf-8")
        manifest["tenants"].append(
            {
                "key": key,
                "name": tenant.get("name"),
                "portfolio_sql": str((out_dir / "portfolio.sql").relative_to(ROOT)),
                "market_history_sql": str((out_dir / "market_history.sql").relative_to(ROOT)),
            }
        )
    repo_readme = (
        "# Repo Data Snapshot\n\n"
        "This folder stores Git-tracked SQL snapshots of all tenant SQLite databases.\n"
        "The application updates these files so repository backups stay current without committing live .db binaries.\n"
    )
    (REPO_DATA_DIR / "README.md").write_text(repo_readme, encoding="utf-8")
    (REPO_DATA_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest


def sync_repo_data_snapshots_to_git():
    manifest = export_repo_data_snapshots()
    status = git_run(["status", "--porcelain", "--", "repo_data"], check=True)
    if not str(status.stdout or "").strip():
        return {"ok": True, "changed": False, "manifest": manifest}
    git_run(["add", "repo_data"], check=True)
    commit_msg = f"Sync repo data snapshot {now_iso()}"
    git_run(["commit", "-m", commit_msg], check=True)
    git_run(["push", "origin", "main"], check=True)
    return {"ok": True, "changed": True, "manifest": manifest, "commit_message": commit_msg}


def get_live_config(conn):
    rows = conn.execute(
        "SELECT key, value FROM app_config WHERE key IN ('live_refresh_enabled','live_refresh_interval_sec','live_quote_sources','live_quote_max_deviation_pct','live_quote_top_k','live_quote_explore_ratio')"
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("live_refresh_enabled", "1")) == "1"
    try:
        interval = int(float(cfg.get("live_refresh_interval_sec", "10")))
    except ValueError:
        interval = 10
    interval = max(LIVE_REFRESH_MIN_SEC, interval)
    sources = parse_source_list(cfg.get("live_quote_sources"), DEFAULT_LIVE_QUOTE_SOURCES)
    try:
        max_dev = float(cfg.get("live_quote_max_deviation_pct", LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT))
    except (ValueError, TypeError):
        max_dev = LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT
    max_dev = clamp(max_dev, 1.0, 20.0)
    try:
        top_k = int(float(cfg.get("live_quote_top_k", LIVE_QUOTE_TOP_K_DEFAULT)))
    except (ValueError, TypeError):
        top_k = LIVE_QUOTE_TOP_K_DEFAULT
    top_k = max(1, min(8, top_k))
    try:
        explore_ratio = float(cfg.get("live_quote_explore_ratio", LIVE_QUOTE_EXPLORE_RATIO_DEFAULT))
    except (ValueError, TypeError):
        explore_ratio = LIVE_QUOTE_EXPLORE_RATIO_DEFAULT
    explore_ratio = clamp(explore_ratio, 0.0, 0.8)
    return {
        "enabled": enabled,
        "interval_seconds": interval,
        "quote_sources": sources,
        "quote_max_deviation_pct": round(max_dev, 2),
        "quote_top_k": top_k,
        "quote_explore_ratio": round(explore_ratio, 3),
    }


def get_live_quote_policy(conn):
    cfg = get_live_config(conn)
    return {
        "sources": cfg.get("quote_sources", parse_source_list(DEFAULT_LIVE_QUOTE_SOURCES)),
        "max_deviation_pct": clamp(
            parse_float(cfg.get("quote_max_deviation_pct"), LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT),
            1.0,
            20.0,
        ),
        "top_k": max(1, min(8, int(parse_float(cfg.get("quote_top_k"), LIVE_QUOTE_TOP_K_DEFAULT)))),
        "explore_ratio": clamp(parse_float(cfg.get("quote_explore_ratio"), LIVE_QUOTE_EXPLORE_RATIO_DEFAULT), 0.0, 0.8),
    }


def discovered_quote_sources():
    sources = [s for s in BUILTIN_LIVE_QUOTE_SOURCES]
    try:
        method_names = [m for m in dir(MarketDataClient) if m.startswith("fetch_") and m.endswith("_quote")]
    except Exception:
        method_names = []
    aliases = {
        "nse": "nse_api",
        "bse": "bse_api",
        "nsetools": "nsetools_api",
        "stock_nse_india": "stock_nse_india_api",
        "yahoo": "yahoo_finance",
        "google": "google_scrape",
        "screener": "screener_scrape",
        "trendlyne": "trendlyne_scrape",
        "cnbc": "cnbc_scrape",
    }
    for m in method_names:
        core = str(m)[6:-6]
        if core in ("", "quote", "multi_source", "source"):
            continue
        src = aliases.get(core)
        if not src:
            src = core
            if not src.endswith(("_api", "_finance", "_scrape")):
                src = f"{src}_scrape"
        if re.match(r"^[a-z0-9_:-]{2,64}$", src) and src not in sources:
            sources.append(src)
    return sources


def ensure_quote_source_registry(conn, candidate_sources=None):
    now = now_iso()
    catalog = []
    for s in discovered_quote_sources():
        if s not in catalog:
            catalog.append(s)
    for s in parse_source_list(",".join(candidate_sources or []), DEFAULT_LIVE_QUOTE_SOURCES):
        if s not in catalog:
            catalog.append(s)
    for source in catalog:
        conn.execute(
            """
            INSERT INTO quote_source_registry(source, adapter, enabled, created_at, updated_at, notes)
            VALUES (?, ?, 1, ?, ?, NULL)
            ON CONFLICT(source) DO UPDATE SET updated_at=excluded.updated_at
            """,
            (source, source, now, now),
        )


def quote_source_score_from_totals(attempts, successes, total_latency_ms, total_accuracy_error_pct, accuracy_samples):
    attempts_n = max(0, int(attempts))
    succ_n = max(0, int(successes))
    if attempts_n <= 0 or succ_n <= 0:
        return 0.0
    success_rate = (succ_n / attempts_n)
    avg_latency = (parse_float(total_latency_ms, 0.0) / succ_n) if succ_n > 0 else 2000.0
    latency_score = 1.0 / (1.0 + (max(1.0, avg_latency) / 1200.0))
    if int(accuracy_samples or 0) > 0:
        avg_err = parse_float(total_accuracy_error_pct, 0.0) / max(1, int(accuracy_samples))
        accuracy_score = 1.0 / (1.0 + (max(0.0, avg_err) / 3.0))
    else:
        accuracy_score = 0.5
    quality = (0.55 * accuracy_score) + (0.45 * latency_score)
    score = success_rate * quality
    return round(score * 100.0, 4)


def recompute_quote_source_scores(conn, sources=None):
    params = []
    where = ""
    if sources:
        uniq = sorted(set(str(s).strip().lower() for s in sources if str(s).strip()))
        if uniq:
            ph = ",".join("?" for _ in uniq)
            where = f"WHERE source IN ({ph})"
            params = uniq
    rows = conn.execute(
        f"""
        SELECT source, attempts, successes, total_latency_ms, total_accuracy_error_pct, accuracy_samples
        FROM quote_source_stats
        {where}
        """,
        params,
    ).fetchall()
    for r in rows:
        score = quote_source_score_from_totals(
            r["attempts"],
            r["successes"],
            r["total_latency_ms"],
            r["total_accuracy_error_pct"],
            r["accuracy_samples"],
        )
        conn.execute(
            "UPDATE quote_source_stats SET score = ?, updated_at = ? WHERE source = ?",
            (score, now_iso(), r["source"]),
        )


def apply_quote_source_metrics(conn, metric_events):
    if not metric_events:
        return
    ensure_quote_source_registry(conn, [e.get("source") for e in metric_events])
    by_source = {}
    for ev in metric_events:
        src = str(ev.get("source") or "").strip().lower()
        if not src:
            continue
        item = by_source.setdefault(
            src,
            {
                "attempts": 0,
                "successes": 0,
                "failures": 0,
                "total_latency_ms": 0.0,
                "total_accuracy_error_pct": 0.0,
                "accuracy_samples": 0,
                "last_success_at": None,
                "last_error_at": None,
            },
        )
        ok = bool(ev.get("success"))
        ts = str(ev.get("fetched_at") or now_iso())
        item["attempts"] += 1
        if ok:
            item["successes"] += 1
            item["last_success_at"] = ts
        else:
            item["failures"] += 1
            item["last_error_at"] = ts
        item["total_latency_ms"] += max(0.0, parse_float(ev.get("latency_ms"), 0.0))
        err = ev.get("accuracy_error_pct")
        if err is not None:
            item["total_accuracy_error_pct"] += max(0.0, parse_float(err, 0.0))
            item["accuracy_samples"] += 1

    for source, m in by_source.items():
        conn.execute(
            """
            INSERT INTO quote_source_stats(
              source, attempts, successes, failures, total_latency_ms,
              total_accuracy_error_pct, accuracy_samples, score, last_success_at, last_error_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?)
            ON CONFLICT(source) DO UPDATE SET
              attempts = quote_source_stats.attempts + excluded.attempts,
              successes = quote_source_stats.successes + excluded.successes,
              failures = quote_source_stats.failures + excluded.failures,
              total_latency_ms = quote_source_stats.total_latency_ms + excluded.total_latency_ms,
              total_accuracy_error_pct = quote_source_stats.total_accuracy_error_pct + excluded.total_accuracy_error_pct,
              accuracy_samples = quote_source_stats.accuracy_samples + excluded.accuracy_samples,
              last_success_at = COALESCE(excluded.last_success_at, quote_source_stats.last_success_at),
              last_error_at = COALESCE(excluded.last_error_at, quote_source_stats.last_error_at),
              updated_at = excluded.updated_at
            """,
            (
                source,
                int(m["attempts"]),
                int(m["successes"]),
                int(m["failures"]),
                round(parse_float(m["total_latency_ms"], 0.0), 4),
                round(parse_float(m["total_accuracy_error_pct"], 0.0), 6),
                int(m["accuracy_samples"]),
                m["last_success_at"],
                m["last_error_at"],
                now_iso(),
            ),
        )
    recompute_quote_source_scores(conn, list(by_source.keys()))


def get_ranked_quote_sources(conn, policy, exchange=None):
    configured = parse_source_list(",".join(policy.get("sources") or []), DEFAULT_LIVE_QUOTE_SOURCES)
    ex = str(exchange or "").upper()
    if ex in ("NSE", "BSE"):
        configured = [s for s in configured if str(s or "").strip().lower() != "gold_rate_scrape"]
    ensure_quote_source_registry(conn, configured)
    rows = conn.execute(
        """
        SELECT
          r.source,
          r.adapter,
          r.enabled,
          COALESCE(s.score, 0) AS score,
          COALESCE(s.attempts, 0) AS attempts,
          COALESCE(s.successes, 0) AS successes,
          COALESCE(s.failures, 0) AS failures,
          s.last_error_at
        FROM quote_source_registry r
        LEFT JOIN quote_source_stats s ON s.source = r.source
        WHERE r.enabled = 1
        ORDER BY score DESC, attempts DESC, r.source
        """
    ).fetchall()
    if ex in ("NSE", "BSE"):
        rows = [r for r in rows if str(r["source"] or "").strip().lower() != "gold_rate_scrape"]
    now_dt = dt.datetime.now()

    def _is_hard_failed(row):
        attempts = int(row["attempts"] or 0)
        successes = int(row["successes"] or 0)
        if attempts < 15 or successes > 0:
            return False
        last_error_at = str(row["last_error_at"] or "").strip()
        if not last_error_at:
            return True
        try:
            err_dt = dt.datetime.fromisoformat(last_error_at[:19])
        except Exception:
            return True
        return (now_dt - err_dt).total_seconds() < QUOTE_SOURCE_HARD_FAIL_COOLDOWN_SEC

    positives = [str(r["source"]) for r in rows if parse_float(r["score"], 0.0) > 0]
    fresh = [str(r["source"]) for r in rows if parse_float(r["score"], 0.0) <= 0 and int(r["attempts"] or 0) == 0]
    soft_failed = [
        str(r["source"])
        for r in rows
        if parse_float(r["score"], 0.0) <= 0
        and int(r["attempts"] or 0) > 0
        and not _is_hard_failed(r)
    ]
    hard_failed = [
        str(r["source"])
        for r in rows
        if _is_hard_failed(r)
    ]
    hard_failed_set = set(hard_failed)
    ordered = positives + fresh + soft_failed
    for src in configured:
        if src not in ordered and src not in hard_failed_set:
            ordered.append(src)

    top_k = max(1, min(8, int(parse_float(policy.get("top_k"), LIVE_QUOTE_TOP_K_DEFAULT))))
    selected = ordered[:top_k]
    if len(selected) < 2 and hard_failed:
        ex_pref = "nse_api" if str(exchange or "").upper() == "NSE" else ("bse_api" if str(exchange or "").upper() == "BSE" else None)
        backup = ex_pref if ex_pref in hard_failed else hard_failed[0]
        if backup and backup not in selected:
            selected.append(backup)
    explore_ratio = clamp(parse_float(policy.get("explore_ratio"), LIVE_QUOTE_EXPLORE_RATIO_DEFAULT), 0.0, 0.8)
    remainder = ordered[top_k:] + [s for s in hard_failed if s not in ordered[top_k:]]
    if remainder and explore_ratio > 0 and random.random() < explore_ratio:
        probe = random.choice(remainder)
        if probe not in selected:
            selected.append(probe)
    must_include = "nse_api" if ex == "NSE" else ("bse_api" if ex == "BSE" else None)
    if must_include and must_include in ordered and must_include not in selected:
        must_row = next((r for r in rows if str(r["source"]) == must_include), None)
        must_score = parse_float(must_row["score"], 0.0) if must_row else 0.0
        must_attempts = int(must_row["attempts"] or 0) if must_row else 0
        must_successes = int(must_row["successes"] or 0) if must_row else 0
        # Always include exchange source if it has any working signal, or if no positive source exists yet.
        if must_score > 0 or not positives or (must_attempts < 15 and must_successes <= 0):
            selected.insert(0, must_include)
    if not selected:
        selected = [s for s in BUILTIN_LIVE_QUOTE_SOURCES]
    if ex == "NSE" and "screener_scrape" in configured and "screener_scrape" not in selected:
        selected.append("screener_scrape")
    out = []
    for s in selected:
        if s not in out:
            out.append(s)
    return out


def quote_source_ranking(conn):
    ensure_quote_source_registry(conn, [])
    rows = conn.execute(
        """
        SELECT
          r.source,
          r.adapter,
          r.enabled,
          COALESCE(s.score, 0) AS score,
          COALESCE(s.attempts, 0) AS attempts,
          COALESCE(s.successes, 0) AS successes,
          COALESCE(s.failures, 0) AS failures,
          COALESCE(s.total_latency_ms, 0) AS total_latency_ms,
          COALESCE(s.total_accuracy_error_pct, 0) AS total_accuracy_error_pct,
          COALESCE(s.accuracy_samples, 0) AS accuracy_samples,
          s.last_success_at,
          s.last_error_at,
          r.updated_at
        FROM quote_source_registry r
        LEFT JOIN quote_source_stats s ON s.source = r.source
        ORDER BY score DESC, attempts DESC, r.source
        """
    ).fetchall()
    out = []
    for r in rows:
        attempts = int(r["attempts"] or 0)
        succ = int(r["successes"] or 0)
        latency = parse_float(r["total_latency_ms"], 0.0)
        avg_latency = (latency / succ) if succ > 0 else None
        err_samples = int(r["accuracy_samples"] or 0)
        avg_err = (
            parse_float(r["total_accuracy_error_pct"], 0.0) / err_samples
            if err_samples > 0
            else None
        )
        out.append(
            {
                "source": r["source"],
                "adapter": r["adapter"],
                "enabled": bool(int(r["enabled"] or 0)),
                "score": round(parse_float(r["score"], 0.0), 4),
                "attempts": attempts,
                "successes": succ,
                "failures": int(r["failures"] or 0),
                "success_rate_pct": round((succ * 100.0 / attempts), 2) if attempts > 0 else 0.0,
                "avg_latency_ms": round(avg_latency, 2) if avg_latency is not None else None,
                "avg_accuracy_error_pct": round(avg_err, 4) if avg_err is not None else None,
                "last_success_at": r["last_success_at"],
                "last_error_at": r["last_error_at"],
                "updated_at": r["updated_at"],
            }
        )
    return out


def set_quote_source_enabled(conn, source, enabled=True, notes=None):
    src = str(source or "").strip().lower()
    if not src or not re.match(r"^[a-z0-9_:-]{2,64}$", src):
        raise ValueError("invalid_source")
    ensure_quote_source_registry(conn, [src])
    conn.execute(
        """
        UPDATE quote_source_registry
        SET enabled = ?, updated_at = ?, notes = COALESCE(?, notes)
        WHERE source = ?
        """,
        (1 if bool(enabled) else 0, now_iso(), notes, src),
    )


def _quote_corroboration_count(candidates, selected_source, ltp, tolerance_pct=2.0):
    anchor = parse_float(ltp, 0.0)
    if anchor <= 0:
        return 0
    tol = max(0.01, anchor * max(0.2, parse_float(tolerance_pct, 2.0)) / 100.0)
    src = str(selected_source or "").lower()
    cnt = 0
    for c in candidates or []:
        csrc = str(c.get("source") or "").lower()
        if csrc == src:
            continue
        cpx = parse_float(c.get("ltp"), 0.0)
        if cpx > 0 and abs(cpx - anchor) <= tol:
            cnt += 1
    return cnt


def _quote_is_plausible(selected_source, ltp, prev_ltp, qty, avg_cost, candidates, max_dev_pct, asset_class=None, symbol=None):
    px = parse_float(ltp, 0.0)
    if px <= 0:
        return False
    source = str(selected_source or "").lower()
    ac = normalize_asset_class(asset_class, fallback=infer_asset_class(symbol=symbol, name=symbol))
    if source == "gold_rate_scrape" and ac != ASSET_CLASS_GOLD:
        return False
    is_scrape = source.endswith("_scrape")
    corroborated = _quote_corroboration_count(
        candidates,
        selected_source,
        px,
        tolerance_pct=max(1.5, parse_float(max_dev_pct, LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT)),
    )
    prev = parse_float(prev_ltp, 0.0)
    if prev > 0:
        jump_pct = abs(px - prev) / prev * 100.0
        jump_limit = 30.0 if is_scrape else 60.0
        if jump_pct > jump_limit and corroborated <= 0:
            return False
    q = parse_float(qty, 0.0)
    avg = parse_float(avg_cost, 0.0)
    if q > 0 and avg > 0 and corroborated <= 0:
        if px > (avg * 80.0):
            return False
        if px < (avg / 30.0):
            return False
    return True


def sanitize_latest_price_outliers(conn, fetched_at):
    rows = conn.execute(
        """
        SELECT
          lp.symbol,
          lp.ltp,
          COALESCE(h.qty,0) AS qty,
          COALESCE(h.avg_cost,0) AS avg_cost
        FROM latest_prices lp
        LEFT JOIN holdings h ON h.symbol = lp.symbol
        """
    ).fetchall()
    last_trade_rows = conn.execute(
        """
        SELECT t.symbol, t.price
        FROM trades t
        JOIN (
          SELECT symbol, MAX(id) AS max_id
          FROM trades
          GROUP BY symbol
        ) x ON x.symbol = t.symbol AND x.max_id = t.id
        """
    ).fetchall()
    last_trade_px = {symbol_upper(r["symbol"]): parse_float(r["price"], 0.0) for r in last_trade_rows}
    replacements = []
    for r in rows:
        symbol = symbol_upper(r["symbol"])
        ltp = parse_float(r["ltp"], 0.0)
        qty = parse_float(r["qty"], 0.0)
        avg_cost = parse_float(r["avg_cost"], 0.0)
        if ltp <= 0 or qty <= 0 or avg_cost <= 0:
            continue
        if (ltp <= avg_cost * 80.0) and (ltp >= avg_cost / 30.0):
            continue
        fallback = parse_float(last_trade_px.get(symbol), 0.0)
        if fallback <= 0:
            fallback = avg_cost
        if fallback <= 0:
            continue
        replacements.append((symbol, fallback, 0.0, fetched_at, "plausibility_guard"))

    if replacements:
        conn.executemany(
            """
            INSERT INTO latest_prices(symbol, ltp, change_abs, updated_at) VALUES (?, ?, ?, ?)
            ON CONFLICT(symbol) DO UPDATE SET
              ltp=excluded.ltp,
              change_abs=excluded.change_abs,
              updated_at=excluded.updated_at
            """,
            [(r[0], r[1], r[2], r[3]) for r in replacements],
        )
        conn.executemany(
            "INSERT INTO price_ticks(symbol, ltp, change_abs, fetched_at, source) VALUES (?, ?, ?, ?, ?)",
            replacements,
        )
    return replacements


def sanitize_latest_price_day_change_outliers(conn, fetched_at=None):
    rows = conn.execute(
        """
        SELECT
          lp.symbol,
          lp.ltp,
          lp.change_abs,
          COALESCE(i.asset_class, 'EQUITY') AS asset_class
        FROM latest_prices lp
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(lp.symbol)
        WHERE COALESCE(lp.ltp, 0) > 0
        """
    ).fetchall()
    if not rows:
        return []
    prev_close_map = load_prev_close_map(conn, [r["symbol"] for r in rows])
    replacements = []
    for r in rows:
        symbol = symbol_upper(r["symbol"])
        ltp = parse_float(r["ltp"], 0.0)
        raw_change = parse_float(r["change_abs"], 0.0)
        asset_class = normalize_asset_class(
            r["asset_class"],
            fallback=infer_asset_class(symbol=symbol, name=symbol),
        )
        if ltp <= 0 or not symbol:
            continue
        if asset_class == ASSET_CLASS_GOLD:
            # GOLD uses a dedicated previous-close baseline and should not be recast
            # from equity intraday tick behavior here.
            continue
        raw_prev_close = ltp - raw_change
        if abs(raw_change) <= 1e-9:
            continue
        if raw_prev_close > 0 and is_plausible_day_reference_price(ltp, raw_prev_close):
            continue
        fixed_change = resolve_effective_change_abs(conn, symbol, ltp, 0.0, prev_close_map=prev_close_map)
        if abs(fixed_change - raw_change) <= 1e-9:
            continue
        replacements.append((fixed_change, fetched_at or now_iso(), symbol))
    if replacements:
        conn.executemany(
            """
            UPDATE latest_prices
            SET change_abs = ?, updated_at = ?
            WHERE UPPER(symbol) = ?
            """,
            replacements,
        )
    return replacements


def purge_non_gold_quote_source_contamination(conn, source_name="gold_rate_scrape"):
    src = str(source_name or "").strip().lower()
    if not src:
        return {"price_ticks": 0, "quote_samples": 0}
    rows = conn.execute(
        """
        SELECT symbol
        FROM instruments
        WHERE UPPER(COALESCE(asset_class, 'EQUITY')) <> 'GOLD'
        """
    ).fetchall()
    symbols = [symbol_upper(r["symbol"]) for r in rows if symbol_upper(r["symbol"])]
    if not symbols:
        return {"price_ticks": 0, "quote_samples": 0}
    total_ticks = 0
    total_samples = 0
    for i in range(0, len(symbols), 400):
        chunk = symbols[i : i + 400]
        placeholders = ",".join(["?"] * len(chunk))
        params = chunk + [src]
        total_ticks += conn.execute(
            f"""
            DELETE FROM price_ticks
            WHERE UPPER(symbol) IN ({placeholders})
              AND LOWER(COALESCE(source,'')) = ?
            """,
            params,
        ).rowcount
        total_samples += conn.execute(
            f"""
            DELETE FROM quote_samples
            WHERE UPPER(symbol) IN ({placeholders})
              AND LOWER(COALESCE(source,'')) = ?
            """,
            params,
        ).rowcount
    return {"price_ticks": total_ticks, "quote_samples": total_samples}


def repair_current_tenant_market_data(conn=None, fetched_at=None):
    owns_conn = False
    if conn is None:
        conn = db_connect()
        owns_conn = True
    stamp = fetched_at or now_iso()
    try:
        purged = purge_non_gold_quote_source_contamination(conn, "gold_rate_scrape")
        ltp_replacements = sanitize_latest_price_outliers(conn, stamp)
        day_change_replacements = sanitize_latest_price_day_change_outliers(conn, stamp)
        refresh_holdings_mark_to_market(conn)
        conn.commit()
        return {
            "purged_price_ticks": int(purged.get("price_ticks", 0)),
            "purged_quote_samples": int(purged.get("quote_samples", 0)),
            "ltp_repairs": len(ltp_replacements),
            "day_change_repairs": len(day_change_replacements),
        }
    finally:
        if owns_conn:
            conn.close()


def repair_all_tenants_market_data_once():
    global _startup_market_data_repair_done
    with _startup_market_data_repair_lock:
        if _startup_market_data_repair_done:
            return []
        _startup_market_data_repair_done = True
    results = []
    ensure_tenant_bootstrap()
    for tenant in list_tenants():
        key = tenant.get("key")
        if not key:
            continue
        with tenant_context(key):
            init_db()
            with db_connect() as conn:
                summary = repair_current_tenant_market_data(conn, fetched_at=now_iso())
            results.append({"tenant": key, **summary})
    return results


def get_history_sync_config(conn):
    rows = conn.execute(
        "SELECT key, value FROM app_config WHERE key IN ('history_refresh_enabled','history_refresh_interval_sec','history_last_sync_at')"
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("history_refresh_enabled", "1")) == "1"
    try:
        interval = int(float(cfg.get("history_refresh_interval_sec", str(HISTORY_REFRESH_DEFAULT_SEC))))
    except ValueError:
        interval = HISTORY_REFRESH_DEFAULT_SEC
    interval = max(HISTORY_REFRESH_MIN_SEC, interval)
    return {
        "enabled": enabled,
        "interval_seconds": interval,
        "last_sync_at": str(cfg.get("history_last_sync_at", "") or ""),
    }


def set_history_sync_config(enabled=None, interval_seconds=None):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('history_refresh_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_seconds is not None:
            interval_seconds = max(HISTORY_REFRESH_MIN_SEC, int(interval_seconds))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('history_refresh_interval_sec', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(interval_seconds), now_iso()),
            )
        conn.commit()


def set_live_config(
    enabled=None,
    interval_seconds=None,
    quote_sources=None,
    quote_max_deviation_pct=None,
    quote_top_k=None,
    quote_explore_ratio=None,
):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_refresh_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_seconds is not None:
            interval_seconds = max(LIVE_REFRESH_MIN_SEC, int(interval_seconds))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_refresh_interval_sec', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(interval_seconds), now_iso()),
            )
        if quote_sources is not None:
            if isinstance(quote_sources, list):
                raw = ",".join(str(s or "").strip() for s in quote_sources)
            else:
                raw = str(quote_sources or "")
            norm = ",".join(parse_source_list(raw, DEFAULT_LIVE_QUOTE_SOURCES))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_quote_sources', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (norm, now_iso()),
            )
        if quote_max_deviation_pct is not None:
            val = clamp(parse_float(quote_max_deviation_pct, LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT), 1.0, 20.0)
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_quote_max_deviation_pct', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(round(val, 3)), now_iso()),
            )
        if quote_top_k is not None:
            val = max(1, min(8, int(parse_float(quote_top_k, LIVE_QUOTE_TOP_K_DEFAULT))))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_quote_top_k', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(val), now_iso()),
            )
        if quote_explore_ratio is not None:
            val = clamp(parse_float(quote_explore_ratio, LIVE_QUOTE_EXPLORE_RATIO_DEFAULT), 0.0, 0.8)
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_quote_explore_ratio', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(round(val, 4)), now_iso()),
            )
        conn.commit()


def get_strategy_refresh_interval(conn):
    cfg = get_strategy_config(conn)
    return int(cfg["interval_seconds"])


def get_strategy_config(conn):
    rows = conn.execute(
        "SELECT key, value FROM app_config WHERE key IN ('strategy_refresh_enabled','strategy_refresh_interval_sec')"
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("strategy_refresh_enabled", "1")) == "1"
    try:
        interval = int(float(cfg.get("strategy_refresh_interval_sec", str(STRATEGY_REFRESH_MIN_SEC))))
    except ValueError:
        interval = STRATEGY_REFRESH_MIN_SEC
    return {
        "enabled": enabled,
        "interval_seconds": max(300, interval),
    }


def set_strategy_config(enabled=None, interval_seconds=None):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('strategy_refresh_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_seconds is not None:
            interval_seconds = max(300, int(interval_seconds))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('strategy_refresh_interval_sec', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(interval_seconds), now_iso()),
            )
        conn.commit()


def latest_backup_file_info():
    bdir = tenant_paths(get_current_tenant_key())["backup_dir"]
    if not bdir.exists():
        return {"last_run_at": "", "last_file": ""}
    files = [p for p in bdir.glob("portfolio_*.db") if p.is_file()]
    if not files:
        return {"last_run_at": "", "last_file": ""}
    latest = max(files, key=lambda p: p.stat().st_mtime)
    try:
        run_at = dt.datetime.fromtimestamp(latest.stat().st_mtime).replace(microsecond=0).isoformat()
    except Exception:
        run_at = ""
    return {"last_run_at": run_at, "last_file": str(latest)}


def get_backup_config(conn):
    rows = conn.execute(
        "SELECT key, value FROM app_config WHERE key IN ('backup_agent_enabled','backup_last_run_at','backup_last_file')"
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    fs = latest_backup_file_info()
    last_run = str(cfg.get("backup_last_run_at", "") or "")
    last_file = str(cfg.get("backup_last_file", "") or "")
    if not last_run:
        last_run = fs["last_run_at"]
    if not last_file:
        last_file = fs["last_file"]
    return {
        "enabled": str(cfg.get("backup_agent_enabled", "1")) == "1",
        "interval_seconds": DB_BACKUP_INTERVAL_SEC,
        "last_run_at": last_run,
        "last_file": last_file,
    }


def get_repo_sync_config(conn):
    rows = conn.execute(
        "SELECT key, value FROM app_config WHERE key IN ('repo_sync_enabled','repo_sync_interval_sec','repo_sync_auto_push','repo_sync_last_run_at','repo_sync_last_error')"
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    try:
        interval = int(float(cfg.get("repo_sync_interval_sec", str(REPO_SYNC_INTERVAL_DEFAULT_SEC))))
    except Exception:
        interval = REPO_SYNC_INTERVAL_DEFAULT_SEC
    return {
        "enabled": str(cfg.get("repo_sync_enabled", "1")) == "1",
        "auto_push": str(cfg.get("repo_sync_auto_push", "1")) == "1",
        "interval_seconds": max(REPO_SYNC_MIN_INTERVAL_SEC, interval),
        "last_run_at": str(cfg.get("repo_sync_last_run_at", "") or ""),
        "last_error": str(cfg.get("repo_sync_last_error", "") or ""),
    }


def set_backup_config(enabled=None):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('backup_agent_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        conn.commit()


def get_self_learning_config(conn):
    rows = conn.execute(
        """
        SELECT key, value
        FROM app_config
        WHERE key IN ('self_learning_enabled','self_learning_interval_days','self_learning_last_run_at','self_learning_min_samples')
        """
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("self_learning_enabled", "1")) == "1"
    try:
        interval_days = int(float(cfg.get("self_learning_interval_days", "7")))
    except Exception:
        interval_days = 7
    try:
        min_samples = int(float(cfg.get("self_learning_min_samples", "30")))
    except Exception:
        min_samples = 30
    return {
        "enabled": enabled,
        "interval_days": max(1, min(90, interval_days)),
        "last_run_at": str(cfg.get("self_learning_last_run_at", "") or ""),
        "min_samples": max(5, min(5000, min_samples)),
    }


def set_self_learning_config(enabled=None, interval_days=None, min_samples=None, last_run_at=None):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('self_learning_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_days is not None:
            v = max(1, min(90, int(interval_days)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('self_learning_interval_days', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if min_samples is not None:
            v = max(5, min(5000, int(min_samples)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('self_learning_min_samples', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if last_run_at is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('self_learning_last_run_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_run_at or ""), now_iso()),
            )
        conn.commit()


def get_intel_autopilot_config(conn):
    rows = conn.execute(
        """
        SELECT key, value
        FROM app_config
        WHERE key IN (
          'intel_autopilot_enabled',
          'intel_autopilot_interval_sec',
          'intel_autopilot_last_run_at',
          'intel_autopilot_max_docs',
          'intel_autopilot_symbols_limit',
          'intel_autopilot_sources',
          'intel_autopilot_query_seed'
        )
        """
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("intel_autopilot_enabled", "1")) == "1"
    try:
        interval_sec = int(float(cfg.get("intel_autopilot_interval_sec", str(12 * 60 * 60))))
    except Exception:
        interval_sec = 12 * 60 * 60
    try:
        max_docs = int(float(cfg.get("intel_autopilot_max_docs", "24")))
    except Exception:
        max_docs = 24
    try:
        symbols_limit = int(float(cfg.get("intel_autopilot_symbols_limit", "20")))
    except Exception:
        symbols_limit = 20
    raw_sources = str(cfg.get("intel_autopilot_sources", "") or "").strip()
    default_sources = "google_news_rss,screener_financials,nse_announcements,company_site_ir"
    sources = parse_token_list(raw_sources, default_sources)
    # Auto-upgrade legacy default while preserving explicit custom source sets.
    if (not raw_sources) or (raw_sources.lower() == "google_news_rss"):
        for s in ("screener_financials", "nse_announcements", "company_site_ir"):
            if s not in sources:
                sources.append(s)
    if not sources:
        sources = parse_token_list(default_sources, default_sources)
    query_seed = str(cfg.get("intel_autopilot_query_seed", "") or "").strip()
    if not query_seed:
        query_seed = (
            "India union budget policy stocks\n"
            "RBI policy impact equities India\n"
            "SEBI circular equity market impact\n"
            "India sector rotation stocks"
        )
    return {
        "enabled": enabled,
        "interval_seconds": max(60 * 15, min(7 * 24 * 60 * 60, interval_sec)),
        "last_run_at": str(cfg.get("intel_autopilot_last_run_at", "") or ""),
        "max_docs": max(1, min(200, max_docs)),
        "symbols_limit": max(1, min(100, symbols_limit)),
        "sources": sources,
        "query_seed": query_seed,
    }


def set_intel_autopilot_config(
    enabled=None,
    interval_seconds=None,
    max_docs=None,
    symbols_limit=None,
    sources=None,
    query_seed=None,
    last_run_at=None,
):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_seconds is not None:
            v = max(60 * 15, min(7 * 24 * 60 * 60, int(interval_seconds)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_interval_sec', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if max_docs is not None:
            v = max(1, min(200, int(max_docs)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_max_docs', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if symbols_limit is not None:
            v = max(1, min(100, int(symbols_limit)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_symbols_limit', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if sources is not None:
            if isinstance(sources, list):
                src = ",".join(str(x or "").strip() for x in sources if str(x or "").strip())
            else:
                src = str(sources or "")
            src_parts = parse_token_list(
                src,
                "google_news_rss,screener_financials,nse_announcements,company_site_ir",
            )
            if not src_parts:
                src_parts = ["google_news_rss", "screener_financials", "nse_announcements", "company_site_ir"]
            src = ",".join(src_parts)
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_sources', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (src, now_iso()),
            )
        if query_seed is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_query_seed', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(query_seed or ""), now_iso()),
            )
        if last_run_at is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('intel_autopilot_last_run_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_run_at or ""), now_iso()),
            )
        conn.commit()


def get_chart_agent_config(conn):
    rows = conn.execute(
        """
        SELECT key, value
        FROM app_config
        WHERE key IN (
          'chart_agent_enabled',
          'chart_agent_interval_sec',
          'chart_agent_last_run_at',
          'chart_agent_sources'
        )
        """
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("chart_agent_enabled", "1")) == "1"
    try:
        interval_sec = int(float(cfg.get("chart_agent_interval_sec", str(CHART_AGENT_INTERVAL_DEFAULT_SEC))))
    except Exception:
        interval_sec = CHART_AGENT_INTERVAL_DEFAULT_SEC
    sources = parse_chart_source_list(cfg.get("chart_agent_sources"), DEFAULT_CHART_AGENT_SOURCES)
    return {
        "enabled": enabled,
        "interval_seconds": max(CHART_AGENT_MIN_INTERVAL_SEC, min(7 * 24 * 60 * 60, interval_sec)),
        "last_run_at": str(cfg.get("chart_agent_last_run_at", "") or ""),
        "sources": sources,
    }


def set_chart_agent_config(enabled=None, interval_seconds=None, sources=None, last_run_at=None):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('chart_agent_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_seconds is not None:
            v = max(CHART_AGENT_MIN_INTERVAL_SEC, min(7 * 24 * 60 * 60, int(interval_seconds)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('chart_agent_interval_sec', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if sources is not None:
            if isinstance(sources, list):
                src = ",".join(str(x or "").strip() for x in sources if str(x or "").strip())
            else:
                src = str(sources or "")
            src = ",".join(parse_chart_source_list(src, DEFAULT_CHART_AGENT_SOURCES))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('chart_agent_sources', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (src, now_iso()),
            )
        if last_run_at is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('chart_agent_last_run_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_run_at or ""), now_iso()),
            )
        conn.commit()


def get_software_perf_agent_config(conn):
    rows = conn.execute(
        """
        SELECT key, value
        FROM app_config
        WHERE key IN (
          'software_perf_agent_enabled',
          'software_perf_agent_interval_sec',
          'software_perf_agent_last_run_at',
          'software_perf_agent_last_heal_at',
          'software_perf_agent_last_improvement_at',
          'software_perf_agent_auto_tune',
          'software_perf_agent_write_changes',
          'software_perf_agent_core_objective'
        )
        """
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    enabled = str(cfg.get("software_perf_agent_enabled", "1")) == "1"
    auto_tune = str(cfg.get("software_perf_agent_auto_tune", "1")) == "1"
    write_changes = str(cfg.get("software_perf_agent_write_changes", "1")) == "1"
    try:
        interval_sec = int(float(cfg.get("software_perf_agent_interval_sec", str(SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC))))
    except Exception:
        interval_sec = SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC
    core_objective = str(
        cfg.get("software_perf_agent_core_objective", "Preserve portfolio data integrity and strategy objective.") or ""
    ).strip()
    if not core_objective:
        core_objective = "Preserve portfolio data integrity and strategy objective."
    return {
        "enabled": enabled,
        "interval_seconds": max(SOFTWARE_PERF_AGENT_MIN_INTERVAL_SEC, min(7 * 24 * 60 * 60, interval_sec)),
        "last_run_at": str(cfg.get("software_perf_agent_last_run_at", "") or ""),
        "last_heal_at": str(cfg.get("software_perf_agent_last_heal_at", "") or ""),
        "last_improvement_at": str(cfg.get("software_perf_agent_last_improvement_at", "") or ""),
        "auto_tune": auto_tune,
        "write_changes": write_changes,
        "core_objective": core_objective,
    }


def set_software_perf_agent_config(
    enabled=None,
    interval_seconds=None,
    auto_tune=None,
    write_changes=None,
    core_objective=None,
    last_run_at=None,
    last_heal_at=None,
    last_improvement_at=None,
):
    with db_connect() as conn:
        if enabled is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_enabled', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if enabled else "0", now_iso()),
            )
        if interval_seconds is not None:
            v = max(SOFTWARE_PERF_AGENT_MIN_INTERVAL_SEC, min(7 * 24 * 60 * 60, int(interval_seconds)))
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_interval_sec', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(v), now_iso()),
            )
        if auto_tune is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_auto_tune', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if auto_tune else "0", now_iso()),
            )
        if write_changes is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_write_changes', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                ("1" if write_changes else "0", now_iso()),
            )
        if core_objective is not None:
            val = str(core_objective or "").strip()
            if not val:
                val = "Preserve portfolio data integrity and strategy objective."
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_core_objective', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (val, now_iso()),
            )
        if last_run_at is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_last_run_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_run_at or ""), now_iso()),
            )
        if last_heal_at is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_last_heal_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_heal_at or ""), now_iso()),
            )
        if last_improvement_at is not None:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('software_perf_agent_last_improvement_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_improvement_at or ""), now_iso()),
            )
        conn.commit()


def _mask_secret(secret, keep=4):
    raw = str(secret or "").strip()
    if not raw:
        return ""
    if len(raw) <= keep:
        return "*" * len(raw)
    return ("*" * max(4, len(raw) - keep)) + raw[-keep:]


def _read_llm_runtime_config(conn, include_secret=False):
    rows = conn.execute(
        """
        SELECT key, value
        FROM app_config
        WHERE key IN ('llm_api_key', 'llm_model', 'llm_api_url', 'llm_last_status', 'llm_last_error', 'llm_last_checked_at')
        """
    ).fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    api_key_db = str(cfg.get("llm_api_key", "") or "").strip()
    api_key_env = str(os.environ.get("OPENAI_API_KEY", "") or "").strip()
    api_key = api_key_db or api_key_env
    model = str(cfg.get("llm_model", "") or "").strip() or str(os.environ.get("OPENAI_MODEL", "") or "").strip() or LLM_DEFAULT_MODEL
    api_url = str(cfg.get("llm_api_url", "") or "").strip() or str(os.environ.get("OPENAI_API_URL", "") or "").strip() or LLM_DEFAULT_API_URL
    configured = bool(api_key)
    source = "app_config" if api_key_db else ("environment" if api_key_env else "not_configured")
    last_status = str(cfg.get("llm_last_status", "") or "").strip()
    if not last_status:
        last_status = "ready" if configured else "not_configured"
    out = {
        "configured": configured,
        "source": source,
        "model": model,
        "api_url": api_url,
        "api_key_masked": _mask_secret(api_key),
        "last_status": last_status,
        "last_error": str(cfg.get("llm_last_error", "") or "").strip(),
        "last_checked_at": str(cfg.get("llm_last_checked_at", "") or "").strip(),
    }
    if include_secret:
        out["api_key"] = api_key
    return out


def get_llm_runtime_config(conn=None, include_secret=False):
    if conn is not None:
        return _read_llm_runtime_config(conn, include_secret=include_secret)
    with db_connect() as conn2:
        return _read_llm_runtime_config(conn2, include_secret=include_secret)


def set_llm_runtime_config(
    api_key="__UNCHANGED__",
    model="__UNCHANGED__",
    api_url="__UNCHANGED__",
    last_status="__UNCHANGED__",
    last_error="__UNCHANGED__",
    last_checked_at="__UNCHANGED__",
):
    with db_connect() as conn:
        if api_key != "__UNCHANGED__":
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('llm_api_key', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(api_key or "").strip(), now_iso()),
            )
        if model != "__UNCHANGED__":
            val = str(model or "").strip() or LLM_DEFAULT_MODEL
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('llm_model', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (val, now_iso()),
            )
        if api_url != "__UNCHANGED__":
            val = str(api_url or "").strip() or LLM_DEFAULT_API_URL
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('llm_api_url', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (val, now_iso()),
            )
        if last_status != "__UNCHANGED__":
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('llm_last_status', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_status or "").strip(), now_iso()),
            )
        if last_error != "__UNCHANGED__":
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('llm_last_error', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_error or "").strip(), now_iso()),
            )
        if last_checked_at != "__UNCHANGED__":
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('llm_last_checked_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(last_checked_at or "").strip(), now_iso()),
            )
        conn.commit()


def _update_llm_runtime_status(conn, status, error="", checked_at=None):
    stamp = str(checked_at or now_iso())
    conn.execute(
        """
        INSERT INTO app_config(key, value, updated_at) VALUES ('llm_last_status', ?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        """,
        (str(status or "").strip(), now_iso()),
    )
    conn.execute(
        """
        INSERT INTO app_config(key, value, updated_at) VALUES ('llm_last_error', ?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        """,
        (str(error or "").strip(), now_iso()),
    )
    conn.execute(
        """
        INSERT INTO app_config(key, value, updated_at) VALUES ('llm_last_checked_at', ?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        """,
        (stamp, now_iso()),
    )


def _extract_llm_output_text(raw):
    text = str((raw or {}).get("output_text") or "").strip()
    if text:
        return text
    parts = []
    for item in (raw or {}).get("output") or []:
        for content in item.get("content") or []:
            if str(content.get("type") or "") == "output_text":
                t = str(content.get("text") or "").strip()
                if t:
                    parts.append(t)
    return "\n".join(parts).strip()


def try_parse_json_object(text):
    raw = str(text or "").strip()
    if not raw:
        return {}
    candidates = [raw]
    if raw.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.I)
        cleaned = re.sub(r"\s*```$", "", cleaned)
        candidates.append(cleaned.strip())
    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        candidates.append(raw[start : end + 1].strip())
    seen = set()
    for cand in candidates:
        if not cand or cand in seen:
            continue
        seen.add(cand)
        try:
            obj = json.loads(cand)
            if isinstance(obj, dict):
                return obj
        except Exception:
            pass
        try:
            obj = ast.literal_eval(cand)
            if isinstance(obj, dict):
                return obj
        except Exception:
            pass
    return {}


def call_llm_responses_api(system_prompt, user_payload, conn=None, max_output_tokens=500, timeout=12):
    owns_conn = conn is None
    if owns_conn:
        conn = db_connect()
    try:
        cfg = _read_llm_runtime_config(conn, include_secret=True)
        api_key = str(cfg.get("api_key") or "").strip()
        if not api_key:
            _update_llm_runtime_status(conn, "not_configured", "LLM API key not configured.")
            conn.commit()
            raise RuntimeError("LLM API key not configured.")
        payload = {
            "model": str(cfg.get("model") or LLM_DEFAULT_MODEL),
            "input": [
                {"role": "system", "content": str(system_prompt or "").strip()},
                {
                    "role": "user",
                    "content": json.dumps(user_payload, ensure_ascii=True) if not isinstance(user_payload, str) else str(user_payload),
                },
            ],
            "max_output_tokens": max(32, int(max_output_tokens)),
        }
        req = urllib.request.Request(
            str(cfg.get("api_url") or LLM_DEFAULT_API_URL),
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=max(3, int(timeout))) as resp:
                raw = json.loads(resp.read().decode("utf-8"))
        except Exception as ex:
            _update_llm_runtime_status(conn, "error", str(ex))
            conn.commit()
            raise RuntimeError(f"LLM request failed: {str(ex)}")
        text = _extract_llm_output_text(raw)
        if not text:
            _update_llm_runtime_status(conn, "error", "LLM returned empty analysis.")
            conn.commit()
            raise RuntimeError("LLM returned empty analysis.")
        _update_llm_runtime_status(conn, "ok", "")
        conn.commit()
        return {
            "provider": "openai_responses",
            "model": str(cfg.get("model") or LLM_DEFAULT_MODEL),
            "api_url": str(cfg.get("api_url") or LLM_DEFAULT_API_URL),
            "text": text,
            "raw": raw,
        }
    finally:
        if owns_conn and conn is not None:
            conn.close()


def test_llm_runtime(conn=None):
    result = call_llm_responses_api(
        system_prompt="Return a short response confirming the model is reachable.",
        user_payload={"task": "connectivity_test", "reply_format": "one short line"},
        conn=conn,
        max_output_tokens=40,
        timeout=10,
    )
    return {
        "ok": True,
        "status": "ok",
        "provider": result.get("provider"),
        "model": result.get("model"),
        "message": result.get("text"),
        "config": get_llm_runtime_config(conn=conn, include_secret=False),
    }


def _clean_rss_text(text):
    s = str(text or "")
    s = re.sub(r"<!\[CDATA\[(.*?)\]\]>", r"\1", s, flags=re.S)
    s = re.sub(r"<[^>]+>", " ", s, flags=re.S)
    s = html.unescape(s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _rss_extract_tag(block, tag):
    m = re.search(rf"<{tag}[^>]*>(.*?)</{tag}>", str(block or ""), flags=re.S | re.IGNORECASE)
    if not m:
        return ""
    return _clean_rss_text(m.group(1))


def fetch_google_news_rss(query, limit=6, timeout=5):
    q = str(query or "").strip()
    if not q:
        return []
    opener = urllib.request.build_opener()
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/rss+xml,application/xml,text/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://news.google.com/",
    }
    # Include market context to focus equity-relevant items.
    q_full = f"{q} stock market India"
    url = "https://news.google.com/rss/search?q=" + urllib.parse.quote(q_full) + "&hl=en-IN&gl=IN&ceid=IN:en"
    xml = _http_text(opener, url, headers, timeout=timeout)
    items_raw = re.findall(r"<item>(.*?)</item>", xml, flags=re.S | re.IGNORECASE)
    out = []
    for it in items_raw[: max(1, min(30, int(limit) * 3))]:
        title = _rss_extract_tag(it, "title")
        link = _rss_extract_tag(it, "link")
        pub = _rss_extract_tag(it, "pubDate")
        desc = _rss_extract_tag(it, "description")
        if not title:
            continue
        out.append(
            {
                "title": title,
                "link": link,
                "published_at": pub,
                "snippet": desc,
                "source": "google_news_rss",
            }
        )
        if len(out) >= int(limit):
            break
    return out


def _infer_doc_type_from_text(text):
    t = str(text or "").lower()
    if any(k in t for k in ("budget", "policy", "rbi", "sebi", "government", "regulation", "tax", "duty")):
        return "policy"
    if any(k in t for k in ("qoq", "quarter", "q1", "q2", "q3", "q4", "earnings", "results", "financial statement", "cash flow")):
        return "financial_statement"
    return "commentary"


def _collect_autopilot_symbols(conn, symbols_limit=20):
    rows = conn.execute(
        """
        SELECT i.symbol, COALESCE(h.market_value,0) AS mv
        FROM instruments i
        LEFT JOIN holdings h ON UPPER(h.symbol) = UPPER(i.symbol)
        WHERE i.active = 1
        ORDER BY mv DESC, i.symbol
        LIMIT ?
        """,
        (max(1, min(100, int(symbols_limit))),),
    ).fetchall()
    out = []
    seen = set()
    for r in rows:
        sym = symbol_upper(r["symbol"])
        if not sym or sym in seen:
            continue
        seen.add(sym)
        out.append(sym)
    return out


def _collect_autopilot_symbol_queries(conn, symbols_limit=20):
    symbols = _collect_autopilot_symbols(conn, symbols_limit=symbols_limit)
    queries = []
    for sym in symbols:
        queries.append(f"{sym} NSE quarterly results")
        queries.append(f"{sym} NSE guidance commentary")
        queries.append(f"{sym} India policy impact")
    return queries


def _doc_already_exists(conn, source_ref, title, doc_date):
    ref = str(source_ref or "").strip()
    ttl = str(title or "").strip()
    d = str(doc_date or "").strip()
    if ref:
        c = int(
            conn.execute(
                "SELECT COUNT(*) AS c FROM intelligence_documents WHERE source_ref = ?",
                (ref,),
            ).fetchone()["c"]
        )
        if c > 0:
            return True
    if ttl and d:
        c = int(
            conn.execute(
                "SELECT COUNT(*) AS c FROM intelligence_documents WHERE title = ? AND doc_date = ?",
                (ttl, d),
            ).fetchone()["c"]
        )
        if c > 0:
            return True
    return False


def _parse_html_table_rows(html_text):
    rows = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", str(html_text or ""), flags=re.IGNORECASE | re.S):
        cols = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, flags=re.IGNORECASE | re.S)
        cells = [_clean_rss_text(c) for c in cols]
        cells = [c for c in cells if c]
        if not cells:
            continue
        label = str(cells[0]).strip().lower()
        nums = []
        for c in cells[1:]:
            v = parse_float(c, None)
            if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                continue
            nums.append(float(v))
        rows.append({"label": label, "numbers": nums, "cells": cells})
    return rows


def _pick_latest_metric(rows, label_hints):
    hints = [str(h or "").strip().lower() for h in (label_hints or []) if str(h or "").strip()]
    best_val = None
    best_score = None
    for r in (rows or []):
        label = str(r.get("label") or "")
        if not label:
            continue
        if hints and not any(h in label for h in hints):
            continue
        nums = [parse_float(x, None) for x in (r.get("numbers") or [])]
        nums = [x for x in nums if x is not None]
        if not nums:
            continue
        cand = float(nums[-1])
        label_score = 0
        for h in hints:
            if label == h:
                label_score += 8
            elif label.startswith(h):
                label_score += 5
            elif h in label:
                label_score += 2
        # Prefer rows with richer history columns to avoid one-off ratio snippets.
        score = (label_score * 100) + min(24, len(nums))
        if (best_score is None) or (score > best_score):
            best_score = score
            best_val = cand
    return best_val


def _infer_financial_period_from_text(text, fallback_date):
    now_year = dt.date.today().year
    years = [
        int(y)
        for y in re.findall(r"(20[0-9]{2})", str(text or ""))
        if 2000 <= int(y) <= now_year
    ]
    if years:
        yr = max(years)
        return f"FY{str(yr)[2:]}_AUTO", f"{yr}-03-31"
    d = fallback_date or dt.date.today()
    return f"FY{str(d.year)[2:]}_AUTO", d.isoformat()


def _resolve_screener_company_urls(client, symbol):
    sym = symbol_upper(symbol)
    if not sym:
        return []
    urls = []
    fallback_urls = [
        f"https://www.screener.in/company/{urllib.parse.quote(sym)}/",
        f"https://www.screener.in/company/{urllib.parse.quote(sym)}/consolidated/",
    ]
    for u in fallback_urls:
        if u not in urls:
            urls.append(u)
    try:
        search_url = f"https://www.screener.in/api/company/search/?q={urllib.parse.quote(sym)}"
        req = urllib.request.Request(
            search_url,
            headers={**client.screener_headers, "Accept": "application/json,text/plain,*/*"},
        )
        with client.screener_opener.open(req, timeout=5) as resp:
            arr = json.loads(resp.read().decode("utf-8", errors="ignore"))
        if isinstance(arr, list):
            ranked = []
            for idx, it in enumerate(arr[:10]):
                rel = str((it or {}).get("url") or "").strip()
                if not rel:
                    continue
                full = urllib.parse.urljoin("https://www.screener.in", rel)
                name = str((it or {}).get("name") or "").strip().upper()
                m = re.search(r"/company/([^/]+)/?", rel, flags=re.IGNORECASE)
                code = symbol_upper(m.group(1) if m else "")
                score = 0
                if code == sym:
                    score += 200
                elif code and (sym.startswith(code) or code.startswith(sym)):
                    score += 120
                elif sym and sym in name:
                    score += 60
                ranked.append((score, -idx, full))
            ranked.sort(reverse=True)
            for _, __, full in ranked:
                if full not in urls:
                    urls.append(full)
    except Exception:
        pass
    return urls


def fetch_screener_financial_snapshot(symbol, timeout=6):
    client = MarketDataClient()
    urls = _resolve_screener_company_urls(client, symbol)
    sym = symbol_upper(symbol)
    if not sym:
        return None
    for u in urls:
        try:
            html = _http_text(client.screener_opener, u, client.screener_headers, timeout=timeout)
        except Exception:
            continue
        rows = _parse_html_table_rows(html)
        if not rows:
            continue
        revenue = _pick_latest_metric(
            rows,
            [
                "sales +",
                "sales",
                "revenue from operations +",
                "revenue from operations",
                "total income",
                "total income +",
                "interest earned",
                "interest income",
                "interest",
            ],
        )
        pat = _pick_latest_metric(rows, ["net profit +", "net profit", "profit after tax", "pat"])
        debt = _pick_latest_metric(rows, ["borrowings +", "borrowings", "debt"])
        fii = _pick_latest_metric(rows, ["fii", "fiis", "fii +", "foreign institutions"])
        dii = _pick_latest_metric(rows, ["dii", "diis", "dii +", "domestic institutions"])
        promoter = _pick_latest_metric(rows, ["promoter", "promoters", "promoters +"])
        if revenue is None and pat is None and debt is None and fii is None and dii is None and promoter is None:
            continue
        fiscal_period, report_date = _infer_financial_period_from_text(html, fallback_date=dt.date.today())
        return {
            "symbol": sym,
            "fiscal_period": fiscal_period,
            "report_date": report_date,
            "revenue": revenue,
            "pat": pat,
            "debt": debt,
            "fii_holding_pct": fii,
            "dii_holding_pct": dii,
            "promoter_holding_pct": promoter,
            "source": "screener_auto",
            "source_ref": u,
            "notes": "Auto-collected from Screener public company page.",
        }
    return None


def _nse_quote_payload(symbol):
    sym = symbol_upper(symbol)
    if not sym:
        return {}
    client = MarketDataClient()
    try:
        client._bootstrap_nse()
        q = urllib.parse.quote(sym)
        url = f"https://www.nseindia.com/api/quote-equity?symbol={q}"
        data = _http_json(client.nse_opener, url, client.nse_headers, timeout=5)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def fetch_nse_financial_docs(symbol, limit=6):
    sym = symbol_upper(symbol)
    if not sym:
        return []
    client = MarketDataClient()
    out = []
    try:
        client._bootstrap_nse()
    except Exception:
        return out
    urls = [
        f"https://www.nseindia.com/api/corporate-announcements?symbol={urllib.parse.quote(sym)}",
        f"https://www.nseindia.com/api/corporates-corporateAnnouncements?symbol={urllib.parse.quote(sym)}",
    ]
    for url in urls:
        data = None
        try:
            data = _http_json(client.nse_opener, url, client.nse_headers, timeout=5)
        except Exception:
            continue
        rows = []
        if isinstance(data, list):
            rows = data
        elif isinstance(data, dict):
            for key in ("data", "announcements", "corporateAnnouncements"):
                if isinstance(data.get(key), list):
                    rows = data.get(key) or []
                    break
        for r in rows[: max(1, min(20, int(limit) * 3))]:
            title = str(
                (r or {}).get("subject")
                or (r or {}).get("headline")
                or (r or {}).get("desc")
                or (r or {}).get("purpose")
                or ""
            ).strip()
            if not title:
                continue
            snippet = str((r or {}).get("attchmntText") or (r or {}).get("details") or "").strip()
            link = str(
                (r or {}).get("attchmntFile")
                or (r or {}).get("attachment")
                or (r or {}).get("pdf")
                or (r or {}).get("pdffile")
                or ""
            ).strip()
            if link and not link.startswith("http"):
                link = urllib.parse.urljoin("https://www.nseindia.com", link)
            pub = str(
                (r or {}).get("an_dt")
                or (r or {}).get("broadcastDateTime")
                or (r or {}).get("date")
                or ""
            ).strip()
            out.append(
                {
                    "title": title[:220],
                    "snippet": snippet[:1200],
                    "link": link,
                    "published_at": pub,
                    "source": "nse_announcements",
                }
            )
            if len(out) >= int(limit):
                return out
    return out[: int(limit)]


def _extract_company_website(payload):
    candidates = []

    def walk(node):
        if isinstance(node, dict):
            for k, v in node.items():
                lk = str(k or "").strip().lower()
                if isinstance(v, str):
                    if ("website" in lk) or (lk in ("url", "companyurl", "site")):
                        candidates.append(v.strip())
                else:
                    walk(v)
        elif isinstance(node, list):
            for it in node:
                walk(it)

    walk(payload if isinstance(payload, dict) else {})
    for c in candidates:
        url = str(c or "").strip()
        if not url:
            continue
        if not url.startswith("http"):
            url = "https://" + url.lstrip("/")
        if "." in url:
            return url
    return ""


def fetch_company_site_ir_docs(symbol, limit=4):
    sym = symbol_upper(symbol)
    if not sym:
        return []
    payload = _nse_quote_payload(sym)
    website = _extract_company_website(payload)
    if not website:
        return []
    opener = urllib.request.build_opener()
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "text/html,*/*;q=0.9"}
    try:
        html = _http_text(opener, website, headers, timeout=5)
    except Exception:
        return []
    out = []
    seen = set()
    for m in re.finditer(
        r"<a[^>]*href=[\"']([^\"']+)[\"'][^>]*>(.*?)</a>",
        html,
        flags=re.IGNORECASE | re.S,
    ):
        href = str(m.group(1) or "").strip()
        label = _clean_rss_text(m.group(2) or "")
        if not href:
            continue
        low = (href + " " + label).lower()
        if not any(k in low for k in ("investor", "financial", "result", "annual", "quarter", "report", "shareholder")):
            continue
        link = urllib.parse.urljoin(website, href)
        if link in seen:
            continue
        seen.add(link)
        out.append(
            {
                "title": (label or "Investor Relations Link")[:220],
                "snippet": f"Company IR link discovered for {sym}",
                "link": link,
                "published_at": dt.date.today().isoformat(),
                "source": "company_site_ir",
            }
        )
        if len(out) >= int(limit):
            break
    return out


def collect_online_financial_data(conn, symbols, sources, max_items=30, max_runtime_sec=30):
    source_set = {str(s or "").strip().lower() for s in (sources or []) if str(s or "").strip()}
    symbols_eff = [symbol_upper(s) for s in (symbols or []) if symbol_upper(s)]
    max_items = max(1, min(200, int(max_items)))
    t0 = time.time()
    stats = {
        "symbols_considered": len(symbols_eff),
        "inserted_financial_rows": 0,
        "updated_financial_rows": 0,
        "inserted_docs": 0,
        "skipped_docs": 0,
        "errors": [],
        "docs": [],
    }
    for sym in symbols_eff:
        if (time.time() - t0) > max(8, int(max_runtime_sec)):
            break

        if "screener_financials" in source_set:
            try:
                snap = fetch_screener_financial_snapshot(sym, timeout=6)
            except Exception as ex:
                snap = None
                if len(stats["errors"]) < 40:
                    stats["errors"].append(f"screener_financials:{sym}:{str(ex)}")
            if snap:
                existed = conn.execute(
                    """
                    SELECT id
                    FROM company_financials
                    WHERE UPPER(symbol)=? AND fiscal_period=? AND source=?
                    LIMIT 1
                    """,
                    (symbol_upper(sym), str(snap.get("fiscal_period") or ""), str(snap.get("source") or "")),
                ).fetchone()
                upsert_company_financial_row(
                    conn,
                    {
                        "symbol": sym,
                        "fiscal_period": snap.get("fiscal_period"),
                        "report_date": snap.get("report_date"),
                        "revenue": snap.get("revenue"),
                        "pat": snap.get("pat"),
                        "debt": snap.get("debt"),
                        "fii_holding_pct": snap.get("fii_holding_pct"),
                        "dii_holding_pct": snap.get("dii_holding_pct"),
                        "promoter_holding_pct": snap.get("promoter_holding_pct"),
                        "source": snap.get("source"),
                        "notes": snap.get("notes"),
                    },
                )
                if existed:
                    stats["updated_financial_rows"] += 1
                else:
                    stats["inserted_financial_rows"] += 1
                title = f"{sym} auto financial snapshot ({snap.get('fiscal_period')})"
                source_ref = str(snap.get("source_ref") or "")
                doc_date = _parse_iso_date_safe(snap.get("report_date"), fallback=dt.date.today()).isoformat()
                content = (
                    f"{title}\n"
                    f"Revenue: {snap.get('revenue')}\n"
                    f"PAT: {snap.get('pat')}\n"
                    f"Debt: {snap.get('debt')}\n"
                    f"FII: {snap.get('fii_holding_pct')}\n"
                    f"DII: {snap.get('dii_holding_pct')}\n"
                    f"Promoter: {snap.get('promoter_holding_pct')}\n"
                    f"Source: {source_ref}"
                )
                if not _doc_already_exists(conn, source_ref, title, doc_date):
                    try:
                        res = analyze_and_store_intelligence_document(
                            conn=conn,
                            doc_type="financial_statement",
                            source="screener_financials",
                            source_ref=source_ref,
                            doc_date=doc_date,
                            title=title,
                            content=content,
                        )
                        stats["inserted_docs"] += 1
                        stats["docs"].append(
                            {
                                "doc_id": res.get("doc_id"),
                                "doc_type": "financial_statement",
                                "title": title[:160],
                                "source_ref": source_ref,
                            }
                        )
                    except Exception as ex:
                        if len(stats["errors"]) < 40:
                            stats["errors"].append(f"screener_doc:{sym}:{str(ex)}")

        news_docs = []
        if "nse_announcements" in source_set:
            try:
                news_docs.extend(fetch_nse_financial_docs(sym, limit=4))
            except Exception as ex:
                if len(stats["errors"]) < 40:
                    stats["errors"].append(f"nse_announcements:{sym}:{str(ex)}")
        if "company_site_ir" in source_set:
            try:
                news_docs.extend(fetch_company_site_ir_docs(sym, limit=3))
            except Exception as ex:
                if len(stats["errors"]) < 40:
                    stats["errors"].append(f"company_site_ir:{sym}:{str(ex)}")
        for it in news_docs:
            link = str(it.get("link") or "").strip()
            title = str(it.get("title") or "").strip()
            doc_date = _parse_iso_date_safe(str(it.get("published_at") or ""), fallback=dt.date.today()).isoformat()
            if _doc_already_exists(conn, link, title, doc_date):
                stats["skipped_docs"] += 1
                continue
            content = f"{title}\n{str(it.get('snippet') or '').strip()}\nsymbol:{sym}".strip()
            doc_type = _infer_doc_type_from_text(f"{title} {content}")
            try:
                res = analyze_and_store_intelligence_document(
                    conn=conn,
                    doc_type=doc_type,
                    source=str(it.get("source") or "online_financial_source"),
                    source_ref=link,
                    doc_date=doc_date,
                    title=title,
                    content=content,
                )
                stats["inserted_docs"] += 1
                stats["docs"].append(
                    {
                        "doc_id": res.get("doc_id"),
                        "doc_type": doc_type,
                        "title": title[:160],
                        "source_ref": link,
                    }
                )
            except Exception as ex:
                if len(stats["errors"]) < 40:
                    stats["errors"].append(f"online_doc:{sym}:{str(ex)}")
            if (stats["inserted_docs"] + stats["inserted_financial_rows"]) >= max_items:
                break

        if (stats["inserted_docs"] + stats["inserted_financial_rows"]) >= max_items:
            break
    stats["docs"] = stats["docs"][: max_items]
    return stats


def maybe_backfill_missing_financial_rows(conn, symbols, force=False):
    syms = []
    seen = set()
    for s in (symbols or []):
        su = symbol_upper(s)
        if not su or su in seen:
            continue
        seen.add(su)
        syms.append(su)
    if not syms:
        return {"executed": False, "reason": "no_symbols"}

    last_run_at = ""
    row = conn.execute(
        "SELECT value FROM app_config WHERE key = 'intel_financial_backfill_last_run_at'"
    ).fetchone()
    if row:
        last_run_at = str(row["value"] or "").strip()
    if (not force) and last_run_at:
        try:
            age = max(0.0, (dt.datetime.now() - dt.datetime.fromisoformat(last_run_at)).total_seconds())
            if age < INTEL_FIN_BACKFILL_MIN_INTERVAL_SEC:
                return {
                    "executed": False,
                    "reason": "cooldown",
                    "last_run_at": last_run_at,
                    "cooldown_seconds": INTEL_FIN_BACKFILL_MIN_INTERVAL_SEC,
                    "age_seconds": round(age, 3),
                }
        except Exception:
            pass

    report_ceiling = _financial_report_date_ceiling()
    ph = ",".join(["?"] * len(syms))
    existing_rows = conn.execute(
        f"""
        SELECT UPPER(symbol) AS symbol, COUNT(*) AS c
        FROM company_financials
        WHERE UPPER(symbol) IN ({ph})
          AND report_date <= ?
        GROUP BY UPPER(symbol)
        """,
        syms + [report_ceiling],
    ).fetchall()
    existing = {symbol_upper(r["symbol"]): int(parse_float(r["c"], 0.0)) for r in existing_rows}
    targets = [s for s in syms if int(existing.get(s, 0)) < 2]
    if not targets:
        return {"executed": False, "reason": "already_populated", "symbols_considered": len(syms)}
    targets = targets[:INTEL_FIN_BACKFILL_MAX_SYMBOLS]

    # Fast-path backfill for financial scores: Screener snapshot is the primary
    # quantitative source and avoids slower IR/news crawling on summary refresh.
    online_fin_sources = ["screener_financials"]

    stats = collect_online_financial_data(
        conn,
        targets,
        sources=online_fin_sources,
        max_items=max(INTEL_FIN_BACKFILL_MAX_SYMBOLS, len(targets)),
        max_runtime_sec=INTEL_FIN_BACKFILL_MAX_RUNTIME_SEC,
    )
    stamp = now_iso()
    conn.execute(
        """
        INSERT INTO app_config(key, value, updated_at) VALUES ('intel_financial_backfill_last_run_at', ?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        """,
        (stamp, stamp),
    )
    conn.commit()
    stats["executed"] = True
    stats["target_symbols"] = targets
    stats["sources"] = online_fin_sources
    stats["last_run_at"] = stamp
    return stats


def prioritized_symbols_for_financial_backfill(conn, limit=200):
    lim = max(1, min(500, int(limit)))
    out = []
    seen = set()

    # 1) Symbols with strongest recent intelligence impact first.
    impact_rows = conn.execute(
        """
        SELECT UPPER(ii.symbol) AS symbol,
               SUM(ABS(COALESCE(ii.impact_score, 0)) * COALESCE(ii.confidence, 0.5)) AS impact_weight
        FROM intelligence_impacts ii
        JOIN intelligence_documents idoc ON idoc.id = ii.doc_id
        WHERE idoc.doc_date >= ?
        GROUP BY UPPER(ii.symbol)
        ORDER BY impact_weight DESC, UPPER(ii.symbol)
        LIMIT 300
        """,
        ((dt.date.today() - dt.timedelta(days=180)).isoformat(),),
    ).fetchall()
    for r in impact_rows:
        sym = symbol_upper(r["symbol"])
        if not sym or sym in seen:
            continue
        seen.add(sym)
        out.append(sym)
        if len(out) >= lim:
            return out

    # 2) Then currently held/high-value symbols.
    holding_rows = conn.execute(
        """
        SELECT UPPER(symbol) AS symbol, COALESCE(market_value, 0) AS mv
        FROM holdings
        ORDER BY mv DESC, UPPER(symbol)
        LIMIT 300
        """
    ).fetchall()
    for r in holding_rows:
        sym = symbol_upper(r["symbol"])
        if not sym or sym in seen:
            continue
        seen.add(sym)
        out.append(sym)
        if len(out) >= lim:
            return out

    # 3) Then recently traded symbols (including currently zero-qty names).
    trade_rows = conn.execute(
        """
        SELECT UPPER(symbol) AS symbol, MAX(trade_date) AS last_trade_date
        FROM trades
        GROUP BY UPPER(symbol)
        ORDER BY last_trade_date DESC, UPPER(symbol)
        LIMIT 400
        """
    ).fetchall()
    for r in trade_rows:
        sym = symbol_upper(r["symbol"])
        if not sym or sym in seen:
            continue
        seen.add(sym)
        out.append(sym)
        if len(out) >= lim:
            return out

    # 4) Fallback: all active instruments.
    active_rows = conn.execute(
        """
        SELECT UPPER(symbol) AS symbol
        FROM instruments
        WHERE active = 1
        ORDER BY UPPER(symbol)
        LIMIT 500
        """
    ).fetchall()
    for r in active_rows:
        sym = symbol_upper(r["symbol"])
        if not sym or sym in seen:
            continue
        seen.add(sym)
        out.append(sym)
        if len(out) >= lim:
            return out
    return out


def run_intelligence_autopilot_once(max_runtime_sec=45, force=False):
    t0 = time.time()
    with db_connect() as conn:
        cfg = get_intel_autopilot_config(conn)
        if (not force) and (not cfg.get("enabled")):
            return {"ok": True, "executed": False, "reason": "disabled"}
        last = str(cfg.get("last_run_at") or "").strip()
        if (not force) and last:
            try:
                dt_last = dt.datetime.fromisoformat(last)
                if (dt.datetime.now() - dt_last).total_seconds() < int(cfg.get("interval_seconds", 12 * 60 * 60)):
                    return {"ok": True, "executed": False, "reason": "interval_not_elapsed"}
            except Exception:
                pass

        seed_lines = [x.strip() for x in str(cfg.get("query_seed") or "").splitlines() if x.strip()]
        symbols = _collect_autopilot_symbols(conn, symbols_limit=int(cfg.get("symbols_limit", 20)))
        symbol_queries = _collect_autopilot_symbol_queries(conn, symbols_limit=int(cfg.get("symbols_limit", 20)))
        queries = []
        seen_q = set()
        for q in (seed_lines + symbol_queries):
            key = q.lower()
            if key in seen_q:
                continue
            seen_q.add(key)
            queries.append(q)
        max_docs = int(cfg.get("max_docs", 24))
        sources = [str(s or "").strip().lower() for s in (cfg.get("sources") or []) if str(s or "").strip()]
        if not sources:
            sources = ["google_news_rss", "screener_financials", "nse_announcements", "company_site_ir"]
        news_sources = [s for s in sources if s in ("google_news_rss",)]
        online_fin_sources = [s for s in sources if s in ("screener_financials", "nse_announcements", "company_site_ir")]

        inserted = 0
        skipped = 0
        errors = []
        seen_links = set()
        docs = []
        for q in queries:
            if (time.time() - t0) > max(10, int(max_runtime_sec)):
                break
            per_query_limit = max(1, min(8, max_docs // 3 if max_docs >= 3 else 1))
            items = []
            for src in news_sources:
                src = str(src or "").strip().lower()
                try:
                    if src == "google_news_rss":
                        items.extend(fetch_google_news_rss(q, limit=per_query_limit, timeout=5))
                    else:
                        # Unrecognized source names are ignored but retained in config for future adapters.
                        continue
                except Exception as ex:
                    if len(errors) < 40:
                        errors.append(f"{src}:{q} -> {str(ex)}")
            for it in items:
                link = str(it.get("link") or "").strip()
                if link and link in seen_links:
                    skipped += 1
                    continue
                if link:
                    seen_links.add(link)
                title = str(it.get("title") or "").strip()
                snippet = str(it.get("snippet") or "").strip()
                content = f"{title}\n{snippet}\nquery:{q}".strip()
                doc_date = _parse_iso_date_safe(str(it.get("published_at") or ""), fallback=dt.date.today()).isoformat()
                if _doc_already_exists(conn, link, title, doc_date):
                    skipped += 1
                    continue
                doc_type = _infer_doc_type_from_text(f"{q} {title} {snippet}")
                try:
                    res = analyze_and_store_intelligence_document(
                        conn=conn,
                        doc_type=doc_type,
                        source=str(it.get("source") or "google_news_rss"),
                        source_ref=link,
                        doc_date=doc_date,
                        title=title,
                        content=content,
                    )
                    inserted += 1
                    docs.append(
                        {
                            "doc_id": res.get("doc_id"),
                            "doc_type": doc_type,
                            "title": title[:160],
                            "source_ref": link,
                        }
                    )
                except Exception as ex:
                    if len(errors) < 40:
                        errors.append(f"ingest:{title[:80]} -> {str(ex)}")
                if inserted >= max_docs:
                    break
            if inserted >= max_docs:
                break

        online_fin_stats = {
            "symbols_considered": 0,
            "inserted_financial_rows": 0,
            "updated_financial_rows": 0,
            "inserted_docs": 0,
            "skipped_docs": 0,
            "errors": [],
            "docs": [],
        }
        if online_fin_sources and (time.time() - t0) < max(10, int(max_runtime_sec)):
            try:
                remaining = max(8, int(max_runtime_sec - (time.time() - t0)))
                online_fin_stats = collect_online_financial_data(
                    conn=conn,
                    symbols=symbols,
                    sources=online_fin_sources,
                    max_items=max(6, max_docs),
                    max_runtime_sec=remaining,
                )
                skipped += int(parse_float(online_fin_stats.get("skipped_docs"), 0.0))
                docs.extend(list(online_fin_stats.get("docs") or []))
                if online_fin_stats.get("errors"):
                    errors.extend([str(e) for e in (online_fin_stats.get("errors") or [])[:40]])
            except Exception as ex:
                if len(errors) < 40:
                    errors.append(f"online_financial_collection:{str(ex)}")

        if inserted > 0 or int(parse_float(online_fin_stats.get("inserted_financial_rows"), 0.0)) > 0 or int(
            parse_float(online_fin_stats.get("inserted_docs"), 0.0)
        ) > 0:
            try:
                refresh_strategy_analytics(force=True)
            except Exception as ex:
                if len(errors) < 40:
                    errors.append(f"strategy_refresh:{str(ex)}")

        set_intel_autopilot_config(last_run_at=now_iso())

        summary = intelligence_summary(conn, limit=20)
        return {
            "ok": True,
            "executed": True,
            "inserted_docs": int(inserted),
            "skipped_docs": int(skipped),
            "errors": errors,
            "sources": sources,
            "queries_considered": len(queries),
            "docs": docs[: max(1, max_docs * 2)],
            "online_financial_collection": online_fin_stats,
            "summary": {
                "portfolio_score": summary.get("portfolio_score"),
                "portfolio_confidence": summary.get("portfolio_confidence"),
                "documents_recent": summary.get("documents_recent"),
                "impacts_recent": summary.get("impacts_recent"),
            },
        }


def maybe_run_intel_autopilot_once():
    try:
        return run_intelligence_autopilot_once(max_runtime_sec=45, force=False)
    except Exception:
        return None


def _safe_json_dumps(value):
    try:
        return json.dumps({} if value is None else value, ensure_ascii=False, sort_keys=True)
    except Exception:
        return json.dumps({}, ensure_ascii=False)


def _safe_json_loads(value, fallback):
    try:
        return json.loads(str(value or ""))
    except Exception:
        return fallback


def get_active_strategy_set_id(conn):
    row = conn.execute("SELECT id FROM strategy_sets WHERE is_active = 1 LIMIT 1").fetchone()
    return int(row["id"]) if row else None


def update_active_strategy_parameters(conn, updates):
    set_id = get_active_strategy_set_id(conn)
    if set_id is None:
        return {}
    before = get_active_params(conn)
    changed = {}
    for key, value in (updates or {}).items():
        if key is None:
            continue
        k = str(key).strip()
        if not k:
            continue
        v = float(value)
        row = conn.execute(
            "SELECT id, value FROM strategy_parameters WHERE set_id = ? AND key = ?",
            (set_id, k),
        ).fetchone()
        if row:
            old_v = parse_float(row["value"], 0.0)
            if abs(old_v - v) > 1e-9:
                conn.execute("UPDATE strategy_parameters SET value = ? WHERE id = ?", (v, int(row["id"])))
                changed[k] = {"old": old_v, "new": v}
        else:
            conn.execute("INSERT INTO strategy_parameters(set_id, key, value) VALUES (?, ?, ?)", (set_id, k, v))
            changed[k] = {"old": None, "new": v}
    conn.commit()
    after = get_active_params(conn)
    return {"set_id": set_id, "before": before, "after": after, "changed": changed}


def latest_backtest_run(conn):
    row = conn.execute(
        """
        SELECT id, created_at, from_date, to_date, horizon_days, sample_count, hit_rate, avg_future_return,
               momentum_hit_rate, intel_hit_rate, applied_tuning, suggestions_json, diagnostics_json, errors_json
        FROM agent_backtest_runs
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    if not row:
        return None
    out = dict(row)
    out["suggestions"] = _safe_json_loads(out.get("suggestions_json"), {})
    out["diagnostics"] = _safe_json_loads(out.get("diagnostics_json"), {})
    out["errors"] = _safe_json_loads(out.get("errors_json"), [])
    return out


def list_backtest_runs(conn, limit=25):
    lim = max(1, min(200, int(limit)))
    rows = conn.execute(
        """
        SELECT id, created_at, from_date, to_date, horizon_days, sample_count, hit_rate, avg_future_return,
               momentum_hit_rate, intel_hit_rate, applied_tuning, suggestions_json, diagnostics_json, errors_json
        FROM agent_backtest_runs
        ORDER BY id DESC
        LIMIT ?
        """,
        (lim,),
    ).fetchall()
    out = []
    for r in rows:
        item = dict(r)
        item["suggestions"] = _safe_json_loads(item.get("suggestions_json"), {})
        item["diagnostics"] = _safe_json_loads(item.get("diagnostics_json"), {})
        item["errors"] = _safe_json_loads(item.get("errors_json"), [])
        out.append(item)
    return out


def build_agents_status(conn):
    live_cfg = get_live_config(conn)
    market_row = conn.execute(
        "SELECT MAX(updated_at) AS ts, COUNT(*) AS c FROM latest_prices"
    ).fetchone()
    strategy_cfg = get_strategy_config(conn)
    strategy_row = conn.execute(
        "SELECT MAX(created_at) AS ts, MAX(run_date) AS run_date FROM strategy_runs"
    ).fetchone()
    history_cfg = get_history_sync_config(conn)
    backup_cfg = get_backup_config(conn)
    self_cfg = get_self_learning_config(conn)
    intel_auto_cfg = get_intel_autopilot_config(conn)
    chart_cfg = get_chart_agent_config(conn)
    perf_cfg = get_software_perf_agent_config(conn)
    risk_cfg = get_risk_agent_config(conn)
    tax_monitor_cfg = get_tax_monitor_config(conn)
    tax_cfg = get_tax_profile_config(conn)
    bt = latest_backtest_run(conn)
    intel_recent = conn.execute(
        """
        SELECT
          COUNT(*) AS docs_recent,
          COALESCE(SUM(CASE WHEN doc_date >= ? THEN 1 ELSE 0 END),0) AS docs_7d
        FROM intelligence_documents
        """,
        ((dt.date.today() - dt.timedelta(days=7)).isoformat(),),
    ).fetchone()
    chart_recent = conn.execute(
        """
        SELECT
          MAX(created_at) AS last_run_at,
          MAX(as_of_date) AS as_of_date,
          COUNT(*) AS c
        FROM chart_analysis_snapshots
        """
    ).fetchone()
    chart_latest = conn.execute(
        """
        SELECT
          COALESCE(SUM(CASE WHEN UPPER(signal) = 'BULLISH' THEN 1 ELSE 0 END),0) AS bullish,
          COALESCE(SUM(CASE WHEN UPPER(signal) = 'BEARISH' THEN 1 ELSE 0 END),0) AS bearish
        FROM (
          SELECT c.symbol, c.signal
          FROM chart_analysis_snapshots c
          JOIN (
            SELECT symbol, MAX(created_at) AS mx
            FROM chart_analysis_snapshots
            GROUP BY symbol
          ) x ON x.symbol = c.symbol AND x.mx = c.created_at
        ) z
        """
    ).fetchone()
    perf_latest = conn.execute(
        """
        SELECT
          created_at,
          issue_count,
          live_stale_symbols,
          live_missing_price_symbols,
          weak_sources_count,
          avg_quote_latency_ms,
          quote_success_rate
        FROM software_perf_snapshots
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    risk_latest = conn.execute(
        """
        SELECT
          created_at,
          symbols_analyzed,
          portfolio_volatility,
          max_drawdown,
          avg_pair_correlation,
          concentration_hhi,
          risk_score,
          risk_level
        FROM risk_analysis_snapshots
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()

    hist_symbols = 0
    hist_to = ""
    try:
        with market_db_connect() as mconn:
            hist = mconn.execute(
                "SELECT COUNT(DISTINCT symbol) AS c, MAX(price_date) AS to_date FROM daily_prices"
            ).fetchone()
            hist_symbols = int(hist["c"] or 0)
            hist_to = str(hist["to_date"] or "")
    except Exception:
        pass

    return [
        {
            "agent": "market",
            "label": "Live Quote Agent",
            "enabled": bool(live_cfg.get("enabled")),
            "interval_seconds": int(live_cfg.get("interval_seconds", LIVE_REFRESH_MIN_SEC)),
            "last_run_at": market_row["ts"],
            "status": "running" if bool(live_cfg.get("enabled")) else "paused",
            "details": f"latest prices for {int(market_row['c'] or 0)} symbols",
        },
        {
            "agent": "history",
            "label": "History Backfill Agent",
            "enabled": bool(history_cfg.get("enabled")),
            "interval_seconds": int(history_cfg.get("interval_seconds", HISTORY_REFRESH_DEFAULT_SEC)),
            "last_run_at": history_cfg.get("last_sync_at") or "",
            "status": "running" if bool(history_cfg.get("enabled")) else "paused",
            "details": f"symbols={hist_symbols}, latest_date={hist_to or '-'}",
        },
        {
            "agent": "strategy",
            "label": "Strategy Agent",
            "enabled": bool(strategy_cfg.get("enabled")),
            "interval_seconds": int(strategy_cfg.get("interval_seconds", STRATEGY_REFRESH_MIN_SEC)),
            "last_run_at": strategy_row["ts"],
            "status": "running" if bool(strategy_cfg.get("enabled")) else "paused",
            "details": f"latest_run_date={strategy_row['run_date'] or '-'}",
        },
        {
            "agent": "backup",
            "label": "DB Backup Agent",
            "enabled": bool(backup_cfg.get("enabled")),
            "interval_seconds": int(backup_cfg.get("interval_seconds", DB_BACKUP_INTERVAL_SEC)),
            "last_run_at": backup_cfg.get("last_run_at") or "",
            "status": "running" if bool(backup_cfg.get("enabled")) else "paused",
            "details": backup_cfg.get("last_file") or "no backup yet",
        },
        {
            "agent": "self_learning",
            "label": "Self-Learning Agent",
            "enabled": bool(self_cfg.get("enabled")),
            "interval_seconds": int(self_cfg.get("interval_days", 7) * 86400),
            "last_run_at": self_cfg.get("last_run_at") or "",
            "status": "running" if bool(self_cfg.get("enabled")) else "paused",
            "details": (
                f"latest_backtest_hit_rate={round(parse_float((bt or {}).get('hit_rate'), 0.0) * 100.0, 1)}%, "
                f"samples={int(parse_float((bt or {}).get('sample_count'), 0.0))}"
                if bt
                else f"min_samples={int(self_cfg.get('min_samples', 30))}"
            ),
        },
        {
            "agent": "intel_autopilot",
            "label": "Intelligence Autopilot Agent",
            "enabled": bool(intel_auto_cfg.get("enabled")),
            "interval_seconds": int(intel_auto_cfg.get("interval_seconds", 12 * 60 * 60)),
            "last_run_at": intel_auto_cfg.get("last_run_at") or "",
            "status": "running" if bool(intel_auto_cfg.get("enabled")) else "paused",
            "details": (
                f"sources={','.join(intel_auto_cfg.get('sources', []))}; "
                f"max_docs={int(intel_auto_cfg.get('max_docs', 24))}; "
                f"docs_total={int(parse_float(intel_recent['docs_recent'], 0.0))}; "
                f"docs_7d={int(parse_float(intel_recent['docs_7d'], 0.0))}"
            ),
        },
        {
            "agent": "chart_intel",
            "label": "Chart Pattern Agent",
            "enabled": bool(chart_cfg.get("enabled")),
            "interval_seconds": int(chart_cfg.get("interval_seconds", CHART_AGENT_INTERVAL_DEFAULT_SEC)),
            "last_run_at": chart_cfg.get("last_run_at") or chart_recent["last_run_at"] or "",
            "status": "running" if bool(chart_cfg.get("enabled")) else "paused",
            "details": (
                f"sources={','.join(chart_cfg.get('sources', []))}; "
                f"snapshots={int(parse_float(chart_recent['c'], 0.0))}; "
                f"bullish={int(parse_float(chart_latest['bullish'], 0.0))}; "
                f"bearish={int(parse_float(chart_latest['bearish'], 0.0))}; "
                f"as_of={chart_recent['as_of_date'] or '-'}"
            ),
        },
        {
            "agent": "software_performance",
            "label": "Software Performance Agent",
            "enabled": bool(perf_cfg.get("enabled")),
            "interval_seconds": int(perf_cfg.get("interval_seconds", SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC)),
            "last_run_at": perf_cfg.get("last_run_at") or (perf_latest["created_at"] if perf_latest else ""),
            "status": "running" if bool(perf_cfg.get("enabled")) else "paused",
            "details": (
                f"issues={int(parse_float(perf_latest['issue_count'], 0.0))}; "
                f"stale={int(parse_float(perf_latest['live_stale_symbols'], 0.0))}; "
                f"missing={int(parse_float(perf_latest['live_missing_price_symbols'], 0.0))}; "
                f"weak_sources={int(parse_float(perf_latest['weak_sources_count'], 0.0))}; "
                f"latency_ms={round(parse_float(perf_latest['avg_quote_latency_ms'], 0.0), 1)}; "
                f"success={round(parse_float(perf_latest['quote_success_rate'], 0.0) * 100.0, 1)}%; "
                f"llm={str(perf_cfg.get('llm_status') or '-').lower()}; "
                f"retention={int(parse_float(perf_cfg.get('retention_days'), 90))}d; "
                f"last_cleanup={perf_cfg.get('last_cleanup_at') or '-'}"
                if perf_latest
                else (
                    "no snapshots yet; tracks runtime health, applies safe self-heal, writes LLM-guided reviewed draft improvements"
                )
            ),
        },
        {
            "agent": "risk_analysis",
            "label": "Risk Analysis Agent",
            "enabled": bool(risk_cfg.get("enabled")),
            "interval_seconds": int(risk_cfg.get("interval_seconds", RISK_AGENT_INTERVAL_DEFAULT_SEC)),
            "last_run_at": risk_cfg.get("last_run_at") or (risk_latest["created_at"] if risk_latest else ""),
            "status": "running" if bool(risk_cfg.get("enabled")) else "paused",
            "details": (
                f"risk_level={risk_latest['risk_level']}; "
                f"risk_score={round(parse_float(risk_latest['risk_score'], 0.0), 1)}; "
                f"symbols={int(parse_float(risk_latest['symbols_analyzed'], 0.0))}; "
                f"vol={round(parse_float(risk_latest['portfolio_volatility'], 0.0) * 100.0, 2)}%; "
                f"mdd={round(parse_float(risk_latest['max_drawdown'], 0.0) * 100.0, 2)}%; "
                f"corr={round(parse_float(risk_latest['avg_pair_correlation'], 0.0), 3)}; "
                f"hhi={round(parse_float(risk_latest['concentration_hhi'], 0.0), 3)}"
                if risk_latest
                else "no snapshots yet; computes volatility, drawdown, VaR/CVaR, correlation and concentration risk"
            ),
        },
        {
            "agent": "tax_monitor",
            "label": "Tax Rate Monitor",
            "enabled": bool(tax_monitor_cfg.get("enabled")),
            "interval_seconds": int(tax_monitor_cfg.get("interval_seconds", TAX_MONITOR_INTERVAL_DEFAULT_SEC)),
            "last_run_at": tax_monitor_cfg.get("last_run_at") or (tax_latest["created_at"] if tax_latest else ""),
            "status": "running" if bool(tax_monitor_cfg.get("enabled")) else "paused",
            "details": (
                f"STCG={round(parse_float(tax_cfg.get('stcg_rate_pct'), 0.0), 2)}%; "
                f"LTCG={round(parse_float(tax_cfg.get('ltcg_rate_pct'), 0.0), 2)}%; "
                f"Exemption={money(tax_cfg.get('ltcg_exemption_limit'))}; "
                f"last_success={tax_monitor_cfg.get('last_success_at') or '-'}; "
                f"last_error={tax_monitor_cfg.get('last_error') or '-'}"
            ),
        },
    ]


def run_tax_rate_monitor_once(conn=None, force=False, timeout=8):
    owns_conn = conn is None
    cm = None
    if owns_conn:
        cm = db_connect()
        conn = cm.__enter__()
    try:
        cfg = get_tax_monitor_config(conn)
        last_run_dt = _parse_iso_datetime_safe(cfg.get("last_run_at"))
        if (not force) and last_run_dt is not None:
            age = max(0.0, (dt.datetime.now() - last_run_dt.replace(tzinfo=None) if last_run_dt.tzinfo else dt.datetime.now() - last_run_dt).total_seconds())
            if age < int(cfg.get("interval_seconds", TAX_MONITOR_INTERVAL_DEFAULT_SEC)):
                return {"ok": True, "executed": False, "reason": "interval_not_elapsed", "config": cfg}
        opener = urllib.request.build_opener()
        headers = {"User-Agent": "Mozilla/5.0", "Accept": "text/html,*/*;q=0.9"}
        run_at = now_iso()
        try:
            tax_text = _http_text(opener, cfg.get("tax_source_url"), headers, timeout=max(4, int(timeout)))
            charges_text = _http_text(opener, cfg.get("charges_source_url"), headers, timeout=max(4, int(timeout)))
            observed = _parse_tax_monitor_snapshot(tax_text, charges_text)
            current = get_tax_profile_config(conn)
            changes = []
            for key in (
                "stcg_rate_pct",
                "ltcg_rate_pct",
                "ltcg_exemption_limit",
                "txn_rate_nse",
                "txn_rate_bse",
                "stt_delivery_rate",
                "stamp_buy_rate",
                "gst_rate",
                "dp_charge_sell_incl_gst",
            ):
                old_v = parse_float(current.get(key), 0.0)
                new_v = parse_float(observed.get(key), 0.0)
                tol = 0.0000005 if "rate" in key else 0.005
                if abs(old_v - new_v) > tol:
                    changes.append({"field": key, "old": old_v, "new": new_v})
            set_tax_profile_config(
                conn,
                stcg_rate_pct=observed.get("stcg_rate_pct"),
                ltcg_rate_pct=observed.get("ltcg_rate_pct"),
                ltcg_exemption_limit=observed.get("ltcg_exemption_limit"),
                txn_rate_nse=observed.get("txn_rate_nse"),
                txn_rate_bse=observed.get("txn_rate_bse"),
                stt_delivery_rate=observed.get("stt_delivery_rate"),
                stamp_buy_rate=observed.get("stamp_buy_rate"),
                gst_rate=observed.get("gst_rate"),
                dp_charge_sell_incl_gst=observed.get("dp_charge_sell_incl_gst"),
            )
            set_tax_monitor_config(
                conn,
                last_run_at=run_at,
                last_success_at=run_at,
                last_error="",
                last_change_at=(run_at if changes else None),
            )
            _insert_tax_rate_sync_run(
                conn,
                "success",
                "zerodha_tax_monitor",
                source_url=cfg.get("tax_source_url"),
                snapshot=observed,
                detail=("changes=" + json.dumps(changes, ensure_ascii=True)) if changes else "no_change",
            )
            if changes:
                upsert_attention_alert(
                    conn,
                    "TAX_RATE_SOURCE_CHANGED",
                    "tax_monitor",
                    92,
                    "critical",
                    "Tax or charge rates changed from monitored source",
                    detail="; ".join([f"{c['field']}: {c['old']} -> {c['new']}" for c in changes[:8]]),
                    source_ref=cfg.get("tax_source_url") or "",
                    meta={"changes": changes},
                )
            else:
                resolve_attention_alert(conn, "TAX_RATE_SOURCE_CHANGED", detail="Latest monitored source matches stored rates.")
            resolve_attention_alert(conn, "TAX_RATE_SYNC_FAILED", detail="Tax rate monitor source refresh succeeded.")
            resolve_attention_alert(conn, "TAX_RATE_MONITOR_STALE", detail="Tax rate monitor is receiving fresh source updates.")
            conn.commit()
            return {"ok": True, "executed": True, "changed": bool(changes), "changes": changes, "observed": observed, "config": get_tax_monitor_config(conn)}
        except Exception as ex:
            err = str(ex)
            set_tax_monitor_config(conn, last_run_at=run_at, last_error=err)
            _insert_tax_rate_sync_run(
                conn,
                "error",
                "zerodha_tax_monitor",
                source_url=(cfg.get("tax_source_url") or ""),
                detail="fetch_or_parse_failed",
                error=err,
            )
            upsert_attention_alert(
                conn,
                "TAX_RATE_SYNC_FAILED",
                "tax_monitor",
                96,
                "critical",
                "Tax rate monitor failed to refresh source rates",
                detail=err,
                source_ref=(cfg.get("tax_source_url") or ""),
                meta={"charges_source_url": cfg.get("charges_source_url")},
            )
            conn.commit()
            return {"ok": False, "executed": True, "error": err, "config": get_tax_monitor_config(conn)}
    finally:
        if owns_conn and cm is not None:
            cm.__exit__(None, None, None)


def refresh_attention_alerts(conn):
    tax_cfg = get_tax_profile_config(conn)
    tax_monitor_cfg = get_tax_monitor_config(conn)
    repo_cfg = get_repo_sync_config(conn)
    llm_cfg = get_llm_runtime_config(conn, include_secret=False)
    perf_latest_row = conn.execute(
        """
        SELECT created_at, issue_count, live_stale_symbols, live_missing_price_symbols, weak_sources_count
        FROM software_perf_snapshots
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    perf_latest = dict(perf_latest_row) if perf_latest_row else {}

    last_success_dt = _parse_iso_datetime_safe(tax_monitor_cfg.get("last_success_at"))
    stale_threshold = max(int(tax_monitor_cfg.get("interval_seconds", TAX_MONITOR_INTERVAL_DEFAULT_SEC)) * 2, 2 * 24 * 60 * 60)
    if last_success_dt is None or (dt.datetime.now() - (last_success_dt.replace(tzinfo=None) if last_success_dt.tzinfo else last_success_dt)).total_seconds() > stale_threshold:
        upsert_attention_alert(
            conn,
            "TAX_RATE_MONITOR_STALE",
            "tax_monitor",
            78,
            "warning",
            "Tax rate monitor has not refreshed recently",
            detail=f"Last success: {tax_monitor_cfg.get('last_success_at') or '-'}; interval={int(tax_monitor_cfg.get('interval_seconds', TAX_MONITOR_INTERVAL_DEFAULT_SEC))} sec.",
            source_ref=tax_monitor_cfg.get("tax_source_url") or "",
        )
    else:
        resolve_attention_alert(conn, "TAX_RATE_MONITOR_STALE")

    if str(repo_cfg.get("last_error") or "").strip():
        upsert_attention_alert(
            conn,
            "REPO_SYNC_ERROR",
            "repo_sync",
            58,
            "warning",
            "Repository sync reported an error",
            detail=str(repo_cfg.get("last_error") or ""),
        )
    else:
        resolve_attention_alert(conn, "REPO_SYNC_ERROR")

    if str(llm_cfg.get("last_status") or "").lower() == "error" and str(llm_cfg.get("last_error") or "").strip():
        upsert_attention_alert(
            conn,
            "LLM_RUNTIME_ERROR",
            "llm",
            52,
            "warning",
            "LLM runtime reported an error",
            detail=str(llm_cfg.get("last_error") or ""),
        )
    else:
        resolve_attention_alert(conn, "LLM_RUNTIME_ERROR")

    perf_issue_count = int(parse_float(perf_latest.get("issue_count"), 0.0)) if perf_latest else 0
    if perf_issue_count > 0:
        upsert_attention_alert(
            conn,
            "SOFTWARE_PERF_OPEN_ISSUES",
            "software_performance",
            min(72, 40 + perf_issue_count),
            "warning",
            "Software Performance Agent still sees open issues",
            detail=(
                f"issues={perf_issue_count}; stale={int(parse_float(perf_latest.get('live_stale_symbols'), 0.0))}; "
                f"missing={int(parse_float(perf_latest.get('live_missing_price_symbols'), 0.0))}; "
                f"weak_sources={int(parse_float(perf_latest.get('weak_sources_count'), 0.0))}"
            ),
            meta={"created_at": str(perf_latest.get("created_at") or "")},
        )
    else:
        resolve_attention_alert(conn, "SOFTWARE_PERF_OPEN_ISSUES")

    return {
        "tax_profile": tax_cfg,
        "tax_monitor": tax_monitor_cfg,
    }


def build_attention_console_payload(conn):
    ctx = refresh_attention_alerts(conn)
    alerts = list_attention_alerts(conn, status=None, limit=120)
    open_alerts = [x for x in alerts if str(x.get("status") or "").lower() == "open"]
    resolved_alerts = [x for x in alerts if str(x.get("status") or "").lower() == "resolved"][:25]
    sev_counts = defaultdict(int)
    for item in open_alerts:
        sev_counts[str(item.get("severity_label") or "info").lower()] += 1
    latest_sync = conn.execute(
        """
        SELECT created_at, status, detail, error
        FROM tax_rate_sync_runs
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    latest_sync_item = dict(latest_sync) if latest_sync else {}
    return {
        "summary": {
            "open_count": len(open_alerts),
            "resolved_count": len(resolved_alerts),
            "critical_open": sev_counts.get("critical", 0),
            "warning_open": sev_counts.get("warning", 0),
            "info_open": sev_counts.get("info", 0),
            "latest_tax_sync_at": str(latest_sync_item.get("created_at") or ""),
            "latest_tax_sync_status": str(latest_sync_item.get("status") or ""),
        },
        "tax_profile": {
            **dict(ctx.get("tax_profile") or {}),
            **compute_realized_equity_tax_summary(conn),
        },
        "tax_monitor": dict(ctx.get("tax_monitor") or {}),
        "open_alerts": open_alerts,
        "resolved_alerts": resolved_alerts,
        "tax_sync_runs": list_tax_rate_sync_runs(conn, limit=25),
    }


def get_active_params(conn):
    row = conn.execute("SELECT id FROM strategy_sets WHERE is_active = 1 LIMIT 1").fetchone()
    if not row:
        return {}
    rows = conn.execute(
        "SELECT key, value FROM strategy_parameters WHERE set_id = ?", (row["id"],)
    ).fetchall()
    return {r["key"]: r["value"] for r in rows}


def load_split_map(conn):
    rows = conn.execute(
        """
        SELECT symbol, effective_date, factor
        FROM corporate_actions
        WHERE action_type = 'SPLIT'
        ORDER BY effective_date
        """
    ).fetchall()
    split_map = defaultdict(list)
    for r in rows:
        split_map[symbol_upper(r["symbol"])].append((r["effective_date"], float(r["factor"])))
    return split_map


def ensure_peak_split_reviews_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS peak_split_reviews (
          corporate_action_id INTEGER PRIMARY KEY,
          decision TEXT NOT NULL CHECK (decision IN ('apply','ignore')),
          decided_at TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS ix_peak_split_reviews_decision ON peak_split_reviews(decision)")


def load_peak_split_map(conn):
    ensure_peak_split_reviews_table(conn)
    rows = conn.execute(
        """
        SELECT ca.symbol, ca.effective_date, ca.factor
        FROM corporate_actions ca
        JOIN peak_split_reviews psr ON psr.corporate_action_id = ca.id
        WHERE ca.action_type = 'SPLIT' AND LOWER(psr.decision) = 'apply'
        ORDER BY ca.effective_date
        """
    ).fetchall()
    split_map = defaultdict(list)
    for r in rows:
        split_map[symbol_upper(r["symbol"])].append((r["effective_date"], float(r["factor"])))
    return split_map


def adjusted_trade_values(symbol, trade_date, quantity, price, split_map):
    symbol_q = symbol_upper(symbol)
    factor = 1.0
    for eff_date, mult in split_map.get(symbol_q, []):
        if trade_date < eff_date:
            factor *= mult
    if factor <= 0:
        factor = 1.0
    adj_qty = float(quantity) * factor
    adj_price = float(price) / factor
    return adj_qty, adj_price


def peak_traded_metrics(conn, symbol, split_map=None):
    if split_map is None:
        split_map = load_peak_split_map(conn)
    symbol_q = symbol_upper(symbol)
    rows = conn.execute(
        "SELECT trade_date, quantity, price FROM trades WHERE UPPER(symbol) = ? AND side IN ('BUY','SELL')",
        (symbol_q,),
    ).fetchall()
    peak = 0.0
    for r in rows:
        _, p = adjusted_trade_values(
            symbol,
            r["trade_date"],
            float(r["quantity"]),
            float(r["price"]),
            split_map,
        )
        if p > peak:
            peak = p
    ltp = get_effective_ltp(conn, symbol, split_map)
    pct_diff = ((ltp - peak) / peak * 100.0) if peak > 0 else 0.0
    return {
        "peak_traded_price": peak,
        "pct_from_peak_traded": pct_diff,
        # Backward-compat aliases for existing UI/API consumers.
        "peak_buy_price": peak,
        "pct_from_peak_buy": pct_diff,
    }


def peak_buy_metrics(conn, symbol, split_map=None):
    return peak_traded_metrics(conn, symbol, split_map=split_map)


def pending_peak_split_candidates(conn, symbol=None):
    ensure_peak_split_reviews_table(conn)
    where = ["ca.action_type = 'SPLIT'", "psr.corporate_action_id IS NULL"]
    params = []
    if symbol:
        where.append("UPPER(ca.symbol) = ?")
        params.append(symbol_upper(symbol))
    rows = conn.execute(
        f"""
        SELECT
          ca.id,
          ca.symbol,
          ca.effective_date,
          ca.factor,
          ca.note,
          COUNT(t.id) AS buys_before_split,
          MIN(t.trade_date) AS first_buy_date,
          MAX(t.trade_date) AS last_buy_date_before_split
        FROM corporate_actions ca
        LEFT JOIN peak_split_reviews psr ON psr.corporate_action_id = ca.id
        LEFT JOIN trades t
          ON UPPER(t.symbol) = UPPER(ca.symbol)
         AND t.side = 'BUY'
         AND t.trade_date < ca.effective_date
        WHERE {' AND '.join(where)}
        GROUP BY ca.id, ca.symbol, ca.effective_date, ca.factor, ca.note
        HAVING COUNT(t.id) > 0
        ORDER BY ca.symbol, ca.effective_date DESC, ca.id DESC
        """,
        params,
    ).fetchall()
    out = []
    for r in rows:
        eff = str(r["effective_date"] or "")
        first_buy = str(r["first_buy_date"] or "")
        out.append(
            {
                "id": int(r["id"]),
                "symbol": symbol_upper(r["symbol"]),
                "effective_date": eff,
                "factor": round(parse_float(r["factor"], 1.0), 6),
                "note": str(r["note"] or ""),
                "split_year": eff[:4] if len(eff) >= 4 else "",
                "buys_before_split": int(r["buys_before_split"] or 0),
                "first_buy_date": first_buy,
                "first_buy_year": first_buy[:4] if len(first_buy) >= 4 else "",
                "last_buy_date_before_split": str(r["last_buy_date_before_split"] or ""),
            }
        )
    return out


def get_last_trade_snapshot(conn, symbol):
    symbol_q = symbol_upper(symbol)
    row = conn.execute(
        """
        SELECT trade_date, quantity, price
        FROM trades
        WHERE UPPER(symbol) = ?
        ORDER BY trade_date DESC, id DESC
        LIMIT 1
        """,
        (symbol_q,),
    ).fetchone()
    tax_latest = conn.execute(
        """
        SELECT created_at, status, detail, error
        FROM tax_rate_sync_runs
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    if not row:
        return None
    return {
        "trade_date": str(row["trade_date"] or ""),
        "quantity": parse_float(row["quantity"], 0.0),
        "price": parse_float(row["price"], 0.0),
    }


def get_last_trade_price(conn, symbol, split_map=None):
    if split_map is None:
        split_map = load_split_map(conn)
    snap = get_last_trade_snapshot(conn, symbol)
    if not snap:
        return 0.0
    _, adj_price = adjusted_trade_values(
        symbol,
        snap["trade_date"],
        float(snap["quantity"]),
        float(snap["price"]),
        split_map,
    )
    return float(adj_price) if adj_price > 0 else 0.0


def get_effective_ltp(conn, symbol, split_map=None):
    row = conn.execute(
        "SELECT ltp FROM latest_prices WHERE UPPER(symbol) = ? ORDER BY updated_at DESC LIMIT 1",
        (symbol_upper(symbol),),
    ).fetchone()
    if row:
        v = parse_float(row["ltp"], 0.0)
        if v > 0:
            return v
    return get_last_trade_price(conn, symbol, split_map)


def get_effective_ltp_for_asset(conn, symbol, asset_class=None, split_map=None):
    sym = symbol_upper(symbol)
    ac = asset_class
    if ac is None:
        row = conn.execute(
            "SELECT COALESCE(asset_class, 'EQUITY') AS asset_class FROM instruments WHERE UPPER(symbol)=? LIMIT 1",
            (sym,),
        ).fetchone()
        ac = row["asset_class"] if row else ASSET_CLASS_EQUITY
    ac_norm = normalize_asset_class(ac, fallback=infer_asset_class(symbol=symbol, name=symbol))
    if ac_norm == ASSET_CLASS_GOLD:
        row = conn.execute(
            "SELECT ltp FROM latest_prices WHERE UPPER(symbol) = ? ORDER BY updated_at DESC LIMIT 1",
            (sym,),
        ).fetchone()
        px = parse_float(row["ltp"], 0.0) if row else 0.0
        return px if px > 0 else 0.0
    return get_effective_ltp(conn, symbol, split_map=split_map)


def latest_tick_source_map(conn, symbols=None):
    syms = sorted({symbol_upper(s) for s in (symbols or []) if symbol_upper(s)})
    params = []
    where = ""
    if syms:
        placeholders = ",".join(["?"] * len(syms))
        where = f"WHERE UPPER(symbol) IN ({placeholders})"
        params = syms
    rows = conn.execute(
        f"""
        SELECT pt.symbol, pt.source, pt.fetched_at
        FROM price_ticks pt
        JOIN (
          SELECT UPPER(symbol) AS sym, MAX(id) AS max_id
          FROM price_ticks
          {where}
          GROUP BY UPPER(symbol)
        ) mx ON mx.max_id = pt.id
        """,
        params,
    ).fetchall()
    out = {}
    for r in rows:
        out[symbol_upper(r["symbol"])] = {
            "source": str(r["source"] or ""),
            "fetched_at": str(r["fetched_at"] or ""),
        }
    return out


def refresh_holdings_mark_to_market(conn):
    split_map = load_split_map(conn)
    rows = conn.execute(
        """
        SELECT
          h.symbol,
          COALESCE(h.qty, 0) AS qty,
          COALESCE(h.invested, 0) AS invested,
          COALESCE(h.realized_pnl, 0) AS realized_pnl,
          COALESCE(i.asset_class, 'EQUITY') AS asset_class
        FROM holdings h
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(h.symbol)
        """
    ).fetchall()
    ts = now_iso()
    updates = []
    for r in rows:
        symbol = symbol_upper(r["symbol"])
        qty = parse_float(r["qty"], 0.0)
        invested = parse_float(r["invested"], 0.0)
        realized = parse_float(r["realized_pnl"], 0.0)
        asset_class = normalize_asset_class(
            r["asset_class"],
            fallback=infer_asset_class(symbol=symbol, name=symbol),
        )
        ltp = get_effective_ltp_for_asset(conn, symbol, asset_class=asset_class, split_map=split_map)
        market_value = qty * ltp
        unrealized = market_value - invested
        base = invested if invested > 0 else max(abs(realized) + abs(unrealized), 1.0)
        total_return_pct = ((realized + unrealized) / base * 100.0) if base > 0 else 0.0
        updates.append((market_value, unrealized, total_return_pct, ts, symbol))
    if updates:
        conn.executemany(
            """
            UPDATE holdings
            SET market_value = ?, unrealized_pnl = ?, total_return_pct = ?, updated_at = ?
            WHERE UPPER(symbol) = ?
            """,
            updates,
        )
    return len(updates)


def load_prev_close_map(conn, symbols, as_of_date=None, lookback_days=20):
    syms = sorted({symbol_upper(s) for s in (symbols or []) if symbol_upper(s)})
    if not syms:
        return {}
    try:
        as_of = dt.date.fromisoformat(str(as_of_date or dt.date.today().isoformat())[:10])
    except Exception:
        as_of = dt.date.today()
    from_s = (as_of - dt.timedelta(days=max(5, int(lookback_days)))).isoformat()
    as_of_s = as_of.isoformat()

    rows = []
    with market_db_connect() as mconn:
        for i in range(0, len(syms), 400):
            chunk = syms[i : i + 400]
            placeholders = ",".join(["?"] * len(chunk))
            rows.extend(
                mconn.execute(
                    f"""
                    SELECT symbol, price_date, close
                    FROM daily_prices
                    WHERE UPPER(symbol) IN ({placeholders})
                      AND price_date <= ?
                      AND price_date >= ?
                      AND close > 0
                                          ORDER BY symbol, price_date DESC
                    """,
                    chunk + [as_of_s, from_s],
                ).fetchall()
            )

    by_symbol = defaultdict(list)
    for r in rows:
        sym = symbol_upper(r["symbol"])
        if len(by_symbol[sym]) >= 2:
            continue
        day_s = str(r["price_date"] or "")
        close = parse_float(r["close"], 0.0)
        if day_s and close > 0:
            by_symbol[sym].append((day_s, close))

    prev_close = {}
    for sym in syms:
        series = by_symbol.get(sym, [])
        px = 0.0
        if series:
            if series[0][0] >= as_of_s:
                if len(series) >= 2:
                    px = series[1][1]
            else:
                px = series[0][1]
        if px > 0:
            prev_close[sym] = px
    return prev_close


def intraday_change_abs_from_ticks(conn, symbol, current_ltp=None, as_of_date=None):
    if conn is None:
        return False, 0.0
    sym = symbol_upper(symbol)
    if not sym:
        return False, 0.0
    as_of_s = str(as_of_date or dt.date.today().isoformat())[:10]
    rows = conn.execute(
        """
        SELECT ltp
        FROM price_ticks
        WHERE UPPER(symbol) = ?
          AND SUBSTR(fetched_at,1,10) = ?
          AND ltp > 0
        ORDER BY fetched_at ASC
        """,
        (sym, as_of_s),
    ).fetchall()
    if len(rows) < 2:
        return False, 0.0
    last = parse_float(current_ltp, 0.0)
    if last <= 0:
        last = parse_float(rows[-1]["ltp"], 0.0)
    first = first_plausible_intraday_ltp([r["ltp"] for r in rows], last)
    if first <= 0 or last <= 0:
        return False, 0.0
    return True, (last - first)


def is_plausible_day_reference_price(current_ltp, reference_price):
    cur = parse_float(current_ltp, 0.0)
    ref = parse_float(reference_price, 0.0)
    if cur <= 0 or ref <= 0:
        return False
    ratio = cur / ref
    return 0.25 <= ratio <= 4.0


def first_plausible_intraday_ltp(prices, current_ltp):
    cur = parse_float(current_ltp, 0.0)
    if cur <= 0:
        return 0.0
    for px in prices or []:
        p = parse_float(px, 0.0)
        if p > 0 and is_plausible_day_reference_price(cur, p):
            return p
    return 0.0


def resolve_preferred_equity_day_change_abs(exchange, selected_ltp, selected_change_abs, candidates):
    ex = str(exchange or "").strip().upper()
    ltp_v = parse_float(selected_ltp, 0.0)
    raw_change = parse_float(selected_change_abs, 0.0)
    if ltp_v <= 0:
        return 0.0
    preferred_sources = []
    if ex == "NSE":
        preferred_sources = ["nse_api", "stock_nse_india_api", "nsetools_api"]
    elif ex == "BSE":
        preferred_sources = ["bse_api"]
    else:
        preferred_sources = ["nse_api", "bse_api", "stock_nse_india_api", "nsetools_api"]

    for src in preferred_sources:
        for cand in candidates or []:
            csrc = str(cand.get("source") or "").strip().lower()
            if csrc != src:
                continue
            c_ltp = parse_float(cand.get("ltp"), 0.0)
            c_change = parse_float(cand.get("change_abs"), 0.0)
            c_prev_close = c_ltp - c_change
            if c_ltp > 0 and abs(c_change) > 1e-9 and c_prev_close > 0 and is_plausible_day_reference_price(ltp_v, c_prev_close):
                return ltp_v - c_prev_close

    raw_prev_close = ltp_v - raw_change
    if abs(raw_change) > 1e-9 and raw_prev_close > 0 and is_plausible_day_reference_price(ltp_v, raw_prev_close):
        return raw_change
    return 0.0


def resolve_effective_change_abs(conn, symbol, ltp, change_abs, prev_close_map=None):
    # Prefer source-provided day change from online feed.
    ltp_v = parse_float(ltp, 0.0)
    raw_change = parse_float(change_abs, 0.0)
    if ltp_v <= 0:
        return 0.0
    prev_close = ltp_v - raw_change
    if abs(raw_change) > 1e-9 and prev_close > 0 and is_plausible_day_reference_price(ltp_v, prev_close):
        return raw_change

    # Fallback 1: previous close from market history DB (exchange-fed EOD), not invested value.
    sym = symbol_upper(symbol)
    prev_map = prev_close_map or {}
    prev_close = parse_float(prev_map.get(sym), 0.0)
    if prev_close <= 0 and conn is not None and sym:
        prev_close = parse_float(load_prev_close_map(conn, [sym]).get(sym), 0.0)
    if prev_close > 0 and is_plausible_day_reference_price(ltp_v, prev_close):
        return ltp_v - prev_close

    # Fallback 2: same-day online tick movement (first tick -> latest tick).
    ok_tick, tick_change = intraday_change_abs_from_ticks(conn, symbol, current_ltp=ltp_v)
    if ok_tick:
        return tick_change
    return 0.0


def ensure_latest_prices_nonzero_from_last_trade():
    with db_connect() as conn:
        symbols = [
            {
                "symbol": r["symbol"],
                "asset_class": normalize_asset_class(
                    r["asset_class"],
                    fallback=infer_asset_class(symbol=r["symbol"], name=r["name"]),
                ),
            }
            for r in conn.execute(
                "SELECT symbol, COALESCE(asset_class, 'EQUITY') AS asset_class, name FROM instruments WHERE active = 1"
            ).fetchall()
        ]
        split_map = load_split_map(conn)
        upserts = []
        ticks = []
        ts = now_iso()
        for row in symbols:
            symbol = row["symbol"]
            if str(row.get("asset_class") or "").upper() == ASSET_CLASS_GOLD:
                # GOLD must come from live 24K per-gram feed, not last-trade fallback.
                continue
            snap = get_last_trade_snapshot(conn, symbol)
            if not snap:
                continue
            trade_date = _parse_iso_date_safe(snap.get("trade_date"))
            if not trade_date:
                continue
            age_days = max(0, (dt.date.today() - trade_date).days)
            if age_days > int(LAST_TRADE_FALLBACK_MAX_AGE_DAYS):
                continue
            _, ltp = adjusted_trade_values(
                symbol,
                snap["trade_date"],
                parse_float(snap["quantity"], 0.0),
                parse_float(snap["price"], 0.0),
                split_map,
            )
            if ltp <= 0:
                continue
            cur = conn.execute(
                "SELECT ltp FROM latest_prices WHERE UPPER(symbol) = ? ORDER BY updated_at DESC LIMIT 1",
                (symbol_upper(symbol),),
            ).fetchone()
            cur_ltp = parse_float(cur["ltp"], 0.0) if cur else 0.0
            if cur_ltp > 0:
                continue
            upserts.append((symbol, ltp, 0.0, ts))
            ticks.append((symbol, ltp, 0.0, ts, "last_trade_fallback"))
        if upserts:
            conn.executemany(
                """
                INSERT INTO latest_prices(symbol, ltp, change_abs, updated_at) VALUES (?, ?, ?, ?)
                ON CONFLICT(symbol) DO UPDATE SET
                  ltp=excluded.ltp,
                  change_abs=excluded.change_abs,
                  updated_at=excluded.updated_at
                """,
                upserts,
            )
            conn.executemany(
                "INSERT INTO price_ticks(symbol, ltp, change_abs, fetched_at, source) VALUES (?, ?, ?, ?, ?)",
                ticks,
            )
            conn.commit()


def latest_signal_map(conn):
    rows = conn.execute(
        """
        SELECT symbol, buy_signal, sell_signal, score, signal_date
        FROM signals
        ORDER BY signal_date DESC, id DESC
        """
    ).fetchall()
    out = {}
    for r in rows:
        key = symbol_upper(r["symbol"])
        if key in out:
            continue
        out[key] = dict(r)
    return out


def price_momentum_pct(conn, symbol, lookback_days=30):
    lookback_days = max(1, int(lookback_days))
    cutoff = (dt.datetime.now() - dt.timedelta(days=lookback_days)).replace(microsecond=0).isoformat()
    rows = conn.execute(
        """
        SELECT ltp
        FROM price_ticks
        WHERE UPPER(symbol) = ? AND fetched_at >= ? AND ltp > 0
        ORDER BY fetched_at ASC
        """,
        (symbol_upper(symbol), cutoff),
    ).fetchall()
    if len(rows) >= 2:
        first = parse_float(rows[0]["ltp"], 0.0)
        last = parse_float(rows[-1]["ltp"], 0.0)
        if first > 0 and last > 0:
            return ((last - first) / first) * 100.0

    row = conn.execute(
        "SELECT ltp, change_abs FROM latest_prices WHERE UPPER(symbol) = ? LIMIT 1",
        (symbol_upper(symbol),),
    ).fetchone()
    if row:
        ltp = parse_float(row["ltp"], 0.0)
        change_abs = parse_float(row["change_abs"], 0.0)
        prev = ltp - change_abs
        if prev > 0:
            return (change_abs / prev) * 100.0
    return 0.0


def collect_strategy_universe(conn, lookback_days=30):
    split_map = load_split_map(conn)
    peak_split_map = load_peak_split_map(conn)
    rows = conn.execute(
        """
        SELECT
          i.symbol,
          i.exchange,
          i.feed_code,
          COALESCE(h.qty,0) AS qty,
          COALESCE(h.avg_cost,0) AS avg_cost,
          COALESCE(h.invested,0) AS invested,
          COALESCE(h.market_value,0) AS market_value,
          COALESCE(h.realized_pnl,0) AS realized_pnl,
          COALESCE(h.unrealized_pnl,0) AS unrealized_pnl,
          COALESCE(h.total_return_pct,0) AS total_return_pct,
          COALESCE(lp.ltp,0) AS ltp,
          COALESCE(lp.change_abs,0) AS change_abs
        FROM instruments i
        LEFT JOIN holdings h ON h.symbol = i.symbol
        LEFT JOIN latest_prices lp ON lp.symbol = i.symbol
        WHERE i.active = 1
          AND UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
        ORDER BY i.symbol
        """
    ).fetchall()
    signals = latest_signal_map(conn)
    symbols = [symbol_upper(r["symbol"]) for r in rows if symbol_upper(r["symbol"])]
    prev_close_map = load_prev_close_map(conn, symbols)
    items = []
    for r in rows:
        item = dict(r)
        symbol = item["symbol"]
        if parse_float(item.get("ltp"), 0.0) <= 0:
            item["ltp"] = round(get_effective_ltp(conn, symbol, split_map), 4)
        ltp = parse_float(item.get("ltp"), 0.0)
        qty = parse_float(item.get("qty"), 0.0)
        avg_cost = parse_float(item.get("avg_cost"), 0.0)
        last_trade_px = get_last_trade_price(conn, symbol, split_map)
        fallback_px = max(last_trade_px, avg_cost, 0.0)
        outlier = False
        if ltp <= 0:
            outlier = True
        if (not outlier) and qty > 0 and avg_cost > 0:
            if ltp > (avg_cost * 80.0) or ltp < (avg_cost / 30.0):
                outlier = True
        if (not outlier) and last_trade_px > 0:
            if ltp > (last_trade_px * 50.0) or ltp < (last_trade_px / 25.0):
                outlier = True
        if outlier and fallback_px > 0:
            item["ltp"] = round(fallback_px, 4)
            item["ltp_adjusted"] = True
        sig = signals.get(symbol_upper(symbol))
        item["buy_signal"] = sig["buy_signal"] if sig else None
        item["sell_signal"] = sig["sell_signal"] if sig else None
        item["signal_score"] = round(parse_float(sig["score"], 0.0), 2) if sig else None
        peak = peak_traded_metrics(conn, symbol, peak_split_map)
        item["peak_traded_price"] = round(peak["peak_traded_price"], 4)
        item["pct_from_peak_traded"] = round(peak["pct_from_peak_traded"], 2)
        item["peak_buy_price"] = round(peak["peak_buy_price"], 4)
        item["pct_from_peak_buy"] = round(peak["pct_from_peak_buy"], 2)
        item["momentum_lookback_pct"] = round(price_momentum_pct(conn, symbol, lookback_days), 2)
        enrich_holding_metrics(item, conn=conn, prev_close_map=prev_close_map)
        items.append(item)
    return items


def _tokenize_intel(text):
    return re.findall(r"[a-z0-9]+", str(text or "").lower())


def _intel_sentiment_score(text):
    tokens = _tokenize_intel(text)
    if not tokens:
        return {"score": 0.0, "confidence": 0.35, "positive_hits": 0, "negative_hits": 0}
    pos = sum(1 for t in tokens if t in INTEL_POSITIVE_WORDS)
    neg = sum(1 for t in tokens if t in INTEL_NEGATIVE_WORDS)
    total_hits = pos + neg
    if total_hits <= 0:
        return {"score": 0.0, "confidence": 0.35, "positive_hits": 0, "negative_hits": 0}
    score = ((pos - neg) / max(1, total_hits)) * 100.0
    confidence = clamp(0.35 + min(0.55, total_hits / 22.0), 0.35, 0.9)
    return {
        "score": clamp(score, -100.0, 100.0),
        "confidence": confidence,
        "positive_hits": pos,
        "negative_hits": neg,
    }


def _normalize_intel_doc_type(doc_type):
    t = str(doc_type or "").strip().lower().replace(" ", "_")
    if t in ("youtube", "youtube_commentary", "commentary", "analyst_commentary", "news_commentary", "earnings_call"):
        return "commentary"
    if t in ("budget", "budget_policy", "policy", "policy_update", "regulatory", "macro_policy"):
        return "policy"
    if t in ("financial_statement", "quarterly_results", "qoq_financials", "financials"):
        return "financial_statement"
    if not t:
        return "commentary"
    return t


def _symbol_mention_count(text, symbol):
    s = symbol_upper(symbol)
    if not s:
        return 0
    pats = [
        rf"(?<![A-Z0-9]){re.escape(s)}(?![A-Z0-9])",
        rf"\bNSE[:\s-]*{re.escape(s)}\b",
        rf"\bBSE[:\s-]*{re.escape(s)}\b",
    ]
    c = 0
    for p in pats:
        try:
            c += len(re.findall(p, str(text or ""), flags=re.IGNORECASE))
        except Exception:
            pass
    return c


def _theme_boost_for_symbol(text_lower, symbol, base_sentiment_score):
    sym = symbol_upper(symbol)
    sign = 1.0 if parse_float(base_sentiment_score, 0.0) >= 0 else -1.0
    boost = 0.0
    for theme, kws in INTEL_POLICY_THEMES.items():
        hit = any(k.lower() in text_lower for k in kws)
        if not hit:
            continue
        hints = INTEL_THEME_SYMBOL_HINTS.get(theme, [])
        if any(h.upper() in sym for h in hints):
            boost += 7.5 * sign
    return clamp(boost, -24.0, 24.0)


def _parse_iso_date_safe(value, fallback=None):
    if fallback is None:
        fallback = dt.date.today()
    if value is None:
        return fallback
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s:
        return fallback
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    try:
        return dt.date.fromisoformat(s[:10])
    except Exception:
        return fallback


def _intelligence_recency_weight(doc_date_str, decay_days):
    d = _parse_iso_date_safe(doc_date_str, fallback=dt.date.today())
    age = max(0, (dt.date.today() - d).days)
    dec = max(1.0, parse_float(decay_days, 45.0))
    return math.exp(-(age / dec))


def analyze_and_store_intelligence_document(
    conn,
    doc_type,
    source,
    source_ref,
    doc_date,
    title,
    content,
):
    doc_type_norm = _normalize_intel_doc_type(doc_type)
    doc_date_iso = _parse_iso_date_safe(doc_date, fallback=dt.date.today()).isoformat()
    text = str(content or "").strip()
    if not text:
        raise ValueError("content_required")

    sent = _intel_sentiment_score(text)
    cur = conn.execute(
        """
        INSERT INTO intelligence_documents(
          doc_type, source, source_ref, doc_date, title, content, sentiment_score, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            doc_type_norm,
            str(source or ""),
            str(source_ref or ""),
            doc_date_iso,
            str(title or ""),
            text,
            round(parse_float(sent["score"], 0.0), 4),
            now_iso(),
        ),
    )
    doc_id = int(cur.lastrowid)

    symbols = [
        r["symbol"]
        for r in conn.execute("SELECT symbol FROM instruments WHERE active = 1 ORDER BY symbol").fetchall()
    ]
    holdings = {
        r["symbol"]
        for r in conn.execute("SELECT symbol FROM holdings WHERE qty > 0").fetchall()
    }
    lower = text.lower()
    impact_rows = []
    for symbol in symbols:
        mentions = _symbol_mention_count(text, symbol)
        theme_boost = _theme_boost_for_symbol(lower, symbol, sent["score"])
        if mentions <= 0 and abs(theme_boost) < 4.0 and symbol not in holdings:
            continue
        base_factor = 0.14 if mentions <= 0 else (0.28 + min(0.9, mentions * 0.22))
        impact_score = clamp((parse_float(sent["score"], 0.0) * base_factor) + theme_boost, -100.0, 100.0)
        if abs(impact_score) < 2.0 and mentions <= 0:
            continue
        confidence = clamp(
            parse_float(sent["confidence"], 0.35) * (0.62 + min(0.34, mentions * 0.08))
            + (0.08 if abs(theme_boost) >= 5.0 else 0.0),
            0.3,
            0.97,
        )
        reason_bits = []
        if mentions > 0:
            reason_bits.append(f"mentions={mentions}")
        if abs(theme_boost) >= 4.0:
            reason_bits.append(f"policy_theme={theme_boost:+.1f}")
        reason_bits.append(
            f"doc_sentiment={parse_float(sent['score'], 0.0):+.1f} ({sent['positive_hits']}p/{sent['negative_hits']}n)"
        )
        impact_rows.append(
            (
                doc_id,
                symbol,
                round(impact_score, 4),
                round(confidence, 4),
                "; ".join(reason_bits),
                now_iso(),
            )
        )

    if impact_rows:
        conn.executemany(
            """
            INSERT INTO intelligence_impacts(doc_id, symbol, impact_score, confidence, reason, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            impact_rows,
        )
    conn.commit()
    return {
        "doc_id": doc_id,
        "doc_type": doc_type_norm,
        "doc_date": doc_date_iso,
        "sentiment_score": round(parse_float(sent["score"], 0.0), 4),
        "impact_count": len(impact_rows),
    }


def upsert_company_financial_row(conn, payload):
    symbol_raw = payload.get("symbol")
    symbol = resolve_symbol(conn, symbol_raw) or symbol_upper(symbol_raw)
    if not symbol:
        raise ValueError("symbol_required")
    fiscal_period = str(payload.get("fiscal_period", "")).strip()
    if not fiscal_period:
        raise ValueError("fiscal_period_required")
    report_date = _parse_iso_date_safe(payload.get("report_date"), fallback=dt.date.today()).isoformat()
    source = str(payload.get("source") or "manual").strip() or "manual"
    conn.execute(
        """
        INSERT INTO company_financials(
          symbol, fiscal_period, report_date, revenue, pat, operating_cash_flow, investing_cash_flow,
          financing_cash_flow, debt, fii_holding_pct, dii_holding_pct, promoter_holding_pct, source, notes, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(symbol, fiscal_period, source) DO UPDATE SET
          report_date=excluded.report_date,
          revenue=excluded.revenue,
          pat=excluded.pat,
          operating_cash_flow=excluded.operating_cash_flow,
          investing_cash_flow=excluded.investing_cash_flow,
          financing_cash_flow=excluded.financing_cash_flow,
          debt=excluded.debt,
          fii_holding_pct=excluded.fii_holding_pct,
          dii_holding_pct=excluded.dii_holding_pct,
          promoter_holding_pct=excluded.promoter_holding_pct,
          notes=excluded.notes,
          created_at=excluded.created_at
        """,
        (
            symbol,
            fiscal_period,
            report_date,
            parse_float(payload.get("revenue"), None),
            parse_float(payload.get("pat"), None),
            parse_float(payload.get("operating_cash_flow"), None),
            parse_float(payload.get("investing_cash_flow"), None),
            parse_float(payload.get("financing_cash_flow"), None),
            parse_float(payload.get("debt"), None),
            parse_float(payload.get("fii_holding_pct"), None),
            parse_float(payload.get("dii_holding_pct"), None),
            parse_float(payload.get("promoter_holding_pct"), None),
            source,
            str(payload.get("notes", "")).strip(),
            now_iso(),
        ),
    )
    conn.commit()
    row = conn.execute(
        """
        SELECT *
        FROM company_financials
        WHERE UPPER(symbol) = ? AND fiscal_period = ? AND source = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (symbol_upper(symbol), fiscal_period, source),
    ).fetchone()
    return dict(row) if row else {"symbol": symbol, "fiscal_period": fiscal_period, "source": source}


def _safe_growth_pct(cur, prev):
    c = parse_float(cur, None)
    p = parse_float(prev, None)
    if c is None or p is None:
        return None
    if abs(p) < 1e-9:
        return None
    return ((c - p) / abs(p)) * 100.0


def _financial_signal_from_rows(rows):
    if not rows:
        return {"score": 0.0, "confidence": 0.0, "summary": "No financial snapshots available.", "components": []}
    cur = rows[0]
    prev = rows[1] if len(rows) > 1 else None
    components = []
    score = 0.0
    available = 0

    if prev:
        rev_g = _safe_growth_pct(cur.get("revenue"), prev.get("revenue"))
        if rev_g is not None:
            available += 1
            w = clamp(rev_g, -60.0, 60.0) * 0.28
            score += w
            components.append(f"Revenue QoQ {rev_g:+.1f}%")
        pat_g = _safe_growth_pct(cur.get("pat"), prev.get("pat"))
        if pat_g is not None:
            available += 1
            w = clamp(pat_g, -60.0, 60.0) * 0.36
            score += w
            components.append(f"PAT QoQ {pat_g:+.1f}%")
        ocf_g = _safe_growth_pct(cur.get("operating_cash_flow"), prev.get("operating_cash_flow"))
        if ocf_g is not None:
            available += 1
            w = clamp(ocf_g, -60.0, 60.0) * 0.22
            score += w
            components.append(f"OCF QoQ {ocf_g:+.1f}%")
        debt_g = _safe_growth_pct(cur.get("debt"), prev.get("debt"))
        if debt_g is not None:
            available += 1
            w = clamp(-debt_g, -50.0, 50.0) * 0.18
            score += w
            components.append(f"Debt QoQ {debt_g:+.1f}%")
        fii_d = parse_float(cur.get("fii_holding_pct"), None)
        fii_p = parse_float(prev.get("fii_holding_pct"), None)
        if fii_d is not None and fii_p is not None:
            available += 1
            delta = fii_d - fii_p
            score += clamp(delta * 2.2, -12.0, 12.0)
            components.append(f"FII holding {delta:+.2f} pts QoQ")
        dii_d = parse_float(cur.get("dii_holding_pct"), None)
        dii_p = parse_float(prev.get("dii_holding_pct"), None)
        if dii_d is not None and dii_p is not None:
            available += 1
            delta = dii_d - dii_p
            score += clamp(delta * 1.6, -10.0, 10.0)
            components.append(f"DII holding {delta:+.2f} pts QoQ")

    pat = parse_float(cur.get("pat"), None)
    ocf = parse_float(cur.get("operating_cash_flow"), None)
    if pat is not None and ocf is not None:
        available += 1
        if pat > 0 and ocf > 0 and ocf >= (0.7 * pat):
            score += 8.0
            components.append("Cash-flow quality supportive (OCF tracks PAT).")
        elif pat > 0 and ocf < 0:
            score -= 10.0
            components.append("Cash-flow quality weak (PAT positive but OCF negative).")

    icf = parse_float(cur.get("investing_cash_flow"), None)
    fcf = parse_float(cur.get("financing_cash_flow"), None)
    if icf is not None and fcf is not None:
        available += 1
        if icf < 0 and fcf > 0:
            score += 3.0
            components.append("Funding + investment activity indicates expansion phase.")
        elif icf > 0 and fcf < 0:
            score -= 2.0
            components.append("Possible balance-sheet defensive posture.")

    # Single-snapshot fallback so financial signal is not always zero when only one
    # recent financial row exists (common for auto-collected online snapshots).
    rev = parse_float(cur.get("revenue"), None)
    debt = parse_float(cur.get("debt"), None)
    if rev is not None and rev > 0 and pat is not None:
        available += 1
        pat_margin = (pat / rev) * 100.0
        score += clamp((pat_margin - 5.0) * 1.1, -12.0, 12.0)
        components.append(f"PAT margin {pat_margin:+.2f}%")
    if rev is not None and rev > 0 and debt is not None:
        available += 1
        debt_to_rev = debt / rev
        score += clamp((1.2 - debt_to_rev) * 6.0, -10.0, 8.0)
        components.append(f"Debt/Revenue {debt_to_rev:.2f}x")
    if pat is not None:
        available += 1
        if pat > 0:
            score += 2.0
            components.append("PAT positive.")
        elif pat < 0:
            score -= 8.0
            components.append("PAT negative.")
    if ocf is not None:
        available += 1
        score += 2.0 if ocf > 0 else -2.5
        components.append("OCF positive." if ocf > 0 else "OCF negative.")
    fii = parse_float(cur.get("fii_holding_pct"), None)
    dii = parse_float(cur.get("dii_holding_pct"), None)
    promoter = parse_float(cur.get("promoter_holding_pct"), None)
    if fii is not None and dii is not None:
        available += 1
        inst_total = fii + dii
        score += clamp((inst_total - 15.0) * 0.35, -5.0, 5.0)
        components.append(f"Institutional holding {inst_total:.2f}%")
    if promoter is not None:
        available += 1
        if promoter < 30.0:
            score -= 3.0
            components.append("Low promoter holding.")
        elif promoter > 85.0:
            score -= 1.0
            components.append("Very high promoter concentration.")
        else:
            score += 1.5
            components.append("Promoter holding stability supportive.")

    score = clamp(score, -100.0, 100.0)
    confidence = clamp(0.35 + min(0.55, available * 0.08), 0.3, 0.92)
    if not components:
        components = ["Insufficient QoQ metrics."]
    return {
        "score": round(score, 4),
        "confidence": round(confidence, 4),
        "summary": " | ".join(components[:4]),
        "components": components,
        "latest_period": str(cur.get("fiscal_period") or ""),
        "latest_report_date": str(cur.get("report_date") or ""),
    }


def _financial_report_date_ceiling():
    # Guard against malformed/future-placeholder periods being treated as latest.
    return (dt.date.today() + dt.timedelta(days=730)).isoformat()


def financial_signal_for_symbol(conn, symbol):
    report_ceiling = _financial_report_date_ceiling()
    rows = [
        dict(r)
        for r in conn.execute(
            """
            SELECT *
            FROM company_financials
            WHERE UPPER(symbol) = ?
              AND report_date <= ?
            ORDER BY report_date DESC, id DESC
            LIMIT 2
            """,
            (symbol_upper(symbol), report_ceiling),
        ).fetchall()
    ]
    sig = _financial_signal_from_rows(rows)
    sig["symbol"] = symbol_upper(symbol)
    return sig


def infer_cross_company_flows(conn, limit=16):
    report_ceiling = _financial_report_date_ceiling()
    symbols = [
        r["symbol"]
        for r in conn.execute(
            "SELECT DISTINCT symbol FROM company_financials ORDER BY symbol"
        ).fetchall()
    ]
    inflow = []
    outflow = []
    for symbol in symbols:
        rows = conn.execute(
            """
            SELECT fii_holding_pct, dii_holding_pct, report_date
            FROM company_financials
            WHERE UPPER(symbol) = ?
              AND report_date <= ?
            ORDER BY report_date DESC, id DESC
            LIMIT 2
            """,
            (symbol_upper(symbol), report_ceiling),
        ).fetchall()
        if len(rows) < 2:
            continue
        c = rows[0]
        p = rows[1]
        c_fii = parse_float(c["fii_holding_pct"], None)
        p_fii = parse_float(p["fii_holding_pct"], None)
        c_dii = parse_float(c["dii_holding_pct"], None)
        p_dii = parse_float(p["dii_holding_pct"], None)
        if c_fii is None or p_fii is None or c_dii is None or p_dii is None:
            continue
        delta = (c_fii - p_fii) + (c_dii - p_dii)
        if delta >= 0.05:
            inflow.append([symbol, delta, str(c["report_date"])])
        elif delta <= -0.05:
            outflow.append([symbol, abs(delta), str(c["report_date"])])

    inflow.sort(key=lambda x: x[1], reverse=True)
    outflow.sort(key=lambda x: x[1], reverse=True)
    edges = []
    i = 0
    j = 0
    while i < len(outflow) and j < len(inflow) and len(edges) < max(1, int(limit)):
        out_s, out_v, out_d = outflow[i]
        in_s, in_v, in_d = inflow[j]
        flow = min(out_v, in_v)
        if flow >= 0.05 and symbol_upper(out_s) != symbol_upper(in_s):
            edges.append(
                {
                    "from_symbol": symbol_upper(out_s),
                    "to_symbol": symbol_upper(in_s),
                    "flow_score": round(flow, 4),
                    "period": max(out_d, in_d),
                    "reason": "Inferred from combined FII+DII holding delta.",
                }
            )
        out_v -= flow
        in_v -= flow
        if out_v <= 1e-9:
            i += 1
        else:
            outflow[i][1] = out_v
        if in_v <= 1e-9:
            j += 1
        else:
            inflow[j][1] = in_v
    return edges


def _avg(values):
    vals = [parse_float(v, 0.0) for v in values]
    vals = [v for v in vals if not math.isnan(v) and not math.isinf(v)]
    if not vals:
        return 0.0
    return sum(vals) / len(vals)


def _sma(values, period):
    p = max(1, int(period))
    vals = [parse_float(v, 0.0) for v in values if parse_float(v, 0.0) > 0]
    if len(vals) < p:
        return 0.0
    return _avg(vals[-p:])


def _ema_series(values, period):
    vals = [parse_float(v, 0.0) for v in values]
    vals = [v for v in vals if v > 0]
    if not vals:
        return []
    p = max(1, int(period))
    alpha = 2.0 / (p + 1.0)
    out = []
    ema = vals[0]
    for v in vals:
        ema = (alpha * v) + ((1.0 - alpha) * ema)
        out.append(ema)
    return out


def _rsi14(values):
    vals = [parse_float(v, 0.0) for v in values]
    vals = [v for v in vals if v > 0]
    if len(vals) < 15:
        return 0.0
    diffs = [vals[i] - vals[i - 1] for i in range(1, len(vals))]
    gains = [max(0.0, d) for d in diffs[-14:]]
    losses = [max(0.0, -d) for d in diffs[-14:]]
    avg_gain = _avg(gains)
    avg_loss = _avg(losses)
    if avg_loss <= 1e-12:
        return 100.0 if avg_gain > 0 else 50.0
    rs = avg_gain / avg_loss
    return 100.0 - (100.0 / (1.0 + rs))


def _series_last_n_return(values, n):
    vals = [parse_float(v, 0.0) for v in values if parse_float(v, 0.0) > 0]
    k = max(1, int(n))
    if len(vals) < (k + 1):
        return 0.0
    first = vals[-(k + 1)]
    last = vals[-1]
    if first <= 0:
        return 0.0
    return (last - first) / first * 100.0


def _daily_return_series(series):
    if not series or len(series) < 2:
        return []
    out = []
    prev = parse_float(series[0][1], 0.0)
    for d, px in series[1:]:
        cur = parse_float(px, 0.0)
        if prev > 0 and cur > 0:
            out.append((str(d), (cur - prev) / prev))
        prev = cur if cur > 0 else prev
    return out


def _corr(a, b):
    if not a or not b:
        return 0.0
    n = min(len(a), len(b))
    if n < 3:
        return 0.0
    xa = [parse_float(v, 0.0) for v in a[-n:]]
    xb = [parse_float(v, 0.0) for v in b[-n:]]
    ma = _avg(xa)
    mb = _avg(xb)
    num = sum((xa[i] - ma) * (xb[i] - mb) for i in range(n))
    den_a = math.sqrt(sum((x - ma) ** 2 for x in xa))
    den_b = math.sqrt(sum((x - mb) ** 2 for x in xb))
    if den_a <= 1e-12 or den_b <= 1e-12:
        return 0.0
    return clamp(num / (den_a * den_b), -1.0, 1.0)


def _load_market_series_batch(symbols, lookback_days=320):
    syms = sorted({symbol_upper(s) for s in (symbols or []) if symbol_upper(s)})
    if not syms:
        return {}
    from_s = (dt.date.today() - dt.timedelta(days=max(40, int(lookback_days)))).isoformat()
    out = defaultdict(list)
    with market_db_connect() as mconn:
        for i in range(0, len(syms), 400):
            chunk = syms[i : i + 400]
            placeholders = ",".join(["?"] * len(chunk))
            rows = mconn.execute(
                f"""
                SELECT symbol, price_date, close
                FROM daily_prices
                WHERE UPPER(symbol) IN ({placeholders})
                  AND price_date >= ?
                  AND close > 0
                ORDER BY symbol, price_date
                """,
                chunk + [from_s],
            ).fetchall()
            for r in rows:
                s = symbol_upper(r["symbol"])
                d = str(r["price_date"] or "")
                px = parse_float(r["close"], 0.0)
                if s and d and px > 0:
                    out[s].append((d, px))
    return out


def _choose_chart_proxy_symbol(series_map, fallback_symbol=None):
    candidates = []
    for sym in CHART_INDEX_PROXY_SYMBOLS:
        arr = series_map.get(symbol_upper(sym), [])
        if len(arr) >= 30:
            candidates.append((len(arr), symbol_upper(sym)))
    if candidates:
        candidates.sort(reverse=True)
        for _, sym in candidates:
            if symbol_upper(sym) != symbol_upper(fallback_symbol):
                return sym
    generic = sorted(
        ((len(v), k) for k, v in (series_map or {}).items() if len(v) >= 30 and symbol_upper(k) != symbol_upper(fallback_symbol)),
        reverse=True,
    )
    return generic[0][1] if generic else ""


def _align_returns_for_corr(series_a, series_b):
    ra = {d: r for d, r in _daily_return_series(series_a)}
    rb = {d: r for d, r in _daily_return_series(series_b)}
    common = sorted(set(ra.keys()) & set(rb.keys()))
    if len(common) < 3:
        return [], []
    aa = [ra[d] for d in common]
    bb = [rb[d] for d in common]
    return aa, bb


def fetch_tradingview_scan_bulk(symbol_exchange_rows, timeout=6):
    rows = []
    for r in (symbol_exchange_rows or []):
        sym = symbol_upper((r or {}).get("symbol"))
        ex = str((r or {}).get("exchange") or "NSE").strip().upper()
        if not sym:
            continue
        rows.append((sym, ex))
    if not rows:
        return {}
    opener = urllib.request.build_opener()
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json,text/plain,*/*",
        "Content-Type": "application/json",
        "Referer": "https://www.tradingview.com/",
        "Origin": "https://www.tradingview.com",
    }
    out = {}
    for i in range(0, len(rows), 60):
        chunk = rows[i : i + 60]
        tickers = [f"{ex}:{sym}" for sym, ex in chunk]
        payload = {
            "symbols": {"tickers": tickers, "query": {"types": []}},
            "columns": ["Recommend.All", "RSI", "MACD.macd", "MACD.signal", "close", "change"],
        }
        try:
            data = _http_json_post(opener, "https://scanner.tradingview.com/india/scan", headers, payload, timeout=timeout)
        except Exception:
            continue
        for item in (data or {}).get("data", []) or []:
            tsym = str(item.get("s") or "")
            vals = item.get("d") or []
            if not tsym:
                continue
            sym = symbol_upper(tsym.split(":")[-1])
            recommend = parse_float(vals[0] if len(vals) > 0 else 0.0, 0.0)
            rsi = parse_float(vals[1] if len(vals) > 1 else 0.0, 0.0)
            close = parse_float(vals[4] if len(vals) > 4 else 0.0, 0.0)
            tv_score = clamp(recommend * 30.0, -30.0, 30.0)
            conf = clamp(0.45 + min(0.35, abs(recommend) * 0.4), 0.2, 0.9)
            out[sym] = {
                "score": round(tv_score, 4),
                "confidence": round(conf, 4),
                "recommend": round(recommend, 6),
                "rsi": round(rsi, 4),
                "close": round(close, 4) if close > 0 else 0.0,
                "source": "tradingview_scan",
            }
    return out


def _micro_quote_signal_map(conn, symbols, lookback_hours=36, max_points=16):
    syms = sorted({symbol_upper(s) for s in (symbols or []) if symbol_upper(s)})
    if not syms:
        return {}
    placeholders = ",".join(["?"] * len(syms))
    rows = conn.execute(
        f"""
        SELECT symbol, ltp
        FROM price_ticks
        WHERE UPPER(symbol) IN ({placeholders})
          AND fetched_at >= datetime('now', ?)
          AND ltp > 0
        ORDER BY symbol, fetched_at
        """,
        syms + [f"-{max(1, int(lookback_hours))} hours"],
    ).fetchall()
    by_symbol = defaultdict(list)
    for r in rows:
        s = symbol_upper(r["symbol"])
        px = parse_float(r["ltp"], 0.0)
        if s and px > 0:
            by_symbol[s].append(px)
    out = {}
    for sym, vals in by_symbol.items():
        seq = vals[-max(3, int(max_points)) :]
        if len(seq) < 3 or seq[0] <= 0:
            continue
        move_pct = (seq[-1] - seq[0]) / seq[0] * 100.0
        score = clamp(move_pct * 2.2, -10.0, 10.0)
        conf = clamp(min(0.35, len(seq) / 40.0), 0.05, 0.35)
        out[sym] = {"score": round(score, 4), "confidence": round(conf, 4), "move_pct": round(move_pct, 4)}
    return out


def _analyze_chart_for_symbol(symbol, series, proxy_series=None, tv_hint=None, micro_hint=None, ltp_hint=0.0):
    sym = symbol_upper(symbol)
    arr = list(series or [])
    if ltp_hint > 0:
        today_s = dt.date.today().isoformat()
        if not arr or str(arr[-1][0]) < today_s:
            arr.append((today_s, parse_float(ltp_hint, 0.0)))
    closes = [parse_float(x[1], 0.0) for x in arr if parse_float(x[1], 0.0) > 0]
    if len(closes) < 8:
        return {
            "symbol": sym,
            "score": 0.0,
            "confidence": 0.2,
            "signal": "NEUTRAL",
            "trend_score": 0.0,
            "momentum_score": 0.0,
            "mean_reversion_score": 0.0,
            "breakout_score": 0.0,
            "relative_strength_score": 0.0,
            "index_correlation": 0.0,
            "source_summary": "insufficient_history",
            "pattern_flags": [],
            "details": {"history_points": len(closes), "reason": "insufficient_history"},
        }

    current = closes[-1]
    sma20 = _sma(closes, 20)
    sma50 = _sma(closes, 50)
    sma200 = _sma(closes, 200)
    ret20 = _series_last_n_return(closes, 20)
    ret60 = _series_last_n_return(closes, 60)
    rsi14 = _rsi14(closes)
    ema12 = _ema_series(closes, 12)
    ema26 = _ema_series(closes, 26)
    macd_val = 0.0
    macd_signal = 0.0
    if ema12 and ema26:
        n = min(len(ema12), len(ema26))
        macd_series = [ema12[-n + i] - ema26[-n + i] for i in range(n)]
        sig_series = _ema_series(macd_series, 9)
        if macd_series:
            macd_val = macd_series[-1]
        if sig_series:
            macd_signal = sig_series[-1]

    trend_score = 0.0
    if sma20 > 0:
        trend_score += 12.0 if current >= sma20 else -12.0
    if sma50 > 0 and sma20 > 0:
        trend_score += 10.0 if sma20 >= sma50 else -10.0
    if sma200 > 0 and sma50 > 0:
        trend_score += 8.0 if sma50 >= sma200 else -8.0
    if len(closes) >= 25:
        slope = (_avg(closes[-5:]) - _avg(closes[-25:-20])) / max(1e-9, _avg(closes[-25:-20])) * 100.0
        trend_score += clamp(slope * 0.8, -8.0, 8.0)
    trend_score = clamp(trend_score, -35.0, 35.0)

    momentum_score = clamp((ret20 * 1.0) + (ret60 * 0.45), -30.0, 30.0)
    mean_rev_score = 0.0
    if rsi14 > 0:
        if rsi14 <= 30:
            mean_rev_score = 16.0
        elif rsi14 >= 70:
            mean_rev_score = -16.0
        else:
            mean_rev_score = clamp((50.0 - rsi14) * 0.45, -10.0, 10.0)
    if macd_val != 0.0 or macd_signal != 0.0:
        mean_rev_score += clamp((macd_val - macd_signal) * 2.5, -8.0, 8.0)
    mean_rev_score = clamp(mean_rev_score, -22.0, 22.0)

    breakout_score = 0.0
    if len(closes) >= 22:
        recent_high = max(closes[-21:-1])
        recent_low = min(closes[-21:-1])
        if recent_high > 0 and current >= recent_high * 1.003:
            breakout_score = 18.0
        elif recent_low > 0 and current <= recent_low * 0.997:
            breakout_score = -18.0

    rel_strength_score = 0.0
    index_corr = 0.0
    proxy_ret20 = 0.0
    if proxy_series and len(proxy_series) >= 25:
        proxy_closes = [parse_float(x[1], 0.0) for x in proxy_series if parse_float(x[1], 0.0) > 0]
        if len(proxy_closes) >= 21:
            proxy_ret20 = _series_last_n_return(proxy_closes, 20)
            rel_strength_score = clamp((ret20 - proxy_ret20) * 1.4, -22.0, 22.0)
        aa, bb = _align_returns_for_corr(arr, proxy_series)
        index_corr = _corr(aa, bb) if aa and bb else 0.0

    tv_score = clamp(parse_float((tv_hint or {}).get("score"), 0.0), -35.0, 35.0)
    tv_conf = clamp(parse_float((tv_hint or {}).get("confidence"), 0.0), 0.0, 0.95)
    micro_score = clamp(parse_float((micro_hint or {}).get("score"), 0.0), -12.0, 12.0)
    micro_conf = clamp(parse_float((micro_hint or {}).get("confidence"), 0.0), 0.0, 0.5)

    combined = (
        (0.30 * trend_score)
        + (0.20 * momentum_score)
        + (0.14 * mean_rev_score)
        + (0.12 * breakout_score)
        + (0.16 * rel_strength_score)
        + (0.06 * tv_score)
        + (0.08 * micro_score)
    )
    score = clamp(combined, -100.0, 100.0)
    conf = 0.35
    conf += min(0.30, len(closes) / 320.0)
    conf += min(0.16, abs(index_corr) * 0.16)
    conf += min(0.14, tv_conf * 0.14)
    conf += min(0.08, micro_conf * 0.2)
    conf = clamp(conf, 0.2, 0.93)

    signal = "NEUTRAL"
    if score >= 20:
        signal = "BULLISH"
    elif score <= -20:
        signal = "BEARISH"

    flags = []
    if sma20 > 0 and current >= sma20:
        flags.append("above_sma20")
    if sma50 > 0 and current >= sma50:
        flags.append("above_sma50")
    if sma50 > 0 and sma200 > 0 and sma50 >= sma200:
        flags.append("long_trend_up")
    if breakout_score > 0:
        flags.append("breakout_up")
    elif breakout_score < 0:
        flags.append("breakout_down")
    if rsi14 <= 32 and rsi14 > 0:
        flags.append("rsi_oversold")
    if rsi14 >= 68:
        flags.append("rsi_overbought")
    if rel_strength_score >= 6:
        flags.append("outperforming_proxy")
    elif rel_strength_score <= -6:
        flags.append("underperforming_proxy")
    if tv_hint:
        flags.append("tradingview_scan")
    if micro_hint:
        flags.append("micro_quote_flow")

    sources = ["market_history"]
    if tv_hint:
        sources.append("tradingview_scan")
    if proxy_series:
        sources.append("index_proxy")
    if micro_hint:
        sources.append("quote_samples")
    return {
        "symbol": sym,
        "score": round(score, 4),
        "confidence": round(conf, 4),
        "signal": signal,
        "trend_score": round(trend_score, 4),
        "momentum_score": round(momentum_score, 4),
        "mean_reversion_score": round(mean_rev_score, 4),
        "breakout_score": round(breakout_score, 4),
        "relative_strength_score": round(rel_strength_score, 4),
        "index_correlation": round(index_corr, 4),
        "source_summary": ",".join(sources),
        "pattern_flags": flags,
        "details": {
            "close": round(current, 4),
            "sma20": round(sma20, 4) if sma20 > 0 else 0.0,
            "sma50": round(sma50, 4) if sma50 > 0 else 0.0,
            "sma200": round(sma200, 4) if sma200 > 0 else 0.0,
            "ret20_pct": round(ret20, 4),
            "ret60_pct": round(ret60, 4),
            "proxy_ret20_pct": round(proxy_ret20, 4),
            "rsi14": round(rsi14, 4) if rsi14 > 0 else 0.0,
            "macd": round(macd_val, 6),
            "macd_signal": round(macd_signal, 6),
            "tradingview_recommend": parse_float((tv_hint or {}).get("recommend"), 0.0),
            "micro_move_pct": parse_float((micro_hint or {}).get("move_pct"), 0.0),
            "history_points": len(closes),
        },
    }


def upsert_chart_analysis_snapshots(conn, snapshots, as_of_date=None, created_at=None):
    rows = []
    as_of = str(as_of_date or dt.date.today().isoformat())[:10]
    stamp = str(created_at or now_iso())
    for s in snapshots or []:
        sym = symbol_upper((s or {}).get("symbol"))
        if not sym:
            continue
        rows.append(
            (
                sym,
                as_of,
                parse_float(s.get("score"), 0.0),
                clamp(parse_float(s.get("confidence"), 0.0), 0.0, 0.99),
                str(s.get("signal") or "NEUTRAL").upper(),
                parse_float(s.get("trend_score"), 0.0),
                parse_float(s.get("momentum_score"), 0.0),
                parse_float(s.get("mean_reversion_score"), 0.0),
                parse_float(s.get("breakout_score"), 0.0),
                parse_float(s.get("relative_strength_score"), 0.0),
                parse_float(s.get("index_correlation"), 0.0),
                str(s.get("source_summary") or ""),
                _safe_json_dumps(s.get("pattern_flags", [])),
                _safe_json_dumps(s.get("details", {})),
                stamp,
            )
        )
    if not rows:
        return 0
    conn.executemany(
        """
        INSERT INTO chart_analysis_snapshots(
          symbol, as_of_date, score, confidence, signal,
          trend_score, momentum_score, mean_reversion_score, breakout_score,
          relative_strength_score, index_correlation, source_summary,
          pattern_flags_json, details_json, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(symbol, as_of_date) DO UPDATE SET
          score=excluded.score,
          confidence=excluded.confidence,
          signal=excluded.signal,
          trend_score=excluded.trend_score,
          momentum_score=excluded.momentum_score,
          mean_reversion_score=excluded.mean_reversion_score,
          breakout_score=excluded.breakout_score,
          relative_strength_score=excluded.relative_strength_score,
          index_correlation=excluded.index_correlation,
          source_summary=excluded.source_summary,
          pattern_flags_json=excluded.pattern_flags_json,
          details_json=excluded.details_json,
          created_at=excluded.created_at
        """,
        rows,
    )
    return len(rows)


def list_chart_snapshots(conn, limit=80, symbol=None):
    lim = max(1, min(500, int(limit)))
    sym = symbol_upper(symbol)
    if sym:
        rows = conn.execute(
            """
            SELECT symbol, as_of_date, score, confidence, signal, trend_score, momentum_score,
                   mean_reversion_score, breakout_score, relative_strength_score, index_correlation,
                   source_summary, pattern_flags_json, details_json, created_at
            FROM chart_analysis_snapshots
            WHERE UPPER(symbol) = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ?
            """,
            (sym, lim),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT c.symbol, c.as_of_date, c.score, c.confidence, c.signal, c.trend_score, c.momentum_score,
                   c.mean_reversion_score, c.breakout_score, c.relative_strength_score, c.index_correlation,
                   c.source_summary, c.pattern_flags_json, c.details_json, c.created_at
            FROM chart_analysis_snapshots c
            JOIN (
              SELECT symbol, MAX(created_at) AS mx
              FROM chart_analysis_snapshots
              GROUP BY symbol
            ) x ON x.symbol = c.symbol AND x.mx = c.created_at
            ORDER BY ABS(c.score) DESC, c.symbol
            LIMIT ?
            """,
            (lim,),
        ).fetchall()
    out = []
    for r in rows:
        item = dict(r)
        item["pattern_flags"] = _safe_json_loads(item.get("pattern_flags_json"), [])
        item["details"] = _safe_json_loads(item.get("details_json"), {})
        out.append(item)
    return out


def latest_chart_snapshot_map(conn):
    items = list_chart_snapshots(conn, limit=500, symbol=None)
    out = {}
    for it in items:
        sym = symbol_upper(it.get("symbol"))
        if not sym:
            continue
        out[sym] = {
            "score": round(parse_float(it.get("score"), 0.0), 4),
            "confidence": round(clamp(parse_float(it.get("confidence"), 0.0), 0.0, 0.99), 4),
            "signal": str(it.get("signal") or "NEUTRAL"),
            "summary": str(it.get("source_summary") or "chart_agent"),
            "as_of_date": str(it.get("as_of_date") or ""),
            "details": it.get("details") or {},
            "pattern_flags": it.get("pattern_flags") or [],
        }
    return out


def run_chart_intel_agent_once(max_runtime_sec=60, force=False):
    t0 = time.time()
    with db_connect() as conn:
        cfg = get_chart_agent_config(conn)
        if (not force) and (not cfg.get("enabled")):
            return {"ok": True, "executed": False, "reason": "disabled"}
        last = str(cfg.get("last_run_at") or "").strip()
        if (not force) and last:
            try:
                dt_last = dt.datetime.fromisoformat(last)
                if (dt.datetime.now() - dt_last).total_seconds() < int(cfg.get("interval_seconds", CHART_AGENT_INTERVAL_DEFAULT_SEC)):
                    return {"ok": True, "executed": False, "reason": "interval_not_elapsed"}
            except Exception:
                pass

        inst = conn.execute(
            """
            SELECT symbol, exchange
            FROM instruments
            WHERE active = 1
            ORDER BY symbol
            """
        ).fetchall()
        if not inst:
            set_chart_agent_config(last_run_at=now_iso())
            return {"ok": True, "executed": True, "updated": 0, "reason": "no_symbols"}

        symbol_rows = [{"symbol": r["symbol"], "exchange": str(r["exchange"] or "NSE")} for r in inst]
        symbols = [symbol_upper(r["symbol"]) for r in symbol_rows if symbol_upper(r["symbol"])]
        ltp_rows = (
            conn.execute(
                "SELECT symbol, ltp FROM latest_prices WHERE UPPER(symbol) IN (" + ",".join(["?"] * len(symbols)) + ")",
                symbols,
            ).fetchall()
            if symbols
            else []
        )
        ltp_map = {symbol_upper(r["symbol"]): parse_float(r["ltp"], 0.0) for r in ltp_rows}

        sources = parse_chart_source_list(cfg.get("sources"), DEFAULT_CHART_AGENT_SOURCES)
        series_map = _load_market_series_batch(symbols + list(CHART_INDEX_PROXY_SYMBOLS), lookback_days=340)
        proxy_symbol = _choose_chart_proxy_symbol(series_map)
        proxy_series_global = series_map.get(proxy_symbol, [])

        tv_map = {}
        micro_map = {}
        errors = []
        if "tradingview_scan" in sources and (time.time() - t0) < max(5, int(max_runtime_sec)):
            try:
                tv_map = fetch_tradingview_scan_bulk(symbol_rows, timeout=6)
            except Exception as ex:
                errors.append(f"tradingview_scan:{str(ex)}")
        if "quote_samples" in sources:
            try:
                micro_map = _micro_quote_signal_map(conn, symbols, lookback_hours=36, max_points=16)
            except Exception as ex:
                errors.append(f"quote_samples:{str(ex)}")

        snapshots = []
        for row in symbol_rows:
            if (time.time() - t0) > max(10, int(max_runtime_sec)):
                break
            sym = symbol_upper(row["symbol"])
            if not sym:
                continue
            series = series_map.get(sym, [])
            proxy_series = proxy_series_global
            if sym == symbol_upper(proxy_symbol):
                alt = _choose_chart_proxy_symbol(series_map, fallback_symbol=sym)
                proxy_series = series_map.get(alt, [])
            snap = _analyze_chart_for_symbol(
                symbol=sym,
                series=series,
                proxy_series=proxy_series,
                tv_hint=tv_map.get(sym),
                micro_hint=micro_map.get(sym),
                ltp_hint=ltp_map.get(sym, 0.0),
            )
            if proxy_symbol:
                snap.setdefault("details", {})
                snap["details"]["proxy_symbol"] = proxy_symbol
            snapshots.append(snap)

        updated = upsert_chart_analysis_snapshots(conn, snapshots, as_of_date=dt.date.today().isoformat(), created_at=now_iso())
        conn.commit()

    set_chart_agent_config(last_run_at=now_iso())
    return {
        "ok": True,
        "executed": True,
        "updated": int(updated),
        "symbols_considered": len(symbol_rows),
        "proxy_symbol": proxy_symbol,
        "sources": sources,
        "errors": errors[:20],
        "elapsed_sec": round(time.time() - t0, 3),
    }


def maybe_run_chart_intel_agent_once():
    try:
        return run_chart_intel_agent_once(max_runtime_sec=60, force=False)
    except Exception:
        return None


def get_intel_parameter_bundle(conn):
    params = get_active_params(conn)
    w_commentary = clamp(parse_float(params.get("intel_weight_commentary"), 0.25), 0.0, 1.0)
    w_policy = clamp(parse_float(params.get("intel_weight_policy"), 0.25), 0.0, 1.0)
    w_fin = clamp(parse_float(params.get("intel_weight_financials"), 0.5), 0.0, 1.0)
    w_chart = clamp(parse_float(params.get("intel_weight_chart"), 0.2), 0.0, 1.0)
    decay_days = int(clamp(parse_float(params.get("intel_decay_days"), 45), 7, 180))
    total_w = w_commentary + w_policy + w_fin + w_chart
    if total_w <= 1e-9:
        w_commentary, w_policy, w_fin, w_chart = 0.25, 0.25, 0.4, 0.1
        total_w = 1.0
    return {
        "w_commentary": w_commentary / total_w,
        "w_policy": w_policy / total_w,
        "w_financials": w_fin / total_w,
        "w_chart": w_chart / total_w,
        "decay_days": decay_days,
    }


def build_intelligence_bias_map(conn, decay_days=45, w_commentary=0.25, w_policy=0.25, w_financials=0.5, w_chart=0.0):
    symbols = [
        r["symbol"]
        for r in conn.execute("SELECT symbol FROM instruments WHERE active = 1 ORDER BY symbol").fetchall()
    ]
    sym_map = {
        symbol_upper(s): {
            "symbol": symbol_upper(s),
            "commentary_score": 0.0,
            "commentary_confidence": 0.0,
            "policy_score": 0.0,
            "policy_confidence": 0.0,
            "financial_score": 0.0,
            "financial_confidence": 0.0,
            "chart_score": 0.0,
            "chart_confidence": 0.0,
            "score": 0.0,
            "confidence": 0.0,
            "summary": "",
        }
        for s in symbols
    }

    cutoff = (dt.date.today() - dt.timedelta(days=max(30, int(decay_days) * 4))).isoformat()
    rows = conn.execute(
        """
        SELECT
          ii.symbol, ii.impact_score, ii.confidence, idoc.doc_type, idoc.doc_date
        FROM intelligence_impacts ii
        JOIN intelligence_documents idoc ON idoc.id = ii.doc_id
        WHERE idoc.doc_date >= ?
        ORDER BY idoc.doc_date DESC, ii.id DESC
        """,
        (cutoff,),
    ).fetchall()
    grouped = defaultdict(lambda: defaultdict(lambda: {"wsum": 0.0, "w": 0.0}))
    for r in rows:
        symbol = symbol_upper(r["symbol"])
        if symbol not in sym_map:
            continue
        dt_type = _normalize_intel_doc_type(r["doc_type"])
        bucket = "policy" if dt_type == "policy" else "commentary"
        recency = _intelligence_recency_weight(r["doc_date"], decay_days=decay_days)
        conf = clamp(parse_float(r["confidence"], 0.5), 0.1, 1.0)
        w = recency * conf
        grouped[symbol][bucket]["wsum"] += parse_float(r["impact_score"], 0.0) * w
        grouped[symbol][bucket]["w"] += w

    for symbol in sym_map:
        for bucket in ("commentary", "policy"):
            g = grouped[symbol][bucket]
            if g["w"] > 0:
                score = g["wsum"] / g["w"]
                conf = clamp(min(1.0, g["w"] / 6.0), 0.0, 0.95)
                sym_map[symbol][f"{bucket}_score"] = round(clamp(score, -100.0, 100.0), 4)
                sym_map[symbol][f"{bucket}_confidence"] = round(conf, 4)

    for symbol in sym_map:
        fs = financial_signal_for_symbol(conn, symbol)
        sym_map[symbol]["financial_score"] = round(parse_float(fs.get("score"), 0.0), 4)
        sym_map[symbol]["financial_confidence"] = round(parse_float(fs.get("confidence"), 0.0), 4)
    chart_map = latest_chart_snapshot_map(conn)
    for symbol in sym_map:
        ch = chart_map.get(symbol, {})
        sym_map[symbol]["chart_score"] = round(parse_float(ch.get("score"), 0.0), 4)
        sym_map[symbol]["chart_confidence"] = round(parse_float(ch.get("confidence"), 0.0), 4)

    chart_abs_avg = 0.0
    chart_abs_n = 0
    for symbol in sym_map:
        c_score = parse_float(sym_map[symbol]["commentary_score"], 0.0)
        p_score = parse_float(sym_map[symbol]["policy_score"], 0.0)
        f_score = parse_float(sym_map[symbol]["financial_score"], 0.0)
        ch_score = parse_float(sym_map[symbol]["chart_score"], 0.0)
        c_conf = parse_float(sym_map[symbol]["commentary_confidence"], 0.0)
        p_conf = parse_float(sym_map[symbol]["policy_confidence"], 0.0)
        f_conf = parse_float(sym_map[symbol]["financial_confidence"], 0.0)
        ch_conf = parse_float(sym_map[symbol]["chart_confidence"], 0.0)

        combined = (w_commentary * c_score) + (w_policy * p_score) + (w_financials * f_score) + (w_chart * ch_score)
        combined_conf = (w_commentary * c_conf) + (w_policy * p_conf) + (w_financials * f_conf) + (w_chart * ch_conf)
        sym_map[symbol]["score"] = round(clamp(combined, -100.0, 100.0), 4)
        sym_map[symbol]["confidence"] = round(clamp(combined_conf, 0.0, 0.95), 4)
        if abs(ch_score) > 0:
            chart_abs_avg += abs(ch_score)
            chart_abs_n += 1

        bits = []
        if abs(c_score) >= 4.0:
            bits.append(f"commentary {c_score:+.1f}")
        if abs(p_score) >= 4.0:
            bits.append(f"policy {p_score:+.1f}")
        if abs(f_score) >= 4.0:
            bits.append(f"financials {f_score:+.1f}")
        if abs(ch_score) >= 4.0:
            bits.append(f"chart {ch_score:+.1f}")
        sym_map[symbol]["summary"] = ", ".join(bits) if bits else "No strong intelligence bias."

    holdings = conn.execute(
        """
        SELECT symbol, market_value
        FROM holdings
        WHERE qty > 0
        """
    ).fetchall()
    total_mv = sum(max(0.0, parse_float(r["market_value"], 0.0)) for r in holdings)
    port_score = 0.0
    port_conf = 0.0
    if total_mv > 0:
        for r in holdings:
            s = symbol_upper(r["symbol"])
            w = max(0.0, parse_float(r["market_value"], 0.0)) / total_mv
            port_score += w * parse_float(sym_map.get(s, {}).get("score"), 0.0)
            port_conf += w * parse_float(sym_map.get(s, {}).get("confidence"), 0.0)

    docs_recent = int(
        conn.execute(
            "SELECT COUNT(*) AS c FROM intelligence_documents WHERE doc_date >= ?",
            ((dt.date.today() - dt.timedelta(days=max(30, int(decay_days) * 2))).isoformat(),),
        ).fetchone()["c"]
    )
    impacts_recent = int(
        conn.execute(
            """
            SELECT COUNT(*) AS c
            FROM intelligence_impacts ii
            JOIN intelligence_documents idoc ON idoc.id = ii.doc_id
            WHERE idoc.doc_date >= ?
            """,
            ((dt.date.today() - dt.timedelta(days=max(30, int(decay_days) * 2))).isoformat(),),
        ).fetchone()["c"]
    )
    cross_flows = infer_cross_company_flows(conn, limit=14)
    chart_abs_avg = (chart_abs_avg / chart_abs_n) if chart_abs_n > 0 else 0.0
    if abs(port_score) >= 8:
        thought = f"Intelligence overlay {port_score:+.1f} (weighted by holdings)."
    else:
        thought = "Intelligence overlay neutral to mild; no dominant cross-signal."
    if chart_abs_avg >= 8.0:
        thought += f" Chart-pattern layer active (avg abs score {chart_abs_avg:.1f})."
    return {
        "symbols": sym_map,
        "portfolio_score": round(port_score, 4),
        "portfolio_confidence": round(clamp(port_conf, 0.0, 0.95), 4),
        "documents_recent": docs_recent,
        "impacts_recent": impacts_recent,
        "cross_flows": cross_flows,
        "thought": thought,
        "weights": {
            "commentary": round(w_commentary, 4),
            "policy": round(w_policy, 4),
            "financials": round(w_financials, 4),
            "chart": round(w_chart, 4),
            "decay_days": int(decay_days),
        },
    }


def intelligence_summary(conn, limit=30):
    bundle = get_intel_parameter_bundle(conn)
    pre_bias = build_intelligence_bias_map(
        conn,
        decay_days=bundle["decay_days"],
        w_commentary=bundle["w_commentary"],
        w_policy=bundle["w_policy"],
        w_financials=bundle["w_financials"],
        w_chart=bundle["w_chart"],
    )
    backfill_stats = {"executed": False, "reason": "not_attempted"}
    try:
        pre_symbols = list((pre_bias or {}).get("symbols", {}).values())
        pre_symbols.sort(
            key=lambda x: abs(parse_float(x.get("commentary_score"), 0.0)) + abs(parse_float(x.get("policy_score"), 0.0)),
            reverse=True,
        )
        high_non_fin_missing = []
        for it in pre_symbols:
            sym = symbol_upper(it.get("symbol"))
            if not sym:
                continue
            f_conf = parse_float(it.get("financial_confidence"), 0.0)
            if f_conf > 0:
                continue
            high_non_fin_missing.append(sym)
            if len(high_non_fin_missing) >= 120:
                break
        fallback_targets = prioritized_symbols_for_financial_backfill(conn, limit=240)
        targets = []
        seen = set()
        for s in high_non_fin_missing + fallback_targets:
            su = symbol_upper(s)
            if not su or su in seen:
                continue
            seen.add(su)
            targets.append(su)
        backfill_stats = maybe_backfill_missing_financial_rows(conn, targets, force=False)
    except Exception as ex:
        backfill_stats = {"executed": False, "reason": f"error:{str(ex)}"}

    bias = build_intelligence_bias_map(
        conn,
        decay_days=bundle["decay_days"],
        w_commentary=bundle["w_commentary"],
        w_policy=bundle["w_policy"],
        w_financials=bundle["w_financials"],
        w_chart=bundle["w_chart"],
    )
    docs = [
        dict(r)
        for r in conn.execute(
            """
            SELECT id, doc_type, source, source_ref, doc_date, title, sentiment_score, created_at
            FROM intelligence_documents
            ORDER BY doc_date DESC, id DESC
            LIMIT ?
            """,
            (max(1, min(200, int(limit))),),
        ).fetchall()
    ]
    impacts = [
        dict(r)
        for r in conn.execute(
            """
            SELECT ii.id, ii.symbol, ii.impact_score, ii.confidence, ii.reason, ii.created_at, idoc.doc_type, idoc.doc_date
            FROM intelligence_impacts ii
            JOIN intelligence_documents idoc ON idoc.id = ii.doc_id
            ORDER BY ii.created_at DESC, ii.id DESC
            LIMIT ?
            """,
            (max(1, min(400, int(limit) * 2)),),
        ).fetchall()
    ]
    symbol_scores = sorted(
        [v for v in bias["symbols"].values()],
        key=lambda x: abs(parse_float(x.get("score"), 0.0)),
        reverse=True,
    )
    return {
        "portfolio_score": bias["portfolio_score"],
        "portfolio_confidence": bias["portfolio_confidence"],
        "thought": bias["thought"],
        "weights": bias["weights"],
        "documents_recent": bias["documents_recent"],
        "impacts_recent": bias["impacts_recent"],
        "cross_flows": bias["cross_flows"],
        "financial_backfill": backfill_stats,
        "symbol_scores": symbol_scores[:200],
        "recent_documents": docs,
        "recent_impacts": impacts,
    }


def _sgn(v, eps=1e-9):
    x = parse_float(v, 0.0)
    if x > eps:
        return 1
    if x < -eps:
        return -1
    return 0


def _stale_value_case_scan(conn, staleness_sec=900):
    staleness_sec = max(300, int(parse_float(staleness_sec, 900)))
    rows = conn.execute(
        """
        SELECT
          i.symbol,
          COALESCE(i.asset_class, 'EQUITY') AS asset_class,
          COALESCE(h.qty, 0) AS qty,
          COALESCE(h.market_value, 0) AS market_value,
          COALESCE(lp.ltp, 0) AS ltp,
          lp.updated_at
        FROM instruments i
        LEFT JOIN holdings h ON UPPER(h.symbol) = UPPER(i.symbol)
        LEFT JOIN latest_prices lp ON UPPER(lp.symbol) = UPPER(i.symbol)
        WHERE i.active = 1
        ORDER BY i.symbol
        """
    ).fetchall()
    symbols = [symbol_upper(r["symbol"]) for r in rows if symbol_upper(r["symbol"])]
    source_map = latest_tick_source_map(conn, symbols)
    fallback_symbols = []
    gold_source_mismatch_symbols = []
    stale_fallback_symbols = []
    holdings_mtm_drift_symbols = []
    stale_value_symbols = []
    scan_items = []

    for r in rows:
        symbol = symbol_upper(r["symbol"])
        asset_class = normalize_asset_class(
            r["asset_class"],
            fallback=infer_asset_class(symbol=symbol, name=symbol),
        )
        qty = parse_float(r["qty"], 0.0)
        ltp = parse_float(r["ltp"], 0.0)
        ts = str(r["updated_at"] or "").strip()
        age_sec = _iso_age_seconds(ts)
        src_meta = source_map.get(symbol, {})
        src = str(src_meta.get("source") or "").strip().lower()
        src_age_sec = _iso_age_seconds(src_meta.get("fetched_at"))
        if src == "last_trade_fallback":
            fallback_symbols.append(symbol)
            if age_sec is None or age_sec > staleness_sec:
                stale_fallback_symbols.append(symbol)
        if asset_class == ASSET_CLASS_GOLD and qty > 0:
            if src != "gold_rate_scrape" and ltp > 0:
                gold_source_mismatch_symbols.append(symbol)
            if src == "last_trade_fallback" or (age_sec is not None and age_sec > GOLD_QUOTE_STALE_MAX_AGE_SEC):
                stale_value_symbols.append(symbol)
        if qty > 0 and ltp > 0:
            expected_mv = qty * ltp
            current_mv = parse_float(r["market_value"], 0.0)
            drift_abs = abs(expected_mv - current_mv)
            drift_pct = (drift_abs / expected_mv * 100.0) if expected_mv > 0 else 0.0
            if drift_abs > 5.0 and drift_pct > 0.2:
                holdings_mtm_drift_symbols.append(symbol)
        if src == "last_trade_fallback" and qty > 0 and (age_sec is None or age_sec > staleness_sec):
            stale_value_symbols.append(symbol)
        scan_items.append(
            {
                "symbol": symbol,
                "asset_class": asset_class,
                "qty": round(qty, 6),
                "ltp": round(ltp, 4),
                "updated_at": ts,
                "price_age_sec": round(parse_float(age_sec, 0.0), 3) if age_sec is not None else None,
                "source": src,
                "source_age_sec": round(parse_float(src_age_sec, 0.0), 3) if src_age_sec is not None else None,
            }
        )

    # de-dup symbol lists preserving first-seen order
    def _uniq(seq):
        out = []
        seen = set()
        for x in seq:
            if x in seen:
                continue
            seen.add(x)
            out.append(x)
        return out

    return {
        "items": scan_items,
        "fallback_symbols": _uniq(fallback_symbols),
        "stale_fallback_symbols": _uniq(stale_fallback_symbols),
        "gold_source_mismatch_symbols": _uniq(gold_source_mismatch_symbols),
        "holdings_mtm_drift_symbols": _uniq(holdings_mtm_drift_symbols),
        "stale_value_symbols": _uniq(stale_value_symbols),
    }


def build_data_pipe_diagnostics(conn):
    now_dt = dt.datetime.now()
    active_symbols = [
        symbol_upper(r["symbol"])
        for r in conn.execute("SELECT symbol FROM instruments WHERE active = 1 ORDER BY symbol").fetchall()
    ]
    live_cfg = get_live_config(conn)
    staleness_sec = max(900, int(parse_float(live_cfg.get("interval_seconds"), 10)) * 6)
    stale_cutoff = now_dt - dt.timedelta(seconds=staleness_sec)
    stale_scan = _stale_value_case_scan(conn, staleness_sec=staleness_sec)

    lp_rows = conn.execute(
        """
        SELECT i.symbol, lp.ltp, lp.updated_at
        FROM instruments i
        LEFT JOIN latest_prices lp ON UPPER(lp.symbol) = UPPER(i.symbol)
        WHERE i.active = 1
        """
    ).fetchall()
    stale_prices = 0
    zero_ltp = 0
    with_price = 0
    for r in lp_rows:
        ltp = parse_float(r["ltp"], 0.0)
        if ltp > 0:
            with_price += 1
        else:
            zero_ltp += 1
        ts = str(r["updated_at"] or "").strip()
        if not ts:
            stale_prices += 1
            continue
        try:
            dt_ts = dt.datetime.fromisoformat(ts)
            if dt_ts < stale_cutoff:
                stale_prices += 1
        except Exception:
            stale_prices += 1

    missing_price = max(0, len(active_symbols) - with_price)
    qstats = [
        dict(r)
        for r in conn.execute(
            """
            SELECT source, attempts, successes, failures, score, last_error_at, last_success_at
            FROM quote_source_stats
            ORDER BY failures DESC, attempts DESC
            LIMIT 12
            """
        ).fetchall()
    ]
    weak_sources = []
    for r in qstats:
        attempts = int(parse_float(r.get("attempts"), 0))
        failures = int(parse_float(r.get("failures"), 0))
        fail_ratio = (failures / attempts) if attempts > 0 else 0.0
        if attempts >= 10 and fail_ratio >= 0.45:
            weak_sources.append(
                {
                    "source": r.get("source"),
                    "attempts": attempts,
                    "failures": failures,
                    "fail_ratio": round(fail_ratio, 4),
                    "last_error_at": r.get("last_error_at"),
                }
            )

    history_covered = 0
    history_latest_date = ""
    history_errors = []
    if active_symbols:
        try:
            placeholders = ",".join(["?"] * len(active_symbols))
            with market_db_connect() as mconn:
                hrows = mconn.execute(
                    f"""
                    SELECT UPPER(symbol) AS symbol, MAX(price_date) AS mx
                    FROM daily_prices
                    WHERE UPPER(symbol) IN ({placeholders})
                    GROUP BY UPPER(symbol)
                    """,
                    active_symbols,
                ).fetchall()
            mx_map = {symbol_upper(r["symbol"]): str(r["mx"] or "") for r in hrows}
            for s in active_symbols:
                mx = mx_map.get(s, "")
                if not mx:
                    continue
                history_latest_date = max(history_latest_date, mx)
                try:
                    dmx = dt.date.fromisoformat(mx)
                    if (dt.date.today() - dmx).days <= 5:
                        history_covered += 1
                except Exception:
                    pass
        except Exception as ex:
            history_errors.append(str(ex))
    history_coverage_ratio = (history_covered / len(active_symbols)) if active_symbols else 1.0

    issues = []
    if zero_ltp > 0:
        issues.append(f"Zero/invalid LTP rows: {zero_ltp}")
    if stale_prices > 0:
        issues.append(f"Stale/missing latest_prices rows: {stale_prices}")
    if missing_price > 0:
        issues.append(f"Active symbols without valid latest price: {missing_price}")
    if history_coverage_ratio < 0.75:
        issues.append(f"Market history coverage low: {round(history_coverage_ratio * 100.0, 1)}%")
    if weak_sources:
        issues.append(f"Weak quote sources detected: {len(weak_sources)}")
    if stale_scan["fallback_symbols"]:
        issues.append(f"Fallback-sourced prices in use: {len(stale_scan['fallback_symbols'])}")
    if stale_scan["stale_fallback_symbols"]:
        issues.append(f"Stale fallback prices detected: {len(stale_scan['stale_fallback_symbols'])}")
    if stale_scan["gold_source_mismatch_symbols"]:
        issues.append(f"Gold symbols on non-gold source: {len(stale_scan['gold_source_mismatch_symbols'])}")
    if stale_scan["holdings_mtm_drift_symbols"]:
        issues.append(f"Holdings mark-to-market drift rows: {len(stale_scan['holdings_mtm_drift_symbols'])}")
    if history_errors:
        issues.append("Market history diagnostics error.")
    if not issues:
        issues.append("No critical data-pipe issues detected.")

    return {
        "as_of": now_iso(),
        "active_symbols": len(active_symbols),
        "live_interval_seconds": int(parse_float(live_cfg.get("interval_seconds"), 10)),
        "staleness_threshold_seconds": staleness_sec,
        "with_price_symbols": with_price,
        "zero_ltp_symbols": zero_ltp,
        "stale_price_symbols": stale_prices,
        "missing_price_symbols": missing_price,
        "history_coverage_ratio": round(history_coverage_ratio, 4),
        "history_latest_date": history_latest_date,
        "history_errors": history_errors,
        "weak_sources": weak_sources,
        "fallback_price_symbols": len(stale_scan["fallback_symbols"]),
        "stale_fallback_symbols": len(stale_scan["stale_fallback_symbols"]),
        "gold_source_mismatch_symbols": len(stale_scan["gold_source_mismatch_symbols"]),
        "holdings_mtm_drift_symbols": len(stale_scan["holdings_mtm_drift_symbols"]),
        "stale_value_symbols": len(stale_scan["stale_value_symbols"]),
        "stale_value_scan": stale_scan["items"][:120],
        "issues": issues,
    }


def heal_stale_value_cases(max_runtime_sec=25):
    started = time.time()
    runtime_cap = max(5, int(parse_float(max_runtime_sec, 25)))
    actions = []
    errors = []
    details = {"cleared_symbols": [], "remaining_stale_symbols": [], "scan_before": {}, "scan_after": {}}

    with db_connect() as conn:
        scan_before = _stale_value_case_scan(conn, staleness_sec=max(900, runtime_cap * 6))
        details["scan_before"] = {
            "fallback_symbols": scan_before["fallback_symbols"],
            "stale_fallback_symbols": scan_before["stale_fallback_symbols"],
            "gold_source_mismatch_symbols": scan_before["gold_source_mismatch_symbols"],
            "holdings_mtm_drift_symbols": scan_before["holdings_mtm_drift_symbols"],
            "stale_value_symbols": scan_before["stale_value_symbols"],
        }
        candidates = list(scan_before["stale_value_symbols"])
        to_clear = set(scan_before["stale_fallback_symbols"])
        to_clear.update(scan_before["gold_source_mismatch_symbols"])
        if to_clear:
            placeholders = ",".join(["?"] * len(to_clear))
            sym_params = sorted(to_clear)
            conn.execute(
                f"DELETE FROM latest_prices WHERE UPPER(symbol) IN ({placeholders})",
                sym_params,
            )
            conn.execute(
                f"DELETE FROM price_ticks WHERE UPPER(symbol) IN ({placeholders}) AND LOWER(COALESCE(source,'')) = 'last_trade_fallback'",
                sym_params,
            )
            refresh_holdings_mark_to_market(conn)
            conn.commit()
            actions.append(f"clear_stale_prices:{len(sym_params)}")
            details["cleared_symbols"] = sym_params
        else:
            conn.commit()

    if (time.time() - started) < runtime_cap:
        try:
            refresh_latest_prices_from_exchange(max_runtime_sec=max(8, runtime_cap))
            actions.append("refresh_latest_prices_after_stale_clear")
        except Exception as ex:
            errors.append(f"stale_value_refresh_failed:{str(ex)}")

    with db_connect() as conn:
        scan_after = _stale_value_case_scan(conn, staleness_sec=max(900, runtime_cap * 6))
        details["scan_after"] = {
            "fallback_symbols": scan_after["fallback_symbols"],
            "stale_fallback_symbols": scan_after["stale_fallback_symbols"],
            "gold_source_mismatch_symbols": scan_after["gold_source_mismatch_symbols"],
            "holdings_mtm_drift_symbols": scan_after["holdings_mtm_drift_symbols"],
            "stale_value_symbols": scan_after["stale_value_symbols"],
        }
        details["remaining_stale_symbols"] = list(scan_after["stale_value_symbols"])
        refresh_holdings_mark_to_market(conn)
        conn.commit()
        actions.append("refresh_holdings_mark_to_market")

    return {"actions": actions, "errors": errors, "details": details}


def apply_data_pipe_fixes():
    actions = []
    errors = []
    try:
        refresh_latest_prices_from_exchange(max_runtime_sec=20)
        actions.append("refresh_latest_prices_from_exchange")
    except Exception as ex:
        errors.append(f"price_refresh_failed: {str(ex)}")
    try:
        ensure_latest_prices_nonzero_from_last_trade()
        actions.append("ensure_latest_prices_nonzero_from_last_trade")
    except Exception as ex:
        errors.append(f"nonzero_ltp_guard_failed: {str(ex)}")
    try:
        stale_fix = heal_stale_value_cases(max_runtime_sec=25)
        actions.extend([str(x) for x in stale_fix.get("actions", [])])
        errors.extend([str(x) for x in stale_fix.get("errors", [])])
        actions.append(
            "stale_value_case_summary:"
            + _safe_json_dumps(
                {
                    "cleared_symbols": (stale_fix.get("details") or {}).get("cleared_symbols", []),
                    "remaining_stale_symbols": (stale_fix.get("details") or {}).get("remaining_stale_symbols", []),
                }
            )
        )
    except Exception as ex:
        errors.append(f"stale_value_self_heal_failed: {str(ex)}")
    try:
        sync_market_history(backfill_all=False, max_runtime_sec=240, max_symbols=120)
        actions.append("sync_market_history_incremental")
    except Exception as ex:
        errors.append(f"history_sync_failed: {str(ex)}")
    try:
        recompute_holdings_and_signals(force_strategy=False)
        actions.append("recompute_holdings_and_signals")
    except Exception as ex:
        errors.append(f"recompute_failed: {str(ex)}")
    return {"actions": actions, "errors": errors}


def _iso_age_seconds(ts_text):
    raw = str(ts_text or "").strip()
    if not raw:
        return None
    try:
        dt_ts = dt.datetime.fromisoformat(raw)
        return max(0.0, (dt.datetime.now() - dt_ts).total_seconds())
    except Exception:
        return None


def _software_perf_issue_count(diag):
    issues = [str(x or "").strip() for x in (diag or {}).get("issues", [])]
    if not issues:
        return 0
    filtered = [x for x in issues if x and ("no critical data-pipe issues detected" not in x.lower())]
    return len(filtered)


def _software_perf_recent_quote_metrics(conn, lookback_minutes=60):
    cutoff = (dt.datetime.now() - dt.timedelta(minutes=max(5, int(lookback_minutes)))).replace(microsecond=0).isoformat()
    qrow = conn.execute(
        """
        SELECT
          COUNT(*) AS c,
          AVG(CASE WHEN latency_ms > 0 THEN latency_ms END) AS avg_latency_ms
        FROM quote_samples
        WHERE fetched_at >= ?
        """,
        (cutoff,),
    ).fetchone()
    stats = conn.execute(
        """
        SELECT
          COALESCE(SUM(successes), 0) AS succ,
          COALESCE(SUM(attempts), 0) AS att
        FROM quote_source_stats
        """
    ).fetchone()
    succ = parse_float(stats["succ"], 0.0) if stats else 0.0
    att = parse_float(stats["att"], 0.0) if stats else 0.0
    return {
        "sample_count": int(parse_float(qrow["c"], 0.0)) if qrow else 0,
        "avg_latency_ms": parse_float(qrow["avg_latency_ms"], 0.0) if qrow else 0.0,
        "success_rate": (succ / att) if att > 0 else 0.0,
    }


def _software_perf_collect_snapshot(conn, persist=True):
    diag = build_data_pipe_diagnostics(conn)
    quote_m = _software_perf_recent_quote_metrics(conn, lookback_minutes=60)
    lp_row = conn.execute("SELECT MAX(updated_at) AS ts FROM latest_prices").fetchone()
    price_age = _iso_age_seconds(lp_row["ts"]) if lp_row else None
    if price_age is None:
        price_age = float(max(900, int(parse_float(diag.get("staleness_threshold_seconds"), 900))))
    issues = [str(x or "") for x in diag.get("issues", [])]
    issue_count = _software_perf_issue_count(diag)
    snap = {
        "created_at": now_iso(),
        "live_stale_symbols": int(parse_float(diag.get("stale_price_symbols"), 0.0)),
        "live_zero_ltp_symbols": int(parse_float(diag.get("zero_ltp_symbols"), 0.0)),
        "live_missing_price_symbols": int(parse_float(diag.get("missing_price_symbols"), 0.0)),
        "weak_sources_count": len(diag.get("weak_sources") or []),
        "avg_quote_latency_ms": round(parse_float(quote_m.get("avg_latency_ms"), 0.0), 3),
        "quote_success_rate": round(parse_float(quote_m.get("success_rate"), 0.0), 6),
        "last_price_age_sec": round(parse_float(price_age, 0.0), 3),
        "history_coverage_ratio": round(parse_float(diag.get("history_coverage_ratio"), 0.0), 6),
        "issue_count": int(issue_count),
        "issues": issues,
        "weak_sources": diag.get("weak_sources") or [],
        "diag": diag,
        "quote_metrics": quote_m,
    }
    if persist:
        conn.execute(
            """
            INSERT INTO software_perf_snapshots(
              created_at,
              live_stale_symbols,
              live_zero_ltp_symbols,
              live_missing_price_symbols,
              weak_sources_count,
              avg_quote_latency_ms,
              quote_success_rate,
              last_price_age_sec,
              history_coverage_ratio,
              issue_count,
              notes_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                snap["created_at"],
                snap["live_stale_symbols"],
                snap["live_zero_ltp_symbols"],
                snap["live_missing_price_symbols"],
                snap["weak_sources_count"],
                snap["avg_quote_latency_ms"],
                snap["quote_success_rate"],
                snap["last_price_age_sec"],
                snap["history_coverage_ratio"],
                snap["issue_count"],
                _safe_json_dumps(
                    {
                        "issues": issues[:12],
                        "weak_sources": (diag.get("weak_sources") or [])[:10],
                        "staleness_threshold_seconds": diag.get("staleness_threshold_seconds"),
                    }
                ),
            ),
        )
    return snap


def _software_perf_log_action(conn, action_type, status, summary, details=None):
    conn.execute(
        """
        INSERT INTO software_perf_actions(created_at, action_type, status, summary, details_json)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            now_iso(),
            str(action_type or ""),
            str(status or ""),
            str(summary or ""),
            _safe_json_dumps(details or {}),
        ),
    )


def _software_perf_core_snapshot(conn):
    row = conn.execute(
        """
        SELECT
          (SELECT COUNT(*) FROM instruments) AS instruments_total,
          (SELECT COUNT(*) FROM trades) AS trades_total,
          (SELECT COUNT(*) FROM strategy_parameters) AS strategy_param_total,
          (SELECT COUNT(*) FROM cash_ledger) AS cash_ledger_total
        """
    ).fetchone()
    return {
        "instruments_total": int(parse_float(row["instruments_total"], 0.0)) if row else 0,
        "trades_total": int(parse_float(row["trades_total"], 0.0)) if row else 0,
        "strategy_param_total": int(parse_float(row["strategy_param_total"], 0.0)) if row else 0,
        "cash_ledger_total": int(parse_float(row["cash_ledger_total"], 0.0)) if row else 0,
    }


def _software_perf_core_guard(before, after):
    protected = ("instruments_total", "trades_total", "strategy_param_total", "cash_ledger_total")
    drift = {}
    for k in protected:
        b = int(parse_float((before or {}).get(k), 0.0))
        a = int(parse_float((after or {}).get(k), 0.0))
        if a != b:
            drift[k] = {"before": b, "after": a}
    return {"ok": len(drift) == 0, "drift": drift}


def _software_perf_write_improvement_draft(snapshot, tune_updates, cfg):
    out_dir = get_current_tenant_data_dir() / "agent_improvements"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    target = out_dir / f"software_improvement_{stamp}.py"
    payload = {
        "generated_at": now_iso(),
        "core_objective": str((cfg or {}).get("core_objective") or ""),
        "issues": list((snapshot or {}).get("issues") or []),
        "suggested_live_config_updates": dict(tune_updates or {}),
        "notes": [
            "Draft only: this file is generated for review and is not auto-imported by the runtime.",
            "Core objective guard: do not edit trade/instrument/strategy schema or destructive paths.",
        ],
    }
    py = (
        "\"\"\"Auto-generated software improvement draft.\n"
        "Review before use. This file is never auto-executed by the server.\n"
        "\"\"\"\n\n"
        f"PROPOSAL = {repr(payload)}\n\n"
        "def proposed_live_config_updates():\n"
        "    return dict(PROPOSAL.get('suggested_live_config_updates', {}))\n\n"
        "def rationale_lines():\n"
        "    return [str(x) for x in PROPOSAL.get('issues', [])]\n"
    )
    target.write_text(py, encoding="utf-8")
    return str(target)


def list_software_perf_snapshots(conn, limit=40):
    lim = max(1, min(500, int(limit)))
    rows = conn.execute(
        """
        SELECT
          id, created_at, live_stale_symbols, live_zero_ltp_symbols, live_missing_price_symbols,
          weak_sources_count, avg_quote_latency_ms, quote_success_rate, last_price_age_sec,
          history_coverage_ratio, issue_count, notes_json
        FROM software_perf_snapshots
        ORDER BY id DESC
        LIMIT ?
        """,
        (lim,),
    ).fetchall()
    out = []
    for r in rows:
        item = dict(r)
        item["notes"] = _safe_json_loads(item.get("notes_json"), {})
        out.append(item)
    return out


def list_software_perf_actions(conn, limit=80):
    lim = max(1, min(500, int(limit)))
    rows = conn.execute(
        """
        SELECT id, created_at, action_type, status, summary, details_json
        FROM software_perf_actions
        ORDER BY id DESC
        LIMIT ?
        """,
        (lim,),
    ).fetchall()
    out = []
    for r in rows:
        item = dict(r)
        item["details"] = _safe_json_loads(item.get("details_json"), {})
        out.append(item)
    return out


def run_software_perf_agent_once(max_runtime_sec=60, force=False):
    with db_connect() as conn:
        cfg = get_software_perf_agent_config(conn)
    if not cfg.get("enabled") and not force:
        return {"ok": True, "skipped": "disabled", "config": cfg}

    last_run_at = str(cfg.get("last_run_at") or "")
    if last_run_at and not force:
        age = _iso_age_seconds(last_run_at)
        if age is not None and age < int(cfg.get("interval_seconds", SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC)):
            return {"ok": True, "skipped": "interval_not_elapsed", "seconds_since_last_run": round(age, 3)}

    t0 = time.time()
    runtime_cap = max(10, int(parse_float(max_runtime_sec, 60)))
    actions = []
    errors = []
    tune_updates = {}
    draft_path = ""
    core_guard = {"ok": True, "drift": {}}

    with db_connect() as conn:
        core_before = _software_perf_core_snapshot(conn)
        before = _software_perf_collect_snapshot(conn, persist=True)
        conn.commit()

    needs_heal = bool(
        int(parse_float(before.get("live_stale_symbols"), 0.0)) > 0
        or int(parse_float(before.get("live_missing_price_symbols"), 0.0)) > 0
        or int(parse_float(before.get("live_zero_ltp_symbols"), 0.0)) > 0
        or int(parse_float(before.get("issue_count"), 0.0)) > 0
    )
    last_heal_age = _iso_age_seconds(cfg.get("last_heal_at"))
    heal_cooldown = max(120, min(3600, int(parse_float(cfg.get("interval_seconds"), 900) // 2)))
    can_heal_now = (last_heal_age is None) or (last_heal_age >= heal_cooldown) or force

    if needs_heal and can_heal_now and (time.time() - t0) < runtime_cap:
        fix = apply_data_pipe_fixes()
        actions.extend([str(x) for x in fix.get("actions", [])])
        errors.extend([str(x) for x in fix.get("errors", [])])
        set_software_perf_agent_config(last_heal_at=now_iso())
        with db_connect() as conn:
            _software_perf_log_action(
                conn,
                action_type="self_heal",
                status="ok" if not fix.get("errors") else "partial",
                summary="Applied data-pipe self-healing actions.",
                details=fix,
            )
            conn.commit()

    if cfg.get("auto_tune") and (time.time() - t0) < runtime_cap:
        with db_connect() as conn:
            live_cfg = get_live_config(conn)
        cur_interval = int(parse_float(live_cfg.get("interval_seconds"), 10))
        cur_top_k = int(parse_float(live_cfg.get("quote_top_k"), LIVE_QUOTE_TOP_K_DEFAULT))
        cur_exp = parse_float(live_cfg.get("quote_explore_ratio"), LIVE_QUOTE_EXPLORE_RATIO_DEFAULT)
        cur_dev = parse_float(live_cfg.get("quote_max_deviation_pct"), LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT)

        if not bool(live_cfg.get("enabled")) and needs_heal:
            tune_updates["enabled"] = True
        if int(parse_float(before.get("live_stale_symbols"), 0.0)) > 0 or int(
            parse_float(before.get("live_missing_price_symbols"), 0.0)
        ) > 0:
            tune_updates["interval_seconds"] = max(LIVE_REFRESH_MIN_SEC, min(cur_interval, 8))
            tune_updates["quote_top_k"] = max(cur_top_k, 4)
            tune_updates["quote_explore_ratio"] = min(cur_exp, 0.15)
        if int(parse_float(before.get("weak_sources_count"), 0.0)) >= 2:
            tune_updates["quote_max_deviation_pct"] = max(cur_dev, 9.5)
        if tune_updates:
            set_live_config(
                enabled=tune_updates.get("enabled"),
                interval_seconds=tune_updates.get("interval_seconds"),
                quote_top_k=tune_updates.get("quote_top_k"),
                quote_explore_ratio=tune_updates.get("quote_explore_ratio"),
                quote_max_deviation_pct=tune_updates.get("quote_max_deviation_pct"),
            )
            actions.append("auto_tune_live_config")
            with db_connect() as conn:
                _software_perf_log_action(
                    conn,
                    action_type="auto_tune",
                    status="ok",
                    summary="Adjusted bounded live quote knobs.",
                    details={"updates": tune_updates},
                )
                conn.commit()

    improve_age = _iso_age_seconds(cfg.get("last_improvement_at"))
    improve_due = improve_age is None or improve_age >= (6 * 60 * 60)
    if cfg.get("write_changes") and (time.time() - t0) < runtime_cap and improve_due and (
        needs_heal or bool(tune_updates) or force
    ):
        try:
            draft_path = _software_perf_write_improvement_draft(before, tune_updates, cfg)
            actions.append("write_improvement_draft")
            set_software_perf_agent_config(last_improvement_at=now_iso())
            with db_connect() as conn:
                _software_perf_log_action(
                    conn,
                    action_type="write_draft",
                    status="ok",
                    summary="Wrote software-improvement draft file.",
                    details={"path": draft_path},
                )
                conn.commit()
        except Exception as ex:
            errors.append(f"write_draft_failed:{str(ex)}")

    with db_connect() as conn:
        after = _software_perf_collect_snapshot(conn, persist=True)
        core_after = _software_perf_core_snapshot(conn)
        core_guard = _software_perf_core_guard(core_before, core_after)
        if not core_guard.get("ok"):
            _software_perf_log_action(
                conn,
                action_type="core_guard",
                status="error",
                summary="Core objective guard detected protected-count drift.",
                details=core_guard,
            )
            errors.append("core_objective_guard_triggered")
        conn.commit()

    set_software_perf_agent_config(last_run_at=now_iso())
    return {
        "ok": len(errors) == 0,
        "agent": "software_performance",
        "core_guard_ok": bool(core_guard.get("ok")),
        "core_guard": core_guard,
        "actions": actions,
        "errors": errors,
        "tune_updates": tune_updates,
        "draft_path": draft_path,
        "before": before,
        "after": after,
    }


def maybe_run_software_perf_agent_once():
    with db_connect() as conn:
        cfg = get_software_perf_agent_config(conn)
    if not cfg.get("enabled"):
        return None
    last = str(cfg.get("last_run_at") or "").strip()
    if last:
        age = _iso_age_seconds(last)
        if age is not None and age < int(cfg.get("interval_seconds", SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC)):
            return None
    return run_software_perf_agent_once(max_runtime_sec=60, force=False)


def _load_symbol_price_series(conn, symbols, from_date, to_date):
    out = {symbol_upper(s): [] for s in symbols if symbol_upper(s)}
    syms = [s for s in out.keys()]
    if not syms:
        return out

    chunks = []
    chunk_size = 400
    for i in range(0, len(syms), chunk_size):
        chunks.append(syms[i : i + chunk_size])
    for chunk in chunks:
        placeholders = ",".join(["?"] * len(chunk))
        try:
            with market_db_connect() as mconn:
                rows = mconn.execute(
                    f"""
                    SELECT UPPER(symbol) AS symbol, price_date, close
                    FROM daily_prices
                    WHERE UPPER(symbol) IN ({placeholders})
                      AND price_date >= ?
                      AND price_date <= ?
                      AND close > 0
                    ORDER BY symbol, price_date
                    """,
                    chunk + [from_date, to_date],
                ).fetchall()
            for r in rows:
                sym = symbol_upper(r["symbol"])
                d = str(r["price_date"] or "")
                px = parse_float(r["close"], 0.0)
                if sym and d and px > 0:
                    out[sym].append((d, px))
        except Exception:
            pass

    # fallback from price_ticks if history not present
    for sym in syms:
        if len(out.get(sym, [])) >= 30:
            continue
        ticks = aggregate_price_ticks_daily(conn, sym, from_s=from_date, to_s=to_date)
        if ticks:
            out[sym] = sorted({(d, px) for _, d, px in ticks if px > 0}, key=lambda x: x[0])
    return out


def _symbol_backtest_stats(series, horizon_days, lookback_days, intel_score):
    n = len(series or [])
    if n <= (horizon_days + lookback_days + 2):
        return None
    sample_count = 0
    hits = 0
    mom_hits = 0
    intel_hits = 0
    sum_future = 0.0
    sum_pred = 0.0
    for i in range(lookback_days, n - horizon_days):
        p0 = parse_float(series[i][1], 0.0)
        p_lb = parse_float(series[i - lookback_days][1], 0.0)
        pf = parse_float(series[i + horizon_days][1], 0.0)
        if p0 <= 0 or p_lb <= 0 or pf <= 0:
            continue
        momentum = ((p0 - p_lb) / p_lb) * 100.0
        fut_ret = ((pf - p0) / p0) * 100.0
        pred = (0.68 * momentum) + (0.32 * parse_float(intel_score, 0.0))
        if abs(pred) < 0.2:
            continue
        sample_count += 1
        sum_future += fut_ret
        sum_pred += pred
        if _sgn(pred) * _sgn(fut_ret) > 0:
            hits += 1
        if _sgn(momentum) * _sgn(fut_ret) > 0:
            mom_hits += 1
        if abs(parse_float(intel_score, 0.0)) >= 4.0 and _sgn(parse_float(intel_score, 0.0)) * _sgn(fut_ret) > 0:
            intel_hits += 1
    if sample_count <= 0:
        return None
    return {
        "sample_count": sample_count,
        "hit_rate": hits / sample_count,
        "momentum_hit_rate": mom_hits / sample_count,
        "intel_hit_rate": intel_hits / sample_count,
        "avg_future_return": sum_future / sample_count,
        "avg_pred_signal": sum_pred / sample_count,
    }


def run_agent_backtest(
    from_date=None,
    to_date=None,
    horizon_days=20,
    apply_tuning=False,
    fix_data_pipes=False,
    min_samples=None,
):
    end_d = _parse_iso_date_safe(to_date, fallback=dt.date.today())
    start_d = _parse_iso_date_safe(from_date, fallback=end_d - dt.timedelta(days=365))
    if start_d > end_d:
        start_d, end_d = end_d, start_d
    horizon_days = max(3, min(90, int(horizon_days)))
    lookback_days = max(12, min(120, 20 + (horizon_days // 2)))
    errors = []

    with db_connect() as conn:
        self_cfg = get_self_learning_config(conn)
        min_samples_eff = int(min_samples if min_samples is not None else self_cfg.get("min_samples", 30))
        min_samples_eff = max(5, min(5000, min_samples_eff))
        diagnostics_before = build_data_pipe_diagnostics(conn)

    fixes = {"actions": [], "errors": []}
    if fix_data_pipes:
        fixes = apply_data_pipe_fixes()
        if fixes.get("errors"):
            errors.extend([str(e) for e in fixes.get("errors", [])])

    with db_connect() as conn:
        diagnostics_after = build_data_pipe_diagnostics(conn)
        intel_bundle = get_intel_parameter_bundle(conn)
        intel_map = build_intelligence_bias_map(
            conn,
            decay_days=intel_bundle["decay_days"],
            w_commentary=intel_bundle["w_commentary"],
            w_policy=intel_bundle["w_policy"],
            w_financials=intel_bundle["w_financials"],
            w_chart=intel_bundle["w_chart"],
        )["symbols"]

        symbols = [
            symbol_upper(r["symbol"])
            for r in conn.execute("SELECT symbol FROM instruments WHERE active = 1 ORDER BY symbol").fetchall()
        ]
        price_map = _load_symbol_price_series(conn, symbols, start_d.isoformat(), end_d.isoformat())
        symbol_rows = []
        total_samples = 0
        total_hits = 0.0
        total_mom_hits = 0.0
        total_intel_hits = 0.0
        weighted_future = 0.0

        for sym in symbols:
            series = price_map.get(sym, [])
            intel_item = intel_map.get(sym, {})
            intel_score = parse_float(intel_item.get("score"), 0.0)
            s = _symbol_backtest_stats(series, horizon_days=horizon_days, lookback_days=lookback_days, intel_score=intel_score)
            if s is None:
                continue
            samples = int(s["sample_count"])
            total_samples += samples
            total_hits += s["hit_rate"] * samples
            total_mom_hits += s["momentum_hit_rate"] * samples
            total_intel_hits += s["intel_hit_rate"] * samples
            weighted_future += s["avg_future_return"] * samples
            symbol_rows.append(
                {
                    "symbol": sym,
                    "sample_count": samples,
                    "hit_rate": round(s["hit_rate"], 4),
                    "momentum_hit_rate": round(s["momentum_hit_rate"], 4),
                    "intel_hit_rate": round(s["intel_hit_rate"], 4),
                    "avg_future_return": round(s["avg_future_return"], 4),
                    "avg_pred_signal": round(s["avg_pred_signal"], 4),
                    "intel_score": round(intel_score, 4),
                    "commentary_score": round(parse_float(intel_item.get("commentary_score"), 0.0), 4),
                    "policy_score": round(parse_float(intel_item.get("policy_score"), 0.0), 4),
                    "financial_score": round(parse_float(intel_item.get("financial_score"), 0.0), 4),
                    "chart_score": round(parse_float(intel_item.get("chart_score"), 0.0), 4),
                }
            )

        symbol_rows.sort(key=lambda x: (-x["sample_count"], x["symbol"]))
        overall_hit = (total_hits / total_samples) if total_samples > 0 else 0.0
        overall_mom_hit = (total_mom_hits / total_samples) if total_samples > 0 else 0.0
        overall_intel_hit = (total_intel_hits / total_samples) if total_samples > 0 else 0.0
        avg_future = (weighted_future / total_samples) if total_samples > 0 else 0.0

        params_before = get_active_params(conn)
        suggestions = {
            "min_samples_required": min_samples_eff,
            "eligible_samples": total_samples,
            "updates": {},
            "notes": [],
            "component_hit_rates": {},
        }
        params_after = dict(params_before)
        tuning_applied = False

        comp_hits = {"commentary": [0, 0], "policy": [0, 0], "financials": [0, 0], "chart": [0, 0]}
        for r in symbol_rows:
            fut_sign = _sgn(r.get("avg_future_return"))
            if fut_sign == 0:
                continue
            for key, field in (
                ("commentary", "commentary_score"),
                ("policy", "policy_score"),
                ("financials", "financial_score"),
                ("chart", "chart_score"),
            ):
                sc = parse_float(r.get(field), 0.0)
                if abs(sc) < 4.0:
                    continue
                comp_hits[key][1] += 1
                if _sgn(sc) == fut_sign:
                    comp_hits[key][0] += 1
        for k, (hit_n, tot_n) in comp_hits.items():
            suggestions["component_hit_rates"][k] = round((hit_n / tot_n), 4) if tot_n > 0 else 0.0

        if total_samples < min_samples_eff:
            suggestions["notes"].append(
                f"Not enough samples for auto-tuning ({total_samples} < {min_samples_eff})."
            )
        else:
            c_rate = parse_float(suggestions["component_hit_rates"].get("commentary"), 0.0)
            p_rate = parse_float(suggestions["component_hit_rates"].get("policy"), 0.0)
            f_rate = parse_float(suggestions["component_hit_rates"].get("financials"), 0.0)
            ch_rate = parse_float(suggestions["component_hit_rates"].get("chart"), 0.0)
            scores = {
                "intel_weight_commentary": max(0.05, c_rate),
                "intel_weight_policy": max(0.05, p_rate),
                "intel_weight_financials": max(0.05, f_rate),
                "intel_weight_chart": max(0.05, ch_rate),
            }
            ssum = sum(scores.values())
            if ssum > 0:
                for k in list(scores.keys()):
                    scores[k] = scores[k] / ssum
            suggestions["updates"]["intel_weight_commentary"] = round(scores["intel_weight_commentary"], 4)
            suggestions["updates"]["intel_weight_policy"] = round(scores["intel_weight_policy"], 4)
            suggestions["updates"]["intel_weight_financials"] = round(scores["intel_weight_financials"], 4)
            suggestions["updates"]["intel_weight_chart"] = round(scores["intel_weight_chart"], 4)

            cur_decay = int(clamp(parse_float(params_before.get("intel_decay_days"), 45), 7, 180))
            if overall_hit < 0.48:
                new_decay = max(14, cur_decay - 10)
            elif overall_hit > 0.58:
                new_decay = min(120, cur_decay + 10)
            else:
                new_decay = cur_decay
            suggestions["updates"]["intel_decay_days"] = int(new_decay)

            cur_add = clamp(parse_float(params_before.get("add_trigger_underweight"), 0.02), 0.0, 0.5)
            cur_trim = clamp(parse_float(params_before.get("trim_trigger_overweight"), 0.02), 0.0, 0.5)
            cur_maxw = clamp(parse_float(params_before.get("max_position_weight"), 0.12), 0.03, 0.9)
            if overall_hit < 0.48:
                suggestions["updates"]["add_trigger_underweight"] = round(clamp(cur_add + 0.003, 0.0, 0.5), 6)
                suggestions["updates"]["trim_trigger_overweight"] = round(clamp(cur_trim - 0.003, 0.0, 0.5), 6)
                suggestions["updates"]["max_position_weight"] = round(clamp(cur_maxw - 0.005, 0.03, 0.9), 6)
                suggestions["notes"].append("Defensive shift: tighten adds and concentration due to low hit rate.")
            elif overall_hit > 0.56:
                suggestions["updates"]["add_trigger_underweight"] = round(clamp(cur_add - 0.003, 0.0, 0.5), 6)
                suggestions["updates"]["trim_trigger_overweight"] = round(clamp(cur_trim + 0.002, 0.0, 0.5), 6)
                suggestions["updates"]["max_position_weight"] = round(clamp(cur_maxw + 0.004, 0.03, 0.9), 6)
                suggestions["notes"].append("Constructive shift: allow faster adds and slightly wider trim bands.")
            else:
                suggestions["notes"].append("Hit rate neutral; only intelligence mix recalibration suggested.")

            if apply_tuning:
                upd = update_active_strategy_parameters(conn, suggestions["updates"])
                params_after = upd.get("after", params_after)
                tuning_applied = bool(upd.get("changed"))
                if tuning_applied:
                    suggestions["notes"].append(f"Applied {len(upd.get('changed', {}))} parameter updates.")

        if total_samples <= 0:
            errors.append("insufficient_backtest_data")

        created = now_iso()
        conn.execute(
            """
            INSERT INTO agent_backtest_runs(
              created_at, from_date, to_date, horizon_days, sample_count, hit_rate, avg_future_return,
              momentum_hit_rate, intel_hit_rate, applied_tuning,
              params_before_json, params_after_json, suggestions_json, diagnostics_json, errors_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                created,
                start_d.isoformat(),
                end_d.isoformat(),
                int(horizon_days),
                int(total_samples),
                float(overall_hit),
                float(avg_future),
                float(overall_mom_hit),
                float(overall_intel_hit),
                1 if tuning_applied else 0,
                _safe_json_dumps(params_before),
                _safe_json_dumps(params_after),
                _safe_json_dumps(suggestions),
                _safe_json_dumps(
                    {
                        "before": diagnostics_before,
                        "after": diagnostics_after,
                        "fixes": fixes,
                    }
                ),
                _safe_json_dumps(errors),
            ),
        )
        run_id = int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
        conn.commit()

    if tuning_applied:
        refresh_strategy_analytics(force=True)

    return {
        "ok": True,
        "run_id": run_id,
        "created_at": now_iso(),
        "from_date": start_d.isoformat(),
        "to_date": end_d.isoformat(),
        "horizon_days": int(horizon_days),
        "sample_count": int(total_samples),
        "hit_rate": round(overall_hit, 6),
        "avg_future_return": round(avg_future, 6),
        "momentum_hit_rate": round(overall_mom_hit, 6),
        "intel_hit_rate": round(overall_intel_hit, 6),
        "symbol_stats": symbol_rows[:300],
        "suggestions": suggestions,
        "applied_tuning": bool(tuning_applied),
        "params_before": params_before,
        "params_after": params_after,
        "diagnostics": {
            "before": diagnostics_before,
            "after": diagnostics_after,
            "fixes": fixes,
        },
        "errors": errors,
    }


def maybe_run_self_learning_once():
    with db_connect() as conn:
        cfg = get_self_learning_config(conn)
    if not cfg.get("enabled"):
        return None
    last = str(cfg.get("last_run_at") or "").strip()
    if last:
        try:
            dt_last = dt.datetime.fromisoformat(last)
            if (dt.datetime.now() - dt_last).total_seconds() < (int(cfg.get("interval_days", 7)) * 86400):
                return None
        except Exception:
            pass

    res = run_agent_backtest(
        from_date=(dt.date.today() - dt.timedelta(days=420)).isoformat(),
        to_date=dt.date.today().isoformat(),
        horizon_days=20,
        apply_tuning=True,
        fix_data_pipes=True,
        min_samples=int(cfg.get("min_samples", 30)),
    )
    set_self_learning_config(last_run_at=now_iso())
    return res


def portfolio_daily_returns(conn, lookback_days=365):
    end = dt.date.today()
    start = end - dt.timedelta(days=max(30, int(lookback_days)))
    points = portfolio_timeseries(conn, start.isoformat(), end.isoformat())
    returns = []
    prev_val = None
    for p in points:
        cur_val = parse_float(p.get("total_value"), 0.0)
        if cur_val <= 0:
            continue
        if prev_val is not None and prev_val > 0:
            returns.append((cur_val - prev_val) / prev_val)
        prev_val = cur_val
    return returns


def build_strategy_insights(conn, run_date=None):
    run_date = run_date or dt.date.today().isoformat()
    params = get_active_params(conn)

    buy_l1 = clamp(parse_float(params.get("buy_l1_discount"), 0.03), 0.0, 0.5)
    buy_l2 = clamp(parse_float(params.get("buy_l2_discount"), 0.06), 0.0, 0.7)
    sell_s1 = clamp(parse_float(params.get("sell_s1_markup"), 0.12), 0.0, 2.0)
    sell_s2 = clamp(parse_float(params.get("sell_s2_markup"), 0.5), 0.0, 3.0)
    allocation_limit = clamp(parse_float(params.get("allocation_limit"), 0.25), 0.02, 0.8)
    trim_trigger = clamp(parse_float(params.get("trim_trigger_overweight"), 0.02), 0.0, 0.5)
    add_trigger = clamp(parse_float(params.get("add_trigger_underweight"), 0.02), 0.0, 0.5)
    max_position_weight = clamp(parse_float(params.get("max_position_weight"), 0.12), 0.03, 0.9)
    max_new_ideas = int(clamp(parse_float(params.get("max_new_ideas"), 2), 1, 2))
    lookback_days = int(clamp(parse_float(params.get("momentum_lookback_days"), 30), 5, 120))
    projection_years = int(clamp(parse_float(params.get("projection_years"), 5), 1, 15))
    projection_base_return = clamp(parse_float(params.get("projection_base_return"), 0.12), -0.3, 0.4)
    projection_conservative_delta = clamp(parse_float(params.get("projection_conservative_delta"), 0.04), 0.0, 0.3)
    projection_aggressive_delta = clamp(parse_float(params.get("projection_aggressive_delta"), 0.04), 0.0, 0.3)
    confidence_floor = clamp(parse_float(params.get("strategy_confidence_floor"), 0.45), 0.2, 0.95)
    confidence_ceiling = clamp(parse_float(params.get("strategy_confidence_ceiling"), 0.92), confidence_floor, 0.99)
    intel_bundle = get_intel_parameter_bundle(conn)
    intel_overlay = build_intelligence_bias_map(
        conn,
        decay_days=intel_bundle["decay_days"],
        w_commentary=intel_bundle["w_commentary"],
        w_policy=intel_bundle["w_policy"],
        w_financials=intel_bundle["w_financials"],
        w_chart=intel_bundle["w_chart"],
    )
    macro = build_macro_thoughts()
    macro_regime = str(macro.get("regime", "neutral")).lower()
    macro_score = parse_float(macro.get("score"), 0.0)
    macro_confidence = parse_float(macro.get("confidence"), 0.45)
    macro_thought = str(macro.get("thought", "Macro bias neutral."))

    items = collect_strategy_universe(conn, lookback_days=lookback_days)
    holdings = [i for i in items if parse_float(i.get("qty"), 0.0) > 0]
    total_market = sum(max(parse_float(i.get("market_value"), 0.0), 0.0) for i in holdings)
    holding_count = len(holdings)
    target_weight_base = (1.0 / holding_count) if holding_count > 0 else 0.0
    target_weight = min(allocation_limit, target_weight_base * 1.1) if holding_count > 0 else 0.0

    recommendations = []

    for item in holdings:
        symbol = item["symbol"]
        ltp = parse_float(item.get("ltp"), 0.0)
        avg_cost = parse_float(item.get("avg_cost"), 0.0)
        market_value = parse_float(item.get("market_value"), 0.0)
        current_weight = (market_value / total_market) if total_market > 0 else 0.0
        delta_weight = target_weight - current_weight
        buy_signal = str(item.get("buy_signal") or "").upper()
        sell_signal = str(item.get("sell_signal") or "").upper()
        upl_pct = parse_float(item.get("upl_pct"), 0.0)
        day_change_pct = parse_float(item.get("day_change_pct"), 0.0)
        pct_from_peak = parse_float(
            item.get("pct_from_peak_traded"),
            parse_float(item.get("pct_from_peak_buy"), 0.0),
        )
        momentum = parse_float(item.get("momentum_lookback_pct"), 0.0)
        total_return_pct = parse_float(item.get("total_return_pct"), 0.0)
        intel_item = intel_overlay["symbols"].get(symbol_upper(symbol), {})
        intel_score = parse_float(intel_item.get("score"), 0.0)
        intel_conf = parse_float(intel_item.get("confidence"), 0.0)
        intel_detail = str(intel_item.get("summary") or "")
        if intel_score >= 12:
            intel_summary = f"Bullish bias ({intel_score:+.1f})"
        elif intel_score <= -12:
            intel_summary = f"Cautious bias ({intel_score:+.1f})"
        else:
            intel_summary = f"Neutral bias ({intel_score:+.1f})"
        if intel_detail and intel_detail.lower() != "no strong intelligence bias.":
            intel_summary = f"{intel_summary}: {intel_detail}"

        buy_price_1 = max(0.01, (avg_cost * (1.0 - buy_l1)) if avg_cost > 0 else (ltp * 0.98))
        buy_price_2 = max(0.01, (avg_cost * (1.0 - buy_l2)) if avg_cost > 0 else (ltp * 0.94))
        sell_price_1 = max(0.01, (avg_cost * (1.0 + sell_s1)) if avg_cost > 0 else (ltp * 1.08))
        sell_price_2 = max(
            0.01,
            (avg_cost * (1.0 + min(0.35, sell_s2))) if avg_cost > 0 else (ltp * 1.18),
        )

        action = "HOLD"
        reasons = []

        if current_weight > max_position_weight or (current_weight - target_weight) > trim_trigger:
            action = "TRIM"
            reasons.append(
                f"Position overweight ({current_weight:.2%}) versus target ({target_weight:.2%})."
            )

        if sell_signal in ("S2", "S3"):
            action = "TRIM"
            reasons.append(f"Workbook signal {sell_signal} indicates profit-booking zone.")

        if action == "HOLD" and (
            (delta_weight > add_trigger and (ltp <= buy_price_1 or pct_from_peak <= -8.0))
            or buy_signal in ("BUY", "B1")
        ):
            action = "ADD"
            reasons.append(
                f"Under target by {delta_weight:.2%} with pullback ({pct_from_peak:.2f}% from peak traded)."
            )

        if action == "HOLD" and momentum <= -16.0 and current_weight > target_weight:
            action = "REVIEW"
            reasons.append(f"Weak {lookback_days}-day momentum ({momentum:.2f}%) while overweight.")

        if macro_regime == "risk_off":
            if action == "ADD" and delta_weight <= add_trigger * 2:
                action = "HOLD"
                reasons.append("Macro risk-off: delayed non-urgent adds.")
            if action == "HOLD" and momentum < -10:
                action = "REVIEW"
                reasons.append("Macro risk-off plus weak momentum.")
        elif macro_regime == "risk_on":
            if action == "TRIM" and current_weight <= max_position_weight and momentum > 4:
                action = "HOLD"
                reasons.append("Macro risk-on: let strength run unless overweight.")

        if action == "HOLD":
            if intel_score >= 20 and delta_weight > (add_trigger * 0.4) and current_weight < max_position_weight:
                action = "ADD"
                reasons.append(f"Intelligence overlay positive ({intel_score:+.1f}) supports accumulation.")
            elif intel_score <= -24 and current_weight > target_weight:
                action = "REVIEW"
                reasons.append(f"Intelligence overlay negative ({intel_score:+.1f}) flags caution.")
        elif action == "ADD" and intel_score <= -20:
            action = "HOLD"
            reasons.append(f"ADD tempered by negative intelligence overlay ({intel_score:+.1f}).")
        elif action == "TRIM" and intel_score >= 24 and current_weight <= max_position_weight and momentum > 0:
            action = "HOLD"
            reasons.append(f"TRIM relaxed due to strong positive intelligence overlay ({intel_score:+.1f}).")

        if action == "HOLD":
            reasons.append("Within allocation and signal thresholds.")

        confidence = 0.58
        confidence += min(0.22, abs(delta_weight) * 2.2)
        confidence += min(0.08, abs(momentum) / 120.0)
        confidence += min(0.06, abs(macro_score) * 0.03)
        confidence += min(0.04, macro_confidence * 0.05)
        confidence += min(0.08, (abs(intel_score) / 100.0) * (0.6 + intel_conf))
        if buy_signal:
            confidence += 0.04
        if sell_signal:
            confidence += 0.04
        if action in ("ADD", "TRIM", "REVIEW"):
            confidence += 0.04
        if action == "HOLD":
            confidence -= 0.07
        confidence = clamp(confidence, confidence_floor, confidence_ceiling)

        expected_annual_return = clamp(
            (momentum / 100.0) * 1.8
            + (day_change_pct / 100.0) * 8.0
            + (total_return_pct / 100.0) * 0.25
            + (macro_score / 100.0) * 15.0
            + (intel_score / 100.0) * 12.0,
            -0.35,
            0.45,
        )

        priority = int(
            round(
                abs(delta_weight) * 1000.0
                + abs(pct_from_peak) * 1.1
                + abs(upl_pct) * 0.5
                + abs(momentum) * 0.35
                + abs(intel_score) * 0.5
            )
        )
        if action == "TRIM":
            priority += 60
        elif action == "ADD":
            priority += 55
        elif action == "REVIEW":
            priority += 50
        else:
            priority += 20

        recommendations.append(
            {
                "symbol": symbol,
                "action": action,
                "priority": priority,
                "weight_current": round(current_weight, 6),
                "weight_target": round(target_weight, 6),
                "delta_weight": round(delta_weight, 6),
                "confidence": round(confidence, 4),
                "price_now": round(ltp, 4),
                "buy_price_1": round(buy_price_1, 4),
                "buy_price_2": round(buy_price_2, 4),
                "sell_price_1": round(sell_price_1, 4),
                "sell_price_2": round(sell_price_2, 4),
                "expected_annual_return": round(expected_annual_return, 6),
                "intel_score": round(intel_score, 4),
                "intel_confidence": round(intel_conf, 4),
                "intel_summary": intel_summary,
                "reason": " ".join(reasons[:2]).strip(),
                "source": "rotation_existing",
            }
        )

    fresh_candidates = []
    for item in items:
        if parse_float(item.get("qty"), 0.0) > 0:
            continue
        ltp = parse_float(item.get("ltp"), 0.0)
        if ltp <= 0:
            continue
        momentum = parse_float(item.get("momentum_lookback_pct"), 0.0)
        day_change_pct = parse_float(item.get("day_change_pct"), 0.0)
        score = momentum + (0.8 * day_change_pct)
        fresh_candidates.append((score, item))

    fresh_candidates.sort(key=lambda x: x[0], reverse=True)
    for score, item in fresh_candidates[:max_new_ideas]:
        symbol = item["symbol"]
        ltp = parse_float(item.get("ltp"), 0.0)
        momentum = parse_float(item.get("momentum_lookback_pct"), 0.0)
        intel_item = intel_overlay["symbols"].get(symbol_upper(symbol), {})
        intel_score = parse_float(intel_item.get("score"), 0.0)
        intel_conf = parse_float(intel_item.get("confidence"), 0.0)
        intel_detail = str(intel_item.get("summary") or "")
        if intel_score >= 12:
            intel_summary = f"Bullish bias ({intel_score:+.1f})"
        elif intel_score <= -12:
            intel_summary = f"Cautious bias ({intel_score:+.1f})"
        else:
            intel_summary = f"Neutral bias ({intel_score:+.1f})"
        if intel_detail and intel_detail.lower() != "no strong intelligence bias.":
            intel_summary = f"{intel_summary}: {intel_detail}"
        confidence = clamp(
            0.5
            + min(0.22, max(0.0, momentum) / 40.0)
            + min(0.05, abs(macro_score) * 0.02)
            + min(0.07, (abs(intel_score) / 100.0) * (0.6 + intel_conf)),
            confidence_floor,
            confidence_ceiling,
        )
        expected_annual_return = clamp(
            (momentum / 100.0) * 1.5 + (macro_score / 100.0) * 12.0 + (intel_score / 100.0) * 10.0,
            -0.2,
            0.45,
        )
        recommendations.append(
            {
                "symbol": symbol,
                "action": "WATCH_ADD",
                "priority": int(round(90 + max(0.0, score) * 4.5)),
                "weight_current": 0.0,
                "weight_target": round(min(allocation_limit, target_weight_base if target_weight_base > 0 else 0.08), 6),
                "delta_weight": round(min(allocation_limit, target_weight_base if target_weight_base > 0 else 0.08), 6),
                "confidence": round(confidence, 4),
                "price_now": round(ltp, 4),
                "buy_price_1": round(max(0.01, ltp * 0.98), 4),
                "buy_price_2": round(max(0.01, ltp * 0.94), 4),
                "sell_price_1": round(max(0.01, ltp * 1.12), 4),
                "sell_price_2": round(max(0.01, ltp * 1.22), 4),
                "expected_annual_return": round(expected_annual_return, 6),
                "intel_score": round(intel_score, 4),
                "intel_confidence": round(intel_conf, 4),
                "intel_summary": intel_summary,
                "reason": (
                    f"Not in current holdings. Positive {lookback_days}-day momentum ({momentum:.2f}%). "
                    f"Suggested as optional rotation candidate."
                ),
                "source": "rotation_new_idea",
            }
        )

    recommendations.sort(
        key=lambda r: (
            0 if r["source"] == "rotation_existing" else 1,
            -parse_float(r.get("priority"), 0.0),
            r["symbol"],
        )
    )

    summary = portfolio_summary(conn)
    invested_value = parse_float(summary.get("market_deployment"), parse_float(summary.get("invested"), 0.0))
    market_value = parse_float(summary.get("market_value"), 0.0)
    cash_balance = parse_float(summary.get("cash_balance"), 0.0)
    projected_start_value = market_value + cash_balance
    if projected_start_value <= 0:
        projected_start_value = market_value if market_value > 0 else max(invested_value, 1.0)

    daily_returns = portfolio_daily_returns(conn, lookback_days=730)
    if len(daily_returns) >= 45:
        mean_daily = sum(daily_returns) / len(daily_returns)
        variance = 0.0
        if len(daily_returns) > 1:
            variance = sum((r - mean_daily) ** 2 for r in daily_returns) / (len(daily_returns) - 1)
        annualized_return = clamp(mean_daily * 252.0, -0.35, 0.45)
        annualized_vol = math.sqrt(max(variance, 0.0)) * math.sqrt(252.0)
    else:
        annualized_return = projection_base_return
        annualized_vol = 0.22
    annualized_return = clamp(
        annualized_return + (parse_float(intel_overlay.get("portfolio_score"), 0.0) / 100.0) * 0.06,
        -0.4,
        0.5,
    )

    scenario_returns = {
        "conservative": clamp(annualized_return - projection_conservative_delta, -0.5, 0.5),
        "base": clamp(annualized_return, -0.5, 0.5),
        "aggressive": clamp(annualized_return + projection_aggressive_delta, -0.5, 0.7),
    }
    projection_points = []
    for scenario, ann_ret in scenario_returns.items():
        safe_ann_ret = max(-0.95, ann_ret)
        for y in range(0, projection_years + 1):
            projection_points.append(
                {
                    "scenario": scenario,
                    "year_offset": y,
                    "annual_return": round(safe_ann_ret, 6),
                    "projected_value": round(projected_start_value * ((1.0 + safe_ann_ret) ** y), 2),
                }
            )

    counts = {"ADD": 0, "TRIM": 0, "HOLD": 0, "REVIEW": 0, "WATCH_ADD": 0}
    for r in recommendations:
        a = str(r["action"]).upper()
        if a in counts:
            counts[a] += 1

    return {
        "run_date": run_date,
        "generated_at": now_iso(),
        "market_value": round(market_value, 2),
        "invested_value": round(invested_value, 2),
        "cash_balance": round(cash_balance, 2),
        "projected_start_value": round(projected_start_value, 2),
        "macro": {
            "as_of": macro.get("as_of"),
            "regime": macro_regime,
            "score": round(macro_score, 3),
            "confidence": round(macro_confidence, 4),
            "signals": macro.get("signals", {}),
            "thought": macro_thought,
        },
        "intelligence": {
            "score": round(parse_float(intel_overlay.get("portfolio_score"), 0.0), 4),
            "confidence": round(parse_float(intel_overlay.get("portfolio_confidence"), 0.0), 4),
            "documents_recent": int(intel_overlay.get("documents_recent", 0)),
            "impacts_recent": int(intel_overlay.get("impacts_recent", 0)),
            "thought": str(intel_overlay.get("thought") or ""),
            "weights": intel_overlay.get("weights", {}),
            "cross_flows": intel_overlay.get("cross_flows", []),
        },
        "counts": counts,
        "recommendations": recommendations,
        "projection_points": projection_points,
        "meta": {
            "lookback_days": lookback_days,
            "projection_years": projection_years,
            "daily_returns_samples": len(daily_returns),
            "annualized_return_base": round(annualized_return, 6),
            "annualized_volatility": round(annualized_vol, 6),
            "max_new_ideas": max_new_ideas,
        },
    }


def persist_strategy_insights(conn, insights):
    run_date = insights["run_date"]
    counts = insights["counts"]
    macro = insights.get("macro", {})
    intel = insights.get("intelligence", {})
    conn.execute(
        """
        INSERT INTO strategy_runs(
          run_date, created_at, market_value, invested_value, cash_balance, projected_start_value,
          macro_regime, macro_score, macro_confidence, macro_thought,
          intel_score, intel_confidence, intel_thought,
          add_count, trim_count, hold_count, review_count, watch_add_count
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(run_date) DO UPDATE SET
          created_at=excluded.created_at,
          market_value=excluded.market_value,
          invested_value=excluded.invested_value,
          cash_balance=excluded.cash_balance,
          projected_start_value=excluded.projected_start_value,
          macro_regime=excluded.macro_regime,
          macro_score=excluded.macro_score,
          macro_confidence=excluded.macro_confidence,
          macro_thought=excluded.macro_thought,
          intel_score=excluded.intel_score,
          intel_confidence=excluded.intel_confidence,
          intel_thought=excluded.intel_thought,
          add_count=excluded.add_count,
          trim_count=excluded.trim_count,
          hold_count=excluded.hold_count,
          review_count=excluded.review_count,
          watch_add_count=excluded.watch_add_count
        """,
        (
            run_date,
            insights["generated_at"],
            parse_float(insights.get("market_value"), 0.0),
            parse_float(insights.get("invested_value"), 0.0),
            parse_float(insights.get("cash_balance"), 0.0),
            parse_float(insights.get("projected_start_value"), 0.0),
            str(macro.get("regime", "neutral")),
            parse_float(macro.get("score"), 0.0),
            parse_float(macro.get("confidence"), 0.0),
            str(macro.get("thought", "")),
            parse_float(intel.get("score"), 0.0),
            parse_float(intel.get("confidence"), 0.0),
            str(intel.get("thought", "")),
            int(counts.get("ADD", 0)),
            int(counts.get("TRIM", 0)),
            int(counts.get("HOLD", 0)),
            int(counts.get("REVIEW", 0)),
            int(counts.get("WATCH_ADD", 0)),
        ),
    )
    conn.execute("DELETE FROM strategy_recommendations WHERE run_date = ?", (run_date,))
    rec_rows = []
    for r in insights.get("recommendations", []):
        rec_rows.append(
            (
                run_date,
                r["symbol"],
                r["action"],
                int(parse_float(r.get("priority"), 0.0)),
                parse_float(r.get("weight_current"), 0.0),
                parse_float(r.get("weight_target"), 0.0),
                parse_float(r.get("delta_weight"), 0.0),
                parse_float(r.get("confidence"), 0.0),
                parse_float(r.get("price_now"), 0.0),
                parse_float(r.get("buy_price_1"), 0.0),
                parse_float(r.get("buy_price_2"), 0.0),
                parse_float(r.get("sell_price_1"), 0.0),
                parse_float(r.get("sell_price_2"), 0.0),
                parse_float(r.get("expected_annual_return"), 0.0),
                parse_float(r.get("intel_score"), 0.0),
                parse_float(r.get("intel_confidence"), 0.0),
                str(r.get("intel_summary", "")),
                str(r.get("reason", "")),
                str(r.get("source", "rotation_engine")),
            )
        )
    if rec_rows:
        conn.executemany(
            """
            INSERT INTO strategy_recommendations(
              run_date, symbol, action, priority, weight_current, weight_target, delta_weight,
              confidence, price_now, buy_price_1, buy_price_2, sell_price_1, sell_price_2,
              expected_annual_return, intel_score, intel_confidence, intel_summary, reason, source
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rec_rows,
        )

    conn.execute("DELETE FROM strategy_projection_points WHERE run_date = ?", (run_date,))
    proj_rows = []
    for p in insights.get("projection_points", []):
        proj_rows.append(
            (
                run_date,
                str(p["scenario"]),
                int(parse_float(p.get("year_offset"), 0.0)),
                parse_float(p.get("annual_return"), 0.0),
                parse_float(p.get("projected_value"), 0.0),
            )
        )
    if proj_rows:
        conn.executemany(
            """
            INSERT INTO strategy_projection_points(run_date, scenario, year_offset, annual_return, projected_value)
            VALUES (?, ?, ?, ?, ?)
            """,
            proj_rows,
        )


def load_latest_strategy_insights(conn):
    run = conn.execute(
        """
        SELECT
          run_date, created_at, market_value, invested_value, cash_balance, projected_start_value,
          macro_regime, macro_score, macro_confidence, macro_thought,
          intel_score, intel_confidence, intel_thought,
          add_count, trim_count, hold_count, review_count, watch_add_count
        FROM strategy_runs
        ORDER BY run_date DESC
        LIMIT 1
        """
    ).fetchone()
    if not run:
        return None

    run_date = run["run_date"]
    recs = [
        {
            "symbol": r["symbol"],
            "action": r["action"],
            "priority": int(parse_float(r["priority"], 0.0)),
            "weight_current": round(parse_float(r["weight_current"], 0.0), 6),
            "weight_target": round(parse_float(r["weight_target"], 0.0), 6),
            "delta_weight": round(parse_float(r["delta_weight"], 0.0), 6),
            "confidence": round(parse_float(r["confidence"], 0.0), 4),
            "price_now": round(parse_float(r["price_now"], 0.0), 4),
            "buy_price_1": round(parse_float(r["buy_price_1"], 0.0), 4),
            "buy_price_2": round(parse_float(r["buy_price_2"], 0.0), 4),
            "sell_price_1": round(parse_float(r["sell_price_1"], 0.0), 4),
            "sell_price_2": round(parse_float(r["sell_price_2"], 0.0), 4),
            "expected_annual_return": round(parse_float(r["expected_annual_return"], 0.0), 6),
            "intel_score": round(parse_float(r["intel_score"], 0.0), 4),
            "intel_confidence": round(parse_float(r["intel_confidence"], 0.0), 4),
            "intel_summary": str(r["intel_summary"] or ""),
            "reason": r["reason"] or "",
            "source": r["source"] or "rotation_engine",
        }
        for r in conn.execute(
            """
            SELECT
              symbol, action, priority, weight_current, weight_target, delta_weight, confidence,
              price_now, buy_price_1, buy_price_2, sell_price_1, sell_price_2,
              expected_annual_return, intel_score, intel_confidence, intel_summary, reason, source
            FROM strategy_recommendations
            WHERE run_date = ?
            ORDER BY priority DESC, symbol
            """,
            (run_date,),
        ).fetchall()
    ]

    projection_rows = conn.execute(
        """
        SELECT scenario, year_offset, annual_return, projected_value
        FROM strategy_projection_points
        WHERE run_date = ?
        ORDER BY scenario, year_offset
        """,
        (run_date,),
    ).fetchall()
    scenario_map = {}
    for p in projection_rows:
        sc = str(p["scenario"])
        if sc not in scenario_map:
            scenario_map[sc] = {
                "scenario": sc,
                "annual_return": round(parse_float(p["annual_return"], 0.0), 6),
                "points": [],
            }
        scenario_map[sc]["points"].append(
            {
                "year_offset": int(parse_float(p["year_offset"], 0.0)),
                "projected_value": round(parse_float(p["projected_value"], 0.0), 2),
            }
        )
    scenario_order = {"conservative": 0, "base": 1, "aggressive": 2}
    scenarios = sorted(
        scenario_map.values(),
        key=lambda x: scenario_order.get(x["scenario"], 99),
    )
    projection_years = 0
    for sc in scenarios:
        for p in sc["points"]:
            projection_years = max(projection_years, int(p["year_offset"]))

    return {
        "run_date": run_date,
        "generated_at": run["created_at"],
        "market_value": round(parse_float(run["market_value"], 0.0), 2),
        "invested_value": round(parse_float(run["invested_value"], 0.0), 2),
        "cash_balance": round(parse_float(run["cash_balance"], 0.0), 2),
        "projected_start_value": round(parse_float(run["projected_start_value"], 0.0), 2),
        "macro": {
            "regime": str(run["macro_regime"] or "neutral"),
            "score": round(parse_float(run["macro_score"], 0.0), 3),
            "confidence": round(parse_float(run["macro_confidence"], 0.0), 4),
            "thought": str(run["macro_thought"] or ""),
        },
        "intelligence": {
            "score": round(parse_float(run["intel_score"], 0.0), 4),
            "confidence": round(parse_float(run["intel_confidence"], 0.0), 4),
            "thought": str(run["intel_thought"] or ""),
        },
        "counts": {
            "ADD": int(run["add_count"]),
            "TRIM": int(run["trim_count"]),
            "HOLD": int(run["hold_count"]),
            "REVIEW": int(run["review_count"]),
            "WATCH_ADD": int(run["watch_add_count"]),
        },
        "recommendations": recs,
        "projections": {
            "years": projection_years,
            "scenarios": scenarios,
        },
    }


def latest_strategy_recommendation_map(conn):
    run = conn.execute("SELECT MAX(run_date) AS run_date FROM strategy_runs").fetchone()
    run_date = str(run["run_date"] or "") if run else ""
    if not run_date:
        return {}
    rows = conn.execute(
        """
        SELECT symbol, action, priority, confidence, intel_score, intel_confidence, intel_summary, reason
        FROM strategy_recommendations
        WHERE run_date = ?
        ORDER BY priority DESC, id DESC
        """,
        (run_date,),
    ).fetchall()
    out = {}
    for r in rows:
        sym = symbol_upper(r["symbol"])
        if not sym or sym in out:
            continue
        out[sym] = {
            "action": str(r["action"] or "").upper(),
            "priority": int(parse_float(r["priority"], 0.0)),
            "confidence": round(parse_float(r["confidence"], 0.0), 4),
            "intel_score": round(parse_float(r["intel_score"], 0.0), 4),
            "intel_confidence": round(parse_float(r["intel_confidence"], 0.0), 4),
            "intel_summary": str(r["intel_summary"] or ""),
            "reason": str(r["reason"] or ""),
        }
    return out


def _strategy_audit_finding(severity, code, title, detail="", symbol="", metric_value=None, expected_range=""):
    sev = str(severity or "info").strip().lower() or "info"
    if sev not in ("critical", "warn", "info"):
        sev = "info"
    return {
        "severity": sev,
        "code": str(code or "").strip().lower() or "general",
        "title": str(title or "").strip() or "Finding",
        "detail": str(detail or "").strip(),
        "symbol": symbol_upper(symbol) if symbol else "",
        "metric_value": None if metric_value is None else round(parse_float(metric_value, 0.0), 4),
        "expected_range": str(expected_range or "").strip(),
    }


def _persist_strategy_audit_run(conn, payload):
    ts = now_iso()
    findings = list(payload.get("findings") or [])
    stats = dict(payload.get("stats") or {})
    conn.execute(
        """
        INSERT INTO strategy_audit_runs(
          created_at, strategy_run_date, audit_mode, overall_status, overall_score,
          summary, recommendation, findings_count, stats_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            ts,
            str(payload.get("strategy_run_date") or ""),
            str(payload.get("audit_mode") or "heuristic"),
            str(payload.get("overall_status") or "ok"),
            round(parse_float(payload.get("overall_score"), 0.0), 2),
            str(payload.get("summary") or ""),
            str(payload.get("recommendation") or ""),
            len(findings),
            _safe_json_dumps(stats),
        ),
    )
    audit_id = int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
    for f in findings:
        conn.execute(
            """
            INSERT INTO strategy_audit_findings(
              audit_id, severity, code, title, detail, symbol, metric_value, expected_range, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                audit_id,
                str(f.get("severity") or "info"),
                str(f.get("code") or ""),
                str(f.get("title") or ""),
                str(f.get("detail") or ""),
                str(f.get("symbol") or ""),
                (None if f.get("metric_value") is None else round(parse_float(f.get("metric_value"), 0.0), 4)),
                str(f.get("expected_range") or ""),
                ts,
            ),
        )
    payload["audit_id"] = audit_id
    payload["created_at"] = ts
    return payload


def get_strategy_audit_run(conn, audit_id):
    row = conn.execute(
        """
        SELECT id, created_at, strategy_run_date, audit_mode, overall_status, overall_score,
               summary, recommendation, findings_count, stats_json
        FROM strategy_audit_runs
        WHERE id = ?
        """,
        (int(parse_float(audit_id, 0.0)),),
    ).fetchone()
    if not row:
        return None
    findings = [
        {
            "id": int(parse_float(r["id"], 0.0)),
            "severity": str(r["severity"] or "info"),
            "code": str(r["code"] or ""),
            "title": str(r["title"] or ""),
            "detail": str(r["detail"] or ""),
            "symbol": str(r["symbol"] or ""),
            "metric_value": None if r["metric_value"] is None else round(parse_float(r["metric_value"], 0.0), 4),
            "expected_range": str(r["expected_range"] or ""),
            "created_at": str(r["created_at"] or ""),
        }
        for r in conn.execute(
            """
            SELECT id, severity, code, title, detail, symbol, metric_value, expected_range, created_at
            FROM strategy_audit_findings
            WHERE audit_id = ?
            ORDER BY CASE LOWER(severity) WHEN 'critical' THEN 0 WHEN 'warn' THEN 1 ELSE 2 END, id ASC
            """,
            (int(parse_float(audit_id, 0.0)),),
        ).fetchall()
    ]
    out = {
        "audit_id": int(parse_float(row["id"], 0.0)),
        "created_at": str(row["created_at"] or ""),
        "strategy_run_date": str(row["strategy_run_date"] or ""),
        "audit_mode": str(row["audit_mode"] or "heuristic"),
        "overall_status": str(row["overall_status"] or "ok"),
        "overall_score": round(parse_float(row["overall_score"], 0.0), 2),
        "summary": str(row["summary"] or ""),
        "recommendation": str(row["recommendation"] or ""),
        "findings_count": int(parse_float(row["findings_count"], 0.0)),
        "stats": _safe_json_loads(row["stats_json"], {}),
        "findings": findings,
    }
    return out


def list_strategy_audit_runs(conn, limit=25):
    lim = max(1, min(200, int(parse_float(limit, 25))))
    rows = conn.execute(
        """
        SELECT id, created_at, strategy_run_date, audit_mode, overall_status, overall_score,
               summary, recommendation, findings_count, stats_json
        FROM strategy_audit_runs
        ORDER BY id DESC
        LIMIT ?
        """,
        (lim,),
    ).fetchall()
    items = []
    for r in rows:
        stats = _safe_json_loads(r["stats_json"], {})
        items.append(
            {
                "audit_id": int(parse_float(r["id"], 0.0)),
                "created_at": str(r["created_at"] or ""),
                "strategy_run_date": str(r["strategy_run_date"] or ""),
                "audit_mode": str(r["audit_mode"] or "heuristic"),
                "overall_status": str(r["overall_status"] or "ok"),
                "overall_score": round(parse_float(r["overall_score"], 0.0), 2),
                "summary": str(r["summary"] or ""),
                "recommendation": str(r["recommendation"] or ""),
                "findings_count": int(parse_float(r["findings_count"], 0.0)),
                "critical_count": int(parse_float(stats.get("critical_count"), 0.0)),
                "warn_count": int(parse_float(stats.get("warn_count"), 0.0)),
                "info_count": int(parse_float(stats.get("info_count"), 0.0)),
            }
        )
    latest = get_strategy_audit_run(conn, items[0]["audit_id"]) if items else None
    return {"items": items, "latest": latest}


def run_strategy_audit(conn, refresh_strategy=False):
    refresh_flag = bool(refresh_strategy)
    today = dt.date.today().isoformat()
    insights = load_latest_strategy_insights(conn)
    if refresh_flag or insights is None:
        fresh = build_strategy_insights(conn, run_date=today)
        persist_strategy_insights(conn, fresh)
        conn.commit()
        insights = load_latest_strategy_insights(conn)
    if not insights:
        raise ValueError("strategy_insights_unavailable")

    recs = list(insights.get("recommendations") or [])
    counts = dict(insights.get("counts") or {})
    macro = dict(insights.get("macro") or {})
    intel = dict(insights.get("intelligence") or {})
    bt = latest_backtest_run(conn)
    findings = []

    total = len(recs)
    addish = [r for r in recs if str(r.get("action") or "").upper() in ("ADD", "WATCH_ADD")]
    trimish = [r for r in recs if str(r.get("action") or "").upper() == "TRIM"]
    reviewish = [r for r in recs if str(r.get("action") or "").upper() == "REVIEW"]
    low_conf = [r for r in recs if parse_float(r.get("confidence"), 0.0) < 0.45]
    negative_add = [r for r in addish if parse_float(r.get("expected_annual_return"), 0.0) <= 0]
    positive_trim = [r for r in trimish if parse_float(r.get("expected_annual_return"), 0.0) > 0.12]
    missing_reason = [r for r in recs if not str(r.get("reason") or "").strip()]
    overweight = [r for r in recs if parse_float(r.get("weight_target"), 0.0) > 0.18]
    weight_sum = sum(max(0.0, parse_float(r.get("weight_target"), 0.0)) for r in recs)

    if total == 0:
        findings.append(_strategy_audit_finding("critical", "no_recommendations", "No strategy recommendations available", "Run strategy refresh before taking trades."))
    if total > 0 and (len(low_conf) / max(1, total)) >= 0.45:
        findings.append(
            _strategy_audit_finding(
                "warn",
                "low_confidence_mix",
                "Large share of recommendations have low confidence",
                f"{len(low_conf)} of {total} recommendations are below 45% confidence.",
                metric_value=(len(low_conf) / max(1, total)) * 100.0,
                expected_range="< 45%",
            )
        )
    if reviewish and (len(reviewish) / max(1, total)) >= 0.30:
        findings.append(
            _strategy_audit_finding(
                "warn",
                "too_many_review_flags",
                "Too many REVIEW recommendations",
                f"{len(reviewish)} of {total} names are flagged REVIEW, which weakens actionability.",
                metric_value=(len(reviewish) / max(1, total)) * 100.0,
                expected_range="< 30%",
            )
        )
    for r in negative_add[:8]:
        findings.append(
            _strategy_audit_finding(
                "critical",
                "negative_add_expectation",
                "ADD recommendation has non-positive expected return",
                f"{r.get('symbol')} is marked {r.get('action')} even though expected annual return is {parse_float(r.get('expected_annual_return'), 0.0) * 100.0:+.1f}%.",
                symbol=r.get("symbol"),
                metric_value=parse_float(r.get("expected_annual_return"), 0.0) * 100.0,
                expected_range="> 0%",
            )
        )
    for r in positive_trim[:5]:
        findings.append(
            _strategy_audit_finding(
                "warn",
                "trim_positive_expectation",
                "TRIM recommendation still has strong positive expectation",
                f"{r.get('symbol')} is marked TRIM with expected annual return {parse_float(r.get('expected_annual_return'), 0.0) * 100.0:+.1f}%. Validate whether sizing logic is too defensive.",
                symbol=r.get("symbol"),
                metric_value=parse_float(r.get("expected_annual_return"), 0.0) * 100.0,
                expected_range="<= 12%",
            )
        )
    if missing_reason:
        findings.append(
            _strategy_audit_finding(
                "warn",
                "missing_reason_text",
                "Some recommendations are missing reasons",
                f"{len(missing_reason)} recommendations do not include a reason string.",
                metric_value=len(missing_reason),
                expected_range="0",
            )
        )
    if weight_sum > 1.15 or (weight_sum > 0 and weight_sum < 0.85):
        findings.append(
            _strategy_audit_finding(
                "warn",
                "target_weight_sum",
                "Target weights look imbalanced",
                f"Total target weight sums to {weight_sum:.2f}.",
                metric_value=weight_sum,
                expected_range="0.85 to 1.15",
            )
        )
    if overweight:
        top = sorted(overweight, key=lambda x: parse_float(x.get("weight_target"), 0.0), reverse=True)[0]
        findings.append(
            _strategy_audit_finding(
                "warn",
                "single_name_concentration",
                "Single-name concentration is high",
                f"{top.get('symbol')} target weight is {parse_float(top.get('weight_target'), 0.0) * 100.0:.1f}%.",
                symbol=top.get("symbol"),
                metric_value=parse_float(top.get("weight_target"), 0.0) * 100.0,
                expected_range="<= 18%",
            )
        )
    macro_conf = parse_float(macro.get("confidence"), 0.0)
    intel_conf = parse_float(intel.get("confidence"), 0.0)
    if macro_conf < 0.45:
        findings.append(
            _strategy_audit_finding(
                "warn",
                "macro_low_confidence",
                "Macro layer confidence is low",
                f"Macro confidence is only {macro_conf * 100.0:.1f}%.",
                metric_value=macro_conf * 100.0,
                expected_range=">= 45%",
            )
        )
    if intel_conf < 0.45:
        findings.append(
            _strategy_audit_finding(
                "warn",
                "intel_low_confidence",
                "Intelligence layer confidence is low",
                f"Intelligence confidence is only {intel_conf * 100.0:.1f}%.",
                metric_value=intel_conf * 100.0,
                expected_range=">= 45%",
            )
        )
    if bt is None:
        findings.append(_strategy_audit_finding("warn", "backtest_missing", "No recent backtest available", "Run a backtest to validate signal quality before trusting the audit."))
    else:
        hit_rate = parse_float(bt.get("hit_rate"), 0.0)
        sample_count = int(parse_float(bt.get("sample_count"), 0.0))
        age_days = 0
        try:
            bt_dt = dt.datetime.fromisoformat(str(bt.get("created_at") or ""))
            age_days = max(0, int((dt.datetime.now() - bt_dt).total_seconds() // 86400))
        except Exception:
            age_days = 0
        if sample_count < 25:
            findings.append(
                _strategy_audit_finding(
                    "warn",
                    "backtest_small_sample",
                    "Backtest sample size is small",
                    f"Latest backtest has only {sample_count} samples.",
                    metric_value=sample_count,
                    expected_range=">= 25",
                )
            )
        if hit_rate < 0.45:
            sev = "critical" if hit_rate < 0.40 else "warn"
            findings.append(
                _strategy_audit_finding(
                    sev,
                    "backtest_hit_rate_low",
                    "Backtest hit rate is below threshold",
                    f"Latest hit rate is {hit_rate * 100.0:.1f}% over {sample_count} samples.",
                    metric_value=hit_rate * 100.0,
                    expected_range=">= 45%",
                )
            )
        if age_days >= 14:
            findings.append(
                _strategy_audit_finding(
                    "info",
                    "backtest_stale",
                    "Backtest is getting stale",
                    f"Latest backtest was run {age_days} day(s) ago.",
                    metric_value=age_days,
                    expected_range="< 14 days",
                )
            )

    critical_count = sum(1 for f in findings if f.get("severity") == "critical")
    warn_count = sum(1 for f in findings if f.get("severity") == "warn")
    info_count = sum(1 for f in findings if f.get("severity") == "info")
    overall_score = clamp(100.0 - (critical_count * 22.0) - (warn_count * 9.0) - (info_count * 3.0), 0.0, 100.0)
    overall_status = "critical" if critical_count > 0 else "warn" if warn_count > 0 else "ok"
    summary = (
        f"{critical_count} critical, {warn_count} warning, {info_count} info finding(s). "
        f"Audit score {overall_score:.1f}/100."
    )
    recommendation = (
        "Refresh or retune strategy before acting on it."
        if overall_status == "critical"
        else "Use with caution; review flagged sizing/confidence issues."
        if overall_status == "warn"
        else "Strategy health looks acceptable under heuristic audit."
    )
    stats = {
        "recommendation_count": total,
        "add_count": len(addish),
        "trim_count": len(trimish),
        "review_count": len(reviewish),
        "low_confidence_count": len(low_conf),
        "negative_add_count": len(negative_add),
        "positive_trim_count": len(positive_trim),
        "missing_reason_count": len(missing_reason),
        "overweight_count": len(overweight),
        "target_weight_sum": round(weight_sum, 6),
        "macro_confidence": round(macro_conf, 4),
        "intel_confidence": round(intel_conf, 4),
        "critical_count": critical_count,
        "warn_count": warn_count,
        "info_count": info_count,
        "backtest_hit_rate": round(parse_float((bt or {}).get("hit_rate"), 0.0), 4),
        "backtest_sample_count": int(parse_float((bt or {}).get("sample_count"), 0.0)),
    }
    payload = {
        "strategy_run_date": str(insights.get("run_date") or ""),
        "audit_mode": "heuristic",
        "overall_status": overall_status,
        "overall_score": round(overall_score, 2),
        "summary": summary,
        "recommendation": recommendation,
        "stats": stats,
        "findings": findings,
        "strategy_snapshot": {
            "run_date": str(insights.get("run_date") or ""),
            "generated_at": str(insights.get("generated_at") or ""),
            "counts": counts,
            "macro": {
                "regime": str(macro.get("regime") or ""),
                "score": round(parse_float(macro.get("score"), 0.0), 4),
                "confidence": round(macro_conf, 4),
            },
            "intelligence": {
                "score": round(parse_float(intel.get("score"), 0.0), 4),
                "confidence": round(intel_conf, 4),
            },
        },
        "backtest": {
            "created_at": str((bt or {}).get("created_at") or ""),
            "hit_rate": round(parse_float((bt or {}).get("hit_rate"), 0.0), 4),
            "sample_count": int(parse_float((bt or {}).get("sample_count"), 0.0)),
            "avg_future_return": round(parse_float((bt or {}).get("avg_future_return"), 0.0), 4),
        },
    }
    _persist_strategy_audit_run(conn, payload)
    conn.commit()
    return payload


def _harvest_action_bias(action):
    act = str(action or "").strip().upper()
    if act == "TRIM":
        return -8.0
    if act == "REVIEW":
        return -5.0
    if act == "HOLD":
        return 0.0
    if act == "WATCH_ADD":
        return 4.0
    if act == "ADD":
        return 7.0
    return 0.0


def _harvest_signal_bias(buy_signal, sell_signal):
    score = 0.0
    b = str(buy_signal or "").upper()
    s = str(sell_signal or "").upper()
    if b in ("BUY", "B1"):
        score += 5.0
    elif b == "B2":
        score += 3.0
    if s == "S1":
        score -= 4.0
    elif s == "S2":
        score -= 7.0
    elif s == "S3":
        score -= 10.0
    return score


def _harvest_expected_move_score(intel_score, chart_score, fin_score, action_bias, signal_bias):
    score = (0.36 * chart_score) + (0.33 * intel_score) + (0.18 * fin_score) + action_bias + signal_bias
    return round(clamp(score, -100.0, 100.0), 4)


def _harvest_direction_label(expected_move_score):
    score = parse_float(expected_move_score, 0.0)
    if score >= 10.0:
        return "UP"
    if score <= -10.0:
        return "DOWN"
    return "MIXED"


def _harvest_priority_reason(direction, strategy_action, chart_signal, intel_summary, fin_summary):
    bits = []
    if direction == "DOWN":
        bits.append("trend-risk down")
    elif direction == "UP":
        bits.append("trend support up")
    if strategy_action:
        bits.append(f"strategy {strategy_action}")
    if chart_signal:
        bits.append(f"chart {chart_signal}")
    if intel_summary and intel_summary.lower() != "no strong intelligence bias.":
        bits.append(intel_summary)
    elif fin_summary:
        bits.append(fin_summary)
    return ", ".join(bits[:4]) if bits else "Signal mix neutral."


def _harvest_tax_bucket(buy_date, as_of_date=None):
    as_of = parse_history_date(as_of_date) or ist_now().date()
    buy_d = parse_history_date(buy_date)
    if not buy_d:
        return "STCG", 0
    held_days = max(0, (as_of - buy_d).days)
    return ("LTCG" if held_days >= 365 else "STCG"), held_days


def _harvest_tax_bucket_rank(bucket):
    return 0 if str(bucket or "").upper() == "STCG" else 1


def _harvest_tax_bucket_bias(bucket, side="loss"):
    b = str(bucket or "").upper()
    side_s = str(side or "loss").strip().lower()
    if b == "STCG":
        return 6.0 if side_s == "loss" else 3.0
    if b == "LTCG":
        return 2.0 if side_s == "loss" else 0.5
    return 0.0


def _india_fy_bounds(as_of_date=None):
    d = parse_history_date(as_of_date) or ist_now().date()
    start_year = d.year if d.month >= 4 else (d.year - 1)
    start = dt.date(start_year, 4, 1)
    end = dt.date(start_year + 1, 3, 31)
    return {
        "fy_label": f"FY{start_year}-{str(start_year + 1)[-2:]}",
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
    }


def compute_realized_equity_tax_summary(conn, as_of_date=None):
    fy = _india_fy_bounds(as_of_date=as_of_date)
    split_map = load_split_map(conn)
    tax_cfg = get_tax_profile_config(conn)
    start_d = dt.date.fromisoformat(fy["start_date"])
    end_d = dt.date.fromisoformat(fy["end_date"])
    rows = conn.execute(
        """
        SELECT t.id, UPPER(t.symbol) AS symbol, UPPER(t.side) AS side, t.trade_date, t.quantity, t.price,
               UPPER(COALESCE(i.asset_class, 'EQUITY')) AS asset_class
        FROM trades t
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(t.symbol)
        WHERE t.trade_date <= ?
        ORDER BY t.trade_date, t.id
        """,
        (fy["end_date"],),
    ).fetchall()
    lots_by_symbol = defaultdict(deque)
    stcg_gain = 0.0
    stcg_loss = 0.0
    ltcg_gain = 0.0
    ltcg_loss = 0.0
    for r in rows:
        if str(r["asset_class"] or "EQUITY").upper() == "GOLD":
            continue
        symbol = symbol_upper(r["symbol"])
        side = str(r["side"] or "").upper()
        trade_date = str(r["trade_date"] or "")[:10]
        q, p = adjusted_trade_values(symbol, trade_date, float(r["quantity"]), float(r["price"]), split_map)
        if q <= 0 or p <= 0:
            continue
        if side == "BUY":
            lots_by_symbol[symbol].append({"qty": q, "buy_price": p, "buy_date": trade_date})
            continue
        if side != "SELL":
            continue
        sell_d = parse_history_date(trade_date)
        remaining = q
        while remaining > 1e-9 and lots_by_symbol[symbol]:
            lot = lots_by_symbol[symbol][0]
            matched = min(remaining, parse_float(lot.get("qty"), 0.0))
            if matched <= 0:
                lots_by_symbol[symbol].popleft()
                continue
            if sell_d and start_d <= sell_d <= end_d:
                pnl = (p - parse_float(lot.get("buy_price"), 0.0)) * matched
                bucket, _held_days = _harvest_tax_bucket(lot.get("buy_date"), as_of_date=trade_date)
                if pnl >= 0:
                    if bucket == "LTCG":
                        ltcg_gain += pnl
                    else:
                        stcg_gain += pnl
                else:
                    if bucket == "LTCG":
                        ltcg_loss += abs(pnl)
                    else:
                        stcg_loss += abs(pnl)
            lot["qty"] = max(0.0, parse_float(lot.get("qty"), 0.0) - matched)
            if parse_float(lot.get("qty"), 0.0) <= 1e-9:
                lots_by_symbol[symbol].popleft()
            remaining -= matched
    ltcg_net_gain = max(0.0, ltcg_gain - ltcg_loss)
    stcg_net_gain = max(0.0, stcg_gain - stcg_loss)
    ltcg_remaining_exemption = max(0.0, parse_float(tax_cfg.get("ltcg_exemption_limit"), 125000.0) - ltcg_net_gain)
    return {
        "fy_label": fy["fy_label"],
        "fy_start_date": fy["start_date"],
        "fy_end_date": fy["end_date"],
        "stcg_gain": round(stcg_gain, 2),
        "stcg_loss": round(stcg_loss, 2),
        "ltcg_gain": round(ltcg_gain, 2),
        "ltcg_loss": round(ltcg_loss, 2),
        "stcg_net_gain": round(stcg_net_gain, 2),
        "ltcg_net_gain": round(ltcg_net_gain, 2),
        "ltcg_exemption_limit": round(parse_float(tax_cfg.get("ltcg_exemption_limit"), 125000.0), 2),
        "ltcg_remaining_exemption": round(ltcg_remaining_exemption, 2),
    }


def _parse_iso_datetime_safe(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return dt.datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception:
        return None


def _strip_html_to_text(raw_html):
    text = str(raw_html or "")
    text = re.sub(r"(?is)<script.*?>.*?</script>", " ", text)
    text = re.sub(r"(?is)<style.*?>.*?</style>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _parse_tax_monitor_snapshot(tax_text, charges_text):
    tax_clean = _strip_html_to_text(tax_text)
    charges_clean = _strip_html_to_text(charges_text)

    def pick_pct(text, label, default=None):
        m = re.search(rf"{label}.{{0,90}}?(\d+(?:\.\d+)?)\s*%", text, re.IGNORECASE)
        return parse_float(m.group(1), default) if m else default

    def pick_tax_pct(text, label, default=None):
        vals = [
            parse_float(x, 0.0)
            for x in re.findall(rf"{label}.{{0,140}}?(\d+(?:\.\d+)?)\s*%", text, re.IGNORECASE)
        ]
        vals = [v for v in vals if 0 < v <= 50]
        return (max(vals) if vals else default)

    def pick_money(text, pattern, default=None):
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            return default
        raw_source = m.group(1) if (m.lastindex or 0) >= 1 else m.group(0)
        raw_text = str(raw_source or "")
        raw = re.sub(r"[^\d.]", "", raw_text)
        value = parse_float(raw, default)
        if "lakh" in raw_text.lower() and value > 0:
            value *= 100000.0
        return value

    stcg_pct = pick_tax_pct(tax_clean, "STCG", None)
    ltcg_pct = pick_tax_pct(tax_clean, "LTCG", None)
    exemption_limit = pick_money(
        tax_clean,
        r"(?:1\.25\s*lakh|1,25,000|125000)",
        125000.0,
    )
    if exemption_limit <= 0:
        exemption_limit = 125000.0
    nse_txn_pct = pick_pct(charges_clean, r"NSE:\s*", None)
    bse_txn_pct = pick_pct(charges_clean, r"BSE:\s*", None)
    stt_pct = pick_pct(charges_clean, r"STT/CTT.{0,80}?equity delivery.{0,80}?", None)
    if stt_pct is None:
        stt_pct = pick_pct(charges_clean, r"0\.1% on buy and sell|0\.1% on buy & sell", 0.1)
    stamp_pct = pick_pct(charges_clean, r"Stamp charges.{0,80}?buy side", None)
    gst_pct = pick_pct(charges_clean, r"GST", None)
    dp_charge = pick_money(charges_clean, r"₹\s*([\d.]+)\s*per scrip", 15.34)
    if None in (stcg_pct, ltcg_pct, nse_txn_pct, bse_txn_pct, stt_pct, stamp_pct, gst_pct):
        raise ValueError("tax_rate_source_parse_failed")
    return {
        "stcg_rate_pct": round(stcg_pct, 4),
        "ltcg_rate_pct": round(ltcg_pct, 4),
        "ltcg_exemption_limit": round(exemption_limit, 2),
        "txn_rate_nse": round(nse_txn_pct / 100.0, 8),
        "txn_rate_bse": round(bse_txn_pct / 100.0, 8),
        "stt_delivery_rate": round(stt_pct / 100.0, 6),
        "stamp_buy_rate": round(stamp_pct / 100.0, 6),
        "gst_rate": round(gst_pct / 100.0, 6),
        "dp_charge_sell_incl_gst": round(dp_charge, 2),
    }


def _insert_tax_rate_sync_run(conn, status, source_label, source_url=None, snapshot=None, detail=None, error=None):
    snap = dict(snapshot or {})
    conn.execute(
        """
        INSERT INTO tax_rate_sync_runs(
          created_at, status, source_label, source_url, stcg_rate_pct, ltcg_rate_pct, ltcg_exemption_limit,
          stt_delivery_rate, stamp_buy_rate, gst_rate, dp_charge_sell, detail, error
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            now_iso(),
            str(status or "error"),
            str(source_label or "tax_monitor"),
            str(source_url or ""),
            parse_float(snap.get("stcg_rate_pct"), 0.0) or None,
            parse_float(snap.get("ltcg_rate_pct"), 0.0) or None,
            parse_float(snap.get("ltcg_exemption_limit"), 0.0) or None,
            parse_float(snap.get("stt_delivery_rate"), 0.0) or None,
            parse_float(snap.get("stamp_buy_rate"), 0.0) or None,
            parse_float(snap.get("gst_rate"), 0.0) or None,
            parse_float(snap.get("dp_charge_sell_incl_gst"), 0.0) or None,
            str(detail or ""),
            str(error or ""),
        ),
    )


def upsert_attention_alert(conn, code, category, severity_rank, severity_label, title, detail="", source_ref="", meta=None):
    code_s = str(code or "").strip().upper()
    if not code_s:
        raise ValueError("attention_code_required")
    row = conn.execute("SELECT id, occurrence_count FROM attention_alerts WHERE code = ?", (code_s,)).fetchone()
    stamp = now_iso()
    meta_json = _safe_json_dumps(meta or {})
    if row:
        conn.execute(
            """
            UPDATE attention_alerts
            SET category = ?, severity_rank = ?, severity_label = ?, status = 'open', title = ?, detail = ?,
                source_ref = ?, last_seen_at = ?, resolved_at = NULL, meta_json = ?, occurrence_count = COALESCE(occurrence_count, 0) + 1
            WHERE code = ?
            """,
            (
                str(category or "system"),
                int(severity_rank),
                str(severity_label or "info"),
                str(title or code_s),
                str(detail or ""),
                str(source_ref or ""),
                stamp,
                meta_json,
                code_s,
            ),
        )
        return code_s
    conn.execute(
        """
        INSERT INTO attention_alerts(
          code, category, severity_rank, severity_label, status, title, detail, source_ref, detected_at,
          last_seen_at, resolved_at, meta_json, occurrence_count
        ) VALUES (?, ?, ?, ?, 'open', ?, ?, ?, ?, ?, NULL, ?, 1)
        """,
        (
            code_s,
            str(category or "system"),
            int(severity_rank),
            str(severity_label or "info"),
            str(title or code_s),
            str(detail or ""),
            str(source_ref or ""),
            stamp,
            stamp,
            meta_json,
        ),
    )
    return code_s


def resolve_attention_alert(conn, code, detail=None):
    code_s = str(code or "").strip().upper()
    if not code_s:
        return
    row = conn.execute("SELECT id, status, detail FROM attention_alerts WHERE code = ?", (code_s,)).fetchone()
    if not row or str(row["status"] or "").lower() == "resolved":
        return
    resolved_detail = str(detail if detail is not None else row["detail"] or "")
    conn.execute(
        """
        UPDATE attention_alerts
        SET status='resolved', detail=?, resolved_at=?, last_seen_at=?
        WHERE code = ?
        """,
        (resolved_detail, now_iso(), now_iso(), code_s),
    )


def list_attention_alerts(conn, status=None, limit=80):
    lim = max(1, min(300, int(limit)))
    sql = "SELECT * FROM attention_alerts"
    params = []
    if status:
        sql += " WHERE LOWER(status) = ?"
        params.append(str(status).strip().lower())
    sql += " ORDER BY CASE WHEN LOWER(status)='open' THEN 0 ELSE 1 END, severity_rank DESC, last_seen_at DESC, id DESC LIMIT ?"
    params.append(lim)
    rows = conn.execute(sql, params).fetchall()
    out = []
    for r in rows:
        item = dict(r)
        item["meta"] = _safe_json_loads(item.get("meta_json"), {})
        out.append(item)
    return out


def list_tax_rate_sync_runs(conn, limit=20):
    lim = max(1, min(120, int(limit)))
    rows = conn.execute(
        """
        SELECT *
        FROM tax_rate_sync_runs
        ORDER BY id DESC
        LIMIT ?
        """,
        (lim,),
    ).fetchall()
    return [dict(r) for r in rows]


def open_lot_tax_bucket_summary(conn, symbol, ltp=None, as_of_date=None, split_map=None):
    px = parse_float(ltp, 0.0)
    if px <= 0:
        px = get_effective_ltp(conn, symbol, split_map)
    lots = open_lots_for_symbol(conn, symbol, split_map=split_map)
    out = {}
    for lot in lots:
        qty = parse_float(lot.get("qty"), 0.0)
        buy_price = parse_float(lot.get("buy_price"), 0.0)
        buy_date = lot.get("buy_date")
        if qty <= 0 or buy_price <= 0:
            continue
        bucket, held_days = _harvest_tax_bucket(buy_date, as_of_date=as_of_date)
        row = out.setdefault(
            bucket,
            {
                "tax_bucket": bucket,
                "qty": 0.0,
                "invested": 0.0,
                "market_value": 0.0,
                "lot_count": 0,
                "held_days_min": None,
                "held_days_max": None,
                "days_to_ltcg_min": None,
            },
        )
        row["qty"] += qty
        row["invested"] += qty * buy_price
        row["market_value"] += qty * px
        row["lot_count"] += 1
        row["held_days_min"] = held_days if row["held_days_min"] is None else min(row["held_days_min"], held_days)
        row["held_days_max"] = held_days if row["held_days_max"] is None else max(row["held_days_max"], held_days)
        if bucket == "STCG":
            days_to_ltcg = max(0, 365 - held_days)
            row["days_to_ltcg_min"] = (
                days_to_ltcg
                if row["days_to_ltcg_min"] is None
                else min(row["days_to_ltcg_min"], days_to_ltcg)
            )
    rows = []
    for bucket in sorted(out.keys(), key=_harvest_tax_bucket_rank):
        row = out[bucket]
        qty = parse_float(row.get("qty"), 0.0)
        invested = parse_float(row.get("invested"), 0.0)
        market_value = parse_float(row.get("market_value"), 0.0)
        avg_cost = (invested / qty) if qty > 0 else 0.0
        unrealized = market_value - invested
        upl_pct = (unrealized / invested * 100.0) if invested > 0 else 0.0
        rows.append(
            {
                "tax_bucket": bucket,
                "qty": round(qty, 4),
                "avg_cost": round(avg_cost, 4),
                "invested": round(invested, 2),
                "market_value": round(market_value, 2),
                "unrealized_pnl": round(unrealized, 2),
                "upl_pct": round(upl_pct, 2),
                "lot_count": int(parse_float(row.get("lot_count"), 0.0)),
                "held_days_min": int(parse_float(row.get("held_days_min"), 0.0)) if row.get("held_days_min") is not None else 0,
                "held_days_max": int(parse_float(row.get("held_days_max"), 0.0)) if row.get("held_days_max") is not None else 0,
                "days_to_ltcg_min": (
                    int(parse_float(row.get("days_to_ltcg_min"), 0.0))
                    if row.get("days_to_ltcg_min") is not None
                    else None
                ),
            }
        )
    return rows


def open_lot_tax_bucket_rows(conn, symbol, ltp=None, as_of_date=None, split_map=None):
    px = parse_float(ltp, 0.0)
    if px <= 0:
        px = get_effective_ltp(conn, symbol, split_map)
    rows = []
    for lot in open_lots_for_symbol(conn, symbol, split_map=split_map):
        qty = parse_float(lot.get("qty"), 0.0)
        buy_price = parse_float(lot.get("buy_price"), 0.0)
        buy_date = lot.get("buy_date")
        trade_id = int(parse_float(lot.get("trade_id"), 0.0))
        if qty <= 0 or buy_price <= 0:
            continue
        bucket, held_days = _harvest_tax_bucket(buy_date, as_of_date=as_of_date)
        invested = qty * buy_price
        market_value = qty * px
        unrealized = market_value - invested
        upl_pct = (unrealized / invested * 100.0) if invested > 0 else 0.0
        rows.append(
            {
                "buy_trade_id": trade_id,
                "buy_date": str(buy_date or ""),
                "tax_bucket": bucket,
                "qty": round(qty, 4),
                "avg_cost": round(buy_price, 4),
                "invested": round(invested, 2),
                "market_value": round(market_value, 2),
                "unrealized_pnl": round(unrealized, 2),
                "upl_pct": round(upl_pct, 2),
                "lot_count": 1,
                "held_days_min": int(held_days),
                "held_days_max": int(held_days),
                "days_to_ltcg_min": (max(0, 365 - held_days) if bucket == "STCG" else None),
            }
        )
    rows.sort(
        key=lambda r: (
            _harvest_tax_bucket_rank(r.get("tax_bucket")),
            str(r.get("buy_date") or ""),
            int(parse_float(r.get("buy_trade_id"), 0.0)),
        )
    )
    return rows


def build_tax_harvest_heuristic_analysis(plan):
    summary = plan.get("summary") or {}
    harvest = list(plan.get("harvest_candidates") or [])
    offsets = list(plan.get("profit_offset_candidates") or [])
    top_harvest = [x for x in harvest if parse_float(x.get("suggested_qty"), 0.0) > 0][:4]
    top_offsets = [x for x in offsets if parse_float(x.get("suggested_qty"), 0.0) > 0][:4]
    macro = plan.get("macro") or {}
    lines = [
        f"Target loss: {round(parse_float(summary.get('target_loss'), 0.0), 2)} | available loss: {round(parse_float(summary.get('total_loss_available'), 0.0), 2)}.",
        (
            "Bucket split: "
            f"STCG loss {round(parse_float(summary.get('total_loss_available_stcg'), 0.0), 2)}, "
            f"LTCG loss {round(parse_float(summary.get('total_loss_available_ltcg'), 0.0), 2)}, "
            f"STCG profit {round(parse_float(summary.get('total_profit_available_stcg'), 0.0), 2)}, "
            f"LTCG profit {round(parse_float(summary.get('total_profit_available_ltcg'), 0.0), 2)}."
        ),
        f"Suggested harvest now: {round(parse_float(summary.get('suggested_harvest_loss'), 0.0), 2)} across {int(parse_float(summary.get('suggested_harvest_count'), 0.0))} line(s).",
        (
            "Suggested bucket mix: "
            f"STCG harvest {round(parse_float(summary.get('suggested_harvest_loss_stcg'), 0.0), 2)} vs "
            f"LTCG harvest {round(parse_float(summary.get('suggested_harvest_loss_ltcg'), 0.0), 2)}; "
            f"STCG offset {round(parse_float(summary.get('suggested_offset_profit_stcg'), 0.0), 2)} vs "
            f"LTCG offset {round(parse_float(summary.get('suggested_offset_profit_ltcg'), 0.0), 2)}."
        ),
        f"Macro: {str(macro.get('regime') or 'neutral').upper()} ({round(parse_float(macro.get('score'), 0.0), 2)}).",
    ]
    if top_harvest:
        lines.append(
            "Best loss-harvest candidates: "
            + "; ".join(
                f"{x['symbol']} {x.get('tax_bucket','-')} qty {round(parse_float(x.get('suggested_qty'), 0.0), 4)} ({x.get('likely_direction')}, {x.get('decision_hint')})"
                for x in top_harvest
            )
            + "."
        )
    else:
        lines.append("No immediate loss-harvest suggestion from current target.")
    if top_offsets:
        lines.append(
            "Likely profit-offset sells if gains must be booked later: "
            + "; ".join(
                f"{x['symbol']} {x.get('tax_bucket','-')} qty {round(parse_float(x.get('suggested_qty'), 0.0), 4)} ({x.get('likely_direction')}, {x.get('decision_hint')})"
                for x in top_offsets
            )
            + "."
        )
    else:
        lines.append("No profit-offset quantity suggestion required yet.")
    return {
        "mode": "heuristic",
        "provider": "local_rules",
        "model": None,
        "text": "\n".join(lines),
        "error": None,
    }


def run_tax_harvest_llm_analysis(plan):
    fallback = build_tax_harvest_heuristic_analysis(plan)
    summary = plan.get("summary") or {}
    harvest = list(plan.get("harvest_candidates") or [])[:8]
    offsets = list(plan.get("profit_offset_candidates") or [])[:8]
    macro = plan.get("macro") or {}
    try:
        result = call_llm_responses_api(
            system_prompt=(
                "You are a portfolio tax-loss harvesting analyst. "
                "Be concise, practical, and explain which positions are better to harvest now "
                "versus defer, using only the supplied portfolio intelligence."
            ),
            user_payload={
                "task": "Advise on current tax-loss harvesting and profit-offset candidates.",
                "summary": summary,
                "macro": macro,
                "harvest_candidates": harvest,
                "profit_offset_candidates": offsets,
            },
            max_output_tokens=500,
            timeout=12,
        )
    except Exception as ex:
        fallback["error"] = str(ex)
        return fallback
    text = str(result.get("text") or "").strip()
    if not text:
        fallback["error"] = "LLM returned empty analysis."
        return fallback
    return {
        "mode": "llm",
        "provider": str(result.get("provider") or "openai_responses"),
        "model": str(result.get("model") or LLM_DEFAULT_MODEL),
        "text": text,
        "error": None,
    }


def build_tax_harvest_plan(conn, target_loss=0.0, run_llm=False):
    try:
        target_loss_v = max(0.0, parse_float(target_loss, 0.0))
    except Exception:
        target_loss_v = 0.0
    items = [x for x in collect_strategy_universe(conn, lookback_days=30) if parse_float(x.get("qty"), 0.0) > 0]
    bundle = get_intel_parameter_bundle(conn)
    intel_overlay = build_intelligence_bias_map(
        conn,
        decay_days=bundle["decay_days"],
        w_commentary=bundle["w_commentary"],
        w_policy=bundle["w_policy"],
        w_financials=bundle["w_financials"],
        w_chart=bundle["w_chart"],
    )
    intel_map = intel_overlay.get("symbols", {})
    chart_map = latest_chart_snapshot_map(conn)
    strategy_map = latest_strategy_recommendation_map(conn)
    macro = build_macro_thoughts()
    split_map = load_split_map(conn)
    as_of_s = ist_now().date().isoformat()

    harvest_rows = []
    profit_rows = []
    total_loss_available = 0.0
    total_profit_available = 0.0
    total_loss_available_by_bucket = {"STCG": 0.0, "LTCG": 0.0}
    total_profit_available_by_bucket = {"STCG": 0.0, "LTCG": 0.0}
    for item in items:
        symbol = symbol_upper(item.get("symbol"))
        qty = max(0.0, parse_float(item.get("qty"), 0.0))
        ltp = max(0.0, parse_float(item.get("ltp"), 0.0))
        if not symbol or qty <= 0 or ltp <= 0:
            continue
        intel = intel_map.get(symbol, {})
        chart = chart_map.get(symbol, {})
        fin = financial_signal_for_symbol(conn, symbol)
        strat = strategy_map.get(symbol, {})
        action_bias = _harvest_action_bias(strat.get("action"))
        signal_bias = _harvest_signal_bias(item.get("buy_signal"), item.get("sell_signal"))
        expected_move_score = _harvest_expected_move_score(
            parse_float(intel.get("score"), 0.0),
            parse_float(chart.get("score"), 0.0),
            parse_float(fin.get("score"), 0.0),
            action_bias,
            signal_bias,
        )
        likely_direction = _harvest_direction_label(expected_move_score)
        direction_bias = -expected_move_score
        decision_hint = "Hold candidate"
        if likely_direction == "DOWN":
            decision_hint = "Likely weaker"
        elif likely_direction == "UP":
            decision_hint = "Likely stronger"
        reason = _harvest_priority_reason(
            likely_direction,
            strat.get("action"),
            chart.get("signal"),
            intel.get("summary"),
            fin.get("summary"),
        )
        bucket_rows = open_lot_tax_bucket_rows(
            conn,
            symbol,
            ltp=ltp,
            as_of_date=as_of_s,
            split_map=split_map,
        )
        if not bucket_rows:
            avg_cost = max(0.0, parse_float(item.get("avg_cost"), 0.0))
            invested = max(0.0, parse_float(item.get("invested"), 0.0))
            market_value = max(0.0, parse_float(item.get("market_value"), 0.0))
            unrealized = parse_float(item.get("unrealized_pnl"), 0.0)
            bucket_rows = [
                {
                    "buy_trade_id": None,
                    "buy_date": "",
                    "tax_bucket": "STCG",
                    "qty": round(qty, 4),
                    "avg_cost": round(avg_cost, 4),
                    "invested": round(invested, 2),
                    "market_value": round(market_value, 2),
                    "unrealized_pnl": round(unrealized, 2),
                    "upl_pct": round(parse_float(item.get("upl_pct"), 0.0), 2),
                    "lot_count": 0,
                    "held_days_min": 0,
                    "held_days_max": 0,
                    "days_to_ltcg_min": None,
                }
            ]
        for bucket_row in bucket_rows:
            bucket = str(bucket_row.get("tax_bucket") or "STCG").upper()
            bucket_qty = max(0.0, parse_float(bucket_row.get("qty"), 0.0))
            avg_cost = max(0.0, parse_float(bucket_row.get("avg_cost"), 0.0))
            invested = max(0.0, parse_float(bucket_row.get("invested"), 0.0))
            market_value = max(0.0, parse_float(bucket_row.get("market_value"), 0.0))
            unrealized = parse_float(bucket_row.get("unrealized_pnl"), 0.0)
            if bucket_qty <= 0 or avg_cost <= 0:
                continue
            per_unit_pnl = ltp - avg_cost
            if abs(per_unit_pnl) <= 1e-9:
                continue
            held_min = int(parse_float(bucket_row.get("held_days_min"), 0.0))
            held_max = int(parse_float(bucket_row.get("held_days_max"), 0.0))
            held_label = f"{held_min}" + (f"-{held_max}" if held_max != held_min else "")
            row = {
                "symbol": symbol,
                "buy_trade_id": (
                    None if bucket_row.get("buy_trade_id") in (None, "", 0) else int(parse_float(bucket_row.get("buy_trade_id"), 0.0))
                ),
                "buy_date": str(bucket_row.get("buy_date") or ""),
                "tax_bucket": bucket,
                "qty": round(bucket_qty, 4),
                "ltp": round(ltp, 4),
                "avg_cost": round(avg_cost, 4),
                "invested": round(invested, 2),
                "market_value": round(market_value, 2),
                "unrealized_pnl": round(unrealized, 2),
                "upl_pct": round(parse_float(bucket_row.get("upl_pct"), 0.0), 2),
                "lot_count": int(parse_float(bucket_row.get("lot_count"), 0.0)),
                "held_days_min": held_min,
                "held_days_max": held_max,
                "days_to_ltcg_min": (
                    None
                    if bucket_row.get("days_to_ltcg_min") is None
                    else int(parse_float(bucket_row.get("days_to_ltcg_min"), 0.0))
                ),
                "strategy_action": str(strat.get("action") or item.get("strategy_action") or "").upper(),
                "chart_signal": str(chart.get("signal") or ""),
                "intel_score": round(parse_float(intel.get("score"), 0.0), 2),
                "chart_score": round(parse_float(chart.get("score"), 0.0), 2),
                "financial_score": round(parse_float(fin.get("score"), 0.0), 2),
                "expected_move_score": round(expected_move_score, 2),
                "likely_direction": likely_direction,
                "decision_hint": decision_hint,
                "reason": (
                    f"buy {str(bucket_row.get('buy_date') or '-')}, {bucket} bucket, held {held_label}d, {reason}"
                ),
                "suggested_qty": 0.0,
                "suggested_value": 0.0,
            }
            if per_unit_pnl < 0:
                loss_per_unit = abs(per_unit_pnl)
                loss_available = loss_per_unit * bucket_qty
                total_loss_available += loss_available
                total_loss_available_by_bucket[bucket] = total_loss_available_by_bucket.get(bucket, 0.0) + loss_available
                row["per_unit_loss"] = round(loss_per_unit, 4)
                row["loss_available"] = round(loss_available, 2)
                row["priority_score"] = round(
                    clamp(
                        (direction_bias * 0.65)
                        + min(30.0, abs(row["upl_pct"]) * 0.35)
                        + _harvest_tax_bucket_bias(bucket, side="loss"),
                        -100.0,
                        100.0,
                    ),
                    2,
                )
                harvest_rows.append(row)
            else:
                profit_per_unit = per_unit_pnl
                profit_available = profit_per_unit * bucket_qty
                total_profit_available += profit_available
                total_profit_available_by_bucket[bucket] = total_profit_available_by_bucket.get(bucket, 0.0) + profit_available
                row["per_unit_profit"] = round(profit_per_unit, 4)
                row["profit_available"] = round(profit_available, 2)
                row["priority_score"] = round(
                    clamp(
                        (direction_bias * 0.70)
                        + min(25.0, row["upl_pct"] * 0.20)
                        + _harvest_tax_bucket_bias(bucket, side="profit"),
                        -100.0,
                        100.0,
                    ),
                    2,
                )
                profit_rows.append(row)

    harvest_rows.sort(
        key=lambda x: (
            _harvest_tax_bucket_rank(x.get("tax_bucket")),
            -parse_float(x.get("priority_score"), 0.0),
            -parse_float(x.get("loss_available"), 0.0),
            str(x.get("buy_date") or ""),
            x["symbol"],
        )
    )
    profit_rows.sort(
        key=lambda x: (
            _harvest_tax_bucket_rank(x.get("tax_bucket")),
            -parse_float(x.get("priority_score"), 0.0),
            -parse_float(x.get("profit_available"), 0.0),
            str(x.get("buy_date") or ""),
            x["symbol"],
        )
    )

    remaining_loss_target = min(target_loss_v, total_loss_available)
    suggested_harvest_loss = 0.0
    suggested_harvest_count = 0
    suggested_harvest_by_bucket = {"STCG": 0.0, "LTCG": 0.0}
    for row in harvest_rows:
        loss_per_unit = parse_float(row.get("per_unit_loss"), 0.0)
        qty = parse_float(row.get("qty"), 0.0)
        if remaining_loss_target <= 1e-9 or loss_per_unit <= 0 or qty <= 0:
            continue
        qty_suggest = min(qty, _floor_qty(remaining_loss_target / loss_per_unit, step=0.0001))
        if qty_suggest <= 0 and remaining_loss_target >= (loss_per_unit * 0.25):
            qty_suggest = min(qty, 0.0001)
        realized = qty_suggest * loss_per_unit
        if qty_suggest <= 0 or realized <= 0:
            continue
        row["suggested_qty"] = round(qty_suggest, 4)
        row["suggested_value"] = round(qty_suggest * parse_float(row.get("ltp"), 0.0), 2)
        row["suggested_realized_loss"] = round(realized, 2)
        suggested_harvest_loss += realized
        bucket = str(row.get("tax_bucket") or "STCG").upper()
        suggested_harvest_by_bucket[bucket] = suggested_harvest_by_bucket.get(bucket, 0.0) + realized
        suggested_harvest_count += 1
        remaining_loss_target = max(0.0, remaining_loss_target - realized)
    for row in harvest_rows:
        if "suggested_realized_loss" not in row:
            row["suggested_realized_loss"] = 0.0

    remaining_profit_target_by_bucket = dict(suggested_harvest_by_bucket)
    suggested_offset_profit = 0.0
    suggested_offset_count = 0
    suggested_offset_by_bucket = {"STCG": 0.0, "LTCG": 0.0}
    for row in profit_rows:
        profit_per_unit = parse_float(row.get("per_unit_profit"), 0.0)
        qty = parse_float(row.get("qty"), 0.0)
        bucket = str(row.get("tax_bucket") or "STCG").upper()
        remaining_profit_target = max(0.0, parse_float(remaining_profit_target_by_bucket.get(bucket), 0.0))
        if remaining_profit_target <= 1e-9 or profit_per_unit <= 0 or qty <= 0:
            continue
        qty_suggest = min(qty, _floor_qty(remaining_profit_target / profit_per_unit, step=0.0001))
        if qty_suggest <= 0 and remaining_profit_target >= (profit_per_unit * 0.25):
            qty_suggest = min(qty, 0.0001)
        realized = qty_suggest * profit_per_unit
        if qty_suggest <= 0 or realized <= 0:
            continue
        row["suggested_qty"] = round(qty_suggest, 4)
        row["suggested_value"] = round(qty_suggest * parse_float(row.get("ltp"), 0.0), 2)
        row["suggested_realized_profit"] = round(realized, 2)
        suggested_offset_profit += realized
        suggested_offset_by_bucket[bucket] = suggested_offset_by_bucket.get(bucket, 0.0) + realized
        suggested_offset_count += 1
        remaining_profit_target_by_bucket[bucket] = max(0.0, remaining_profit_target - realized)
    for row in profit_rows:
        if "suggested_realized_profit" not in row:
            row["suggested_realized_profit"] = 0.0

    plan = {
        "summary": {
            "target_loss": round(target_loss_v, 2),
            "total_loss_available": round(total_loss_available, 2),
            "total_loss_available_stcg": round(total_loss_available_by_bucket.get("STCG", 0.0), 2),
            "total_loss_available_ltcg": round(total_loss_available_by_bucket.get("LTCG", 0.0), 2),
            "total_profit_available": round(total_profit_available, 2),
            "total_profit_available_stcg": round(total_profit_available_by_bucket.get("STCG", 0.0), 2),
            "total_profit_available_ltcg": round(total_profit_available_by_bucket.get("LTCG", 0.0), 2),
            "suggested_harvest_loss": round(suggested_harvest_loss, 2),
            "suggested_harvest_loss_stcg": round(suggested_harvest_by_bucket.get("STCG", 0.0), 2),
            "suggested_harvest_loss_ltcg": round(suggested_harvest_by_bucket.get("LTCG", 0.0), 2),
            "suggested_harvest_count": int(suggested_harvest_count),
            "unfilled_harvest_loss": round(max(0.0, target_loss_v - suggested_harvest_loss), 2),
            "suggested_offset_profit": round(suggested_offset_profit, 2),
            "suggested_offset_profit_stcg": round(suggested_offset_by_bucket.get("STCG", 0.0), 2),
            "suggested_offset_profit_ltcg": round(suggested_offset_by_bucket.get("LTCG", 0.0), 2),
            "suggested_offset_count": int(suggested_offset_count),
            "unfilled_offset_profit": round(max(0.0, suggested_harvest_loss - suggested_offset_profit), 2),
            "unfilled_offset_profit_stcg": round(
                max(0.0, suggested_harvest_by_bucket.get("STCG", 0.0) - suggested_offset_by_bucket.get("STCG", 0.0)),
                2,
            ),
            "unfilled_offset_profit_ltcg": round(
                max(0.0, suggested_harvest_by_bucket.get("LTCG", 0.0) - suggested_offset_by_bucket.get("LTCG", 0.0)),
                2,
            ),
            "harvest_candidates_count": len(harvest_rows),
            "profit_candidates_count": len(profit_rows),
        },
        "macro": macro,
        "harvest_candidates": harvest_rows,
        "profit_offset_candidates": profit_rows,
        "analysis": build_tax_harvest_heuristic_analysis(
            {
                "summary": {
                    "target_loss": round(target_loss_v, 2),
                    "total_loss_available": round(total_loss_available, 2),
                    "suggested_harvest_loss": round(suggested_harvest_loss, 2),
                    "suggested_harvest_count": int(suggested_harvest_count),
                },
                "macro": macro,
                "harvest_candidates": harvest_rows,
                "profit_offset_candidates": profit_rows,
            }
        ),
        "llm_enabled": bool(get_llm_runtime_config(conn, include_secret=False).get("configured")),
    }
    if run_llm:
        plan["analysis"] = run_tax_harvest_llm_analysis(plan)
    return plan


def build_loss_lot_analysis(conn):
    plan = build_tax_harvest_plan(conn, target_loss=0.0, run_llm=False)
    harvest = list(plan.get("harvest_candidates") or [])
    profits = list(plan.get("profit_offset_candidates") or [])
    stcg_rows = [r for r in harvest if str(r.get("tax_bucket") or "").upper() == "STCG"]
    ltcg_rows = [r for r in harvest if str(r.get("tax_bucket") or "").upper() == "LTCG"]
    stcg_profit_rows = [r for r in profits if str(r.get("tax_bucket") or "").upper() == "STCG"]
    ltcg_profit_rows = [r for r in profits if str(r.get("tax_bucket") or "").upper() == "LTCG"]
    stcg_rows.sort(key=lambda r: (-parse_float(r.get("loss_available"), 0.0), str(r.get("buy_date") or ""), str(r.get("symbol") or "")))
    ltcg_rows.sort(key=lambda r: (-parse_float(r.get("loss_available"), 0.0), str(r.get("buy_date") or ""), str(r.get("symbol") or "")))
    stcg_profit_rows.sort(key=lambda r: (-parse_float(r.get("profit_available"), 0.0), str(r.get("buy_date") or ""), str(r.get("symbol") or "")))
    ltcg_profit_rows.sort(key=lambda r: (-parse_float(r.get("profit_available"), 0.0), str(r.get("buy_date") or ""), str(r.get("symbol") or "")))
    total_rows = len(harvest)
    total_qty = sum(parse_float(r.get("qty"), 0.0) for r in harvest)
    summary = {
        "total_loss_lots": total_rows,
        "total_loss_qty": round(total_qty, 4),
        "total_loss_available": round(sum(parse_float(r.get("loss_available"), 0.0) for r in harvest), 2),
        "stcg_loss_lots": len(stcg_rows),
        "stcg_loss_qty": round(sum(parse_float(r.get("qty"), 0.0) for r in stcg_rows), 4),
        "stcg_loss_available": round(sum(parse_float(r.get("loss_available"), 0.0) for r in stcg_rows), 2),
        "ltcg_loss_lots": len(ltcg_rows),
        "ltcg_loss_qty": round(sum(parse_float(r.get("qty"), 0.0) for r in ltcg_rows), 4),
        "ltcg_loss_available": round(sum(parse_float(r.get("loss_available"), 0.0) for r in ltcg_rows), 2),
        "total_profit_lots": len(profits),
        "total_profit_qty": round(sum(parse_float(r.get("qty"), 0.0) for r in profits), 4),
        "total_profit_available": round(sum(parse_float(r.get("profit_available"), 0.0) for r in profits), 2),
        "stcg_profit_lots": len(stcg_profit_rows),
        "stcg_profit_qty": round(sum(parse_float(r.get("qty"), 0.0) for r in stcg_profit_rows), 4),
        "stcg_profit_available": round(sum(parse_float(r.get("profit_available"), 0.0) for r in stcg_profit_rows), 2),
        "ltcg_profit_lots": len(ltcg_profit_rows),
        "ltcg_profit_qty": round(sum(parse_float(r.get("qty"), 0.0) for r in ltcg_profit_rows), 4),
        "ltcg_profit_available": round(sum(parse_float(r.get("profit_available"), 0.0) for r in ltcg_profit_rows), 2),
    }
    return {
        "summary": summary,
        "stcg_items": stcg_rows[:10],
        "ltcg_items": ltcg_rows[:10],
        "stcg_profit_items": stcg_profit_rows[:10],
        "ltcg_profit_items": ltcg_profit_rows[:10],
        "generated_at": now_iso(),
    }


def refresh_strategy_analytics(force=False):
    with db_connect() as conn:
        latest = conn.execute(
            "SELECT run_date, created_at FROM strategy_runs ORDER BY run_date DESC LIMIT 1"
        ).fetchone()
        today = dt.date.today().isoformat()
        if not force and latest and latest["run_date"] == today:
            try:
                created = dt.datetime.fromisoformat(str(latest["created_at"]))
                age_sec = (dt.datetime.now() - created).total_seconds()
                min_interval = get_strategy_refresh_interval(conn)
                if age_sec < min_interval:
                    current = load_latest_strategy_insights(conn)
                    if current is not None:
                        return current
            except Exception:
                pass

        insights = build_strategy_insights(conn, run_date=today)
        persist_strategy_insights(conn, insights)
        conn.commit()
        return load_latest_strategy_insights(conn)


def clear_core_tables(conn):
    conn.execute("DELETE FROM trades")
    conn.execute("DELETE FROM instruments")
    conn.execute("DELETE FROM latest_prices")
    conn.execute("DELETE FROM holdings")
    conn.execute("DELETE FROM lot_closures")
    conn.execute("DELETE FROM cash_ledger")
    conn.execute("DELETE FROM dividends")
    conn.execute("DELETE FROM signals")


def _clean_symbol(raw):
    if raw is None:
        return None
    symbol = str(raw).strip()
    if symbol == "":
        return None
    return symbol.upper()


def symbol_upper(raw):
    if raw is None:
        return ""
    return str(raw).strip().upper()


def normalize_asset_class(raw, fallback=ASSET_CLASS_EQUITY):
    s = str(raw or "").strip().upper()
    if not s:
        s = str(fallback or ASSET_CLASS_EQUITY).strip().upper() or ASSET_CLASS_EQUITY
    if s not in (ASSET_CLASS_EQUITY, ASSET_CLASS_GOLD):
        raise ValueError("asset_class_must_be_equity_or_gold")
    return s


def infer_asset_class(symbol=None, name=None, notes=None, fallback=ASSET_CLASS_EQUITY):
    text = " ".join([str(symbol or ""), str(name or ""), str(notes or "")]).strip()
    if text:
        for pat in GOLD_HINT_PATTERNS:
            if pat.search(text):
                return ASSET_CLASS_GOLD
    return normalize_asset_class(fallback, fallback=ASSET_CLASS_EQUITY)


def sync_instrument_asset_class(conn, symbol, asset_class=None, fallback=ASSET_CLASS_EQUITY):
    sym = symbol_upper(symbol)
    if not sym:
        return None
    if asset_class is None:
        row = conn.execute("SELECT symbol, name FROM instruments WHERE UPPER(symbol)=?", (sym,)).fetchone()
        if not row:
            return None
        resolved = infer_asset_class(symbol=row["symbol"], name=row["name"], fallback=fallback)
    else:
        resolved = normalize_asset_class(asset_class, fallback=fallback)
    conn.execute("UPDATE instruments SET asset_class=? WHERE UPPER(symbol)=?", (resolved, sym))
    return resolved


def normalize_external_trade_id(raw):
    if raw is None:
        return None
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = str(raw).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s.upper()


def normalize_header_key(raw):
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    s = re.sub(r"[\s_\-/]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def resolve_symbol(conn, raw_symbol):
    s_up = symbol_upper(raw_symbol)
    if not s_up:
        return None
    row = conn.execute(
        "SELECT symbol FROM instruments WHERE UPPER(symbol) = ? ORDER BY active DESC, symbol LIMIT 1",
        (s_up,),
    ).fetchone()
    if row:
        return row["symbol"]
    row = conn.execute(
        "SELECT symbol FROM trades WHERE UPPER(symbol) = ? ORDER BY id DESC LIMIT 1",
        (s_up,),
    ).fetchone()
    if row:
        return row["symbol"]
    return None




def _coerce_guard_value(raw, allow_none=False):
    if raw is None:
        return None if allow_none else 0.0
    if isinstance(raw, str) and str(raw).strip() == "":
        return None if allow_none else 0.0
    val = parse_float(raw, 0.0)
    if not math.isfinite(val) or val < 0:
        raise ValueError("guard_value_must_be_non_negative")
    return round(float(val), 4)


def list_scrip_position_guards(conn, symbols=None):
    params = []
    where = ["1=1"]
    if symbols:
        norm = [symbol_upper(s) for s in symbols if symbol_upper(s)]
        if norm:
            placeholders = ",".join(["?"] * len(norm))
            where.append(f"UPPER(symbol) IN ({placeholders})")
            params.extend(norm)
    rows = conn.execute(
        "SELECT symbol, min_value, max_value, updated_at FROM scrip_position_guards WHERE "
        + " AND ".join(where)
        + " ORDER BY symbol",
        params,
    ).fetchall()
    return [
        {
            "symbol": r["symbol"],
            "min_value": round(parse_float(r["min_value"], 0.0), 4),
            "max_value": (round(parse_float(r["max_value"], 0.0), 4) if r["max_value"] is not None else None),
            "updated_at": r["updated_at"],
        }
        for r in rows
    ]


def set_scrip_position_guard(conn, symbol, min_value=0.0, max_value=None):
    sym = symbol_upper(symbol)
    if not sym:
        raise ValueError("symbol_required")
    if not resolve_symbol(conn, sym):
        raise ValueError("symbol_not_found")
    min_v = _coerce_guard_value(min_value, allow_none=False)
    max_v = _coerce_guard_value(max_value, allow_none=True)
    if max_v is not None and max_v < min_v:
        raise ValueError("max_value_must_be_greater_or_equal_min_value")
    # default row can be removed to keep table compact
    if min_v <= 0 and max_v is None:
        conn.execute("DELETE FROM scrip_position_guards WHERE UPPER(symbol)=?", (sym,))
    else:
        conn.execute(
            """
            INSERT INTO scrip_position_guards(symbol, min_value, max_value, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(symbol) DO UPDATE SET
              min_value = excluded.min_value,
              max_value = excluded.max_value,
              updated_at = excluded.updated_at
            """,
            (sym, min_v, max_v, now_iso()),
        )
    row = conn.execute(
        "SELECT symbol, min_value, max_value, updated_at FROM scrip_position_guards WHERE UPPER(symbol)=?",
        (sym,),
    ).fetchone()
    if not row:
        return {
            "symbol": sym,
            "min_value": 0.0,
            "max_value": None,
            "updated_at": now_iso(),
        }
    return {
        "symbol": row["symbol"],
        "min_value": round(parse_float(row["min_value"], 0.0), 4),
        "max_value": (round(parse_float(row["max_value"], 0.0), 4) if row["max_value"] is not None else None),
        "updated_at": row["updated_at"],
    }


def _floor_qty(value, step=0.0001):
    s = max(1e-8, parse_float(step, 0.0001))
    return round(math.floor(max(0.0, parse_float(value, 0.0)) / s) * s, 4)


def build_rebalance_suggestions(conn, side, percent):
    side_u = str(side or "SELL").strip().upper()
    if side_u not in ("SELL", "BUY"):
        raise ValueError("side_must_be_buy_or_sell")
    pct = parse_float(percent, -1.0)
    if not math.isfinite(pct) or pct <= 0:
        raise ValueError("percent_must_be_positive")
    pct = min(100.0, pct)
    rows = conn.execute(
        """
        SELECT
          i.symbol,
          COALESCE(h.qty, 0) AS qty,
          COALESCE(h.market_value, 0) AS market_value,
          COALESCE(lp.ltp, 0) AS ltp,
          COALESCE(g.min_value, 0) AS min_value,
          g.max_value
        FROM instruments i
        JOIN holdings h ON h.symbol = i.symbol
        LEFT JOIN latest_prices lp ON lp.symbol = i.symbol
        LEFT JOIN scrip_position_guards g ON UPPER(g.symbol) = UPPER(i.symbol)
        WHERE i.active = 1
          AND COALESCE(h.qty, 0) > 0
          AND UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
        ORDER BY COALESCE(h.market_value, 0) DESC, i.symbol
        """
    ).fetchall()
    prepared = []
    total_current_value = 0.0
    for r in rows:
        symbol = r["symbol"]
        qty = max(0.0, parse_float(r["qty"], 0.0))
        market_value = max(0.0, parse_float(r["market_value"], 0.0))
        ltp = parse_float(r["ltp"], 0.0)
        if ltp <= 0:
            ltp = parse_float(get_effective_ltp(conn, symbol), 0.0)
        min_value = max(0.0, parse_float(r["min_value"], 0.0))
        max_value = parse_float(r["max_value"], 0.0) if r["max_value"] is not None else None
        if side_u == "SELL":
            guard_cap_value = max(0.0, market_value - min_value)
            blocked_note = "blocked_by_min_value"
            capped_note = "capped_by_min_value"
        else:
            guard_cap_value = max(0.0, max_value - market_value) if max_value is not None else None
            blocked_note = "blocked_by_max_value"
            capped_note = "capped_by_max_value"
        prepared.append(
            {
                "symbol": symbol,
                "qty": qty,
                "market_value": market_value,
                "ltp": max(0.0, ltp),
                "min_value": min_value,
                "max_value": max_value,
                "guard_cap_value": guard_cap_value,
                "blocked_note": blocked_note,
                "capped_note": capped_note,
                "note": "",
                "desired_trade_value": 0.0,
                "allocated_trade_value": 0.0,
            }
        )
        total_current_value += market_value

    target_total_trade_value = total_current_value * pct / 100.0
    if target_total_trade_value <= 0 or not prepared:
        return {
            "side": side_u,
            "percent": round(pct, 4),
            "allocation_basis": "portfolio_market_value",
            "target_trade_value": round(target_total_trade_value, 2),
            "items": [],
            "lot": None,
            "completed_items": 0,
            "remaining_items": 0,
            "total_remaining_trade_value": 0.0,
            "active_nonzero_count": 0,
            "total_current_market_value": round(total_current_value, 2),
            "total_suggested_trade_value": 0.0,
        }

    # Portfolio-level targeting: assign row desired value from portfolio market-value share.
    denom = max(total_current_value, 1e-9)
    for item in prepared:
        w = item["market_value"] / denom if denom > 0 else 0.0
        item["desired_trade_value"] = target_total_trade_value * w
        cap = item["guard_cap_value"]
        if cap is not None and cap <= 1e-9:
            item["note"] = item["blocked_note"]
            item["allocated_trade_value"] = 0.0
        elif cap is None:
            item["allocated_trade_value"] = item["desired_trade_value"]
        else:
            item["allocated_trade_value"] = min(item["desired_trade_value"], cap)

    allocated_total = sum(max(0.0, x["allocated_trade_value"]) for x in prepared)
    remaining = max(0.0, target_total_trade_value - allocated_total)

    # Redistribute remaining target to rows with available room.
    for _ in range(8):
        if remaining <= 1e-6:
            break
        eligible = []
        weight_sum = 0.0
        for item in prepared:
            cap = item["guard_cap_value"]
            room = float("inf") if cap is None else max(0.0, cap - item["allocated_trade_value"])
            if room <= 1e-9:
                continue
            w = max(0.0, item["market_value"])
            if w <= 1e-9:
                w = 1.0
            eligible.append((item, room, w))
            weight_sum += w
        if not eligible or weight_sum <= 0:
            break
        moved = 0.0
        for item, room, w in eligible:
            share = remaining * (w / weight_sum)
            add = min(room, share)
            if add > 0:
                item["allocated_trade_value"] += add
                moved += add
        if moved <= 1e-9:
            break
        remaining = max(0.0, remaining - moved)

    out = []
    total_suggested_value = 0.0
    for item in prepared:
        symbol = item["symbol"]
        qty = item["qty"]
        ltp = item["ltp"]
        market_value = item["market_value"]
        min_value = item["min_value"]
        max_value = item["max_value"]
        desired_value = max(0.0, item["desired_trade_value"])
        alloc_value = max(0.0, item["allocated_trade_value"])
        guard_cap_value = item["guard_cap_value"]
        note = item["note"]

        if not note and guard_cap_value is not None and alloc_value + 1e-6 < desired_value:
            note = item["capped_note"]

        if ltp <= 0 or qty <= 0:
            suggested_qty = 0.0
            suggested_trade_value = 0.0
            post_trade_market_value = market_value
            if ltp <= 0:
                note = "missing_ltp"
        else:
            suggested_qty = _floor_qty(alloc_value / ltp)
            if side_u == "SELL":
                suggested_qty = min(suggested_qty, qty)
            suggested_trade_value = suggested_qty * ltp
            post_trade_market_value = market_value - suggested_trade_value if side_u == "SELL" else market_value + suggested_trade_value
            if side_u == "SELL" and post_trade_market_value + 1e-6 < min_value:
                post_trade_market_value = min_value

        total_suggested_value += suggested_trade_value
        out.append(
            {
                "symbol": symbol,
                "side": side_u,
                "percent": round(pct, 4),
                "qty": round(qty, 4),
                "ltp": round(ltp, 4),
                "market_value": round(market_value, 2),
                "min_value": round(min_value, 2),
                "max_value": round(max_value, 2) if max_value is not None else None,
                "desired_trade_value": round(desired_value, 2),
                "guard_cap_value": round(guard_cap_value, 2) if guard_cap_value is not None else None,
                "suggested_qty": round(max(0.0, suggested_qty), 4),
                "locked_qty": round(max(0.0, suggested_qty), 4),
                "remaining_qty": round(max(0.0, suggested_qty), 4),
                "completed": False,
                "lot_item_id": None,
                "suggested_trade_value": round(max(0.0, suggested_trade_value), 2),
                "post_trade_market_value": round(max(0.0, post_trade_market_value), 2),
                "note": note,
            }
        )

    return {
        "side": side_u,
        "percent": round(pct, 4),
        "allocation_basis": "portfolio_market_value",
        "target_trade_value": round(target_total_trade_value, 2),
        "lot": None,
        "items": out,
        "completed_items": 0,
        "remaining_items": len(out),
        "total_remaining_trade_value": round(total_suggested_value, 2),
        "active_nonzero_count": len(out),
        "total_current_market_value": round(total_current_value, 2),
        "total_suggested_trade_value": round(total_suggested_value, 2),
    }



def get_active_rebalance_lot(conn):
    row = conn.execute(
        """
        SELECT id, side, percent, allocation_basis, target_trade_value, total_current_market_value,
               status, created_at, completed_at, reset_at
        FROM rebalance_lots
        WHERE LOWER(status) = 'active'
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    return dict(row) if row else None


def _normalize_rebalance_execution_state(raw, default="pending"):
    s = str(raw or "").strip().lower()
    if not s:
        s = str(default or "pending").strip().lower() or "pending"
    if s not in ("pending", "closed", "skipped"):
        raise ValueError("invalid_rebalance_execution_state")
    return s


def _normalize_execution_timestamp(raw, fallback_now=True):
    s = str(raw or "").strip()
    if not s:
        return now_iso() if fallback_now else None
    try:
        return dt.datetime.fromisoformat(s.replace("Z", "+00:00")).replace(microsecond=0).isoformat()
    except Exception:
        pass
    try:
        d = dt.date.fromisoformat(s[:10])
        return dt.datetime.combine(d, dt.time.min).replace(microsecond=0).isoformat()
    except Exception:
        raise ValueError("executed_at_must_be_iso_datetime_or_date")


def list_rebalance_closed_history(conn, limit=250, include_buyback_completed=False):
    lim = int(parse_float(limit, 250))
    lim = max(1, min(1000, lim))
    include_done = bool(include_buyback_completed)
    extra_where = "" if include_done else " AND COALESCE(li.buyback_completed, 0) = 0 "
    rows = conn.execute(
        f"""
        SELECT
          li.id AS lot_item_id,
          li.lot_id,
          li.symbol,
          rl.side,
          li.planned_qty,
          li.execution_state,
          li.executed_price,
          COALESCE(li.executed_at, li.completed_at) AS executed_at,
          li.completion_note,
          li.buyback_completed,
          li.buyback_completed_at,
          li.buyback_price,
          li.buyback_note,
          COALESCE(lp.ltp, 0) AS ltp_now
        FROM rebalance_lot_items li
        JOIN rebalance_lots rl ON rl.id = li.lot_id
        LEFT JOIN latest_prices lp ON UPPER(lp.symbol) = UPPER(li.symbol)
        WHERE LOWER(COALESCE(li.execution_state, '')) = 'closed'
          {extra_where}
        ORDER BY COALESCE(li.executed_at, li.completed_at, rl.created_at) DESC, li.id DESC
        LIMIT ?
        """,
        (lim,),
    ).fetchall()
    out = []
    projected_total = 0.0
    realized_total = 0.0
    tracked_items = 0
    completed_items = 0
    for r in rows:
        qty = max(0.0, parse_float(r["planned_qty"], 0.0))
        executed_price = max(0.0, parse_float(r["executed_price"], 0.0))
        buyback_completed = bool(int(parse_float(r["buyback_completed"], 0.0)))
        buyback_price = parse_float(r["buyback_price"], 0.0) if r["buyback_price"] is not None else None
        buyback_price = buyback_price if (buyback_price is not None and buyback_price > 0) else None
        ltp_now = parse_float(r["ltp_now"], 0.0)
        if ltp_now <= 0:
            ltp_now = parse_float(get_effective_ltp(conn, r["symbol"]), 0.0)
        buyback_ref = ltp_now if ltp_now > 0 else None
        projected_saved_now = (
            (executed_price - buyback_ref) * qty
            if (not buyback_completed and buyback_ref is not None and executed_price > 0)
            else None
        )
        realized_saved = (
            (executed_price - buyback_price) * qty
            if (buyback_completed and buyback_price is not None and executed_price > 0)
            else None
        )
        if buyback_completed:
            completed_items += 1
        else:
            tracked_items += 1
        if projected_saved_now is not None:
            projected_total += projected_saved_now
        if realized_saved is not None:
            realized_total += realized_saved
        out.append(
            {
                "lot_item_id": int(parse_float(r["lot_item_id"], 0.0)),
                "lot_id": int(parse_float(r["lot_id"], 0.0)),
                "symbol": str(r["symbol"] or ""),
                "side": str(r["side"] or "").upper(),
                "qty": round(qty, 4),
                "executed_price": round(executed_price, 4),
                "executed_at": r["executed_at"],
                "note": str(r["completion_note"] or ""),
                "buyback_reference_price": round(buyback_ref, 4) if buyback_ref is not None else None,
                "projected_saved_now": round(projected_saved_now, 2) if projected_saved_now is not None else None,
                "buyback_completed": buyback_completed,
                "buyback_completed_at": r["buyback_completed_at"],
                "buyback_price": round(buyback_price, 4) if buyback_price is not None else None,
                "buyback_note": str(r["buyback_note"] or ""),
                "realized_saved": round(realized_saved, 2) if realized_saved is not None else None,
            }
        )
    return {
        "items": out,
        "count": len(out),
        "summary": {
            "tracked_items": int(tracked_items),
            "completed_items": int(completed_items),
            "projected_saved_now_total": round(projected_total, 2),
            "realized_saved_total": round(realized_total, 2),
            "net_saved_total": round(projected_total + realized_total, 2),
        },
    }


def _rebalance_lot_payload(conn, lot_row):
    closed_history = list_rebalance_closed_history(conn, limit=300)
    if not lot_row:
        return {
            "lot": None,
            "items": [],
            "side": "SELL",
            "percent": 0.0,
            "allocation_basis": "portfolio_market_value",
            "target_trade_value": 0.0,
            "active_nonzero_count": 0,
            "completed_items": 0,
            "remaining_items": 0,
            "total_current_market_value": 0.0,
            "total_suggested_trade_value": 0.0,
            "total_remaining_trade_value": 0.0,
            "closed_items": 0,
            "skipped_items": 0,
            "closed_history": closed_history,
        }
    lot_id = int(parse_float(lot_row.get("id"), 0.0))
    rows = conn.execute(
        """
        SELECT
          li.id AS lot_item_id,
          li.symbol,
          li.planned_qty,
          li.planned_trade_value,
          li.ltp_at_lock,
          li.market_value_at_lock,
          li.min_value_at_lock,
          li.max_value_at_lock,
          li.note,
          li.completed,
          li.completed_at,
          li.completion_note,
          li.execution_state,
          li.executed_price,
          li.executed_at,
          COALESCE(h.qty, 0) AS qty_now,
          COALESCE(h.market_value, 0) AS market_value_now,
          COALESCE(lp.ltp, 0) AS ltp_now,
          COALESCE(g.min_value, 0) AS min_value_now,
          g.max_value AS max_value_now
        FROM rebalance_lot_items li
        LEFT JOIN holdings h ON UPPER(h.symbol) = UPPER(li.symbol)
        LEFT JOIN latest_prices lp ON UPPER(lp.symbol) = UPPER(li.symbol)
        LEFT JOIN scrip_position_guards g ON UPPER(g.symbol) = UPPER(li.symbol)
        WHERE li.lot_id = ?
        ORDER BY li.id
        """,
        (lot_id,),
    ).fetchall()
    items = []
    completed_items = 0
    closed_items = 0
    skipped_items = 0
    remaining_items = 0
    total_planned = 0.0
    total_remaining = 0.0
    for r in rows:
        planned_qty = max(0.0, parse_float(r["planned_qty"], 0.0))
        legacy_done = bool(int(parse_float(r["completed"], 0.0)))
        exec_state = _normalize_rebalance_execution_state(
            r["execution_state"], default=("closed" if legacy_done else "pending")
        )
        is_done = exec_state in ("closed", "skipped")
        remaining_qty = 0.0 if is_done else planned_qty
        if is_done:
            completed_items += 1
            if exec_state == "closed":
                closed_items += 1
            elif exec_state == "skipped":
                skipped_items += 1
        else:
            remaining_items += 1
        total_planned += max(0.0, parse_float(r["planned_trade_value"], 0.0))
        total_remaining += 0.0 if is_done else max(0.0, parse_float(r["planned_trade_value"], 0.0))
        ltp_now = parse_float(r["ltp_now"], 0.0)
        if ltp_now <= 0:
            ltp_now = parse_float(get_effective_ltp(conn, r["symbol"]), 0.0)
        market_now = max(0.0, parse_float(r["market_value_now"], 0.0))
        min_now = max(0.0, parse_float(r["min_value_now"], 0.0))
        max_now = parse_float(r["max_value_now"], 0.0) if r["max_value_now"] is not None else None
        executed_price = max(0.0, parse_float(r["executed_price"], 0.0))
        guard_cap = max(0.0, market_now - min_now) if str(lot_row.get("side") or "").upper() == "SELL" else (
            max(0.0, max_now - market_now) if max_now is not None else None
        )
        post_market = market_now - (remaining_qty * ltp_now) if str(lot_row.get("side") or "").upper() == "SELL" else market_now + (remaining_qty * ltp_now)
        projected_saved_now = (executed_price - ltp_now) * planned_qty if exec_state == "closed" and executed_price > 0 and ltp_now > 0 else None
        items.append(
            {
                "lot_item_id": int(parse_float(r["lot_item_id"], 0.0)),
                "symbol": r["symbol"],
                "side": str(lot_row.get("side") or "SELL").upper(),
                "percent": round(parse_float(lot_row.get("percent"), 0.0), 4),
                "qty": round(max(0.0, parse_float(r["qty_now"], 0.0)), 4),
                "ltp": round(max(0.0, ltp_now), 4),
                "market_value": round(market_now, 2),
                "min_value": round(min_now, 2),
                "max_value": round(max_now, 2) if max_now is not None else None,
                "desired_trade_value": round(max(0.0, parse_float(r["planned_trade_value"], 0.0)), 2),
                "guard_cap_value": round(guard_cap, 2) if guard_cap is not None else None,
                "suggested_qty": round(remaining_qty, 4),
                "locked_qty": round(planned_qty, 4),
                "remaining_qty": round(remaining_qty, 4),
                "suggested_trade_value": round(max(0.0, remaining_qty * max(0.0, ltp_now)), 2),
                "post_trade_market_value": round(max(0.0, post_market), 2),
                "note": str(r["note"] or ""),
                "completed": bool(is_done),
                "completed_at": r["completed_at"],
                "completion_note": r["completion_note"],
                "execution_state": exec_state,
                "executed_price": round(executed_price, 4) if executed_price > 0 else None,
                "executed_at": r["executed_at"] or r["completed_at"],
                "buyback_reference_price": round(ltp_now, 4) if ltp_now > 0 else None,
                "projected_saved_now": round(projected_saved_now, 2) if projected_saved_now is not None else None,
            }
        )
    return {
        "lot": {
            "id": lot_id,
            "status": str(lot_row.get("status") or "active"),
            "created_at": lot_row.get("created_at"),
            "completed_at": lot_row.get("completed_at"),
            "reset_at": lot_row.get("reset_at"),
        },
        "side": str(lot_row.get("side") or "SELL").upper(),
        "percent": round(parse_float(lot_row.get("percent"), 0.0), 4),
        "allocation_basis": str(lot_row.get("allocation_basis") or "portfolio_market_value"),
        "target_trade_value": round(parse_float(lot_row.get("target_trade_value"), 0.0), 2),
        "items": items,
        "active_nonzero_count": len(items),
        "completed_items": completed_items,
        "closed_items": closed_items,
        "skipped_items": skipped_items,
        "remaining_items": remaining_items,
        "total_current_market_value": round(parse_float(lot_row.get("total_current_market_value"), 0.0), 2),
        "total_suggested_trade_value": round(total_planned, 2),
        "total_remaining_trade_value": round(total_remaining, 2),
        "closed_history": closed_history,
    }


def lock_rebalance_lot(conn, side, percent):
    active = get_active_rebalance_lot(conn)
    if active:
        raise ValueError("active_rebalance_lot_exists")
    suggestion = build_rebalance_suggestions(conn, side=side, percent=percent)
    items = [x for x in (suggestion.get("items") or []) if parse_float(x.get("suggested_qty"), 0.0) > 0]
    if not items:
        raise ValueError("no_suggestions_to_lock")
    cur = conn.execute(
        """
        INSERT INTO rebalance_lots(side, percent, allocation_basis, target_trade_value, total_current_market_value, status, created_at)
        VALUES (?, ?, ?, ?, ?, 'active', ?)
        """,
        (
            str(suggestion.get("side") or "SELL").upper(),
            parse_float(suggestion.get("percent"), 0.0),
            str(suggestion.get("allocation_basis") or "portfolio_market_value"),
            parse_float(suggestion.get("target_trade_value"), 0.0),
            parse_float(suggestion.get("total_current_market_value"), 0.0),
            now_iso(),
        ),
    )
    lot_id = int(cur.lastrowid)
    for item in items:
        conn.execute(
            """
            INSERT INTO rebalance_lot_items(
              lot_id, symbol, planned_qty, planned_trade_value, ltp_at_lock, market_value_at_lock,
              min_value_at_lock, max_value_at_lock, note, completed, execution_state
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 'pending')
            """,
            (
                lot_id,
                symbol_upper(item.get("symbol")),
                parse_float(item.get("suggested_qty"), 0.0),
                parse_float(item.get("suggested_trade_value"), 0.0),
                parse_float(item.get("ltp"), 0.0),
                parse_float(item.get("market_value"), 0.0),
                parse_float(item.get("min_value"), 0.0),
                (parse_float(item.get("max_value"), 0.0) if item.get("max_value") is not None else None),
                str(item.get("note") or ""),
            ),
        )
    row = conn.execute(
        "SELECT id, side, percent, allocation_basis, target_trade_value, total_current_market_value, status, created_at, completed_at, reset_at FROM rebalance_lots WHERE id = ?",
        (lot_id,),
    ).fetchone()
    return _rebalance_lot_payload(conn, dict(row))


def set_rebalance_lot_item_planned_qty(conn, item_id, planned_qty):
    iid = int(parse_float(item_id, 0.0))
    if iid <= 0:
        raise ValueError("invalid_lot_item_id")
    row = conn.execute(
        """
        SELECT
          li.id,
          li.lot_id,
          li.symbol,
          li.execution_state,
          li.ltp_at_lock,
          rl.side,
          rl.status,
          COALESCE(h.qty, 0) AS qty_now,
          COALESCE(lp.ltp, 0) AS ltp_now
        FROM rebalance_lot_items li
        JOIN rebalance_lots rl ON rl.id = li.lot_id
        LEFT JOIN holdings h ON UPPER(h.symbol) = UPPER(li.symbol)
        LEFT JOIN latest_prices lp ON UPPER(lp.symbol) = UPPER(li.symbol)
        WHERE li.id = ?
        """,
        (iid,),
    ).fetchone()
    if not row:
        raise ValueError("lot_item_not_found")
    if str(row["status"] or "").lower() != "active":
        raise ValueError("lot_not_active")
    exec_state = _normalize_rebalance_execution_state(row["execution_state"], default="pending")
    if exec_state != "pending":
        raise ValueError("planned_qty_edit_allowed_only_for_pending")

    q = parse_float(planned_qty, -1.0)
    if not math.isfinite(q) or q < 0:
        raise ValueError("planned_qty_must_be_non_negative")
    q = _floor_qty(q, step=0.0001)

    side = str(row["side"] or "").upper()
    qty_now = max(0.0, parse_float(row["qty_now"], 0.0))
    if side == "SELL" and q > (qty_now + 1e-6):
        raise ValueError("planned_qty_exceeds_current_qty")

    ltp_ref = parse_float(row["ltp_at_lock"], 0.0)
    if ltp_ref <= 0:
        ltp_ref = parse_float(row["ltp_now"], 0.0)
    if ltp_ref <= 0:
        ltp_ref = parse_float(get_effective_ltp(conn, row["symbol"]), 0.0)
    planned_trade_value = max(0.0, q * max(0.0, ltp_ref))
    conn.execute(
        """
        UPDATE rebalance_lot_items
        SET planned_qty = ?, planned_trade_value = ?
        WHERE id = ?
        """,
        (q, planned_trade_value, iid),
    )
    lot_id = int(parse_float(row["lot_id"], 0.0))
    lot_row = conn.execute(
        "SELECT id, side, percent, allocation_basis, target_trade_value, total_current_market_value, status, created_at, completed_at, reset_at FROM rebalance_lots WHERE id = ?",
        (lot_id,),
    ).fetchone()
    return _rebalance_lot_payload(conn, dict(lot_row) if lot_row else None)


def set_rebalance_lot_item_status(conn, item_id, state="pending", note="", executed_price=None, executed_at=None):
    iid = int(parse_float(item_id, 0.0))
    if iid <= 0:
        raise ValueError("invalid_lot_item_id")
    row = conn.execute(
        """
        SELECT li.id, li.lot_id, li.symbol, li.ltp_at_lock, li.executed_price, rl.status
        FROM rebalance_lot_items li
        JOIN rebalance_lots rl ON rl.id = li.lot_id
        WHERE li.id = ?
        """,
        (iid,),
    ).fetchone()
    if not row:
        raise ValueError("lot_item_not_found")
    if str(row["status"] or "").lower() != "active":
        raise ValueError("lot_not_active")
    state_norm = _normalize_rebalance_execution_state(state, default="pending")
    note_norm = str(note or "").strip() or None
    done = state_norm in ("closed", "skipped")
    exec_price = None
    exec_ts = None
    if state_norm == "closed":
        px_raw = executed_price
        if px_raw is None or str(px_raw).strip() == "":
            px_raw = row["executed_price"] if row["executed_price"] is not None else row["ltp_at_lock"]
            if parse_float(px_raw, 0.0) <= 0:
                px_raw = get_effective_ltp(conn, row["symbol"])
        exec_price = parse_float(px_raw, 0.0)
        if not math.isfinite(exec_price) or exec_price <= 0:
            raise ValueError("executed_price_required_for_closed")
        exec_ts = _normalize_execution_timestamp(executed_at, fallback_now=True)
    elif state_norm == "skipped":
        exec_ts = _normalize_execution_timestamp(executed_at, fallback_now=True)
    conn.execute(
        """
        UPDATE rebalance_lot_items
        SET completed = ?, completed_at = ?, completion_note = ?, execution_state = ?, executed_price = ?, executed_at = ?
        WHERE id = ?
        """,
        (
            1 if done else 0,
            exec_ts if done else None,
            note_norm if done else None,
            state_norm,
            exec_price if state_norm == "closed" else None,
            exec_ts if done else None,
            iid,
        ),
    )
    lot_id = int(parse_float(row["lot_id"], 0.0))
    pending = int(
        conn.execute(
            "SELECT COUNT(*) AS c FROM rebalance_lot_items WHERE lot_id = ? AND planned_qty > 0 AND LOWER(COALESCE(execution_state,'pending')) = 'pending'",
            (lot_id,),
        ).fetchone()["c"]
    )
    if pending == 0:
        conn.execute(
            "UPDATE rebalance_lots SET status='completed', completed_at=? WHERE id = ? AND LOWER(status)='active'",
            (now_iso(), lot_id),
        )
    lot_row = conn.execute(
        "SELECT id, side, percent, allocation_basis, target_trade_value, total_current_market_value, status, created_at, completed_at, reset_at FROM rebalance_lots WHERE id = ?",
        (lot_id,),
    ).fetchone()
    return _rebalance_lot_payload(conn, dict(lot_row) if lot_row else None)


def set_rebalance_buyback_status(conn, item_id, buyback_completed=True, buyback_price=None, buyback_at=None, note=""):
    iid = int(parse_float(item_id, 0.0))
    if iid <= 0:
        raise ValueError("invalid_lot_item_id")
    row = conn.execute(
        """
        SELECT id, symbol, execution_state
        FROM rebalance_lot_items
        WHERE id = ?
        """,
        (iid,),
    ).fetchone()
    if not row:
        raise ValueError("lot_item_not_found")
    if str(row["execution_state"] or "").strip().lower() != "closed":
        raise ValueError("lot_item_not_closed")
    done = bool(buyback_completed)
    buy_px = None
    buy_ts = None
    buy_note = None
    if done:
        px_raw = buyback_price
        if px_raw is not None and str(px_raw).strip() != "":
            buy_px = parse_float(px_raw, 0.0)
            if not math.isfinite(buy_px) or buy_px <= 0:
                raise ValueError("buyback_price_must_be_positive")
        buy_ts = _normalize_execution_timestamp(buyback_at, fallback_now=True)
        buy_note = str(note or "").strip() or None
    conn.execute(
        """
        UPDATE rebalance_lot_items
        SET buyback_completed = ?, buyback_completed_at = ?, buyback_price = ?, buyback_note = ?
        WHERE id = ?
        """,
        (1 if done else 0, buy_ts if done else None, buy_px if done else None, buy_note if done else None, iid),
    )
    return {
        "ok": True,
        "lot_item_id": iid,
        "buyback_completed": done,
        "buyback_completed_at": buy_ts,
        "buyback_price": buy_px,
    }


def set_rebalance_lot_item_completed(conn, item_id, completed=False, note=""):
    state = "closed" if bool(completed) else "pending"
    return set_rebalance_lot_item_status(conn, item_id=item_id, state=state, note=note)


def reset_active_rebalance_lot(conn):
    active = get_active_rebalance_lot(conn)
    if not active:
        raise ValueError("no_active_rebalance_lot")
    lot_id = int(parse_float(active.get("id"), 0.0))
    conn.execute(
        "UPDATE rebalance_lots SET status='reset', reset_at=? WHERE id = ? AND LOWER(status)='active'",
        (now_iso(), lot_id),
    )
    return {"ok": True, "lot_id": lot_id, "status": "reset", "reset_at": now_iso()}


def _normalize_daily_target_pair_state(raw, default="pending"):
    state = str(raw or default or "pending").strip().lower()
    if state == "done":
        state = "executed"
    allowed = {"pending", "sell_done", "buy_done", "executed", "skipped", "replaced"}
    if state not in allowed:
        raise ValueError("invalid_daily_target_pair_state")
    return state


def get_active_daily_target_plan(conn):
    row = conn.execute(
        """
        SELECT id, seed_capital, target_profit_pct, target_profit_value, top_n, status, created_at, updated_at,
               last_recalibrated_at, closed_at, notes
        FROM daily_target_plans
        WHERE LOWER(status) = 'active'
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    return dict(row) if row else None


def _daily_target_guard_map(conn):
    return {
        symbol_upper(r["symbol"]): {
            "min_value": round(parse_float(r["min_value"], 0.0), 2),
            "max_value": (
                None if r["max_value"] is None or str(r["max_value"]).strip() == "" else round(parse_float(r["max_value"], 0.0), 2)
            ),
        }
        for r in conn.execute("SELECT symbol, min_value, max_value FROM scrip_position_guards").fetchall()
    }


def _daily_target_pair_progress(buy_ref_price, current_buy_ref_price, buy_target_exit_price):
    entry = parse_float(buy_ref_price, 0.0)
    current = parse_float(current_buy_ref_price, 0.0)
    target = parse_float(buy_target_exit_price, 0.0)
    denom = target - entry
    if entry <= 0 or current <= 0 or target <= 0 or abs(denom) <= 1e-9:
        return 0.0
    return round(clamp(((current - entry) / denom) * 100.0, -250.0, 250.0), 2)


def _daily_target_live_pair_metrics(conn, pair_like, target_profit_pct=None):
    row = dict(pair_like or {})
    sell_symbol = symbol_upper(row.get("sell_symbol"))
    buy_symbol = symbol_upper(row.get("buy_symbol"))
    current_sell = parse_float(row.get("current_sell_ref_price"), parse_float(row.get("sell_ref_price"), 0.0))
    current_buy = parse_float(row.get("current_buy_ref_price"), parse_float(row.get("buy_ref_price"), 0.0))
    sell_ltp = get_effective_ltp(conn, sell_symbol) if sell_symbol else 0.0
    buy_ltp = get_effective_ltp(conn, buy_symbol) if buy_symbol else 0.0
    if sell_ltp > 0:
        current_sell = round(sell_ltp, 4)
    if buy_ltp > 0:
        current_buy = round(buy_ltp, 4)
    target_pct = round(
        parse_float(
            target_profit_pct,
            row.get("target_profit_pct"),
        ),
        2,
    )
    if target_pct <= 0:
        target_pct = 1.0
    buy_entry = parse_float(row.get("executed_buy_price"), 0.0)
    if buy_entry <= 0:
        buy_entry = parse_float(row.get("buy_ref_price"), 0.0)
    target_exit = parse_float(row.get("buy_target_exit_price"), 0.0)
    original_buy_ref = parse_float(row.get("buy_ref_price"), 0.0)
    if buy_entry > 0 and target_exit > 0 and original_buy_ref > 0:
        if parse_float(row.get("executed_buy_price"), 0.0) > 0:
            target_exit = round(buy_entry * (target_exit / max(original_buy_ref, 1e-9)), 4)
    elif buy_entry > 0:
        target_exit = round(buy_entry * (1.0 + (target_pct / 100.0)), 4)
    progress = _daily_target_pair_progress(buy_entry, current_buy, target_exit)
    return {
        "current_sell_ref_price": round(current_sell, 4),
        "current_buy_ref_price": round(current_buy, 4),
        "buy_entry_price": round(buy_entry, 4),
        "buy_target_exit_price": round(target_exit, 4),
        "target_progress_pct": round(progress, 2),
    }


def _daily_target_build_buy_leg(buy_candidate, available_capital, target_profit_pct=1.0, required_net_profit=None, tax_cfg=None):
    cand = dict(buy_candidate or {})
    cfg = dict(tax_cfg or {})
    capital = max(0.0, round(parse_float(available_capital, 0.0), 2))
    ltp = round(parse_float(cand.get("ltp"), 0.0), 4)
    if capital <= 0 or ltp <= 0:
        return None
    capacity = max(0.0, parse_float(cand.get("buy_capacity"), 0.0))
    capital = min(capital, capacity if capacity > 0 else capital)
    qty = math.floor((capital / max(1e-9, ltp)) + 1e-9)
    if qty <= 0:
        return None
    buy_value = round(qty * ltp, 2)
    exchange = str(cand.get("exchange") or "NSE").upper()
    buy_entry_costs = _daily_target_zerodha_delivery_costs(buy_value, exchange=exchange, side="BUY", include_dp_on_sell=False, tax_cfg=cfg)
    target_pct = max(0.1, parse_float(target_profit_pct, 1.0))
    required_net = round(
        parse_float(required_net_profit, 0.0),
        2,
    )
    if required_net <= 0:
        required_net = round((buy_value * target_pct) / 100.0, 2)
    target_exit_price, target_exit_stats = _daily_target_required_exit_price_for_net_goal(
        qty,
        ltp,
        required_net,
        exchange=exchange,
        tax_cfg=cfg,
    )
    expected_profit_value = round(parse_float(target_exit_stats.get("net_profit"), 0.0), 2)
    buy_tax_bits = [
        f"entry cost {money(buy_entry_costs.get('total'))}",
        f"future STCG {money(target_exit_stats.get('future_tax_total'))}",
        f"net target {money(expected_profit_value)}",
    ]
    return {
        "buy_symbol": symbol_upper(cand.get("symbol")),
        "buy_qty": round(qty, 4),
        "buy_ref_price": ltp,
        "buy_trade_value": buy_value,
        "buy_target_exit_price": round(target_exit_price, 4),
        "buy_score": round(parse_float(cand.get("buy_score"), 0.0), 4),
        "buy_reason": f"{str(cand.get('reason') or '').strip()}, {'; '.join(buy_tax_bits)}",
        "expected_profit_value": expected_profit_value,
        "current_buy_ref_price": ltp,
        "buy_exchange": exchange,
        "buy_entry_cost_total": round(parse_float(buy_entry_costs.get("total"), 0.0), 2),
        "buy_exit_cost_total": round(parse_float(target_exit_stats.get("exit_cost_total"), 0.0), 2),
        "buy_exit_tax_total": round(parse_float(target_exit_stats.get("future_tax_total"), 0.0), 2),
        "required_buy_leg_net_profit": required_net,
    }


def _daily_target_should_switch_pipeline_buy(current_row, proposed_buy_leg, buy_candidate_map):
    if not current_row or not proposed_buy_leg:
        return False
    current_symbol = symbol_upper(current_row.get("buy_symbol"))
    proposed_symbol = symbol_upper(proposed_buy_leg.get("buy_symbol"))
    if not proposed_symbol:
        return False
    if current_symbol != proposed_symbol:
        current_candidate = dict(buy_candidate_map.get(current_symbol) or {})
        current_score = parse_float(current_candidate.get("buy_score"), current_row.get("buy_score"))
        proposed_score = parse_float(proposed_buy_leg.get("buy_score"), 0.0)
        current_ltp = parse_float(current_row.get("current_buy_ref_price"), current_row.get("buy_ref_price"))
        proposed_ltp = parse_float(proposed_buy_leg.get("buy_ref_price"), 0.0)
        if current_symbol not in buy_candidate_map:
            return True
        if proposed_score >= (current_score + 1.25):
            return True
        if current_score <= 0 < proposed_score:
            return True
        if proposed_ltp > 0 and current_ltp > 0 and proposed_ltp <= (current_ltp * 0.985):
            return True
        return False
    current_qty = round(parse_float(current_row.get("buy_qty"), 0.0), 4)
    proposed_qty = round(parse_float(proposed_buy_leg.get("buy_qty"), 0.0), 4)
    current_ref = round(parse_float(current_row.get("buy_ref_price"), 0.0), 4)
    proposed_ref = round(parse_float(proposed_buy_leg.get("buy_ref_price"), 0.0), 4)
    current_value = round(parse_float(current_row.get("buy_trade_value"), 0.0), 2)
    proposed_value = round(parse_float(proposed_buy_leg.get("buy_trade_value"), 0.0), 2)
    return (
        abs(current_qty - proposed_qty) > 1e-9
        or abs(current_ref - proposed_ref) > 1e-9
        or abs(current_value - proposed_value) > 0.01
    )


DAILY_TARGET_USER_TAX_BRACKET_PCT = 30.0
DAILY_TARGET_EQUITY_STCG_TAX_PCT = 20.0
DAILY_TARGET_EQUITY_LTCG_TAX_PCT = 12.5
DAILY_TARGET_LTCG_EXEMPTION_LIMIT = 125000.0
DAILY_TARGET_ZERODHA_EQ_DELIVERY_TXN_RATE = {"NSE": 0.0000307, "BSE": 0.0000375}
DAILY_TARGET_ZERODHA_SEBI_RATE = 0.000001
DAILY_TARGET_ZERODHA_STT_DELIVERY_RATE = 0.001
DAILY_TARGET_ZERODHA_STAMP_BUY_RATE = 0.00015
DAILY_TARGET_ZERODHA_GST_RATE = 0.18
DAILY_TARGET_ZERODHA_DP_CHARGE_SELL_INCL_GST = 15.34


def _daily_target_stt_round(value):
    v = max(0.0, parse_float(value, 0.0))
    return float(math.floor(v + 0.5))


def _daily_target_equity_tax_rate(bucket, tax_cfg=None):
    cfg = dict(tax_cfg or {})
    if str(bucket or "").upper() == "LTCG":
        return parse_float(cfg.get("ltcg_rate_pct"), DAILY_TARGET_EQUITY_LTCG_TAX_PCT) / 100.0
    return parse_float(cfg.get("stcg_rate_pct"), DAILY_TARGET_EQUITY_STCG_TAX_PCT) / 100.0


def _daily_target_zerodha_delivery_costs(order_value, exchange="NSE", side="BUY", include_dp_on_sell=False, tax_cfg=None):
    cfg = dict(tax_cfg or {})
    value = max(0.0, round(parse_float(order_value, 0.0), 2))
    ex = str(exchange or "NSE").strip().upper()
    txn_rate_map = {
        "NSE": parse_float(cfg.get("txn_rate_nse"), DAILY_TARGET_ZERODHA_EQ_DELIVERY_TXN_RATE["NSE"]),
        "BSE": parse_float(cfg.get("txn_rate_bse"), DAILY_TARGET_ZERODHA_EQ_DELIVERY_TXN_RATE["BSE"]),
    }
    txn_rate = txn_rate_map.get(ex, txn_rate_map["NSE"])
    txn = value * txn_rate
    sebi = value * parse_float(cfg.get("sebi_rate"), DAILY_TARGET_ZERODHA_SEBI_RATE)
    brokerage = 0.0
    gst = (brokerage + txn + sebi) * parse_float(cfg.get("gst_rate"), DAILY_TARGET_ZERODHA_GST_RATE)
    stt = _daily_target_stt_round(value * parse_float(cfg.get("stt_delivery_rate"), DAILY_TARGET_ZERODHA_STT_DELIVERY_RATE))
    stamp = value * parse_float(cfg.get("stamp_buy_rate"), DAILY_TARGET_ZERODHA_STAMP_BUY_RATE) if str(side or "BUY").strip().upper() == "BUY" else 0.0
    dp = parse_float(cfg.get("dp_charge_sell_incl_gst"), DAILY_TARGET_ZERODHA_DP_CHARGE_SELL_INCL_GST) if include_dp_on_sell and str(side or "BUY").strip().upper() == "SELL" else 0.0
    total = round(brokerage + txn + sebi + gst + stt + stamp + dp, 2)
    return {
        "exchange": ex,
        "side": str(side or "BUY").strip().upper(),
        "brokerage": round(brokerage, 2),
        "transaction_charges": round(txn, 2),
        "sebi_charges": round(sebi, 2),
        "gst": round(gst, 2),
        "stt": round(stt, 2),
        "stamp": round(stamp, 2),
        "dp_charges": round(dp, 2),
        "total": total,
    }


def _daily_target_estimate_sell_tax_profile(conn, symbol, sell_qty, sell_price, split_map=None, as_of_date=None, tax_cfg=None, realized_tax_summary=None):
    cfg = dict(tax_cfg or get_tax_profile_config(conn))
    qty_req = max(0.0, parse_float(sell_qty, 0.0))
    px = max(0.0, parse_float(sell_price, 0.0))
    if qty_req <= 0 or px <= 0:
        return {
            "matched_qty": 0.0,
            "stcg_gain": 0.0,
            "ltcg_gain": 0.0,
            "stcg_loss": 0.0,
            "ltcg_loss": 0.0,
            "tax_payable": 0.0,
            "tax_relief": 0.0,
            "tax_drag": 0.0,
            "tax_bucket_mix": "NA",
            "fy_label": _india_fy_bounds(as_of_date=as_of_date).get("fy_label"),
            "ltcg_remaining_exemption_before": round(parse_float(cfg.get("ltcg_exemption_limit"), DAILY_TARGET_LTCG_EXEMPTION_LIMIT), 2),
            "ltcg_exemption_used": 0.0,
            "ltcg_taxable_gain_after_exemption": 0.0,
        }
    lots = open_lots_for_symbol(conn, symbol, split_map=split_map)
    realized = dict(realized_tax_summary or compute_realized_equity_tax_summary(conn, as_of_date=as_of_date))
    remaining = qty_req
    stcg_gain = 0.0
    ltcg_gain = 0.0
    stcg_loss = 0.0
    ltcg_loss = 0.0
    matched_qty = 0.0
    buckets_seen = set()
    for lot in lots:
        if remaining <= 1e-9:
            break
        q = min(remaining, parse_float(lot.get("qty"), 0.0))
        if q <= 0:
            continue
        bucket, _held_days = _harvest_tax_bucket(lot.get("buy_date"), as_of_date=as_of_date)
        pnl = (px - parse_float(lot.get("buy_price"), 0.0)) * q
        matched_qty += q
        buckets_seen.add(bucket)
        if pnl >= 0:
            if bucket == "LTCG":
                ltcg_gain += pnl
            else:
                stcg_gain += pnl
        else:
            if bucket == "LTCG":
                ltcg_loss += abs(pnl)
            else:
                stcg_loss += abs(pnl)
        remaining -= q
    stcg_net_gain = max(0.0, stcg_gain - stcg_loss)
    stcg_net_loss = max(0.0, stcg_loss - stcg_gain)
    ltcg_net_gain_before_exemption = max(0.0, ltcg_gain - ltcg_loss)
    ltcg_net_loss = max(0.0, ltcg_loss - ltcg_gain)
    ltcg_remaining_exemption_before = max(0.0, parse_float(realized.get("ltcg_remaining_exemption"), parse_float(cfg.get("ltcg_exemption_limit"), DAILY_TARGET_LTCG_EXEMPTION_LIMIT)))
    ltcg_exemption_used = min(ltcg_remaining_exemption_before, ltcg_net_gain_before_exemption)
    ltcg_taxable_gain_after_exemption = max(0.0, ltcg_net_gain_before_exemption - ltcg_exemption_used)
    tax_payable = (
        stcg_net_gain * _daily_target_equity_tax_rate("STCG", cfg)
        + ltcg_taxable_gain_after_exemption * _daily_target_equity_tax_rate("LTCG", cfg)
    )
    tax_relief = (
        stcg_net_loss * _daily_target_equity_tax_rate("STCG", cfg)
        + ltcg_net_loss * _daily_target_equity_tax_rate("LTCG", cfg)
    )
    bucket_mix = "MIXED" if len(buckets_seen) > 1 else (next(iter(buckets_seen)) if buckets_seen else "NA")
    return {
        "matched_qty": round(matched_qty, 4),
        "stcg_gain": round(stcg_gain, 2),
        "ltcg_gain": round(ltcg_gain, 2),
        "stcg_loss": round(stcg_loss, 2),
        "ltcg_loss": round(ltcg_loss, 2),
        "tax_payable": round(tax_payable, 2),
        "tax_relief": round(tax_relief, 2),
        "tax_drag": round(tax_payable - tax_relief, 2),
        "tax_bucket_mix": bucket_mix,
        "stcg_net_gain": round(stcg_net_gain, 2),
        "stcg_net_loss": round(stcg_net_loss, 2),
        "ltcg_net_gain_before_exemption": round(ltcg_net_gain_before_exemption, 2),
        "ltcg_net_loss": round(ltcg_net_loss, 2),
        "ltcg_remaining_exemption_before": round(ltcg_remaining_exemption_before, 2),
        "ltcg_exemption_used": round(ltcg_exemption_used, 2),
        "ltcg_taxable_gain_after_exemption": round(ltcg_taxable_gain_after_exemption, 2),
        "fy_label": str(realized.get("fy_label") or ""),
    }


def _daily_target_buy_leg_net_profit_at_exit(buy_qty, buy_price, exit_price, exchange="NSE", tax_cfg=None):
    cfg = dict(tax_cfg or {})
    qty = max(0.0, parse_float(buy_qty, 0.0))
    entry_px = max(0.0, parse_float(buy_price, 0.0))
    exit_px = max(0.0, parse_float(exit_price, 0.0))
    if qty <= 0 or entry_px <= 0 or exit_px <= 0:
        return {
            "gross_profit": 0.0,
            "entry_cost_total": 0.0,
            "exit_cost_total": 0.0,
            "future_tax_total": 0.0,
            "net_profit": 0.0,
        }
    buy_value = round(qty * entry_px, 2)
    sell_value = round(qty * exit_px, 2)
    entry_costs = _daily_target_zerodha_delivery_costs(buy_value, exchange=exchange, side="BUY", include_dp_on_sell=False, tax_cfg=cfg)
    exit_costs = _daily_target_zerodha_delivery_costs(sell_value, exchange=exchange, side="SELL", include_dp_on_sell=True, tax_cfg=cfg)
    gross_profit = round(sell_value - buy_value, 2)
    future_tax_total = round(max(0.0, gross_profit) * _daily_target_equity_tax_rate("STCG", cfg), 2)
    net_profit = round(gross_profit - parse_float(entry_costs.get("total"), 0.0) - parse_float(exit_costs.get("total"), 0.0) - future_tax_total, 2)
    return {
        "gross_profit": gross_profit,
        "entry_cost_total": round(parse_float(entry_costs.get("total"), 0.0), 2),
        "exit_cost_total": round(parse_float(exit_costs.get("total"), 0.0), 2),
        "future_tax_total": future_tax_total,
        "net_profit": net_profit,
    }


def _daily_target_required_exit_price_for_net_goal(buy_qty, buy_price, required_net_profit, exchange="NSE", tax_cfg=None):
    qty = max(0.0, parse_float(buy_qty, 0.0))
    entry_px = max(0.0, parse_float(buy_price, 0.0))
    required = round(parse_float(required_net_profit, 0.0), 2)
    if qty <= 0 or entry_px <= 0:
        return round(entry_px, 4), _daily_target_buy_leg_net_profit_at_exit(qty, entry_px, entry_px, exchange=exchange, tax_cfg=tax_cfg)
    if required <= 0:
        return round(entry_px, 4), _daily_target_buy_leg_net_profit_at_exit(qty, entry_px, entry_px, exchange=exchange, tax_cfg=tax_cfg)
    lo = entry_px
    hi = max(entry_px * 1.03, entry_px + (required / max(qty, 1e-9)) * 3.0 + 1.0)
    probe = _daily_target_buy_leg_net_profit_at_exit(qty, entry_px, hi, exchange=exchange, tax_cfg=tax_cfg)
    for _ in range(24):
        if parse_float(probe.get("net_profit"), 0.0) >= required:
            break
        hi *= 1.12
        probe = _daily_target_buy_leg_net_profit_at_exit(qty, entry_px, hi, exchange=exchange, tax_cfg=tax_cfg)
    best = probe
    for _ in range(26):
        mid = (lo + hi) / 2.0
        estimate = _daily_target_buy_leg_net_profit_at_exit(qty, entry_px, mid, exchange=exchange, tax_cfg=tax_cfg)
        if parse_float(estimate.get("net_profit"), 0.0) >= required:
            hi = mid
            best = estimate
        else:
            lo = mid
    return round(hi, 4), best


def _daily_target_trade_match(conn, symbol, side, exec_date, qty, price, exclude_ids=None):
    sym = symbol_upper(symbol)
    side_u = str(side or "").strip().upper()
    d = str(exec_date or "").strip()[:10]
    q = parse_float(qty, 0.0)
    p = parse_float(price, 0.0)
    if not sym or side_u not in ("BUY", "SELL") or not d or q <= 0:
        return None
    rows = conn.execute(
        """
        SELECT id, symbol, side, trade_date, quantity, price
        FROM trades
        WHERE UPPER(symbol) = ?
          AND UPPER(side) = ?
          AND trade_date = ?
        ORDER BY id DESC
        """,
        (sym, side_u, d),
    ).fetchall()
    blocked = {int(parse_float(x, 0.0)) for x in (exclude_ids or []) if int(parse_float(x, 0.0)) > 0}
    best = None
    for r in rows:
        tid = int(parse_float(r["id"], 0.0))
        if tid in blocked:
            continue
        q_gap = abs(parse_float(r["quantity"], 0.0) - q)
        p_gap = abs(parse_float(r["price"], 0.0) - p) if p > 0 else 0.0
        if q_gap > max(0.02, q * 0.02):
            continue
        score = (q_gap * 1000.0) + p_gap
        if best is None or score < best[0]:
            best = (
                score,
                {
                    "trade_id": tid,
                    "trade_date": str(r["trade_date"] or ""),
                    "quantity": round(parse_float(r["quantity"], 0.0), 4),
                    "price": round(parse_float(r["price"], 0.0), 4),
                },
            )
    return best[1] if best else None


def reconcile_daily_target_trade_links(conn, plan_id=None):
    where = []
    params = []
    if plan_id is not None:
        where.append("plan_id = ?")
        params.append(int(parse_float(plan_id, 0.0)))
    sql = "SELECT * FROM daily_target_plan_pairs"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY updated_at DESC, id DESC"
    rows = conn.execute(sql, params).fetchall()
    used_trade_ids = set()
    for r in rows:
        pair_id = int(parse_float(r["id"], 0.0))
        sell_match = None
        buy_match = None
        if str(r["executed_sell_at"] or "").strip():
            sell_match = _daily_target_trade_match(
                conn,
                r["sell_symbol"],
                "SELL",
                r["executed_sell_at"],
                r["sell_qty"],
                r["executed_sell_price"],
                exclude_ids=used_trade_ids,
            )
        if sell_match:
            used_trade_ids.add(int(parse_float(sell_match.get("trade_id"), 0.0)))
        if str(r["executed_buy_at"] or "").strip():
            buy_match = _daily_target_trade_match(
                conn,
                r["buy_symbol"],
                "BUY",
                r["executed_buy_at"],
                r["buy_qty"],
                r["executed_buy_price"],
                exclude_ids=used_trade_ids,
            )
        if buy_match:
            used_trade_ids.add(int(parse_float(buy_match.get("trade_id"), 0.0)))
        status = "unmatched"
        if sell_match and buy_match:
            status = "matched"
        elif sell_match or buy_match:
            status = "partial"
        conn.execute(
            """
            UPDATE daily_target_plan_pairs
            SET matched_sell_trade_id = ?, matched_buy_trade_id = ?, reconciliation_status = ?, updated_at = COALESCE(updated_at, ?)
            WHERE id = ?
            """,
            (
                int(parse_float((sell_match or {}).get("trade_id"), 0.0)) or None,
                int(parse_float((buy_match or {}).get("trade_id"), 0.0)) or None,
                status,
                now_iso(),
                pair_id,
            ),
        )


def sync_daily_target_positions(conn, pair_id):
    iid = int(parse_float(pair_id, 0.0))
    row = conn.execute(
        """
        SELECT id, sell_symbol, sell_qty, buy_symbol, buy_qty, executed_sell_price, executed_sell_at,
               executed_buy_price, executed_buy_at, state
        FROM daily_target_plan_pairs
        WHERE id = ?
        """,
        (iid,),
    ).fetchone()
    if not row:
        return
    state = str(row["state"] or "").strip().lower()
    if state != "executed":
        return
    buy_qty = parse_float(row["buy_qty"], 0.0)
    buy_price = parse_float(row["executed_buy_price"], 0.0)
    buy_at = str(row["executed_buy_at"] or "").strip()
    if buy_qty > 0 and buy_price > 0 and buy_at:
        existing_open = conn.execute(
            "SELECT id FROM daily_target_positions WHERE source_pair_id = ? AND LOWER(status) = 'open' ORDER BY id DESC LIMIT 1",
            (iid,),
        ).fetchone()
        if existing_open:
            conn.execute(
                """
                UPDATE daily_target_positions
                SET symbol = ?, qty = ?, initial_qty = ?, closed_qty = 0, entry_price = ?, entry_value = ?, realized_profit = 0, entry_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    symbol_upper(row["buy_symbol"]),
                    round(buy_qty, 4),
                    round(buy_qty, 4),
                    round(buy_price, 4),
                    round(buy_qty * buy_price, 2),
                    buy_at,
                    now_iso(),
                    int(parse_float(existing_open["id"], 0.0)),
                ),
            )
        else:
            conn.execute(
                """
                INSERT INTO daily_target_positions(
                  source_pair_id, symbol, qty, initial_qty, closed_qty, entry_price, entry_value, realized_profit, entry_at, status, created_at, updated_at
                ) VALUES (?, ?, ?, ?, 0, ?, ?, 0, ?, 'open', ?, ?)
                """,
                (
                    iid,
                    symbol_upper(row["buy_symbol"]),
                    round(buy_qty, 4),
                    round(buy_qty, 4),
                    round(buy_price, 4),
                    round(buy_qty * buy_price, 2),
                    buy_at,
                    now_iso(),
                    now_iso(),
                ),
            )
    sell_qty = parse_float(row["sell_qty"], 0.0)
    sell_price = parse_float(row["executed_sell_price"], 0.0)
    sell_at = str(row["executed_sell_at"] or "").strip()
    sell_symbol = symbol_upper(row["sell_symbol"])
    if sell_qty > 0 and sell_price > 0 and sell_at and sell_symbol:
        remaining = sell_qty
        open_positions = conn.execute(
            """
            SELECT id, qty
                 , initial_qty, closed_qty, entry_price, realized_profit
            FROM daily_target_positions
            WHERE UPPER(symbol) = ? AND LOWER(status) = 'open' AND source_pair_id <> ?
            ORDER BY entry_at ASC, id ASC
            """,
            (sell_symbol, iid),
        ).fetchall()
        for pos in open_positions:
            if remaining <= 1e-9:
                break
            pos_id = int(parse_float(pos["id"], 0.0))
            pos_qty = parse_float(pos["qty"], 0.0)
            pos_entry_price = parse_float(pos["entry_price"], 0.0)
            pos_closed_qty = parse_float(pos["closed_qty"], 0.0)
            pos_realized = parse_float(pos["realized_profit"], 0.0)
            if pos_qty <= 1e-9:
                continue
            close_qty = min(pos_qty, remaining)
            realized_add = round((sell_price - pos_entry_price) * close_qty, 2) if pos_entry_price > 0 else 0.0
            if pos_qty <= (remaining + 1e-9):
                conn.execute(
                    """
                    UPDATE daily_target_positions
                    SET qty = 0, closed_qty = ?, realized_profit = ?, exit_pair_id = ?, exit_price = ?, exit_value = ?, exit_at = ?, status = 'closed', updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        round(pos_closed_qty + close_qty, 4),
                        round(pos_realized + realized_add, 2),
                        iid,
                        round(sell_price, 4),
                        round(pos_qty * sell_price, 2),
                        sell_at,
                        now_iso(),
                        pos_id,
                    ),
                )
                remaining -= pos_qty
            else:
                conn.execute(
                    """
                    UPDATE daily_target_positions
                    SET qty = ?, closed_qty = ?, realized_profit = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        round(pos_qty - remaining, 4),
                        round(pos_closed_qty + close_qty, 4),
                        round(pos_realized + realized_add, 2),
                        now_iso(),
                        pos_id,
                    ),
                )
                remaining = 0.0


def _daily_target_virtual_open_positions(conn):
    rows = conn.execute(
        """
        SELECT p.symbol, SUM(p.qty) AS qty, AVG(p.entry_price) AS avg_entry_price
        FROM daily_target_positions p
        JOIN daily_target_plan_pairs dp ON dp.id = p.source_pair_id
        WHERE LOWER(p.status) = 'open'
          AND LOWER(COALESCE(dp.reconciliation_status,'')) <> 'matched'
        GROUP BY p.symbol
        ORDER BY p.symbol
        """
    ).fetchall()
    out = {}
    for r in rows:
        sym = symbol_upper(r["symbol"])
        qty = round(parse_float(r["qty"], 0.0), 4)
        avg_entry = round(parse_float(r["avg_entry_price"], 0.0), 4)
        if sym and qty > 0:
            out[sym] = {"qty": qty, "avg_entry_price": avg_entry}
    return out


def compute_daily_target_performance(conn):
    rows = conn.execute(
        """
        SELECT p.*, dp.created_at AS plan_created_at, dp.seed_capital, dp.target_profit_pct
        FROM daily_target_plan_pairs p
        JOIN daily_target_plans dp ON dp.id = p.plan_id
        WHERE LOWER(dp.status) IN ('active','completed','reset')
        ORDER BY COALESCE(p.executed_buy_at, p.executed_sell_at, p.updated_at) ASC, p.id ASC
        """
    ).fetchall()
    if not rows:
        return {
            "starting_capital": 10000.0,
            "current_compounded_capital": 10000.0,
            "compounded_return_value": 0.0,
            "compounded_return_pct": 0.0,
            "realized_compounded_capital": 10000.0,
            "realized_profit_value": 0.0,
            "realized_profit_pct": 0.0,
            "executed_rotation_count": 0,
            "matched_rotation_count": 0,
            "unmatched_rotation_count": 0,
            "cumulative_sell_value": 0.0,
            "cumulative_buy_value": 0.0,
            "live_mtm_pnl": 0.0,
            "live_mtm_return_pct": 0.0,
            "latest_symbol": "",
            "latest_trade_date": "",
            "suggested_next_seed_capital": 10000.0,
        }
    starting_capital = round(parse_float(rows[0]["seed_capital"], 10000.0), 2)
    cumulative_sell_value = 0.0
    cumulative_buy_value = 0.0
    executed_rotation_count = 0
    matched_rotation_count = 0
    unmatched_rotation_count = 0
    latest_balance = starting_capital
    latest_symbol = ""
    latest_trade_date = ""
    live_mtm_pnl = 0.0
    live_mtm_return_pct = 0.0
    live_mtm_basis_value = 0.0
    latest_row = None
    for r in rows:
        state = str(r["state"] or "").strip().lower()
        if state in ("replaced",):
            continue
        sell_value = round(parse_float(r["executed_sell_price"], 0.0) * parse_float(r["sell_qty"], 0.0), 2) if parse_float(r["executed_sell_price"], 0.0) > 0 else round(parse_float(r["sell_trade_value"], 0.0), 2)
        buy_cost = round(parse_float(r["executed_buy_price"], 0.0) * parse_float(r["buy_qty"], 0.0), 2) if parse_float(r["executed_buy_price"], 0.0) > 0 else round(parse_float(r["buy_trade_value"], 0.0), 2)
        if sell_value > 0:
            cumulative_sell_value += sell_value
        if buy_cost > 0:
            cumulative_buy_value += buy_cost
        if state in ("sell_done", "buy_done", "executed"):
            executed_rotation_count += 1
            latest_row = r
            latest_trade_date = str(r["executed_buy_at"] or r["executed_sell_at"] or r["updated_at"] or "")[:10]
            latest_symbol = str(r["buy_symbol"] or "")
            if str(r["reconciliation_status"] or "").strip().lower() == "matched":
                matched_rotation_count += 1
            else:
                unmatched_rotation_count += 1
    pos_rows = conn.execute(
        """
        SELECT symbol, qty, entry_value, realized_profit, status
        FROM daily_target_positions
        ORDER BY entry_at ASC, id ASC
        """
    ).fetchall()
    realized_profit_value = round(sum(parse_float(r["realized_profit"], 0.0) for r in pos_rows), 2)
    realized_compounded_capital = round(starting_capital + realized_profit_value, 2)
    open_positions = conn.execute(
        """
        SELECT symbol, qty, entry_value
        FROM daily_target_positions
        WHERE LOWER(status) = 'open'
        ORDER BY entry_at ASC, id ASC
        """
    ).fetchall()
    if open_positions:
        latest_balance = realized_compounded_capital
        total_entry_value = 0.0
        live_mtm_pnl = 0.0
        for pos in open_positions:
            px = get_effective_ltp(conn, pos["symbol"])
            qty = parse_float(pos["qty"], 0.0)
            current_value = round(qty * px, 2) if px > 0 and qty > 0 else round(parse_float(pos["entry_value"], 0.0), 2)
            entry_value = round(parse_float(pos["entry_value"], 0.0), 2)
            latest_balance += (current_value - entry_value)
            live_mtm_pnl += round(current_value - entry_value, 2)
            total_entry_value += round(parse_float(pos["entry_value"], 0.0), 2)
        live_mtm_pnl = round(live_mtm_pnl, 2)
        live_mtm_basis_value = round(total_entry_value, 2)
        live_mtm_return_pct = round((live_mtm_pnl / total_entry_value) * 100.0, 2) if total_entry_value > 0 else 0.0
    elif latest_row is None:
        latest_balance = starting_capital
    else:
        latest_balance = realized_compounded_capital
    compounded_return_value = round(latest_balance - starting_capital, 2)
    compounded_return_pct = round((compounded_return_value / starting_capital) * 100.0, 2) if starting_capital > 0 else 0.0
    return {
        "starting_capital": round(starting_capital, 2),
        "current_compounded_capital": round(latest_balance, 2),
        "compounded_return_value": compounded_return_value,
        "compounded_return_pct": compounded_return_pct,
        "realized_compounded_capital": realized_compounded_capital,
        "realized_profit_value": realized_profit_value,
        "realized_profit_pct": round((realized_profit_value / starting_capital) * 100.0, 2) if starting_capital > 0 else 0.0,
        "executed_rotation_count": executed_rotation_count,
        "matched_rotation_count": matched_rotation_count,
        "unmatched_rotation_count": unmatched_rotation_count,
        "open_position_count": len(open_positions),
        "cumulative_sell_value": round(cumulative_sell_value, 2),
        "cumulative_buy_value": round(cumulative_buy_value, 2),
        "live_mtm_pnl": round(live_mtm_pnl, 2),
        "live_mtm_return_pct": round(live_mtm_return_pct, 2),
        "live_mtm_basis_value": round(live_mtm_basis_value, 2),
        "latest_symbol": latest_symbol,
        "latest_trade_date": latest_trade_date,
        "suggested_next_seed_capital": realized_compounded_capital,
    }


def build_daily_target_suggestions(conn, seed_capital=10000.0, target_profit_pct=1.0, top_n=5):
    seed = round(clamp(parse_float(seed_capital, 10000.0), 1000.0, 1_000_000_000.0), 2)
    target_pct = round(clamp(parse_float(target_profit_pct, 1.0), 0.1, 25.0), 2)
    limit = max(1, min(10, int(parse_float(top_n, 5))))
    split_map = load_split_map(conn)
    tax_cfg = get_tax_profile_config(conn)
    realized_tax_summary = compute_realized_equity_tax_summary(conn)

    cfg = get_intel_parameter_bundle(conn)
    strategy_map = latest_strategy_recommendation_map(conn)
    chart_map = latest_chart_snapshot_map(conn)
    intel_map = build_intelligence_bias_map(
        conn,
        decay_days=int(cfg.get("decay_days", 45)),
        w_commentary=parse_float(cfg.get("w_commentary"), 0.25),
        w_policy=parse_float(cfg.get("w_policy"), 0.25),
        w_financials=parse_float(cfg.get("w_financials"), 0.5),
        w_chart=parse_float(cfg.get("w_chart"), 0.0),
    )
    guard_map = _daily_target_guard_map(conn)
    virtual_open = _daily_target_virtual_open_positions(conn)
    holdings = []
    for item in collect_strategy_universe(conn, lookback_days=30):
        symbol = symbol_upper(item.get("symbol"))
        if not symbol:
            continue
        if symbol_upper(item.get("asset_class") or "EQUITY") == "GOLD":
            continue
        qty = parse_float(item.get("qty"), 0.0)
        ltp = parse_float(item.get("ltp"), 0.0)
        vrow = virtual_open.get(symbol)
        if vrow and qty <= 0:
            qty = parse_float(vrow.get("qty"), 0.0)
            item["qty"] = qty
            if parse_float(item.get("avg_cost"), 0.0) <= 0:
                item["avg_cost"] = parse_float(vrow.get("avg_entry_price"), 0.0)
            item["invested"] = round(qty * parse_float(item.get("avg_cost"), 0.0), 2)
            item["market_value"] = round(qty * ltp, 2)
        if qty <= 0 or ltp <= 0:
            continue
        market_value = max(parse_float(item.get("market_value"), 0.0), qty * ltp)
        strat = strategy_map.get(symbol, {})
        intel = intel_map.get(symbol, {})
        chart = chart_map.get(symbol, {})
        fin = financial_signal_for_symbol(conn, symbol)
        action_bias = _harvest_action_bias(strat.get("action"))
        signal_bias = _harvest_signal_bias(item.get("buy_signal"), item.get("sell_signal"))
        expected_move_score = _harvest_expected_move_score(
            parse_float(intel.get("score"), 0.0),
            parse_float(chart.get("score"), 0.0),
            parse_float(fin.get("score"), 0.0),
            action_bias,
            signal_bias,
        )
        direction = _harvest_direction_label(expected_move_score)
        momentum = parse_float(item.get("momentum_lookback_pct"), 0.0)
        total_return = parse_float(item.get("total_return_pct"), 0.0)
        guards = guard_map.get(symbol, {})
        min_value = parse_float(guards.get("min_value"), 0.0)
        max_value = guards.get("max_value")
        max_sell_value = max(0.0, market_value - max(0.0, min_value))
        buy_capacity = (max(0.0, parse_float(max_value, 0.0) - market_value) if max_value not in (None, "") else seed * 3.0)
        tax_rows = open_lot_tax_bucket_rows(conn, symbol, ltp=ltp, split_map=split_map)
        stcg_loss_available = round(sum(max(0.0, -parse_float(x.get("unrealized_pnl"), 0.0)) for x in tax_rows if str(x.get("tax_bucket") or "").upper() == "STCG" and parse_float(x.get("unrealized_pnl"), 0.0) < 0), 2)
        ltcg_loss_available = round(sum(max(0.0, -parse_float(x.get("unrealized_pnl"), 0.0)) for x in tax_rows if str(x.get("tax_bucket") or "").upper() == "LTCG" and parse_float(x.get("unrealized_pnl"), 0.0) < 0), 2)
        stcg_gain_available = round(sum(max(0.0, parse_float(x.get("unrealized_pnl"), 0.0)) for x in tax_rows if str(x.get("tax_bucket") or "").upper() == "STCG" and parse_float(x.get("unrealized_pnl"), 0.0) > 0), 2)
        ltcg_gain_available = round(sum(max(0.0, parse_float(x.get("unrealized_pnl"), 0.0)) for x in tax_rows if str(x.get("tax_bucket") or "").upper() == "LTCG" and parse_float(x.get("unrealized_pnl"), 0.0) > 0), 2)
        tax_alpha_value = round(
            (stcg_loss_available * _daily_target_equity_tax_rate("STCG", tax_cfg))
            + (ltcg_loss_available * _daily_target_equity_tax_rate("LTCG", tax_cfg))
            - (stcg_gain_available * _daily_target_equity_tax_rate("STCG", tax_cfg))
            - (ltcg_gain_available * _daily_target_equity_tax_rate("LTCG", tax_cfg)),
            2,
        )
        tax_alpha_score = round((tax_alpha_value / max(seed, 1.0)) * 500.0, 4)
        sell_score = (
            max(0.0, -expected_move_score)
            + max(0.0, -momentum) * 0.4
            + max(0.0, total_return) * 0.08
            + (6.0 if direction == "DOWN" else (1.0 if direction == "MIXED" else 0.0))
            + tax_alpha_score
        )
        buy_score = (
            max(0.0, expected_move_score)
            + max(0.0, momentum) * 0.4
            + max(0.0, parse_float(intel.get("confidence"), 0.0)) * 10.0
            + (6.0 if direction == "UP" else (1.0 if direction == "MIXED" else 0.0))
        )
        reason = _harvest_priority_reason(
            direction,
            strat.get("action"),
            chart.get("signal"),
            intel.get("summary"),
            fin.get("summary"),
        )
        holdings.append(
            {
                "symbol": symbol,
                "qty": round(qty, 4),
                "ltp": round(ltp, 4),
                "market_value": round(market_value, 2),
                "max_sell_value": round(max_sell_value, 2),
                "buy_capacity": round(buy_capacity, 2),
                "sell_score": round(sell_score, 4),
                "buy_score": round(buy_score, 4),
                "expected_move_score": round(expected_move_score, 4),
                "direction": direction,
                "reason": reason,
                "strategy_action": str(strat.get("action") or "").upper(),
                "buy_signal": str(item.get("buy_signal") or "").upper(),
                "sell_signal": str(item.get("sell_signal") or "").upper(),
                "momentum_lookback_pct": round(momentum, 2),
                "total_return_pct": round(total_return, 2),
                "min_value": round(min_value, 2),
                "max_value": None if max_value in (None, "") else round(parse_float(max_value, 0.0), 2),
                "exchange": str(item.get("exchange") or "NSE").upper(),
                "tax_alpha_value": round(tax_alpha_value, 2),
                "tax_alpha_score": round(tax_alpha_score, 4),
                "stcg_loss_available": stcg_loss_available,
                "ltcg_loss_available": ltcg_loss_available,
                "stcg_gain_available": stcg_gain_available,
                "ltcg_gain_available": ltcg_gain_available,
            }
        )

    sell_candidates = [
        x for x in holdings if x["max_sell_value"] >= max(x["ltp"], min(seed * 0.25, x["market_value"]))
    ]
    buy_candidates = [
        x for x in holdings if x["buy_capacity"] >= x["ltp"]
    ]
    sell_candidates.sort(key=lambda x: (x["sell_score"], x["market_value"]), reverse=True)
    buy_candidates.sort(key=lambda x: (x["buy_score"], x["market_value"]), reverse=True)

    pairs = []
    used_sell = set()
    used_buy = set()
    used_symbols = set()
    for sell in sell_candidates:
        if sell["symbol"] in used_sell or sell["symbol"] in used_symbols:
            continue
        best = None
        for buy in buy_candidates:
            if buy["symbol"] == sell["symbol"] or buy["symbol"] in used_buy or buy["symbol"] in used_symbols or buy["symbol"] in used_sell:
                continue
            allocation = min(
                seed,
                parse_float(sell.get("max_sell_value"), 0.0),
                parse_float(buy.get("buy_capacity"), 0.0),
            )
            if allocation <= 0:
                continue
            sell_qty = min(parse_float(sell.get("qty"), 0.0), math.floor((allocation / max(1e-9, sell["ltp"])) + 1e-9))
            if sell_qty <= 0:
                continue
            sell_value = round(sell_qty * sell["ltp"], 2)
            sell_costs = _daily_target_zerodha_delivery_costs(
                sell_value,
                exchange=sell.get("exchange"),
                side="SELL",
                include_dp_on_sell=True,
                tax_cfg=tax_cfg,
            )
            sell_tax = _daily_target_estimate_sell_tax_profile(
                conn,
                sell["symbol"],
                sell_qty,
                sell["ltp"],
                split_map=split_map,
                tax_cfg=tax_cfg,
                realized_tax_summary=realized_tax_summary,
            )
            redeployable_capital = max(0.0, round(sell_value - parse_float(sell_costs.get("total"), 0.0), 2))
            buy_qty = math.floor((redeployable_capital / max(1e-9, buy["ltp"])) + 1e-9)
            if buy_qty <= 0:
                continue
            buy_value = round(buy_qty * buy["ltp"], 2)
            buy_entry_costs = _daily_target_zerodha_delivery_costs(
                buy_value,
                exchange=buy.get("exchange"),
                side="BUY",
                include_dp_on_sell=False,
                tax_cfg=tax_cfg,
            )
            required_buy_leg_net = round(
                max(
                    0.0,
                    ((buy_value * target_pct) / 100.0)
                    + parse_float(sell_costs.get("total"), 0.0)
                    + parse_float(sell_tax.get("tax_drag"), 0.0),
                ),
                2,
            )
            target_exit_price, target_exit_stats = _daily_target_required_exit_price_for_net_goal(
                buy_qty,
                buy["ltp"],
                required_buy_leg_net,
                exchange=buy.get("exchange"),
                tax_cfg=tax_cfg,
            )
            expected_profit_value = round(
                parse_float(target_exit_stats.get("net_profit"), 0.0)
                - parse_float(sell_costs.get("total"), 0.0)
                - parse_float(sell_tax.get("tax_drag"), 0.0),
                2,
            )
            tax_alpha_score = round(
                (
                    parse_float(sell_tax.get("tax_relief"), 0.0)
                    - parse_float(sell_tax.get("tax_payable"), 0.0)
                    - parse_float(sell_costs.get("total"), 0.0)
                    - parse_float(buy_entry_costs.get("total"), 0.0)
                )
                / max(seed, 1.0)
                * 500.0,
                4,
            )
            rotation_score = round(parse_float(sell.get("sell_score"), 0.0) + parse_float(buy.get("buy_score"), 0.0) + tax_alpha_score, 4)
            sell_tax_bits = []
            if parse_float(sell_tax.get("tax_relief"), 0.0) > 0:
                sell_tax_bits.append(f"tax relief {money(sell_tax.get('tax_relief'))}")
            if parse_float(sell_tax.get("tax_payable"), 0.0) > 0:
                sell_tax_bits.append(f"tax drag {money(sell_tax.get('tax_payable'))}")
            if parse_float(sell_tax.get("ltcg_exemption_used"), 0.0) > 0:
                sell_tax_bits.append(f"LTCG exemption used {money(sell_tax.get('ltcg_exemption_used'))}")
            if parse_float(sell_tax.get("ltcg_remaining_exemption_before"), 0.0) > 0:
                sell_tax_bits.append(f"LTCG exemption left {money(max(0.0, parse_float(sell_tax.get('ltcg_remaining_exemption_before'), 0.0) - parse_float(sell_tax.get('ltcg_exemption_used'), 0.0)))}")
            sell_tax_bits.append(f"cost {money(sell_costs.get('total'))}")
            buy_tax_bits = [
                f"entry cost {money(buy_entry_costs.get('total'))}",
                f"future STCG {money(target_exit_stats.get('future_tax_total'))}",
                f"net target {money(expected_profit_value)}",
            ]
            candidate = {
                "sell_symbol": sell["symbol"],
                "sell_qty": round(sell_qty, 4),
                "sell_ref_price": round(sell["ltp"], 4),
                "sell_trade_value": sell_value,
                "sell_target_price": round(sell["ltp"], 4),
                "sell_score": round(parse_float(sell.get("sell_score"), 0.0), 4),
                "sell_reason": f"{sell['reason']}, {sell_tax.get('tax_bucket_mix')} bucket, {'; '.join(sell_tax_bits)}",
                "buy_symbol": buy["symbol"],
                "buy_qty": round(buy_qty, 4),
                "buy_ref_price": round(buy["ltp"], 4),
                "buy_trade_value": buy_value,
                "buy_target_exit_price": round(target_exit_price, 4),
                "buy_score": round(parse_float(buy.get("buy_score"), 0.0), 4),
                "buy_reason": f"{buy['reason']}, {'; '.join(buy_tax_bits)}",
                "expected_profit_value": expected_profit_value,
                "rotation_score": rotation_score,
                "current_sell_ref_price": round(sell["ltp"], 4),
                "current_buy_ref_price": round(buy["ltp"], 4),
                "target_progress_pct": 0.0,
                "sell_exchange": str(sell.get("exchange") or "NSE").upper(),
                "buy_exchange": str(buy.get("exchange") or "NSE").upper(),
                "sell_cost_total": round(parse_float(sell_costs.get("total"), 0.0), 2),
                "buy_entry_cost_total": round(parse_float(buy_entry_costs.get("total"), 0.0), 2),
                "buy_exit_cost_total": round(parse_float(target_exit_stats.get("exit_cost_total"), 0.0), 2),
                "buy_exit_tax_total": round(parse_float(target_exit_stats.get("future_tax_total"), 0.0), 2),
                "sell_tax_payable": round(parse_float(sell_tax.get("tax_payable"), 0.0), 2),
                "sell_tax_relief": round(parse_float(sell_tax.get("tax_relief"), 0.0), 2),
                "sell_tax_drag": round(parse_float(sell_tax.get("tax_drag"), 0.0), 2),
                "sell_tax_bucket_mix": str(sell_tax.get("tax_bucket_mix") or "NA"),
                "net_target_profit_value": expected_profit_value,
                "required_buy_leg_net_profit": required_buy_leg_net,
                "tax_alpha_score": tax_alpha_score,
                "fy_label": str(realized_tax_summary.get("fy_label") or ""),
                "ltcg_remaining_exemption_before": round(parse_float(sell_tax.get("ltcg_remaining_exemption_before"), 0.0), 2),
                "ltcg_exemption_used": round(parse_float(sell_tax.get("ltcg_exemption_used"), 0.0), 2),
            }
            if best is None or candidate["rotation_score"] > best["rotation_score"]:
                best = candidate
        if not best:
            continue
        pairs.append(best)
        used_sell.add(best["sell_symbol"])
        used_buy.add(best["buy_symbol"])
        used_symbols.add(best["sell_symbol"])
        used_symbols.add(best["buy_symbol"])
        if len(pairs) >= limit:
            break

    for idx, pair in enumerate(pairs, start=1):
        pair["priority_rank"] = idx
        pair["target_progress_pct"] = _daily_target_pair_progress(
            pair["buy_ref_price"], pair["current_buy_ref_price"], pair["buy_target_exit_price"]
        )

    return {
        "summary": {
            "seed_capital": seed,
            "target_profit_pct": target_pct,
            "target_profit_value": round((seed * target_pct) / 100.0, 2),
            "top_n": limit,
            "sell_candidate_count": len(sell_candidates),
            "buy_candidate_count": len(buy_candidates),
            "generated_pairs_count": len(pairs),
            "universe_holdings_count": len(holdings),
            "scope": "existing_non_zero_holdings_only",
            "as_of": now_iso(),
            "tax_mode": "equity_special_rates",
            "investor_tax_bracket_pct": DAILY_TARGET_USER_TAX_BRACKET_PCT,
            "equity_stcg_tax_pct": round(parse_float(tax_cfg.get("stcg_rate_pct"), DAILY_TARGET_EQUITY_STCG_TAX_PCT), 4),
            "equity_ltcg_tax_pct": round(parse_float(tax_cfg.get("ltcg_rate_pct"), DAILY_TARGET_EQUITY_LTCG_TAX_PCT), 4),
            "equity_ltcg_exemption_limit": round(parse_float(tax_cfg.get("ltcg_exemption_limit"), DAILY_TARGET_LTCG_EXEMPTION_LIMIT), 2),
            "fy_label": str(realized_tax_summary.get("fy_label") or ""),
            "fy_start_date": str(realized_tax_summary.get("fy_start_date") or ""),
            "fy_end_date": str(realized_tax_summary.get("fy_end_date") or ""),
            "realized_ltcg_net_gain": round(parse_float(realized_tax_summary.get("ltcg_net_gain"), 0.0), 2),
            "remaining_ltcg_exemption": round(parse_float(realized_tax_summary.get("ltcg_remaining_exemption"), 0.0), 2),
            "zerodha_cost_model": "equity_delivery",
        },
        "pairs": pairs,
        "_holdings": holdings,
        "_sell_candidates": sell_candidates,
        "_buy_candidates": buy_candidates,
    }


def _insert_daily_target_pair_snapshot(conn, pair_id, snapshot_note="recalibrated"):
    row = conn.execute(
        """
        SELECT id, plan_id, sell_ref_price, buy_ref_price, expected_profit_value, rotation_score, buy_target_exit_price
        FROM daily_target_plan_pairs
        WHERE id = ?
        """,
        (int(parse_float(pair_id, 0.0)),),
    ).fetchone()
    if not row:
        return
    conn.execute(
        """
        INSERT INTO daily_target_pair_snapshots(
          plan_id, pair_id, captured_at, sell_ref_price, buy_ref_price, expected_profit_value, rotation_score,
          buy_target_exit_price, snapshot_note
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            int(parse_float(row["plan_id"], 0.0)),
            int(parse_float(row["id"], 0.0)),
            now_iso(),
            round(parse_float(row["sell_ref_price"], 0.0), 4),
            round(parse_float(row["buy_ref_price"], 0.0), 4),
            round(parse_float(row["expected_profit_value"], 0.0), 2),
            round(parse_float(row["rotation_score"], 0.0), 4),
            round(parse_float(row["buy_target_exit_price"], 0.0), 4),
            str(snapshot_note or "").strip() or None,
        ),
    )


def _insert_daily_target_pair(conn, plan_id, pair, state="pending"):
    ts = now_iso()
    conn.execute(
        """
        INSERT INTO daily_target_plan_pairs(
          plan_id, priority_rank, state, sell_symbol, sell_qty, sell_ref_price, sell_trade_value, sell_target_price,
          sell_score, sell_reason, buy_symbol, buy_qty, buy_ref_price, buy_trade_value, buy_target_exit_price,
          buy_score, buy_reason, expected_profit_value, rotation_score, current_sell_ref_price,
          current_buy_ref_price, target_progress_pct, matched_sell_trade_id, matched_buy_trade_id, reconciliation_status,
          created_at, updated_at, last_recalibrated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            int(parse_float(plan_id, 0.0)),
            int(parse_float(pair.get("priority_rank"), 0.0)),
            _normalize_daily_target_pair_state(state),
            symbol_upper(pair.get("sell_symbol")),
            round(parse_float(pair.get("sell_qty"), 0.0), 4),
            round(parse_float(pair.get("sell_ref_price"), 0.0), 4),
            round(parse_float(pair.get("sell_trade_value"), 0.0), 2),
            round(parse_float(pair.get("sell_target_price"), 0.0), 4),
            round(parse_float(pair.get("sell_score"), 0.0), 4),
            str(pair.get("sell_reason") or "").strip() or None,
            symbol_upper(pair.get("buy_symbol")),
            round(parse_float(pair.get("buy_qty"), 0.0), 4),
            round(parse_float(pair.get("buy_ref_price"), 0.0), 4),
            round(parse_float(pair.get("buy_trade_value"), 0.0), 2),
            round(parse_float(pair.get("buy_target_exit_price"), 0.0), 4),
            round(parse_float(pair.get("buy_score"), 0.0), 4),
            str(pair.get("buy_reason") or "").strip() or None,
            round(parse_float(pair.get("expected_profit_value"), 0.0), 2),
            round(parse_float(pair.get("rotation_score"), 0.0), 4),
            round(parse_float(pair.get("current_sell_ref_price"), pair.get("sell_ref_price")), 4),
            round(parse_float(pair.get("current_buy_ref_price"), pair.get("buy_ref_price")), 4),
            round(parse_float(pair.get("target_progress_pct"), 0.0), 2),
            None,
            None,
            "unmatched",
            ts,
            ts,
            ts,
        ),
    )
    pair_id = int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
    _insert_daily_target_pair_snapshot(conn, pair_id, snapshot_note="created")
    return pair_id


def _daily_target_plan_snapshots(conn, plan_id, limit=60):
    rows = conn.execute(
        """
        SELECT s.id, s.captured_at, s.sell_ref_price, s.buy_ref_price, s.expected_profit_value, s.rotation_score,
               s.buy_target_exit_price, s.snapshot_note, p.priority_rank, p.sell_symbol, p.buy_symbol, p.state
        FROM daily_target_pair_snapshots s
        JOIN daily_target_plan_pairs p ON p.id = s.pair_id
        WHERE s.plan_id = ?
        ORDER BY s.captured_at DESC, s.id DESC
        LIMIT ?
        """,
        (int(parse_float(plan_id, 0.0)), max(1, min(300, int(parse_float(limit, 60))))),
    ).fetchall()
    return [
        {
            "snapshot_id": int(parse_float(r["id"], 0.0)),
            "captured_at": str(r["captured_at"] or ""),
            "priority_rank": int(parse_float(r["priority_rank"], 0.0)),
            "sell_symbol": str(r["sell_symbol"] or ""),
            "buy_symbol": str(r["buy_symbol"] or ""),
            "state": str(r["state"] or ""),
            "sell_ref_price": round(parse_float(r["sell_ref_price"], 0.0), 4),
            "buy_ref_price": round(parse_float(r["buy_ref_price"], 0.0), 4),
            "expected_profit_value": round(parse_float(r["expected_profit_value"], 0.0), 2),
            "rotation_score": round(parse_float(r["rotation_score"], 0.0), 4),
            "buy_target_exit_price": round(parse_float(r["buy_target_exit_price"], 0.0), 4),
            "snapshot_note": str(r["snapshot_note"] or ""),
        }
        for r in rows
    ]


def _daily_target_plan_payload(conn, plan_row):
    if not plan_row:
        perf = compute_daily_target_performance(conn)
        tax_cfg = get_tax_profile_config(conn)
        realized_tax = compute_realized_equity_tax_summary(conn)
        return {
            "plan": None,
            "summary": {
                "seed_capital": 10000.0,
                "target_profit_pct": 1.0,
                "target_profit_value": 100.0,
                "top_n": 5,
                "pending_pairs": 0,
                "executed_pairs": 0,
                "sell_done_pairs": 0,
                "buy_done_pairs": 0,
                "skipped_pairs": 0,
                "replaced_pairs": 0,
                "projected_pending_profit": 0.0,
                "suggested_next_seed_capital": round(parse_float(perf.get("suggested_next_seed_capital"), 10000.0), 2),
                "tax_mode": "equity_special_rates",
                "investor_tax_bracket_pct": DAILY_TARGET_USER_TAX_BRACKET_PCT,
                "equity_stcg_tax_pct": round(parse_float(tax_cfg.get("stcg_rate_pct"), DAILY_TARGET_EQUITY_STCG_TAX_PCT), 4),
                "equity_ltcg_tax_pct": round(parse_float(tax_cfg.get("ltcg_rate_pct"), DAILY_TARGET_EQUITY_LTCG_TAX_PCT), 4),
                "equity_ltcg_exemption_limit": round(parse_float(tax_cfg.get("ltcg_exemption_limit"), DAILY_TARGET_LTCG_EXEMPTION_LIMIT), 2),
                "fy_label": str(realized_tax.get("fy_label") or ""),
                "fy_start_date": str(realized_tax.get("fy_start_date") or ""),
                "fy_end_date": str(realized_tax.get("fy_end_date") or ""),
                "realized_ltcg_net_gain": round(parse_float(realized_tax.get("ltcg_net_gain"), 0.0), 2),
                "remaining_ltcg_exemption": round(parse_float(realized_tax.get("ltcg_remaining_exemption"), 0.0), 2),
                "zerodha_cost_model": "equity_delivery",
            },
            "pairs": [],
            "snapshots": [],
            "performance": perf,
        }
    plan_id = int(parse_float(plan_row.get("id"), 0.0))
    reconcile_daily_target_trade_links(conn, plan_id=plan_id)
    pair_rows = conn.execute(
        """
        SELECT *
        FROM daily_target_plan_pairs
        WHERE plan_id = ?
        ORDER BY priority_rank ASC, id ASC
        """,
        (plan_id,),
    ).fetchall()
    pairs = []
    counts = defaultdict(int)
    projected_pending_profit = 0.0
    for r in pair_rows:
        state = _normalize_daily_target_pair_state(r["state"], default="pending")
        counts[state] += 1
        if state == "pending":
            projected_pending_profit += parse_float(r["expected_profit_value"], 0.0)
        live = _daily_target_live_pair_metrics(conn, r, target_profit_pct=plan_row.get("target_profit_pct"))
        pairs.append(
            {
                "pair_id": int(parse_float(r["id"], 0.0)),
                "priority_rank": int(parse_float(r["priority_rank"], 0.0)),
                "state": state,
                "sell_symbol": str(r["sell_symbol"] or ""),
                "sell_qty": round(parse_float(r["sell_qty"], 0.0), 4),
                "sell_ref_price": round(parse_float(r["sell_ref_price"], 0.0), 4),
                "sell_trade_value": round(parse_float(r["sell_trade_value"], 0.0), 2),
                "sell_target_price": round(parse_float(r["sell_target_price"], 0.0), 4),
                "sell_score": round(parse_float(r["sell_score"], 0.0), 4),
                "sell_reason": str(r["sell_reason"] or ""),
                "buy_symbol": str(r["buy_symbol"] or ""),
                "buy_qty": round(parse_float(r["buy_qty"], 0.0), 4),
                "buy_ref_price": round(parse_float(r["buy_ref_price"], 0.0), 4),
                "buy_trade_value": round(parse_float(r["buy_trade_value"], 0.0), 2),
                "buy_target_exit_price": round(parse_float(r["buy_target_exit_price"], 0.0), 4),
                "buy_score": round(parse_float(r["buy_score"], 0.0), 4),
                "buy_reason": str(r["buy_reason"] or ""),
                "expected_profit_value": round(parse_float(r["expected_profit_value"], 0.0), 2),
                "rotation_score": round(parse_float(r["rotation_score"], 0.0), 4),
                "current_sell_ref_price": round(parse_float(live.get("current_sell_ref_price"), r["current_sell_ref_price"]), 4),
                "current_buy_ref_price": round(parse_float(live.get("current_buy_ref_price"), r["current_buy_ref_price"]), 4),
                "target_progress_pct": round(parse_float(live.get("target_progress_pct"), r["target_progress_pct"]), 2),
                "matched_sell_trade_id": int(parse_float(r["matched_sell_trade_id"], 0.0)) or None,
                "matched_buy_trade_id": int(parse_float(r["matched_buy_trade_id"], 0.0)) or None,
                "reconciliation_status": str(r["reconciliation_status"] or "unmatched"),
                "executed_sell_price": (
                    None if r["executed_sell_price"] is None else round(parse_float(r["executed_sell_price"], 0.0), 4)
                ),
                "executed_sell_at": str(r["executed_sell_at"] or ""),
                "executed_sell_value": round(parse_float(r["executed_sell_price"], 0.0) * parse_float(r["sell_qty"], 0.0), 2) if parse_float(r["executed_sell_price"], 0.0) > 0 else None,
                "executed_buy_price": (
                    None if r["executed_buy_price"] is None else round(parse_float(r["executed_buy_price"], 0.0), 4)
                ),
                "executed_buy_at": str(r["executed_buy_at"] or ""),
                "executed_buy_value": round(parse_float(r["executed_buy_price"], 0.0) * parse_float(r["buy_qty"], 0.0), 2) if parse_float(r["executed_buy_price"], 0.0) > 0 else None,
                "completion_note": str(r["completion_note"] or ""),
                "created_at": str(r["created_at"] or ""),
                "updated_at": str(r["updated_at"] or ""),
                "last_recalibrated_at": str(r["last_recalibrated_at"] or ""),
                "buy_entry_price": round(parse_float(live.get("buy_entry_price"), 0.0), 4),
                "buy_target_exit_price": round(parse_float(live.get("buy_target_exit_price"), r["buy_target_exit_price"]), 4),
            }
        )
    summary = {
        "seed_capital": round(parse_float(plan_row.get("seed_capital"), 0.0), 2),
        "target_profit_pct": round(parse_float(plan_row.get("target_profit_pct"), 0.0), 2),
        "target_profit_value": round(parse_float(plan_row.get("target_profit_value"), 0.0), 2),
        "top_n": int(parse_float(plan_row.get("top_n"), 0.0)),
        "pending_pairs": counts.get("pending", 0),
        "sell_done_pairs": counts.get("sell_done", 0),
        "buy_done_pairs": counts.get("buy_done", 0),
        "executed_pairs": counts.get("executed", 0),
        "skipped_pairs": counts.get("skipped", 0),
        "replaced_pairs": counts.get("replaced", 0),
        "projected_pending_profit": round(projected_pending_profit, 2),
        "created_at": str(plan_row.get("created_at") or ""),
        "updated_at": str(plan_row.get("updated_at") or ""),
        "last_recalibrated_at": str(plan_row.get("last_recalibrated_at") or ""),
        "status": str(plan_row.get("status") or ""),
        "tax_mode": "equity_special_rates",
        "investor_tax_bracket_pct": DAILY_TARGET_USER_TAX_BRACKET_PCT,
        "equity_stcg_tax_pct": round(parse_float(get_tax_profile_config(conn).get("stcg_rate_pct"), DAILY_TARGET_EQUITY_STCG_TAX_PCT), 4),
        "equity_ltcg_tax_pct": round(parse_float(get_tax_profile_config(conn).get("ltcg_rate_pct"), DAILY_TARGET_EQUITY_LTCG_TAX_PCT), 4),
        "equity_ltcg_exemption_limit": round(parse_float(get_tax_profile_config(conn).get("ltcg_exemption_limit"), DAILY_TARGET_LTCG_EXEMPTION_LIMIT), 2),
        "zerodha_cost_model": "equity_delivery",
    }
    perf = compute_daily_target_performance(conn)
    realized_tax = compute_realized_equity_tax_summary(conn)
    summary["fy_label"] = str(realized_tax.get("fy_label") or "")
    summary["fy_start_date"] = str(realized_tax.get("fy_start_date") or "")
    summary["fy_end_date"] = str(realized_tax.get("fy_end_date") or "")
    summary["realized_ltcg_net_gain"] = round(parse_float(realized_tax.get("ltcg_net_gain"), 0.0), 2)
    summary["remaining_ltcg_exemption"] = round(parse_float(realized_tax.get("ltcg_remaining_exemption"), 0.0), 2)
    summary["suggested_next_seed_capital"] = round(parse_float(perf.get("suggested_next_seed_capital"), summary["seed_capital"]), 2)
    return {
        "plan": {
            "id": plan_id,
            "seed_capital": summary["seed_capital"],
            "target_profit_pct": summary["target_profit_pct"],
            "target_profit_value": summary["target_profit_value"],
            "top_n": summary["top_n"],
            "status": summary["status"],
            "created_at": summary["created_at"],
            "updated_at": summary["updated_at"],
            "last_recalibrated_at": summary["last_recalibrated_at"],
            "closed_at": str(plan_row.get("closed_at") or ""),
            "notes": str(plan_row.get("notes") or ""),
        },
        "summary": summary,
        "pairs": pairs,
        "snapshots": _daily_target_plan_snapshots(conn, plan_id, limit=80),
        "performance": perf,
    }


def _recalibrate_daily_target_plan(conn, plan_row, seed_capital=None, target_profit_pct=None, top_n=None):
    plan_id = int(parse_float(plan_row.get("id"), 0.0))
    perf = compute_daily_target_performance(conn)
    tax_cfg = get_tax_profile_config(conn)
    realized_tax_summary = compute_realized_equity_tax_summary(conn)
    target_pct_effective = round(
        parse_float((target_profit_pct if target_profit_pct is not None else plan_row.get("target_profit_pct")), 1.0),
        2,
    )
    effective_seed = round(
        parse_float(
            perf.get("suggested_next_seed_capital"),
            (seed_capital if seed_capital is not None else plan_row.get("seed_capital")),
        ),
        2,
    )
    suggestions = build_daily_target_suggestions(
        conn,
        seed_capital=effective_seed,
        target_profit_pct=target_pct_effective,
        top_n=(top_n if top_n is not None else plan_row.get("top_n")),
    )
    pairs = list(suggestions.get("pairs") or [])
    buy_candidates = [dict(x) for x in (suggestions.get("_buy_candidates") or [])]
    buy_candidate_map = {symbol_upper(x.get("symbol")): dict(x) for x in buy_candidates if symbol_upper(x.get("symbol"))}
    now_ts = now_iso()
    active_rows = [
        dict(r)
        for r in conn.execute(
            """
            SELECT *
            FROM daily_target_plan_pairs
            WHERE plan_id = ? AND LOWER(state) <> 'replaced'
            ORDER BY priority_rank, id
            """,
            (plan_id,),
        ).fetchall()
    ]
    pending_rows = {
        int(parse_float(r.get("priority_rank"), 0.0)): dict(r)
        for r in active_rows
        if str(r.get("state") or "").strip().lower() == "pending"
    }
    locked_rows = [r for r in active_rows if str(r.get("state") or "").strip().lower() in ("sell_done", "buy_done", "executed")]
    reserved_symbols = set()
    for row in locked_rows:
        sell_symbol = symbol_upper(row.get("sell_symbol"))
        buy_symbol = symbol_upper(row.get("buy_symbol"))
        state = str(row.get("state") or "").strip().lower()
        if sell_symbol:
            reserved_symbols.add(sell_symbol)
        if buy_symbol and state in ("buy_done", "executed"):
            reserved_symbols.add(buy_symbol)
    recalibration_switches = 0
    pipeline_reviewed = 0
    split_map = load_split_map(conn)
    for row in locked_rows:
        state = str(row.get("state") or "").strip().lower()
        row_id = int(parse_float(row.get("id"), 0.0))
        sell_symbol = symbol_upper(row.get("sell_symbol"))
        live = _daily_target_live_pair_metrics(conn, row, target_profit_pct=target_pct_effective)
        if state in ("buy_done", "executed"):
            pipeline_reviewed += 1
            conn.execute(
                """
                UPDATE daily_target_plan_pairs
                SET current_sell_ref_price = ?, current_buy_ref_price = ?, buy_target_exit_price = ?, target_progress_pct = ?,
                    updated_at = ?, last_recalibrated_at = ?
                WHERE id = ?
                """,
                (
                    round(parse_float(live.get("current_sell_ref_price"), row.get("current_sell_ref_price")), 4),
                    round(parse_float(live.get("current_buy_ref_price"), row.get("current_buy_ref_price")), 4),
                    round(parse_float(live.get("buy_target_exit_price"), row.get("buy_target_exit_price")), 4),
                    round(parse_float(live.get("target_progress_pct"), row.get("target_progress_pct")), 2),
                    now_ts,
                    now_ts,
                    row_id,
                ),
            )
            continue
        if state != "sell_done":
            continue
        pipeline_reviewed += 1
        sell_exec_price = parse_float(row.get("executed_sell_price"), 0.0)
        if sell_exec_price <= 0:
            sell_exec_price = parse_float(row.get("current_sell_ref_price"), row.get("sell_ref_price"))
        available_capital = round(sell_exec_price * parse_float(row.get("sell_qty"), 0.0), 2)
        sell_exchange = str(row.get("sell_exchange") or "NSE").upper()
        sell_costs = _daily_target_zerodha_delivery_costs(
            available_capital,
            exchange=sell_exchange,
            side="SELL",
            include_dp_on_sell=True,
            tax_cfg=tax_cfg,
        )
        sell_tax = _daily_target_estimate_sell_tax_profile(
            conn,
            sell_symbol,
            parse_float(row.get("sell_qty"), 0.0),
            sell_exec_price,
            split_map=split_map,
            as_of_date=(row.get("executed_sell_at") or now_local_date_iso()),
            tax_cfg=tax_cfg,
            realized_tax_summary=realized_tax_summary,
        )
        available_capital = max(0.0, round(available_capital - parse_float(sell_costs.get("total"), 0.0), 2))
        required_buy_leg_net = round(
            max(
                0.0,
                ((available_capital * target_pct_effective) / 100.0)
                + parse_float(sell_costs.get("total"), 0.0)
                + parse_float(sell_tax.get("tax_drag"), 0.0),
            ),
            2,
        )
        blocked_for_row = set(reserved_symbols)
        current_buy_symbol = symbol_upper(row.get("buy_symbol"))
        if current_buy_symbol and current_buy_symbol in blocked_for_row:
            blocked_for_row.remove(current_buy_symbol)
        best_buy_leg = None
        for cand in buy_candidates:
            cand_symbol = symbol_upper(cand.get("symbol"))
            if not cand_symbol or cand_symbol == sell_symbol or cand_symbol in blocked_for_row:
                continue
            leg = _daily_target_build_buy_leg(
                cand,
                available_capital,
                target_profit_pct=target_pct_effective,
                required_net_profit=required_buy_leg_net,
                tax_cfg=tax_cfg,
            )
            if not leg:
                continue
            if best_buy_leg is None or parse_float(leg.get("buy_score"), 0.0) > parse_float(best_buy_leg.get("buy_score"), 0.0):
                best_buy_leg = leg
        if best_buy_leg and _daily_target_should_switch_pipeline_buy(row, best_buy_leg, buy_candidate_map):
            _insert_daily_target_pair_snapshot(conn, row_id, snapshot_note="pipeline_buy_switch_before_update")
            replacement_live = _daily_target_live_pair_metrics(
                conn,
                {
                    **row,
                    **best_buy_leg,
                    "buy_ref_price": best_buy_leg.get("buy_ref_price"),
                    "buy_target_exit_price": best_buy_leg.get("buy_target_exit_price"),
                    "executed_buy_price": None,
                },
                target_profit_pct=target_pct_effective,
            )
            conn.execute(
                """
                UPDATE daily_target_plan_pairs
                SET buy_symbol = ?, buy_qty = ?, buy_ref_price = ?, buy_trade_value = ?, buy_target_exit_price = ?,
                    buy_score = ?, buy_reason = ?, expected_profit_value = ?, rotation_score = ?,
                    current_sell_ref_price = ?, current_buy_ref_price = ?, target_progress_pct = ?, updated_at = ?, last_recalibrated_at = ?
                WHERE id = ?
                """,
                (
                    symbol_upper(best_buy_leg.get("buy_symbol")),
                    round(parse_float(best_buy_leg.get("buy_qty"), 0.0), 4),
                    round(parse_float(best_buy_leg.get("buy_ref_price"), 0.0), 4),
                    round(parse_float(best_buy_leg.get("buy_trade_value"), 0.0), 2),
                    round(parse_float(best_buy_leg.get("buy_target_exit_price"), 0.0), 4),
                    round(parse_float(best_buy_leg.get("buy_score"), 0.0), 4),
                    str(best_buy_leg.get("buy_reason") or "").strip() or None,
                    round(parse_float(best_buy_leg.get("expected_profit_value"), 0.0), 2),
                    round(
                        parse_float(row.get("sell_score"), 0.0)
                        + parse_float(best_buy_leg.get("buy_score"), 0.0)
                        + (
                            (
                                parse_float(sell_tax.get("tax_relief"), 0.0)
                                - parse_float(sell_tax.get("tax_payable"), 0.0)
                                - parse_float(sell_costs.get("total"), 0.0)
                                - parse_float(best_buy_leg.get("buy_entry_cost_total"), 0.0)
                            )
                            / max(effective_seed, 1.0)
                            * 500.0
                        ),
                        4,
                    ),
                    round(parse_float(live.get("current_sell_ref_price"), row.get("current_sell_ref_price")), 4),
                    round(parse_float(replacement_live.get("current_buy_ref_price"), best_buy_leg.get("buy_ref_price")), 4),
                    round(parse_float(replacement_live.get("target_progress_pct"), 0.0), 2),
                    now_ts,
                    now_ts,
                    row_id,
                ),
            )
            _insert_daily_target_pair_snapshot(conn, row_id, snapshot_note="pipeline_buy_switch_after_update")
            recalibration_switches += 1
            reserved_symbols.add(symbol_upper(best_buy_leg.get("buy_symbol")))
        else:
            reserved_symbols.add(current_buy_symbol)
            conn.execute(
                """
                UPDATE daily_target_plan_pairs
                SET current_sell_ref_price = ?, current_buy_ref_price = ?, buy_target_exit_price = ?, target_progress_pct = ?,
                    updated_at = ?, last_recalibrated_at = ?
                WHERE id = ?
                """,
                (
                    round(parse_float(live.get("current_sell_ref_price"), row.get("current_sell_ref_price")), 4),
                    round(parse_float(live.get("current_buy_ref_price"), row.get("current_buy_ref_price")), 4),
                    round(parse_float(live.get("buy_target_exit_price"), row.get("buy_target_exit_price")), 4),
                    round(parse_float(live.get("target_progress_pct"), row.get("target_progress_pct")), 2),
                    now_ts,
                    now_ts,
                    row_id,
                ),
            )
    handled_ranks = set()
    filtered_pairs = []
    temp_reserved = set(reserved_symbols)
    for pair in pairs:
        sell_symbol = symbol_upper(pair.get("sell_symbol"))
        buy_symbol = symbol_upper(pair.get("buy_symbol"))
        if not sell_symbol or not buy_symbol:
            continue
        if sell_symbol in temp_reserved or buy_symbol in temp_reserved:
            continue
        filtered_pairs.append(pair)
        temp_reserved.add(sell_symbol)
        temp_reserved.add(buy_symbol)
    for pair in filtered_pairs:
        rank = int(parse_float(pair.get("priority_rank"), 0.0))
        handled_ranks.add(rank)
        existing = pending_rows.get(rank)
        if existing and symbol_upper(existing.get("sell_symbol")) == symbol_upper(pair.get("sell_symbol")) and symbol_upper(existing.get("buy_symbol")) == symbol_upper(pair.get("buy_symbol")):
            _insert_daily_target_pair_snapshot(conn, existing.get("id"), snapshot_note="recalibrated_before_update")
            conn.execute(
                """
                UPDATE daily_target_plan_pairs
                SET sell_qty = ?, sell_ref_price = ?, sell_trade_value = ?, sell_target_price = ?, sell_score = ?,
                    sell_reason = ?, buy_qty = ?, buy_ref_price = ?, buy_trade_value = ?, buy_target_exit_price = ?,
                    buy_score = ?, buy_reason = ?, expected_profit_value = ?, rotation_score = ?,
                    current_sell_ref_price = ?, current_buy_ref_price = ?, target_progress_pct = ?, updated_at = ?,
                    last_recalibrated_at = ?
                WHERE id = ?
                """,
                (
                    round(parse_float(pair.get("sell_qty"), 0.0), 4),
                    round(parse_float(pair.get("sell_ref_price"), 0.0), 4),
                    round(parse_float(pair.get("sell_trade_value"), 0.0), 2),
                    round(parse_float(pair.get("sell_target_price"), 0.0), 4),
                    round(parse_float(pair.get("sell_score"), 0.0), 4),
                    str(pair.get("sell_reason") or "").strip() or None,
                    round(parse_float(pair.get("buy_qty"), 0.0), 4),
                    round(parse_float(pair.get("buy_ref_price"), 0.0), 4),
                    round(parse_float(pair.get("buy_trade_value"), 0.0), 2),
                    round(parse_float(pair.get("buy_target_exit_price"), 0.0), 4),
                    round(parse_float(pair.get("buy_score"), 0.0), 4),
                    str(pair.get("buy_reason") or "").strip() or None,
                    round(parse_float(pair.get("expected_profit_value"), 0.0), 2),
                    round(parse_float(pair.get("rotation_score"), 0.0), 4),
                    round(parse_float(pair.get("current_sell_ref_price"), pair.get("sell_ref_price")), 4),
                    round(parse_float(pair.get("current_buy_ref_price"), pair.get("buy_ref_price")), 4),
                    round(
                        _daily_target_pair_progress(
                            pair.get("buy_ref_price"), pair.get("current_buy_ref_price"), pair.get("buy_target_exit_price")
                        ),
                        2,
                    ),
                    now_ts,
                    now_ts,
                    int(parse_float(existing.get("id"), 0.0)),
                ),
            )
            _insert_daily_target_pair_snapshot(conn, existing.get("id"), snapshot_note="recalibrated_after_update")
            reserved_symbols.add(symbol_upper(pair.get("sell_symbol")))
            reserved_symbols.add(symbol_upper(pair.get("buy_symbol")))
            continue
        if existing:
            _insert_daily_target_pair_snapshot(conn, existing.get("id"), snapshot_note="replaced_due_to_market_shift")
            conn.execute(
                "UPDATE daily_target_plan_pairs SET state='replaced', updated_at=?, last_recalibrated_at=? WHERE id = ?",
                (now_ts, now_ts, int(parse_float(existing.get("id"), 0.0))),
            )
            recalibration_switches += 1
        _insert_daily_target_pair(conn, plan_id, pair, state="pending")
        reserved_symbols.add(symbol_upper(pair.get("sell_symbol")))
        reserved_symbols.add(symbol_upper(pair.get("buy_symbol")))
    for rank, existing in pending_rows.items():
        if rank in handled_ranks:
            continue
        _insert_daily_target_pair_snapshot(conn, existing.get("id"), snapshot_note="retired_on_recalibration")
        conn.execute(
            "UPDATE daily_target_plan_pairs SET state='replaced', updated_at=?, last_recalibrated_at=? WHERE id = ?",
            (now_ts, now_ts, int(parse_float(existing.get("id"), 0.0))),
        )
        recalibration_switches += 1
    summary = suggestions.get("summary") or {}
    summary["pipeline_rows_reviewed"] = pipeline_reviewed
    summary["pipeline_switches"] = recalibration_switches
    summary["active_live_price_recalibrated_at"] = now_ts
    summary["generated_pairs_count"] = len(filtered_pairs)
    conn.execute(
        """
        UPDATE daily_target_plans
        SET seed_capital = ?, target_profit_pct = ?, target_profit_value = ?, top_n = ?, updated_at = ?, last_recalibrated_at = ?
        WHERE id = ?
        """,
        (
            round(parse_float(summary.get("seed_capital"), 0.0), 2),
            round(parse_float(summary.get("target_profit_pct"), 0.0), 2),
            round(parse_float(summary.get("target_profit_value"), 0.0), 2),
            int(parse_float(summary.get("top_n"), 0.0)),
            now_ts,
            now_ts,
            plan_id,
        ),
    )
    updated = conn.execute(
        """
        SELECT id, seed_capital, target_profit_pct, target_profit_value, top_n, status, created_at, updated_at,
               last_recalibrated_at, closed_at, notes
        FROM daily_target_plans
        WHERE id = ?
        """,
        (plan_id,),
    ).fetchone()
    payload = _daily_target_plan_payload(conn, dict(updated) if updated else None)
    payload["suggestion_meta"] = summary
    return payload


def get_or_create_daily_target_plan(conn, seed_capital=10000.0, target_profit_pct=1.0, top_n=5, recalibrate=True):
    active = get_active_daily_target_plan(conn)
    if active:
        if recalibrate:
            return _recalibrate_daily_target_plan(conn, active, seed_capital=seed_capital, target_profit_pct=target_profit_pct, top_n=top_n)
        payload = _daily_target_plan_payload(conn, active)
        payload["suggestion_meta"] = {
            "seed_capital": round(parse_float(active.get("seed_capital"), 0.0), 2),
            "target_profit_pct": round(parse_float(active.get("target_profit_pct"), 0.0), 2),
            "target_profit_value": round(parse_float(active.get("target_profit_value"), 0.0), 2),
            "top_n": int(parse_float(active.get("top_n"), 0.0)),
        }
        return payload
    perf = compute_daily_target_performance(conn)
    next_seed = round(parse_float(perf.get("suggested_next_seed_capital"), seed_capital), 2)
    effective_seed = next_seed if next_seed > 0 else round(parse_float(seed_capital, 10000.0), 2)
    suggestions = build_daily_target_suggestions(conn, seed_capital=effective_seed, target_profit_pct=target_profit_pct, top_n=top_n)
    summary = suggestions.get("summary") or {}
    ts = now_iso()
    cur = conn.execute(
        """
        INSERT INTO daily_target_plans(
          seed_capital, target_profit_pct, target_profit_value, top_n, status, created_at, updated_at, last_recalibrated_at, notes
        ) VALUES (?, ?, ?, ?, 'active', ?, ?, ?, ?)
        """,
        (
            round(parse_float(summary.get("seed_capital"), 0.0), 2),
            round(parse_float(summary.get("target_profit_pct"), 0.0), 2),
            round(parse_float(summary.get("target_profit_value"), 0.0), 2),
            int(parse_float(summary.get("top_n"), 0.0)),
            ts,
            ts,
            ts,
            "Existing non-zero holdings only; ranked for a 1-day rotation objective.",
        ),
    )
    plan_id = int(cur.lastrowid)
    for pair in suggestions.get("pairs") or []:
        _insert_daily_target_pair(conn, plan_id, pair, state="pending")
    row = conn.execute(
        """
        SELECT id, seed_capital, target_profit_pct, target_profit_value, top_n, status, created_at, updated_at,
               last_recalibrated_at, closed_at, notes
        FROM daily_target_plans
        WHERE id = ?
        """,
        (plan_id,),
    ).fetchone()
    payload = _daily_target_plan_payload(conn, dict(row) if row else None)
    payload["suggestion_meta"] = summary
    return payload


def reset_daily_target_plan(conn):
    active = get_active_daily_target_plan(conn)
    if not active:
        raise ValueError("no_active_daily_target_plan")
    plan_id = int(parse_float(active.get("id"), 0.0))
    ts = now_iso()
    conn.execute(
        "UPDATE daily_target_plans SET status='reset', updated_at=?, closed_at=? WHERE id = ? AND LOWER(status)='active'",
        (ts, ts, plan_id),
    )
    return {"ok": True, "plan_id": plan_id, "status": "reset", "closed_at": ts}


def update_daily_target_pair(conn, pair_id, state="pending", note="", executed_sell_price=None, executed_sell_at=None, executed_buy_price=None, executed_buy_at=None):
    iid = int(parse_float(pair_id, 0.0))
    row = conn.execute(
        """
        SELECT p.id, p.plan_id, p.state, p.executed_sell_price, p.executed_sell_at, p.executed_buy_price, p.executed_buy_at,
               dp.status AS plan_status
        FROM daily_target_plan_pairs p
        JOIN daily_target_plans dp ON dp.id = p.plan_id
        WHERE p.id = ?
        """,
        (iid,),
    ).fetchone()
    if not row:
        raise ValueError("daily_target_pair_not_found")
    if str(row["plan_status"] or "").lower() != "active":
        raise ValueError("daily_target_plan_not_active")
    state_norm = _normalize_daily_target_pair_state(state, default="pending")
    if state_norm == "replaced":
        raise ValueError("daily_target_pair_state_not_user_editable")
    sell_px = None if row["executed_sell_price"] is None else parse_float(row["executed_sell_price"], 0.0)
    buy_px = None if row["executed_buy_price"] is None else parse_float(row["executed_buy_price"], 0.0)
    sell_ts = str(row["executed_sell_at"] or "") or None
    buy_ts = str(row["executed_buy_at"] or "") or None
    if state_norm == "pending":
        sell_px = None
        sell_ts = None
        buy_px = None
        buy_ts = None
    else:
        if executed_sell_price is not None and str(executed_sell_price).strip() != "":
            sell_px = parse_float(executed_sell_price, 0.0)
            if not math.isfinite(sell_px) or sell_px <= 0:
                raise ValueError("executed_sell_price_must_be_positive")
            sell_ts = _normalize_execution_timestamp(executed_sell_at, fallback_now=True)
        elif executed_sell_at is not None and str(executed_sell_at).strip() != "":
            if sell_px is None or sell_px <= 0:
                raise ValueError("executed_sell_price_required")
            sell_ts = _normalize_execution_timestamp(executed_sell_at, fallback_now=True)

        if executed_buy_price is not None and str(executed_buy_price).strip() != "":
            buy_px = parse_float(executed_buy_price, 0.0)
            if not math.isfinite(buy_px) or buy_px <= 0:
                raise ValueError("executed_buy_price_must_be_positive")
            buy_ts = _normalize_execution_timestamp(executed_buy_at, fallback_now=True)
        elif executed_buy_at is not None and str(executed_buy_at).strip() != "":
            if buy_px is None or buy_px <= 0:
                raise ValueError("executed_buy_price_required")
            buy_ts = _normalize_execution_timestamp(executed_buy_at, fallback_now=True)

    sell_present = sell_px is not None and math.isfinite(sell_px) and sell_px > 0
    buy_present = buy_px is not None and math.isfinite(buy_px) and buy_px > 0
    effective_state = state_norm
    if effective_state in ("sell_done", "buy_done", "executed") and not sell_present:
        raise ValueError("executed_sell_price_required")
    if effective_state in ("buy_done", "executed") and not buy_present:
        raise ValueError("executed_buy_price_required")
    if effective_state in ("sell_done", "buy_done") and sell_present and buy_present:
        effective_state = "executed"
    note_text = str(note or "").strip() or None
    conn.execute(
        """
        UPDATE daily_target_plan_pairs
        SET state = ?, executed_sell_price = ?, executed_sell_at = ?, executed_buy_price = ?, executed_buy_at = ?,
            completion_note = ?, updated_at = ?
        WHERE id = ?
        """,
        (effective_state, sell_px, sell_ts, buy_px, buy_ts, note_text, now_iso(), iid),
    )
    if effective_state == "executed":
        sync_daily_target_positions(conn, iid)
    reconcile_daily_target_trade_links(conn, plan_id=int(parse_float(row["plan_id"], 0.0)))
    active = get_active_daily_target_plan(conn)
    payload = _daily_target_plan_payload(conn, active)
    payload["updated_pair_id"] = iid
    return payload


def list_daily_target_plan_history(conn, limit=120, date_from=None, date_to=None, state_filter=None):
    reconcile_daily_target_trade_links(conn)
    lim = max(1, min(500, int(parse_float(limit, 120))))
    df = str(date_from or "").strip()
    dt_to = str(date_to or "").strip()
    sf = str(state_filter or "all").strip().lower()
    where = []
    params = []
    if df:
        where.append("DATE(COALESCE(NULLIF(p.executed_buy_at, ''), NULLIF(p.executed_sell_at, ''), p.updated_at)) >= DATE(?)")
        params.append(df[:10])
    if dt_to:
        where.append("DATE(COALESCE(NULLIF(p.executed_buy_at, ''), NULLIF(p.executed_sell_at, ''), p.updated_at)) <= DATE(?)")
        params.append(dt_to[:10])
    if sf and sf != "all":
        if sf == "active":
            where.append("LOWER(COALESCE(p.state, 'pending')) IN ('pending', 'sell_done', 'buy_done')")
        elif sf == "closed":
            where.append("LOWER(COALESCE(p.state, 'pending')) IN ('executed', 'skipped', 'replaced')")
        else:
            where.append("LOWER(COALESCE(p.state, 'pending')) = ?")
            params.append(sf)
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    rows = conn.execute(
        f"""
        SELECT p.id AS pair_id, p.plan_id, p.priority_rank, p.state, p.sell_symbol, p.sell_qty, p.sell_ref_price,
               p.sell_trade_value, p.buy_symbol, p.buy_qty, p.buy_ref_price, p.buy_trade_value, p.buy_target_exit_price,
               p.expected_profit_value, p.rotation_score, p.matched_sell_trade_id, p.matched_buy_trade_id,
               p.reconciliation_status, p.executed_sell_price, p.executed_sell_at, p.executed_buy_price,
               p.executed_buy_at, p.completion_note, p.updated_at, dp.seed_capital, dp.target_profit_pct, dp.status AS plan_status,
               dp.created_at AS plan_created_at, dp.closed_at AS plan_closed_at
        FROM daily_target_plan_pairs p
        JOIN daily_target_plans dp ON dp.id = p.plan_id
        {where_sql}
        ORDER BY p.updated_at DESC, p.id DESC
        LIMIT ?
        """,
        tuple(params + [lim]),
    ).fetchall()
    total_count = int(
        conn.execute(
            f"""
            SELECT COUNT(*) AS c
            FROM daily_target_plan_pairs p
            JOIN daily_target_plans dp ON dp.id = p.plan_id
            {where_sql}
            """,
            tuple(params),
        ).fetchone()["c"]
    )
    items = []
    for r in rows:
        current_buy_ltp = get_effective_ltp(conn, r["buy_symbol"]) if parse_float(r["buy_qty"], 0.0) > 0 else 0.0
        current_buy_value = round(parse_float(r["buy_qty"], 0.0) * current_buy_ltp, 2) if current_buy_ltp > 0 else 0.0
        buy_cost = round(parse_float(r["executed_buy_price"], 0.0) * parse_float(r["buy_qty"], 0.0), 2) if parse_float(r["executed_buy_price"], 0.0) > 0 else round(parse_float(r["buy_trade_value"], 0.0), 2)
        live_mtm_pnl = round(current_buy_value - buy_cost, 2) if current_buy_value > 0 and buy_cost > 0 else 0.0
        items.append(
            {
                "pair_id": int(parse_float(r["pair_id"], 0.0)),
                "plan_id": int(parse_float(r["plan_id"], 0.0)),
                "priority_rank": int(parse_float(r["priority_rank"], 0.0)),
                "state": str(r["state"] or ""),
                "sell_symbol": str(r["sell_symbol"] or ""),
                "sell_qty": round(parse_float(r["sell_qty"], 0.0), 4),
                "sell_ref_price": round(parse_float(r["sell_ref_price"], 0.0), 4),
                "sell_trade_value": round(parse_float(r["sell_trade_value"], 0.0), 2),
                "buy_symbol": str(r["buy_symbol"] or ""),
                "buy_qty": round(parse_float(r["buy_qty"], 0.0), 4),
                "buy_ref_price": round(parse_float(r["buy_ref_price"], 0.0), 4),
                "buy_trade_value": round(parse_float(r["buy_trade_value"], 0.0), 2),
                "buy_target_exit_price": round(parse_float(r["buy_target_exit_price"], 0.0), 4),
                "expected_profit_value": round(parse_float(r["expected_profit_value"], 0.0), 2),
                "rotation_score": round(parse_float(r["rotation_score"], 0.0), 4),
                "matched_sell_trade_id": int(parse_float(r["matched_sell_trade_id"], 0.0)) or None,
                "matched_buy_trade_id": int(parse_float(r["matched_buy_trade_id"], 0.0)) or None,
                "reconciliation_status": str(r["reconciliation_status"] or "unmatched"),
                "executed_sell_price": (None if r["executed_sell_price"] is None else round(parse_float(r["executed_sell_price"], 0.0), 4)),
                "executed_sell_at": str(r["executed_sell_at"] or ""),
                "executed_sell_value": round(parse_float(r["executed_sell_price"], 0.0) * parse_float(r["sell_qty"], 0.0), 2) if parse_float(r["executed_sell_price"], 0.0) > 0 else None,
                "executed_buy_price": (None if r["executed_buy_price"] is None else round(parse_float(r["executed_buy_price"], 0.0), 4)),
                "executed_buy_at": str(r["executed_buy_at"] or ""),
                "executed_buy_value": round(parse_float(r["executed_buy_price"], 0.0) * parse_float(r["buy_qty"], 0.0), 2) if parse_float(r["executed_buy_price"], 0.0) > 0 else None,
                "current_buy_ltp": round(current_buy_ltp, 4),
                "current_buy_value": current_buy_value,
                "live_mtm_pnl": live_mtm_pnl,
                "completion_note": str(r["completion_note"] or ""),
                "updated_at": str(r["updated_at"] or ""),
                "seed_capital": round(parse_float(r["seed_capital"], 0.0), 2),
                "target_profit_pct": round(parse_float(r["target_profit_pct"], 0.0), 2),
                "plan_status": str(r["plan_status"] or ""),
                "plan_created_at": str(r["plan_created_at"] or ""),
                "plan_closed_at": str(r["plan_closed_at"] or ""),
            }
        )
    summary = {
        "count": len(items),
        "filtered_count": len(items),
        "total_count": total_count,
        "date_from": df or "",
        "date_to": dt_to or "",
        "state_filter": sf or "all",
        "executed": sum(1 for x in items if str(x.get("state") or "").lower() == "executed"),
        "pending": sum(1 for x in items if str(x.get("state") or "").lower() == "pending"),
        "sell_done": sum(1 for x in items if str(x.get("state") or "").lower() == "sell_done"),
        "buy_done": sum(1 for x in items if str(x.get("state") or "").lower() == "buy_done"),
        "skipped": sum(1 for x in items if str(x.get("state") or "").lower() == "skipped"),
        "replaced": sum(1 for x in items if str(x.get("state") or "").lower() == "replaced"),
        "matched": sum(1 for x in items if str(x.get("reconciliation_status") or "").lower() == "matched"),
        "partial": sum(1 for x in items if str(x.get("reconciliation_status") or "").lower() == "partial"),
        "unmatched": sum(1 for x in items if str(x.get("reconciliation_status") or "").lower() == "unmatched"),
    }
    return {"items": items, "summary": summary}


def extract_known_symbol_from_text(conn, text):
    lower = str(text or "").lower()
    symbols = [
        r["symbol"]
        for r in conn.execute("SELECT symbol FROM instruments WHERE active = 1 ORDER BY LENGTH(symbol) DESC, symbol").fetchall()
    ]
    for s in symbols:
        su = symbol_upper(s)
        if not su:
            continue
        if re.search(rf"(?<![a-z0-9]){re.escape(su.lower())}(?![a-z0-9])", lower):
            return su
    return None


def extract_note_pattern_from_chat(message):
    text = str(message or "").strip()
    if not text:
        return None

    # Prefer quoted text if provided.
    quoted = re.findall(r'"([^"]+)"|\'([^\']+)\'', text)
    for q1, q2 in quoted:
        q = (q1 or q2 or "").strip()
        if q:
            return q

    # Common upload marker in notes.
    m = re.search(r"(upload:[^\s,;]+)", text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()

    # Fallback: notes like/contains/= pattern.
    m = re.search(r"notes?\s*(?:like|contains|=)\s*([^\n]+)$", text, flags=re.IGNORECASE)
    if m:
        candidate = m.group(1).strip().strip(".;")
        if candidate:
            return candidate
    return None


def preview_trades_by_note_pattern(conn, note_pattern):
    like_arg = f"%{str(note_pattern or '').upper()}%"
    affected = int(
        conn.execute(
            "SELECT COUNT(*) AS c FROM trades WHERE notes IS NOT NULL AND UPPER(notes) LIKE ?",
            (like_arg,),
        ).fetchone()["c"]
    )
    symbols = [
        r["symbol"]
        for r in conn.execute(
            """
            SELECT DISTINCT symbol
            FROM trades
            WHERE notes IS NOT NULL AND UPPER(notes) LIKE ?
            ORDER BY symbol
            LIMIT 30
            """,
            (like_arg,),
        ).fetchall()
    ]
    notes_examples = [
        r["notes"]
        for r in conn.execute(
            """
            SELECT DISTINCT notes
            FROM trades
            WHERE notes IS NOT NULL AND UPPER(notes) LIKE ?
            ORDER BY notes
            LIMIT 5
            """,
            (like_arg,),
        ).fetchall()
    ]
    return {
        "note_pattern": str(note_pattern or ""),
        "affected_trades": affected,
        "affected_symbols": symbols,
        "notes_examples": notes_examples,
    }


def delete_trades_by_note_pattern(conn, note_pattern):
    preview = preview_trades_by_note_pattern(conn, note_pattern)
    like_arg = f"%{str(note_pattern or '').upper()}%"
    deleted = conn.execute(
        "DELETE FROM trades WHERE notes IS NOT NULL AND UPPER(notes) LIKE ?",
        (like_arg,),
    ).rowcount
    conn.commit()
    return {
        "note_pattern": str(note_pattern or ""),
        "deleted_trades": int(deleted),
        "deleted_symbols": preview["affected_symbols"],
        "notes_examples": preview["notes_examples"],
    }


def create_agent_approval(conn, action_type, query_text, payload, summary):
    try:
        payload_json = json.dumps(payload or {}, ensure_ascii=False, sort_keys=True)
    except Exception:
        payload_json = json.dumps({}, ensure_ascii=False)
    existing = conn.execute(
        """
        SELECT id, created_at, status, action_type, query_text, payload_json, summary
        FROM agent_approvals
        WHERE status = 'pending'
          AND action_type = ?
          AND payload_json = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (str(action_type or ""), payload_json),
    ).fetchone()
    if existing:
        return dict(existing)
    cur = conn.execute(
        """
        INSERT INTO agent_approvals(
          created_at, status, action_type, query_text, payload_json, summary
        ) VALUES (?, 'pending', ?, ?, ?, ?)
        """,
        (now_iso(), str(action_type or ""), str(query_text or ""), payload_json, str(summary or "")),
    )
    conn.commit()
    row = conn.execute(
        """
        SELECT id, created_at, status, action_type, query_text, payload_json, summary
        FROM agent_approvals
        WHERE id = ?
        """,
        (cur.lastrowid,),
    ).fetchone()
    return dict(row)


def list_agent_approvals(conn, status=None, limit=50):
    lim = max(1, min(500, int(limit)))
    params = []
    where = ["1=1"]
    if status:
        where.append("LOWER(status) = ?")
        params.append(str(status).strip().lower())
    sql = (
        "SELECT id, created_at, status, action_type, query_text, payload_json, summary, decided_at, executed_at, decision_note "
        "FROM agent_approvals WHERE "
        + " AND ".join(where)
        + " ORDER BY id DESC LIMIT ?"
    )
    params.append(lim)
    out = []
    for r in conn.execute(sql, params).fetchall():
        item = dict(r)
        try:
            item["payload"] = json.loads(str(item.get("payload_json") or "{}"))
        except Exception:
            item["payload"] = {}
        out.append(item)
    return out


def resolve_agent_approval(conn, approval_id, decision, note=""):
    dec = str(decision or "").strip().lower()
    if dec not in ("approve", "reject"):
        raise ValueError("decision_must_be_approve_or_reject")
    row = conn.execute(
        """
        SELECT id, status, action_type, payload_json, query_text, summary
        FROM agent_approvals
        WHERE id = ?
        """,
        (int(approval_id),),
    ).fetchone()
    if not row:
        raise ValueError("approval_not_found")
    status = str(row["status"] or "").lower()
    if status != "pending":
        raise ValueError("approval_not_pending")
    if dec == "reject":
        conn.execute(
            """
            UPDATE agent_approvals
            SET status = 'rejected', decided_at = ?, decision_note = ?
            WHERE id = ?
            """,
            (now_iso(), str(note or ""), int(approval_id)),
        )
        conn.commit()
        return {
            "ok": True,
            "approval_id": int(approval_id),
            "decision": "rejected",
            "executed": False,
        }

    conn.execute(
        """
        UPDATE agent_approvals
        SET status = 'approved', decided_at = ?, decision_note = ?
        WHERE id = ?
        """,
        (now_iso(), str(note or ""), int(approval_id)),
    )
    action = str(row["action_type"] or "")
    payload = {}
    try:
        payload = json.loads(str(row["payload_json"] or "{}"))
    except Exception:
        payload = {}
    out = {
        "ok": True,
        "approval_id": int(approval_id),
        "decision": "approved",
        "executed": False,
        "action_type": action,
    }
    if action == "delete_by_notes":
        note_pattern = str(payload.get("note_pattern") or "")
        result = delete_trades_by_note_pattern(conn, note_pattern)
        conn.execute(
            """
            UPDATE agent_approvals
            SET status = 'executed', executed_at = ?
            WHERE id = ?
            """,
            (now_iso(), int(approval_id)),
        )
        conn.commit()
        out.update(
            {
                "intent": "delete_by_notes",
                "executed": True,
                "note_pattern": note_pattern,
                "deleted_trades": int(result.get("deleted_trades", 0)),
                "deleted_symbols": result.get("deleted_symbols", []),
                "notes_examples": result.get("notes_examples", []),
            }
        )
        return out

    conn.execute(
        """
        UPDATE agent_approvals
        SET status = 'executed', executed_at = ?
        WHERE id = ?
        """,
        (now_iso(), int(approval_id)),
    )
    conn.commit()
    out["executed"] = True
    return out



def _safe_parse_improvement_proposal(path_s):
    raw = str(path_s or "").strip()
    if not raw:
        return {}
    try:
        path = Path(raw).resolve()
    except Exception:
        return {}
    try:
        safe_root = (get_current_tenant_data_dir() / "agent_improvements").resolve()
    except Exception:
        return {}
    root_s = str(safe_root).rstrip("\/").lower()
    path_s_norm = str(path).lower()
    if not (path_s_norm == root_s or path_s_norm.startswith(root_s + "\\")):
        return {}
    if not path.exists() or not path.is_file():
        return {}
    try:
        module = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
    except Exception:
        return {}
    for node in module.body:
        if not isinstance(node, ast.Assign):
            continue
        if not any(isinstance(t, ast.Name) and t.id == "PROPOSAL" for t in node.targets):
            continue
        try:
            value = ast.literal_eval(node.value)
        except Exception:
            return {}
        if isinstance(value, dict):
            return value
    return {}


def _verification_value_applied(expected, current):
    if isinstance(expected, bool):
        cur = parse_bool(current, default=None)
        return (cur is not None) and (bool(cur) == bool(expected))
    if isinstance(expected, (int, float)) and not isinstance(expected, bool):
        cur = parse_float(current, float("nan"))
        return math.isfinite(cur) and abs(cur - float(expected)) <= 1e-9
    return str(current if current is not None else "").strip().lower() == str(expected if expected is not None else "").strip().lower()


def _build_update_checks(source, updates, live_cfg):
    rows = []
    for key in sorted((updates or {}).keys()):
        expected = updates.get(key)
        current = (live_cfg or {}).get(key)
        rows.append(
            {
                "source": str(source or ""),
                "key": str(key),
                "expected": expected,
                "current": current,
                "applied": bool(_verification_value_applied(expected, current)),
            }
        )
    return rows


def build_approval_verification_payload(conn, approval_limit=120, action_limit=180):
    approval_lim = max(1, min(500, int(approval_limit)))
    action_lim = max(10, min(500, int(action_limit)))
    approvals = list_agent_approvals(conn, status=None, limit=approval_lim)
    counts = {"pending": 0, "approved": 0, "executed": 0, "rejected": 0, "expired": 0}
    for item in approvals:
        status = str(item.get("status") or "").strip().lower()
        if status:
            counts[status] = int(counts.get(status, 0)) + 1
    actions = list_software_perf_actions(conn, limit=action_lim)
    latest_auto_tune = next(
        (
            a
            for a in actions
            if str(a.get("action_type") or "").lower() in ("auto_tune", "llm_improvement", "write_draft")
        ),
        None,
    )
    latest_self_heal = next((a for a in actions if str(a.get("action_type") or "").lower() == "self_heal"), None)
    latest_write_draft = next((a for a in actions if str(a.get("action_type") or "").lower() == "write_draft"), None)
    draft_path = str(((latest_write_draft or {}).get("details") or {}).get("path") or "").strip()
    if not draft_path:
        out_dir = get_current_tenant_data_dir() / "agent_improvements"
        try:
            files = sorted(out_dir.glob("software_improvement_*.py"), reverse=True)
            if files:
                draft_path = str(files[0])
        except Exception:
            draft_path = ""
    draft_proposal = _safe_parse_improvement_proposal(draft_path)
    draft_updates = dict((draft_proposal or {}).get("suggested_live_config_updates") or {})
    auto_updates = dict(
        (((latest_auto_tune or {}).get("details") or {}).get("updates") or {})
        or (((latest_auto_tune or {}).get("details") or {}).get("suggested_live_config_updates") or {})
        or (((latest_write_draft or {}).get("details") or {}).get("suggested_live_config_updates") or {})
    )
    live_cfg = get_live_config(conn)
    draft_checks = _build_update_checks("draft_suggested", draft_updates, live_cfg)
    auto_checks = _build_update_checks("auto_tune", auto_updates, live_cfg)
    checks = auto_checks + draft_checks
    change_summary = []
    for action in actions[:40]:
        details = dict(action.get("details") or {})
        bits = []
        if isinstance(details.get("updates"), dict) and details.get("updates"):
            bits.append("updates=" + ", ".join([f"{k}:{v}" for k, v in list(details["updates"].items())[:8]]))
        if isinstance(details.get("suggested_live_config_updates"), dict) and details.get("suggested_live_config_updates"):
            bits.append(
                "llm_updates="
                + ", ".join([f"{k}:{v}" for k, v in list(details["suggested_live_config_updates"].items())[:8]])
            )
        if isinstance(details.get("actions"), list) and details.get("actions"):
            bits.append("actions=" + ", ".join([str(x) for x in details.get("actions", [])[:8]]))
        if isinstance(details.get("errors"), list) and details.get("errors"):
            bits.append("errors=" + "; ".join([str(x) for x in details.get("errors", [])[:4]]))
        if details.get("llm_status"):
            bits.append(f"llm={details.get('llm_status')}")
        if details.get("model"):
            bits.append(f"model={details.get('model')}")
        if details.get("error"):
            bits.append(f"error={details.get('error')}")
        if details.get("path"):
            bits.append(f"path={details.get('path')}")
        change_summary.append(
            {
                "created_at": str(action.get("created_at") or ""),
                "source": str(action.get("action_type") or ""),
                "status": str(action.get("status") or ""),
                "summary": str(action.get("summary") or ""),
                "details": " | ".join(bits) if bits else "-",
            }
        )
    return {
        "generated_at": now_iso(),
        "approval_process": [
            "Risky assistant actions are queued as pending and require explicit approve/reject.",
            "Approved actions are executed and status moves to executed.",
            "Rejected actions are not executed and status moves to rejected.",
            "Software improvement drafts are review artifacts and are never auto-executed.",
        ],
        "approvals": {
            "counts": counts,
            "pending_count": int(counts.get("pending", 0)),
            "items": approvals,
        },
        "software_performance": {
            "config": get_software_perf_agent_config(conn),
            "live_config": live_cfg,
            "latest_self_heal": latest_self_heal,
            "latest_auto_tune": latest_auto_tune,
            "latest_write_draft": latest_write_draft,
            "latest_draft": {
                "path": draft_path,
                "generated_at": str((draft_proposal or {}).get("generated_at") or ""),
                "issues": list((draft_proposal or {}).get("issues") or []),
                "suggested_live_config_updates": draft_updates,
                "proposal_loaded": bool(draft_proposal),
            },
            "verification_checks": checks,
            "summary": {
                "draft_updates_total": len(draft_checks),
                "draft_updates_applied": sum(1 for x in draft_checks if x.get("applied")),
                "llm_updates_total": len(auto_checks),
                "llm_updates_applied": sum(1 for x in auto_checks if x.get("applied")),
            },
        },
        "change_summary": change_summary,
    }

def _strategy_param_snapshot_for_explain(params):
    return {
        "allocation_limit": clamp(parse_float(params.get("allocation_limit"), 0.25), 0.02, 0.8),
        "trim_trigger_overweight": clamp(parse_float(params.get("trim_trigger_overweight"), 0.02), 0.0, 0.5),
        "add_trigger_underweight": clamp(parse_float(params.get("add_trigger_underweight"), 0.02), 0.0, 0.5),
        "max_position_weight": clamp(parse_float(params.get("max_position_weight"), 0.12), 0.03, 0.9),
        "buy_l1_discount": clamp(parse_float(params.get("buy_l1_discount"), 0.03), 0.0, 0.5),
        "buy_l2_discount": clamp(parse_float(params.get("buy_l2_discount"), 0.06), 0.0, 0.7),
        "momentum_lookback_days": int(clamp(parse_float(params.get("momentum_lookback_days"), 30), 5, 120)),
    }


def explain_strategy_reason(conn, query_text):
    lower = str(query_text or "").strip().lower()
    if not lower:
        return None
    reasonish = any(
        k in lower
        for k in (
            "reason",
            "why",
            "overweight",
            "under target",
            "macro",
            "trim",
            "add",
            "hold",
            "review",
        )
    )
    if not reasonish:
        return None

    insights = load_latest_strategy_insights(conn)
    if insights is None:
        insights = build_strategy_insights(conn, run_date=dt.date.today().isoformat())
        persist_strategy_insights(conn, insights)
        conn.commit()
        insights = load_latest_strategy_insights(conn)
    if insights is None:
        return None

    recs = insights.get("recommendations", []) or []
    if not recs:
        return None

    target = None
    for r in recs:
        sym = str(r.get("symbol", "")).strip().lower()
        if sym and sym in lower:
            target = r
            break
    if target is None and "overweight" in lower:
        for r in recs:
            if "overweight" in str(r.get("reason") or "").lower():
                target = r
                break
    if target is None and "under target" in lower:
        for r in recs:
            if "under target" in str(r.get("reason") or "").lower():
                target = r
                break
    if target is None and "signal" in lower:
        for r in recs:
            if "signal" in str(r.get("reason") or "").lower():
                target = r
                break
    if target is None and "overweight" in lower:
        for r in recs:
            if str(r.get("action", "")).upper() == "TRIM":
                target = r
                break
    if target is None:
        target = recs[0]

    params = get_active_params(conn)
    p = _strategy_param_snapshot_for_explain(params)
    action = str(target.get("action") or "HOLD").upper()
    weight_current = parse_float(target.get("weight_current"), 0.0)
    weight_target = parse_float(target.get("weight_target"), 0.0)
    delta_weight = parse_float(target.get("delta_weight"), 0.0)
    reason = str(target.get("reason") or "").strip()
    macro = insights.get("macro") or {}
    macro_thought = str(macro.get("thought") or "")
    macro_regime = str(macro.get("regime") or "neutral").lower()

    affecting = []
    suggestions = []
    if ("overweight" in reason.lower()) or action == "TRIM":
        affecting.extend(
            [
                ("max_position_weight", p["max_position_weight"], "Hard cap for single-scrip concentration."),
                ("trim_trigger_overweight", p["trim_trigger_overweight"], "Extra overweight band above target before TRIM."),
                ("allocation_limit", p["allocation_limit"], "Caps target weight used for balancing."),
            ]
        )
        suggestions.extend(
            [
                f"Increase `max_position_weight` slightly (e.g. {p['max_position_weight']:.3f} -> {min(0.9, p['max_position_weight'] + 0.01):.3f}) to reduce forced trims.",
                f"Increase `trim_trigger_overweight` (e.g. {p['trim_trigger_overweight']:.3f} -> {min(0.5, p['trim_trigger_overweight'] + 0.005):.3f}) to allow more drift before TRIM.",
            ]
        )
    if ("under target" in reason.lower()) or action == "ADD":
        affecting.extend(
            [
                ("add_trigger_underweight", p["add_trigger_underweight"], "Minimum underweight gap needed to trigger ADD."),
                ("buy_l1_discount", p["buy_l1_discount"], "Primary pullback threshold for ADD logic."),
                ("buy_l2_discount", p["buy_l2_discount"], "Deeper pullback threshold for stronger adds."),
            ]
        )
        suggestions.extend(
            [
                f"Decrease `add_trigger_underweight` (e.g. {p['add_trigger_underweight']:.3f} -> {max(0.0, p['add_trigger_underweight'] - 0.005):.3f}) to allow more ADD candidates.",
                f"Reduce `buy_l1_discount` / `buy_l2_discount` if you want earlier adds on smaller dips.",
            ]
        )
    if ("momentum" in reason.lower()) or action == "REVIEW":
        affecting.append(
            ("momentum_lookback_days", p["momentum_lookback_days"], "Window used to compute momentum weakness/strength.")
        )
        suggestions.append(
            f"Adjust `momentum_lookback_days` (currently {p['momentum_lookback_days']}) to make REVIEW triggers more/less sensitive."
        )

    macro_line = (
        f"Macro regime: {macro_regime.upper()} | score={round(parse_float(macro.get('score'), 0.0), 3)} | confidence={round(parse_float(macro.get('confidence'), 0.0), 4)}."
    )
    if "macro feed unavailable" in macro_thought.lower():
        suggestions.append(
            "Macro feed unavailable is connectivity/source dependent, not a strategy parameter. Once feed is reachable, macro bias will auto-update."
        )

    seen = set()
    affecting_lines = []
    for k, v, d in affecting:
        if k in seen:
            continue
        seen.add(k)
        affecting_lines.append(f"- `{k}` = {v}: {d}")

    if not suggestions:
        suggestions.append("No parameter change is needed unless you want a more aggressive or defensive posture.")

    message = (
        f"Strategy reason expanded for `{target.get('symbol')}` ({action}):\n"
        f"- Current weight: {weight_current:.2%}\n"
        f"- Target weight: {weight_target:.2%}\n"
        f"- Delta weight: {delta_weight:.2%}\n"
        f"- Confidence: {parse_float(target.get('confidence'), 0.0):.2f}\n"
        f"- Base reason: {reason}\n"
        f"- {macro_line}\n"
        f"- Macro thought: {macro_thought}\n"
        f"Parameters affecting this outcome:\n"
        + ("\n".join(affecting_lines) if affecting_lines else "- (No direct parameter drivers identified for this line.)")
        + "\nHow to modify if needed:\n"
        + "\n".join(f"- {s}" for s in suggestions)
    )
    return {
        "ok": True,
        "intent": "strategy_reason_explain",
        "executed": False,
        "symbol": target.get("symbol"),
        "action": action,
        "reason": reason,
        "macro": macro,
        "parameters": {k: v for k, v, _ in affecting},
        "message": message,
    }


def assistant_chat_response(conn, message):
    text = str(message or "").strip()
    if not text:
        return {
            "ok": False,
            "intent": "error",
            "message": "Empty message. Try: help",
        }

    query_catalog = [
        {
            "query": "portfolio summary",
            "explanation": "Shows invested, market value, realized/unrealized P&L, total return, today P&L, and cash balance.",
        },
        {
            "query": "how is cash balance calculated",
            "explanation": "Explains cash-balance formula using cash ledger entries (deposits/withdrawals) and current values.",
        },
        {
            "query": "price status",
            "explanation": "Shows latest price refresh timestamp and number of scrips priced.",
        },
        {
            "query": "refresh gold price / gold ltp / gold rate",
            "explanation": "Refreshes GOLD live price feed and shows qty, ltp, source, updated time, and market value.",
        },
        {
            "query": "software performance status / run software performance agent",
            "explanation": "Shows software-health telemetry and can trigger self-heal + guarded improvement draft generation.",
        },
        {
            "query": "upload summary",
            "explanation": "Lists upload batches grouped by notes (upload file marker), with trade counts and date range.",
        },
        {
            "query": "cashflow summary",
            "explanation": "Shows payin/payout totals, cash balance, and entry count from uploaded cashflow sheet.",
        },
        {
            "query": "cashflow duplicates",
            "explanation": "Shows duplicate cashflow rows by entry id and by same date/type/amount groups.",
        },
        {
            "query": "dividend summary",
            "explanation": "Shows uploaded dividend totals by period and top scrip contributors.",
        },
        {
            "query": "show duplicates",
            "explanation": "Shows duplicate groups by Trade ID and by same symbol/side/date/qty/price.",
        },
        {
            "query": "top gainers / top losers / top day gainers",
            "explanation": "Shows top movers using total return % or day change %. ",
        },
        {
            "query": "strategy summary",
            "explanation": "Shows latest trim/add/hold/review mix and top strategy recommendations.",
        },
        {
            "query": "strategy projection",
            "explanation": "Shows conservative/base/aggressive projected portfolio values over years.",
        },
        {
            "query": "intel summary",
            "explanation": "Shows intelligence overlay score from commentary/policy docs, financial QoQ signals, chart-pattern agent signals, and inferred cross-company flows.",
        },
        {
            "query": "chart summary / chart signal for <symbol>",
            "explanation": "Shows latest chart-agent pattern score, signal, and key technical drivers.",
        },
        {
            "query": "policy impact for <symbol>",
            "explanation": "Explains latest policy/commentary/financial intelligence effect for a specific scrip.",
        },
        {
            "query": "fund flow links",
            "explanation": "Shows inferred cross-company fund rotation links from FII/DII holding deltas.",
        },
        {
            "query": "intel autopilot status / run intel autopilot",
            "explanation": "Shows autopilot settings and triggers online sweep (news + Screener/NSE/company-site financial context).",
        },
        {
            "query": "refresh strategy",
            "explanation": "Forces strategy recomputation immediately using latest holdings and prices.",
        },
        {
            "query": "explain strategy reason for <symbol>",
            "explanation": "Expands recommendation reason with parameter drivers and what to change to tune outcomes.",
        },
        {
            "query": "what is max_position_weight",
            "explanation": "Explains strategy parameters and shows current active value.",
        },
        {
            "query": "preview notes like \"upload:tradebook-OWY330.xlsx\"",
            "explanation": "Counts matching trades across scrips without deleting.",
        },
        {
            "query": "erase trades notes like \"upload:tradebook-OWY330.xlsx\"",
            "explanation": "Queues a delete request; execution happens only after notifier approval.",
        },
        {
            "query": "show pending approvals",
            "explanation": "Lists pending agent actions waiting for explicit approval/rejection.",
        },
    ]

    metric_explanations = {
        "invested": "Investment = Deposits - Withdrawals from cashflow ledger.",
        "market deployment": "Market Deployment = sum(open_qty * avg_cost) across current holdings.",
        "market value": "Market Value = sum(open_qty * LTP) across holdings.",
        "realized": "Realized P/L is from closed quantities; if dividend exceeds remaining invested for a fully sold scrip, excess is treated as realized.",
        "unrealized": "Unrealized P/L = Market Value - Invested for currently open holdings.",
        "total pnl": "Total P/L = Realized P/L + Unrealized P/L.",
        "total return": "Total Return % = (Total P/L / Investment) * 100 (when investment > 0).",
        "cagr": "CAGR is an annualized return estimate using cashflow timing (Modified Dietz style) and current account value.",
        "xirr": "XIRR is money-weighted annualized return solved from dated deposit/withdrawal cashflows and current account value.",
        "today pnl": "Today P/L = sum(open_qty * source day change_abs) across holdings from live feeds (NSE/BSE/scraped).",
        "today change": "Today Change % = Today P/L divided by previous-day portfolio value estimate.",
        "upl %": "UPL % = (Unrealized P/L / Invested) * 100 for the scrip.",
        "peak traded": "Peak Traded = highest adjusted historical traded price (BUY/SELL) for the scrip (split-adjusted).",
        "peak buy": "Peak Buy (legacy label) maps to Peak Traded in current builds.",
        "cash balance": "Cash Balance = SUM(amount) from cash_ledger across deposits, withdrawals, trade credits, investments, and charges.",
    }
    portfolio_field_explanations = {
        "symbol": "Trading symbol/ticker of the scrip.",
        "qty": "Current open quantity (BUY minus SELL) for the scrip.",
        "ltp": "Latest traded price used for valuation (live/fallback guarded).",
        "avg cost": "Average buy cost per open unit after matched sells.",
        "invested": "Open capital deployed after adjusting for total dividend credits on that scrip.",
        "dividend": "Total dividend credited for the scrip from uploaded dividend statements.",
        "mkt value": "Current market value = open_qty * LTP.",
        "market value": "Current market value = open_qty * LTP.",
        "rpl": "Realized P/L from closed quantities matched against historical buy lots (dividends shown separately).",
        "realized pnl": "Realized P/L from closed quantities matched against historical buy lots (dividends shown separately).",
        "upl": "Unrealized P/L on open holdings = Market Value - Invested.",
        "unrealized pnl": "Unrealized P/L on open holdings = Market Value - Invested.",
        "abs p/l": "Absolute P/L = Realized P/L + Unrealized P/L.",
        "upl %": "UPL % = Unrealized P/L divided by Invested, in percent.",
        "day p/l": "Day P/L = open_qty * day absolute price change.",
        "day %": "Day % = day change relative to previous close-based valuation.",
        "return %": "Return % = (Realized + Unrealized) / cumulative historical BUY value * 100 for the scrip.",
        "peak traded": "Highest historical adjusted traded price (BUY/SELL) for the scrip.",
        "% from peak traded": "Percent gap between current LTP and Peak Traded price.",
        "peak buy": "Legacy alias for Peak Traded.",
        "% from peak buy": "Legacy alias for % from Peak Traded.",
        "signal": "Strategy signal band (e.g., B1/B2 for buy zones, S1/S2/S3 for sell zones).",
        "investment": "Hand investment = deposits - withdrawals from cash ledger.",
        "market deployment": "Open deployed capital in current holdings.",
    }
    strategy_param_explanations = {
        "buy_l1_discount": "Primary BUY trigger discount from avg cost (fraction, e.g. 0.03 = 3%).",
        "buy_l2_discount": "Deeper BUY trigger discount from avg cost for stronger accumulation.",
        "sell_s1_markup": "First SELL trigger markup above avg cost.",
        "sell_s2_markup": "Second SELL trigger markup above avg cost.",
        "sell_s3_markup": "Aggressive SELL trigger markup above avg cost.",
        "mux_factor_default": "Default coefficient multiplier applied in strategy math.",
        "brokerage_rate": "Estimated transaction cost rate used for net return and sizing.",
        "allocation_limit": "Upper allocation cap for total deployable capital.",
        "exposure_cap": "Per-idea exposure cap when adding new positions.",
        "trim_trigger_overweight": "Weight delta above target that triggers TRIM action.",
        "add_trigger_underweight": "Weight delta below target that triggers ADD action.",
        "max_position_weight": "Maximum portfolio weight allowed for any single scrip (fraction of portfolio).",
        "max_new_ideas": "Maximum count of new non-portfolio ideas allowed in one strategy run.",
        "momentum_lookback_days": "Lookback window (days) used by momentum features.",
        "projection_years": "How many years to project in strategy projection chart.",
        "intel_weight_commentary": "Weight of commentary/transcript sentiment in intelligence overlay.",
        "intel_weight_policy": "Weight of budget/policy text effect in intelligence overlay.",
        "intel_weight_financials": "Weight of QoQ company financial signal in intelligence overlay.",
        "intel_weight_chart": "Weight of chart-pattern intelligence signal (trend/momentum/breakout/relative-strength) in intelligence overlay.",
        "intel_decay_days": "Recency decay window for commentary/policy intelligence documents.",
    }

    lower = text.lower()

    def contains_term(term):
        tok = str(term or "").strip().lower()
        if not tok:
            return False
        return re.search(rf"(?<![a-z0-9]){re.escape(tok)}(?![a-z0-9])", lower) is not None

    gold_alias_pattern = r"\b(gold|24\s*k|24k|24\s*carat|24\s*karat|24carat|24karat|bullion)\b"
    gold_price_pattern = r"\b(price|ltp|rate|value|spot|per\s*gram|gram|gm|gms)\b"
    gold_refresh_pattern = r"\b(refresh|update|fetch|re[-\s]*fetch|sync|re[-\s]*sync|reload|scrape|pull|run)\b"
    has_gold_alias = re.search(gold_alias_pattern, lower) is not None
    has_gold_price_hint = re.search(gold_price_pattern, lower) is not None
    has_gold_refresh_verb = re.search(gold_refresh_pattern, lower) is not None
    has_gold_support_phrase = ("not supported in assistant" in lower) or ("unsupported in assistant" in lower)
    is_gold_query = has_gold_alias and (has_gold_price_hint or has_gold_refresh_verb or has_gold_support_phrase)

    if any(k in lower for k in ("help", "commands", "what can you do", "features", "what can i ask", "sample queries", "queries")):
        return {
            "ok": True,
            "intent": "query_catalog",
            "executed": False,
            "query_catalog": query_catalog,
            "message": "Available assistant queries and actions:",
        }

    if any(k in lower for k in ("pending approvals", "show approvals", "approvals", "notifier")):
        items = list_agent_approvals(conn, status="pending", limit=20)
        return {
            "ok": True,
            "intent": "approvals_list",
            "executed": False,
            "items": items,
            "pending_count": len(items),
            "message": f"Pending approvals: {len(items)}",
        }

    if ("cash balance" in lower) and any(k in lower for k in ("how", "calculate", "calculated", "formula", "explain", "what is")):
        cash_row = conn.execute(
            """
            SELECT
              COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'deposit' THEN amount ELSE 0 END),0) AS deposits,
              COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'withdrawal' THEN -amount ELSE 0 END),0) AS withdrawals,
              COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'trade_credit' THEN amount ELSE 0 END),0) AS trade_credits,
              COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'investment' THEN -amount ELSE 0 END),0) AS investments,
              COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'charge' THEN -amount ELSE 0 END),0) AS charges,
              COALESCE(SUM(amount),0) AS cash_balance,
              COUNT(*) AS entries
            FROM cash_ledger
            """
        ).fetchone()
        deposits = float(cash_row["deposits"])
        withdrawals = float(cash_row["withdrawals"])
        trade_credits = float(cash_row["trade_credits"])
        investments = float(cash_row["investments"])
        charges = float(cash_row["charges"])
        cash_balance = float(cash_row["cash_balance"])
        return {
            "ok": True,
            "intent": "cash_balance_explain",
            "executed": False,
            "cash_breakdown": {
                "entries": int(cash_row["entries"]),
                "deposits_total": round(deposits, 2),
                "withdrawals_total": round(withdrawals, 2),
                "net_hand_investment_total": round(deposits - withdrawals, 2),
                "trade_credit_total": round(trade_credits, 2),
                "investment_spend_total": round(investments, 2),
                "charges_total": round(charges, 2),
                "cash_balance": round(cash_balance, 2),
            },
            "formula": (
                "cash_balance = deposits + trade_credits - withdrawals - investments - charges"
            ),
            "note": (
                "Instant payout / withdrawal-request entries are withdrawals. "
                "Net settlement debits are investment spend; settlement credits are trade credits. "
                "DP/AMC rows are charges."
            ),
            "message": (
                f"Cash Balance formula: deposits + trade credits - withdrawals - investments - charges.\n"
                f"Deposits: {round(deposits, 2)}\n"
                f"Investment (deposits - withdrawals): {round(deposits - withdrawals, 2)}\n"
                f"Trade Credits: {round(trade_credits, 2)}\n"
                f"Withdrawals: {round(withdrawals, 2)}\n"
                f"Investments: {round(investments, 2)}\n"
                f"Charges: {round(charges, 2)}\n"
                f"Cash Balance: {round(cash_balance, 2)}"
            ),
        }

    if ("cashflow summary" in lower) or ("cash summary" in lower):
        s = cashflow_summary(conn)
        return {
            "ok": True,
            "intent": "cashflow_summary",
            "executed": False,
            "cashflow_summary": s,
            "message": (
                f"Cashflow summary:\n"
                f"Entries: {s['entries']}\n"
                f"Deposits: {s['deposits_total']}\n"
                f"Withdrawals: {s['withdrawals_total']}\n"
                f"Investment: {s['net_hand_investment_total']}\n"
                f"Trade Credits: {s['trade_credit_total']}\n"
                f"Investment Spend: {s['investment_spend_total']}\n"
                f"Charges: {s['charges_total']}\n"
                f"Cash Balance: {s['cash_balance']}"
            ),
        }

    if ("dividend summary" in lower) or ("dividends summary" in lower):
        s = dividend_summary(conn)
        top_txt = ", ".join([f"{x['symbol']}={x['total_dividend']}" for x in s.get("top_symbols", [])]) or "-"
        return {
            "ok": True,
            "intent": "dividend_summary",
            "executed": False,
            "dividend_summary": s,
            "message": (
                f"Dividend summary:\n"
                f"Entries: {s['entries']}\n"
                f"Total Dividend Profit: {s['total_dividend']}\n"
                f"From: {s['from_date']}\n"
                f"To: {s['to_date']}\n"
                f"Top Symbols: {top_txt}"
            ),
        }

    if any(
        k in lower
        for k in (
            "chart summary",
            "chart signal",
            "technical summary",
            "technical signal",
            "chart agent",
        )
    ):
        run_now = any(k in lower for k in ("run", "refresh", "now", "trigger"))
        run_result = None
        if run_now:
            run_result = run_chart_intel_agent_once(max_runtime_sec=60, force=True)
        symbol = extract_known_symbol_from_text(conn, text)
        items = list_chart_snapshots(conn, limit=(30 if not symbol else 8), symbol=symbol)
        top = items[0] if items else None
        msg = "Chart agent has no snapshots yet."
        if top:
            msg = (
                f"{symbol_upper(top.get('symbol'))} chart signal: {top.get('signal')} "
                f"score {round(parse_float(top.get('score'), 0.0), 2)} "
                f"(confidence {round(parse_float(top.get('confidence'), 0.0) * 100.0, 1)}%). "
                f"Drivers: {str((top.get('pattern_flags') or [])[:6])}."
            )
        return {
            "ok": True,
            "intent": "chart_summary",
            "executed": bool(run_now),
            "symbol": symbol,
            "run_result": run_result,
            "items": items[:12],
            "message": msg,
        }

    if any(
        k in lower
        for k in (
            "intel summary",
            "intelligence summary",
            "intelligence overlay",
            "intel overlay",
            "fund flow links",
            "cross company flow",
            "cross-company flow",
            "policy impact",
            "financial signal",
        )
    ):
        intel = intelligence_summary(conn, limit=40)
        symbol = extract_known_symbol_from_text(conn, text)
        if symbol:
            item = None
            for s in intel.get("symbol_scores", []):
                if symbol_upper(s.get("symbol")) == symbol:
                    item = s
                    break
            if item is None:
                item = {
                    "symbol": symbol,
                    "score": 0.0,
                    "confidence": 0.0,
                    "summary": "No intelligence signals available yet.",
                }
            return {
                "ok": True,
                "intent": "intel_symbol",
                "executed": False,
                "symbol": symbol,
                "item": item,
                "message": (
                    f"{symbol} intelligence impact:\n"
                    f"Score: {round(parse_float(item.get('score'), 0.0), 2)}\n"
                    f"Confidence: {round(parse_float(item.get('confidence'), 0.0) * 100.0, 1)}%\n"
                    f"Drivers: {item.get('summary') or '-'}"
                ),
            }
        return {
            "ok": True,
            "intent": "intel_summary",
            "executed": False,
            "portfolio_score": intel.get("portfolio_score"),
            "portfolio_confidence": intel.get("portfolio_confidence"),
            "cross_flows": intel.get("cross_flows", [])[:12],
            "top_symbols": intel.get("symbol_scores", [])[:12],
            "weights": intel.get("weights", {}),
            "message": (
                f"Intelligence overlay: {round(parse_float(intel.get('portfolio_score'), 0.0), 2)} "
                f"(confidence {round(parse_float(intel.get('portfolio_confidence'), 0.0) * 100.0, 1)}%).\n"
                f"Docs analyzed: {intel.get('documents_recent', 0)}, impacts: {intel.get('impacts_recent', 0)}.\n"
                f"{intel.get('thought', '')}"
            ),
        }

    if "intel autopilot" in lower or "autopilot status" in lower:
        cfg = get_intel_autopilot_config(conn)
        run_now = any(k in lower for k in ("run", "refresh", "now", "trigger", "start sweep"))
        run_result = None
        if run_now:
            run_result = run_intelligence_autopilot_once(max_runtime_sec=60, force=True)
        fc = (run_result or {}).get("online_financial_collection") or {}
        return {
            "ok": True,
            "intent": "intel_autopilot_status",
            "executed": bool(run_now),
            "config": cfg,
            "run_result": run_result,
            "message": (
                f"Intelligence autopilot: enabled={cfg.get('enabled')}, "
                f"interval={cfg.get('interval_seconds')} sec, max_docs={cfg.get('max_docs')}, "
                f"symbols_limit={cfg.get('symbols_limit')}. "
                + (
                    f"Run inserted {int(parse_float((run_result or {}).get('inserted_docs'), 0.0))} docs; "
                    f"online financial rows +{int(parse_float(fc.get('inserted_financial_rows'), 0.0))} "
                    f"(updated {int(parse_float(fc.get('updated_financial_rows'), 0.0))})."
                    if run_result
                    else ""
                )
            ),
        }

    if any(k in lower for k in ("how is", "how do you calculate", "explain", "formula")):
        for key, expl in metric_explanations.items():
            if contains_term(key):
                return {
                    "ok": True,
                    "intent": "metric_explain",
                    "executed": False,
                    "metric": key,
                    "explanation": expl,
                    "message": f"{key.title()} calculation:\n{expl}",
                }

    if not is_gold_query:
        for key, expl in portfolio_field_explanations.items():
            if contains_term(key):
                return {
                    "ok": True,
                    "intent": "field_explain",
                    "executed": False,
                    "field": key,
                    "explanation": expl,
                    "message": f"{key.title()}:\n{expl}",
                }

    for key, expl in strategy_param_explanations.items():
        key_spaced = key.replace("_", " ")
        if (key in lower) or (key_spaced in lower):
            params = get_active_params(conn)
            current = params.get(key)
            current_num = parse_float(current, None) if current is not None else None
            if current_num is None:
                value_txt = "N/A"
            else:
                value_txt = str(current_num)
            return {
                "ok": True,
                "intent": "strategy_param_explain",
                "executed": False,
                "parameter": key,
                "current_value": current_num,
                "explanation": expl,
                "message": (
                    f"{key}: {expl}\n"
                    f"Current active value: {value_txt}"
                ),
            }

    if (
        any(k in lower for k in ("portfolio summary", "portfolio snapshot", "snapshot", "summary"))
        and "upload summary" not in lower
        and "strategy summary" not in lower
        and "strategy" not in lower
    ):
        s = portfolio_summary(conn)
        return {
            "ok": True,
            "intent": "portfolio_summary",
            "executed": False,
            "summary": s,
            "message": (
                f"Portfolio snapshot:\n"
                f"Investment: {round(parse_float(s.get('hand_invested'), s['invested']), 2)}\n"
                f"Market Deployment: {round(parse_float(s.get('market_deployment'), 0.0), 2)}\n"
                f"Market Value: {round(s['market_value'], 2)}\n"
                f"Total P/L: {round(s['total_pnl'], 2)} ({round(s['total_return_pct'], 2)}%)\n"
                f"Today: {round(s['today_pnl'], 2)} ({round(s['today_change_pct'], 2)}%)\n"
                f"CAGR: {round(parse_float(s.get('cagr_pct'), 0.0), 2)}%\n"
                f"XIRR: {round(parse_float(s.get('xirr_pct'), 0.0), 2)}%"
            ),
        }

    if ("price status" in lower) or ("prices status" in lower) or ("ltp status" in lower):
        st = conn.execute("SELECT MAX(updated_at) AS updated_at, COUNT(*) AS c FROM latest_prices").fetchone()
        return {
            "ok": True,
            "intent": "price_status",
            "executed": False,
            "price_status": {"updated_at": st["updated_at"], "scrips_with_price": st["c"]},
            "message": f"Prices status: updated_at={st['updated_at']}, scrips_with_price={st['c']}",
        }

    if is_gold_query:
        force_refresh = bool(has_gold_refresh_verb)
        refresh_error = None
        if force_refresh:
            try:
                refresh_latest_prices_from_exchange(max_runtime_sec=25)
                recompute_holdings_and_signals(force_strategy=False)
            except Exception as ex:
                refresh_error = str(ex)
        rows = conn.execute(
            """
            SELECT
              i.symbol,
              COALESCE(h.qty,0) AS qty,
              COALESCE(h.market_value,0) AS market_value,
              COALESCE(lp.ltp,0) AS ltp,
              lp.updated_at
            FROM instruments i
            LEFT JOIN holdings h ON UPPER(h.symbol) = UPPER(i.symbol)
            LEFT JOIN latest_prices lp ON UPPER(lp.symbol) = UPPER(i.symbol)
            WHERE i.active = 1
              AND UPPER(COALESCE(i.asset_class, 'EQUITY')) = 'GOLD'
            ORDER BY COALESCE(h.qty,0) DESC, i.symbol
            """
        ).fetchall()
        items = []
        src_map = latest_tick_source_map(conn, [r["symbol"] for r in rows]) if rows else {}
        for r in rows:
            sym = symbol_upper(r["symbol"])
            ltp = parse_float(r["ltp"], 0.0)
            qty = parse_float(r["qty"], 0.0)
            market_value = parse_float(r["market_value"], 0.0)
            src = str((src_map.get(sym) or {}).get("source") or "")
            src_ts = str((src_map.get(sym) or {}).get("fetched_at") or "")
            if ltp <= 0:
                market_value = 0.0
            items.append(
                {
                    "symbol": sym,
                    "qty": round(qty, 4),
                    "ltp_per_gram": round(ltp, 4),
                    "market_value": round(market_value, 2),
                    "updated_at": r["updated_at"],
                    "source": src,
                    "source_fetched_at": src_ts,
                }
            )
        if not items:
            return {
                "ok": True,
                "intent": "gold_price_status",
                "executed": bool(force_refresh),
                "items": [],
                "message": "No active GOLD symbols found.",
            }
        top = items[0]
        if parse_float(top.get("ltp_per_gram"), 0.0) > 0:
            msg = (
                f"GOLD live price ({top['symbol']}): {top['ltp_per_gram']} per gram; "
                f"qty={top['qty']}; market_value={top['market_value']}; "
                f"updated_at={top.get('updated_at') or '-'}; source={top.get('source') or '-'}."
            )
        else:
            msg = (
                f"GOLD live price unavailable for {top['symbol']} (qty={top['qty']}). "
                f"Last source={top.get('source') or '-'}; updated_at={top.get('updated_at') or '-'}."
            )
        if refresh_error:
            msg += f" Refresh error: {refresh_error}"
        return {
            "ok": True,
            "intent": "gold_price_status",
            "executed": bool(force_refresh),
            "items": items,
            "refresh_error": refresh_error,
            "message": msg,
        }

    if ("software performance" in lower) or ("software health" in lower) or ("self healing status" in lower):
        run_now = any(k in lower for k in ("run", "refresh", "heal now", "run now"))
        run_result = None
        if run_now:
            run_result = run_software_perf_agent_once(max_runtime_sec=90, force=True)
        cfg = get_software_perf_agent_config(conn)
        snaps = list_software_perf_snapshots(conn, limit=5)
        latest = snaps[0] if snaps else None
        return {
            "ok": True,
            "intent": "software_performance_status",
            "executed": bool(run_now),
            "config": cfg,
            "latest": latest,
            "run_result": run_result,
            "message": (
                f"Software performance: issues={int(parse_float((latest or {}).get('issue_count'), 0.0))}, "
                f"stale={int(parse_float((latest or {}).get('live_stale_symbols'), 0.0))}, "
                f"missing={int(parse_float((latest or {}).get('live_missing_price_symbols'), 0.0))}, "
                f"weak_sources={int(parse_float((latest or {}).get('weak_sources_count'), 0.0))}."
                if latest
                else "Software performance: no snapshots yet."
            ),
        }

    strategy_reason = explain_strategy_reason(conn, text)
    if strategy_reason is not None:
        return strategy_reason

    if "strategy" in lower:
        force_refresh = any(k in lower for k in ("refresh", "recompute", "rerun", "run now"))
        if force_refresh:
            insights = build_strategy_insights(conn, run_date=dt.date.today().isoformat())
            persist_strategy_insights(conn, insights)
            conn.commit()
            insights = load_latest_strategy_insights(conn)
        else:
            insights = load_latest_strategy_insights(conn)
            if insights is None:
                insights = build_strategy_insights(conn, run_date=dt.date.today().isoformat())
                persist_strategy_insights(conn, insights)
                conn.commit()
                insights = load_latest_strategy_insights(conn)

        if insights is None:
            return {
                "ok": False,
                "intent": "strategy_summary",
                "executed": False,
                "message": "Strategy data is not available yet.",
            }

        if any(k in lower for k in ("projection", "forecast", "years", "future")):
            scenarios = insights.get("projections", {}).get("scenarios", [])
            snippet = []
            for sc in scenarios:
                pts = sc.get("points", [])
                last = pts[-1] if pts else None
                if last:
                    snippet.append(
                        {
                            "scenario": sc.get("scenario"),
                            "annual_return": sc.get("annual_return"),
                            "year_offset": last.get("year_offset"),
                            "projected_value": last.get("projected_value"),
                        }
                    )
            return {
                "ok": True,
                "intent": "strategy_projection",
                "executed": force_refresh,
                "run_date": insights.get("run_date"),
                "projection": insights.get("projections"),
                "projection_summary": snippet,
                "message": f"Strategy projection ready for run_date={insights.get('run_date')}.",
            }

        top = insights.get("recommendations", [])[:8]
        return {
            "ok": True,
            "intent": "strategy_summary",
            "executed": force_refresh,
            "run_date": insights.get("run_date"),
            "counts": insights.get("counts", {}),
            "macro": insights.get("macro", {}),
            "items": top,
            "message": (
                f"Strategy run {insights.get('run_date')}: "
                f"TRIM={insights.get('counts', {}).get('TRIM', 0)}, "
                f"ADD={insights.get('counts', {}).get('ADD', 0)}, "
                f"HOLD={insights.get('counts', {}).get('HOLD', 0)}, "
                f"REVIEW={insights.get('counts', {}).get('REVIEW', 0)}, "
                f"WATCH_ADD={insights.get('counts', {}).get('WATCH_ADD', 0)}; "
                f"macro={str((insights.get('macro') or {}).get('regime', 'neutral')).upper()}."
            ),
        }

    if ("upload summary" in lower) or ("list uploads" in lower) or ("uploads summary" in lower):
        rows = conn.execute(
            """
            SELECT
              notes,
              COUNT(*) AS trades,
              COUNT(DISTINCT symbol) AS scrips,
              MIN(trade_date) AS from_date,
              MAX(trade_date) AS to_date
            FROM trades
            WHERE notes IS NOT NULL AND UPPER(notes) LIKE 'UPLOAD:%'
            GROUP BY notes
            ORDER BY to_date DESC, notes
            LIMIT 30
            """
        ).fetchall()
        batches = [dict(r) for r in rows]
        return {
            "ok": True,
            "intent": "upload_summary",
            "executed": False,
            "upload_batches": batches,
            "message": f"Upload batches found: {len(batches)}",
        }

    if "duplicate" in lower:
        if "cashflow" in lower or "cash flow" in lower or "cash" in lower:
            by_eid = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT external_entry_id, COUNT(*) AS duplicates
                    FROM cash_ledger
                    WHERE external_entry_id IS NOT NULL AND TRIM(external_entry_id) <> ''
                    GROUP BY external_entry_id
                    HAVING COUNT(*) > 1
                    ORDER BY duplicates DESC, external_entry_id
                    LIMIT 30
                    """
                ).fetchall()
            ]
            by_value = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT entry_date, LOWER(entry_type) AS entry_type, ROUND(amount,2) AS amount, COUNT(*) AS duplicates
                    FROM cash_ledger
                    GROUP BY entry_date, LOWER(entry_type), ROUND(amount,2)
                    HAVING COUNT(*) > 1
                    ORDER BY duplicates DESC, entry_date DESC
                    LIMIT 30
                    """
                ).fetchall()
            ]
            return {
                "ok": True,
                "intent": "cashflow_duplicates",
                "executed": False,
                "cashflow_id_duplicates": by_eid,
                "cashflow_value_duplicates": by_value,
                "message": f"Cashflow duplicates: ID groups={len(by_eid)}, value/date groups={len(by_value)}",
            }

        by_tid = [
            dict(r)
            for r in conn.execute(
                """
                SELECT external_trade_id, COUNT(*) AS duplicates
                FROM trades
                WHERE external_trade_id IS NOT NULL AND TRIM(external_trade_id) <> ''
                GROUP BY external_trade_id
                HAVING COUNT(*) > 1
                ORDER BY duplicates DESC, external_trade_id
                LIMIT 30
                """
            ).fetchall()
        ]
        by_value = [
            dict(r)
            for r in conn.execute(
                """
                SELECT
                  symbol,
                  side,
                  trade_date,
                  quantity,
                  ROUND(price, 4) AS price,
                  COUNT(*) AS duplicates
                FROM trades
                GROUP BY symbol, side, trade_date, quantity, ROUND(price, 4)
                HAVING COUNT(*) > 1
                ORDER BY duplicates DESC, trade_date DESC
                LIMIT 30
                """
            ).fetchall()
        ]
        return {
            "ok": True,
            "intent": "duplicates",
            "executed": False,
            "trade_id_duplicates": by_tid,
            "value_duplicates": by_value,
            "message": f"Duplicates check: Trade-ID groups={len(by_tid)}, value/date groups={len(by_value)}",
        }

    if ("top gainers" in lower) or ("top losers" in lower):
        metric_key = "day_change_pct" if "day" in lower else "total_return_pct"
        rows = conn.execute(
            """
            SELECT
              h.symbol,
              h.qty,
              h.invested,
              h.market_value,
              h.realized_pnl,
              h.unrealized_pnl,
              h.total_return_pct,
              COALESCE(lp.ltp, 0) AS ltp,
              COALESCE(lp.change_abs, 0) AS change_abs
            FROM holdings h
            LEFT JOIN latest_prices lp ON lp.symbol = h.symbol
            WHERE h.qty > 0
            """
        ).fetchall()
        prev_close_map = load_prev_close_map(conn, [r["symbol"] for r in rows])
        items = []
        for r in rows:
            item = dict(r)
            enrich_holding_metrics(item, conn=conn, prev_close_map=prev_close_map)
            items.append(item)
        reverse = "top gainers" in lower
        items.sort(key=lambda x: parse_float(x.get(metric_key), 0.0), reverse=reverse)
        top = [
            {
                "symbol": i["symbol"],
                metric_key: round(parse_float(i.get(metric_key), 0.0), 2),
                "day_pnl": round(parse_float(i.get("day_pnl"), 0.0), 2),
                "total_return_pct": round(parse_float(i.get("total_return_pct"), 0.0), 2),
            }
            for i in items[:10]
        ]
        label = "Top gainers" if reverse else "Top losers"
        return {
            "ok": True,
            "intent": "top_movers",
            "executed": False,
            "metric": metric_key,
            "items": top,
            "message": f"{label} by {metric_key}: {len(top)} rows",
        }

    note_pattern = extract_note_pattern_from_chat(text)
    destructive = any(k in lower for k in ("delete", "erase", "remove", "purge", "clear"))
    wants_preview = any(k in lower for k in ("preview", "dry run", "count only", "count", "find", "show", "list", "search"))
    if note_pattern:
        preview = preview_trades_by_note_pattern(conn, note_pattern)
        affected = int(preview["affected_trades"])
        symbols = preview["affected_symbols"]
        notes_examples = preview["notes_examples"]

        if affected == 0:
            return {
                "ok": True,
                "intent": "notes_query",
                "executed": False,
                "note_pattern": note_pattern,
                "affected_trades": 0,
                "affected_symbols": [],
                "notes_examples": [],
                "message": f"No trades found for notes pattern: {note_pattern}",
            }

        if (not destructive) or wants_preview:
            return {
                "ok": True,
                "intent": "notes_query",
                "executed": False,
                "dry_run": True,
                "note_pattern": note_pattern,
                "affected_trades": affected,
                "affected_symbols": symbols,
                "notes_examples": notes_examples,
                "message": f"Preview: {affected} trades across {len(symbols)} scrip(s) match notes pattern: {note_pattern}",
            }

        approval = create_agent_approval(
            conn,
            action_type="delete_by_notes",
            query_text=text,
            payload={"note_pattern": note_pattern},
            summary=f"Delete {affected} trades for notes pattern: {note_pattern}",
        )
        return {
            "ok": True,
            "intent": "approval_required",
            "executed": False,
            "approval_required": True,
            "approval": approval,
            "note_pattern": note_pattern,
            "affected_trades": affected,
            "affected_symbols": symbols,
            "notes_examples": notes_examples,
            "message": (
                f"Approval required. Request #{approval['id']} queued for deleting {affected} trade(s) "
                f"across {len(symbols)} scrip(s) with notes pattern: {note_pattern}"
            ),
        }

    return {
        "ok": True,
        "intent": "noop",
        "executed": False,
        "message": (
            "Query not recognized. Try:\n"
            "help\n"
            "how is cash balance calculated\n"
            "cashflow summary\n"
            "dividend summary\n"
            "portfolio summary\n"
            "strategy summary\n"
            "strategy projection\n"
            "intel summary\n"
            "chart summary\n"
            "chart signal for <symbol>\n"
            "policy impact for <symbol>\n"
            "fund flow links\n"
            "price status\n"
            "refresh gold price\n"
            "software performance status\n"
            "explain strategy reason for <symbol>\n"
            "upload summary\n"
            "cashflow duplicates\n"
            "show duplicates\n"
            "what is max_position_weight\n"
            "show pending approvals\n"
            "preview notes like \"upload:tradebook-OWY330.xlsx\"\n"
            "erase trades notes like \"upload:tradebook-OWY330.xlsx\""
        ),
    }

def enrich_holding_metrics(item, conn=None, prev_close_map=None):
    symbol = symbol_upper(item.get("symbol"))
    qty = parse_float(item.get("qty"), 0.0)
    ltp = parse_float(item.get("ltp"), 0.0)
    change_abs = parse_float(item.get("change_abs"), 0.0)
    if conn is not None and symbol:
        change_abs = resolve_effective_change_abs(conn, symbol, ltp, change_abs, prev_close_map=prev_close_map)
        item["change_abs"] = round(change_abs, 4)
    invested = parse_float(item.get("invested"), 0.0)
    realized = parse_float(item.get("realized_pnl"), 0.0)
    unrealized = parse_float(item.get("unrealized_pnl"), 0.0)

    prev_close = ltp - change_abs
    day_change_pct = (change_abs / prev_close * 100.0) if prev_close > 0 else 0.0
    day_pnl = qty * change_abs
    abs_pnl = realized + unrealized
    abs_pnl_pct = (abs_pnl / invested * 100.0) if invested > 0 else 0.0
    upl_pct = (unrealized / invested * 100.0) if invested > 0 else 0.0

    item["day_pnl"] = round(day_pnl, 2)
    item["day_change_pct"] = round(day_change_pct, 2)
    item["abs_pnl"] = round(abs_pnl, 2)
    item["abs_pnl_pct"] = round(abs_pnl_pct, 2)
    item["upl_pct"] = round(upl_pct, 2)
    return item


def _trade_dedupe_key(symbol, side, trade_date, amount):
    return (symbol.strip().upper(), side.strip().upper(), trade_date, round(float(amount), 2))


def _is_amount_or_price_close(amount_a, amount_b, price_a, price_b):
    amount_tol = max(20.0, 0.01 * max(abs(amount_a), abs(amount_b), 1.0))
    price_tol = max(2.0, 0.01 * max(abs(price_a), abs(price_b), 1.0))
    return (abs(amount_a - amount_b) <= amount_tol) or (abs(price_a - price_b) <= price_tol)


def _is_duplicate_trade(candidate, existing):
    cand_tid = normalize_external_trade_id(candidate.get("external_trade_id"))
    ex_tid = normalize_external_trade_id(existing.get("external_trade_id"))
    if cand_tid or ex_tid:
        return bool(cand_tid and ex_tid and cand_tid == ex_tid)
    if candidate["symbol"] != existing["symbol"]:
        return False
    if candidate["side"] != existing["side"]:
        return False
    if candidate["trade_date"] != existing["trade_date"]:
        return False
    if abs(candidate["quantity"] - existing["quantity"]) > 1e-6:
        return False
    # Exact amount always duplicate (same or repeated upload).
    if abs(candidate["amount"] - existing["amount"]) <= 0.01:
        return True
    # Cross-source near-match duplicate.
    if candidate.get("source") != existing.get("source"):
        return _is_amount_or_price_close(
            candidate["amount"], existing["amount"], candidate["price"], existing["price"]
        )
    return False


def _is_exact_duplicate_trade(candidate, existing):
    if candidate["symbol"] != existing["symbol"]:
        return False
    if candidate["side"] != existing["side"]:
        return False
    if candidate["trade_date"] != existing["trade_date"]:
        return False
    if abs(candidate["quantity"] - existing["quantity"]) > 1e-6:
        return False
    return abs(candidate["amount"] - existing["amount"]) <= 0.01


def dedupe_cross_source_trades():
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT id, symbol, UPPER(side) AS side, trade_date, quantity, price, amount, source, external_trade_id
            FROM trades
            ORDER BY trade_date, id
            """
        ).fetchall()
        by_trip = defaultdict(list)
        delete_ids = []
        for r in rows:
            cand = {
                "id": r["id"],
                "symbol": r["symbol"].upper(),
                "side": r["side"].upper(),
                "trade_date": r["trade_date"],
                "quantity": float(r["quantity"]),
                "price": float(r["price"]),
                "amount": float(r["amount"]),
                "source": r["source"],
                "external_trade_id": normalize_external_trade_id(r["external_trade_id"]),
            }
            key = (cand["symbol"], cand["side"], cand["trade_date"])
            dup = False
            for existing in by_trip[key]:
                if _is_duplicate_trade(cand, existing):
                    dup = True
                    delete_ids.append(cand["id"])
                    break
            if not dup:
                by_trip[key].append(cand)
        if delete_ids:
            qmarks = ",".join(["?"] * len(delete_ids))
            conn.execute(f"DELETE FROM trades WHERE id IN ({qmarks})", delete_ids)
            conn.commit()
        return len(delete_ids)


def normalize_ref_text(raw):
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def _is_duplicate_cashflow(candidate, existing):
    cand_eid = normalize_external_trade_id(candidate.get("external_entry_id"))
    ex_eid = normalize_external_trade_id(existing.get("external_entry_id"))
    if cand_eid or ex_eid:
        return bool(cand_eid and ex_eid and cand_eid == ex_eid)
    if candidate["entry_date"] != existing["entry_date"]:
        return False
    if candidate["entry_type"] != existing["entry_type"]:
        return False
    cref = normalize_ref_text(candidate.get("reference_text"))
    eref = normalize_ref_text(existing.get("reference_text"))
    if cref and eref and cref != eref:
        return False
    amt_tol = max(1.0, 0.005 * max(abs(candidate["amount"]), abs(existing["amount"]), 1.0))
    if abs(candidate["amount"] - existing["amount"]) <= amt_tol:
        return True
    return False


def dedupe_cross_source_cashflows():
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT id, entry_date, entry_type, amount, reference_text, external_entry_id, source
            FROM cash_ledger
            ORDER BY entry_date, id
            """
        ).fetchall()
        by_day = defaultdict(list)
        delete_ids = []
        for r in rows:
            cand = {
                "id": r["id"],
                "entry_date": str(r["entry_date"]),
                "entry_type": str(r["entry_type"]).strip().lower(),
                "amount": float(r["amount"]),
                "reference_text": normalize_ref_text(r["reference_text"]),
                "external_entry_id": normalize_external_trade_id(r["external_entry_id"]),
                "source": str(r["source"] or "cashflow_upload"),
            }
            dup = False
            for existing in by_day[cand["entry_date"]]:
                if _is_duplicate_cashflow(cand, existing):
                    dup = True
                    delete_ids.append(cand["id"])
                    break
            if not dup:
                by_day[cand["entry_date"]].append(cand)
        if delete_ids:
            qmarks = ",".join(["?"] * len(delete_ids))
            conn.execute(f"DELETE FROM cash_ledger WHERE id IN ({qmarks})", delete_ids)
            conn.commit()
        return len(delete_ids)


def normalize_dividend_symbol(raw):
    s = symbol_upper(raw)
    if not s:
        return ""
    s = s.replace("NSE:", "").replace("BSE:", "")
    if s.endswith(".NS") or s.endswith(".BO"):
        s = s.rsplit(".", 1)[0]
    s = re.sub(r"[^A-Z0-9&\-\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _resolve_dividend_symbol(conn, raw_symbol):
    norm = normalize_dividend_symbol(raw_symbol)
    if not norm:
        return None
    candidates = [norm]
    if norm.endswith("-EQ"):
        candidates.append(norm[:-3].strip())
    if " " in norm:
        candidates.append(norm.split(" ")[0].strip())
    seen = set()
    for cand in candidates:
        c = symbol_upper(cand)
        if not c or c in seen:
            continue
        seen.add(c)
        resolved = resolve_symbol(conn, c)
        if resolved:
            return resolved
    return symbol_upper(candidates[-1] if candidates else norm)


def _is_duplicate_dividend(candidate, existing):
    cand_eid = normalize_external_trade_id(candidate.get("external_entry_id"))
    ex_eid = normalize_external_trade_id(existing.get("external_entry_id"))
    if cand_eid or ex_eid:
        return bool(cand_eid and ex_eid and cand_eid == ex_eid)
    if candidate["entry_date"] != existing["entry_date"]:
        return False
    if symbol_upper(candidate["symbol"]) != symbol_upper(existing["symbol"]):
        return False
    cref = normalize_ref_text(candidate.get("reference_text"))
    eref = normalize_ref_text(existing.get("reference_text"))
    if cref and eref and cref != eref:
        return False
    amt_tol = max(1.0, 0.005 * max(abs(candidate["amount"]), abs(existing["amount"]), 1.0))
    return abs(candidate["amount"] - existing["amount"]) <= amt_tol


def dedupe_cross_source_dividends():
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT id, symbol, entry_date, amount, reference_text, external_entry_id, source
            FROM dividends
            ORDER BY entry_date, id
            """
        ).fetchall()
        by_day = defaultdict(list)
        delete_ids = []
        for r in rows:
            cand = {
                "id": r["id"],
                "symbol": symbol_upper(r["symbol"]),
                "entry_date": str(r["entry_date"]),
                "amount": float(r["amount"]),
                "reference_text": normalize_ref_text(r["reference_text"]),
                "external_entry_id": normalize_external_trade_id(r["external_entry_id"]),
                "source": str(r["source"] or "dividend_upload"),
            }
            dup = False
            for existing in by_day[cand["entry_date"]]:
                if _is_duplicate_dividend(cand, existing):
                    dup = True
                    delete_ids.append(cand["id"])
                    break
            if not dup:
                by_day[cand["entry_date"]].append(cand)
        if delete_ids:
            qmarks = ",".join(["?"] * len(delete_ids))
            conn.execute(f"DELETE FROM dividends WHERE id IN ({qmarks})", delete_ids)
            conn.commit()
        return len(delete_ids)


def _parse_cash_entry_type(raw):
    s = str(raw or "").strip().lower()
    if not s:
        return None
    if any(k in s for k in ("dp charge", "amc", "charge", "charges")):
        return "charge"
    if any(k in s for k in ("settlement", "investment", "book voucher")):
        return "investment"
    if any(k in s for k in ("trade credit", "sell credit", "sale credit")):
        return "trade_credit"
    if s in ("deposit", "payin", "pay in", "credit", "cr", "inflow", "in"):
        return "deposit"
    if s in ("withdrawal", "payout", "pay out", "debit", "dr", "outflow", "out"):
        return "withdrawal"
    if any(k in s for k in ("pay in", "payin", "deposit", "credit", "cr")):
        return "deposit"
    if any(k in s for k in ("pay out", "payout", "withdraw", "debit", "dr")):
        return "withdrawal"
    return None


def _classify_cashflow_entry(entry_type, signed_amount, payin, payout, amount_val, reference_text, voucher_type):
    ref = str(reference_text or "").strip().lower()
    vt = str(voucher_type or "").strip().lower()

    withdrawal_hints = (
        "instant payout",
        "withdrawal request",
        "funds transferred back",
        "bank payments",
        "bank payment",
        "payout",
        "withdrawal",
    )
    deposit_hints = (
        "funds added",
        "bank receipts",
        "bank receipt",
        "payin",
        "pay in",
        "funds received",
        "add funds",
    )
    charge_hints = (
        "dp charges",
        "dp charge",
        "amc for demat",
        "amc",
        "annual maintenance",
        "charge",
        "charges",
    )

    final_type = entry_type
    if any(h in ref for h in charge_hints):
        final_type = "charge"
    elif "net settlement for equity" in ref:
        if signed_amount < 0 or payout > 0:
            final_type = "investment"
        elif signed_amount > 0 or payin > 0:
            final_type = "trade_credit"
        else:
            final_type = "investment"
    elif any(h in ref for h in withdrawal_hints) or "bank payments" in vt:
        final_type = "withdrawal"
    elif any(h in ref for h in deposit_hints) or "bank receipts" in vt:
        final_type = "deposit"
    elif final_type not in ("deposit", "withdrawal", "investment", "trade_credit", "charge"):
        if signed_amount < 0:
            final_type = "investment"
        elif signed_amount > 0:
            final_type = "trade_credit"
        else:
            final_type = None

    amt = float(signed_amount)
    if abs(amt) <= 1e-9:
        amt = float(payin) - float(payout)
    if abs(amt) <= 1e-9:
        amt = float(amount_val)
    if abs(amt) <= 1e-9 or final_type is None:
        return None, 0.0

    if final_type in ("withdrawal", "investment", "charge"):
        amt = -abs(amt)
    elif final_type in ("deposit", "trade_credit"):
        amt = abs(amt)
    return final_type, amt


def import_cashflow_bytes(file_bytes, filename="cashflow.xlsx", replace_existing=False):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Funds Summary"] if "Funds Summary" in wb.sheetnames else wb.worksheets[0]

    req_aliases = {
        "entry_date": (
            "date",
            "entry date",
            "txn date",
            "transaction date",
            "value date",
            "posting date",
            "post date",
        ),
    }
    opt_aliases = {
        "entry_type": ("type", "entry type", "transaction type", "dr/cr", "debit/credit"),
        "amount": ("amount", "value", "net amount"),
        "payin": ("payin", "pay in", "credit", "deposit"),
        "payout": ("payout", "pay out", "debit", "withdrawal"),
        "voucher_type": ("voucher type", "voucher", "txn type"),
        "reference_text": ("remarks", "narration", "description", "reference", "ref", "notes", "particulars"),
        "external_entry_id": ("txn id", "transaction id", "utr", "voucher no", "reference no", "id"),
    }

    header_row = None
    header_map = {}
    for r in range(1, min(ws.max_row, 80) + 1):
        probe = {}
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            key = normalize_header_key(v)
            if key and key not in probe:
                probe[key] = c

        def find_col(aliases):
            for a in aliases:
                idx = probe.get(a)
                if idx:
                    return idx
            return None

        req_cols = {k: find_col(v) for k, v in req_aliases.items()}
        if all(req_cols.values()):
            header_row = r
            header_map = {**req_cols}
            for k, aliases in opt_aliases.items():
                header_map[k] = find_col(aliases)
            break

    if header_row is None:
        raise RuntimeError(
            "Unsupported cashflow format. Expected at least a Date column, with Amount or Payin/Payout columns."
        )

    col_date = header_map["entry_date"]
    col_type = header_map.get("entry_type")
    col_amount = header_map.get("amount")
    col_payin = header_map.get("payin")
    col_payout = header_map.get("payout")
    col_voucher_type = header_map.get("voucher_type")
    col_ref = header_map.get("reference_text")
    col_eid = header_map.get("external_entry_id")

    inserted = 0
    skipped_duplicates = 0
    skipped_id_duplicates = 0
    skipped_invalid = 0
    seen_eids = set()

    with db_connect() as conn:
        if replace_existing:
            conn.execute("DELETE FROM cash_ledger")
            conn.commit()
        existing_rows = conn.execute(
            """
            SELECT entry_date, entry_type, amount, reference_text, external_entry_id, source
            FROM cash_ledger
            ORDER BY entry_date, id
            """
        ).fetchall()
        existing_by_date = defaultdict(list)
        existing_eids = set()
        for r in existing_rows:
            ex = {
                "entry_date": str(r["entry_date"]),
                "entry_type": str(r["entry_type"]).strip().lower(),
                "amount": float(r["amount"]),
                "reference_text": normalize_ref_text(r["reference_text"]),
                "external_entry_id": normalize_external_trade_id(r["external_entry_id"]),
                "source": str(r["source"] or "cashflow_upload"),
            }
            existing_by_date[ex["entry_date"]].append(ex)
            if ex["external_entry_id"]:
                existing_eids.add(ex["external_entry_id"])

        for row in range(header_row + 1, ws.max_row + 1):
            d = parse_excel_date(ws.cell(row, col_date).value)
            if not d:
                skipped_invalid += 1
                continue

            raw_type = ws.cell(row, col_type).value if col_type else None
            entry_type = _parse_cash_entry_type(raw_type)
            voucher_type = ws.cell(row, col_voucher_type).value if col_voucher_type else None

            payin = parse_float(ws.cell(row, col_payin).value, 0.0) if col_payin else 0.0
            payout = parse_float(ws.cell(row, col_payout).value, 0.0) if col_payout else 0.0
            amount_val = parse_float(ws.cell(row, col_amount).value, 0.0) if col_amount else 0.0

            signed_amount = 0.0
            if payin > 0 or payout > 0:
                signed_amount = payin - payout
                if entry_type is None:
                    entry_type = "deposit" if signed_amount >= 0 else "withdrawal"
            elif amount_val != 0:
                if entry_type == "deposit":
                    signed_amount = abs(amount_val)
                elif entry_type == "withdrawal":
                    signed_amount = -abs(amount_val)
                else:
                    signed_amount = amount_val
                    entry_type = "deposit" if signed_amount >= 0 else "withdrawal"

            ext_eid = normalize_external_trade_id(
                ws.cell(row, col_eid).value if col_eid else None
            )
            if ext_eid:
                if ext_eid in seen_eids or ext_eid in existing_eids:
                    skipped_id_duplicates += 1
                    continue

            ref_txt = ""
            if col_ref:
                ref = ws.cell(row, col_ref).value
                ref_txt = "" if ref is None else str(ref).strip()

            entry_type, signed_amount = _classify_cashflow_entry(
                entry_type=entry_type,
                signed_amount=signed_amount,
                payin=payin,
                payout=payout,
                amount_val=amount_val,
                reference_text=ref_txt,
                voucher_type=voucher_type,
            )
            if abs(signed_amount) <= 1e-9 or entry_type not in (
                "deposit",
                "withdrawal",
                "investment",
                "trade_credit",
                "charge",
            ):
                skipped_invalid += 1
                continue

            cand = {
                "entry_date": d.isoformat(),
                "entry_type": entry_type,
                "amount": float(signed_amount),
                "reference_text": normalize_ref_text(ref_txt),
                "external_entry_id": ext_eid,
                "source": "cashflow_upload",
            }
            dup = False
            for ex in existing_by_date[cand["entry_date"]]:
                if _is_duplicate_cashflow(cand, ex):
                    dup = True
                    break
            if dup:
                skipped_duplicates += 1
                continue

            conn.execute(
                """
                INSERT INTO cash_ledger(entry_date, entry_type, amount, reference_text, external_entry_id, source)
                VALUES (?, ?, ?, ?, ?, 'cashflow_upload')
                """,
                (
                    cand["entry_date"],
                    cand["entry_type"],
                    cand["amount"],
                    ref_txt or f"upload:{filename}",
                    ext_eid,
                ),
            )
            existing_by_date[cand["entry_date"]].append(cand)
            if ext_eid:
                existing_eids.add(ext_eid)
                seen_eids.add(ext_eid)
            inserted += 1
        conn.commit()

    dedup_removed = dedupe_cross_source_cashflows()
    refresh_strategy_analytics(force=False)
    return {
        "inserted": inserted,
        "skipped_duplicates": skipped_duplicates,
        "skipped_id_duplicates": skipped_id_duplicates,
        "skipped_invalid": skipped_invalid,
        "cross_source_dedup_removed": dedup_removed,
        "entry_id_column_detected": bool(col_eid),
    }


def import_dividend_bytes(file_bytes, filename="dividends.xlsx", replace_existing=False):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Dividends"] if "Dividends" in wb.sheetnames else wb.worksheets[0]

    req_aliases = {
        "entry_date": (
            "date",
            "credit date",
            "payment date",
            "record date",
            "dividend date",
            "ex-date",
            "ex date",
            "txn date",
            "transaction date",
        ),
        "symbol": (
            "symbol",
            "scrip",
            "stock",
            "security",
            "trading symbol",
            "tradingsymbol",
            "instrument",
        ),
    }
    opt_aliases = {
        "amount": ("amount", "dividend amount", "net dividend amount", "net amount", "credit amount", "paid amount", "payout"),
        "quantity": ("quantity", "qty", "shares", "units", "holding qty"),
        "dividend_per_share": ("dividend per share", "per share", "dps", "rate"),
        "reference_text": ("remarks", "narration", "description", "reference", "ref", "notes", "particulars"),
        "external_entry_id": ("txn id", "transaction id", "reference no", "utr", "dividend id", "id", "voucher no"),
    }

    header_row = None
    header_map = {}
    for r in range(1, min(ws.max_row, 80) + 1):
        probe = {}
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            key = normalize_header_key(v)
            if key and key not in probe:
                probe[key] = c

        def find_col(aliases):
            for a in aliases:
                idx = probe.get(a)
                if idx:
                    return idx
            return None

        req_cols = {k: find_col(v) for k, v in req_aliases.items()}
        if all(req_cols.values()):
            header_row = r
            header_map = {**req_cols}
            for k, aliases in opt_aliases.items():
                header_map[k] = find_col(aliases)
            break

    if header_row is None:
        raise RuntimeError("Unsupported dividend format. Expected Date, Symbol, and Amount columns.")

    col_date = header_map["entry_date"]
    col_symbol = header_map["symbol"]
    col_amount = header_map.get("amount")
    col_qty = header_map.get("quantity")
    col_dps = header_map.get("dividend_per_share")
    col_ref = header_map.get("reference_text")
    col_eid = header_map.get("external_entry_id")

    inserted = 0
    skipped_duplicates = 0
    skipped_id_duplicates = 0
    skipped_invalid = 0
    seen_eids = set()

    with db_connect() as conn:
        if replace_existing:
            conn.execute("DELETE FROM dividends")
            conn.commit()
        existing_rows = conn.execute(
            """
            SELECT symbol, entry_date, amount, reference_text, external_entry_id, source
            FROM dividends
            ORDER BY entry_date, id
            """
        ).fetchall()
        existing_by_date = defaultdict(list)
        existing_eids = set()
        for r in existing_rows:
            ex = {
                "symbol": symbol_upper(r["symbol"]),
                "entry_date": str(r["entry_date"]),
                "amount": float(r["amount"]),
                "reference_text": normalize_ref_text(r["reference_text"]),
                "external_entry_id": normalize_external_trade_id(r["external_entry_id"]),
                "source": str(r["source"] or "dividend_upload"),
            }
            existing_by_date[ex["entry_date"]].append(ex)
            if ex["external_entry_id"]:
                existing_eids.add(ex["external_entry_id"])

        for row in range(header_row + 1, ws.max_row + 1):
            d = parse_excel_date(ws.cell(row, col_date).value)
            symbol = _resolve_dividend_symbol(conn, ws.cell(row, col_symbol).value)
            if not d or not symbol:
                skipped_invalid += 1
                continue

            amount = parse_float(ws.cell(row, col_amount).value, 0.0) if col_amount else 0.0
            if amount <= 0 and col_qty and col_dps:
                qty = parse_float(ws.cell(row, col_qty).value, 0.0)
                dps = parse_float(ws.cell(row, col_dps).value, 0.0)
                amount = qty * dps
            amount = abs(parse_float(amount, 0.0))
            if amount <= 0:
                skipped_invalid += 1
                continue

            ext_eid = normalize_external_trade_id(ws.cell(row, col_eid).value if col_eid else None)
            if ext_eid and (ext_eid in seen_eids or ext_eid in existing_eids):
                skipped_id_duplicates += 1
                continue

            ref_txt = ""
            if col_ref:
                ref = ws.cell(row, col_ref).value
                ref_txt = "" if ref is None else str(ref).strip()

            cand = {
                "symbol": symbol_upper(symbol),
                "entry_date": d.isoformat(),
                "amount": float(amount),
                "reference_text": normalize_ref_text(ref_txt),
                "external_entry_id": ext_eid,
                "source": "dividend_upload",
            }
            dup = False
            for ex in existing_by_date[cand["entry_date"]]:
                if _is_duplicate_dividend(cand, ex):
                    dup = True
                    break
            if dup:
                skipped_duplicates += 1
                continue

            inst = conn.execute(
                "SELECT symbol FROM instruments WHERE UPPER(symbol) = ?",
                (symbol_upper(symbol),),
            ).fetchone()
            if inst is None:
                conn.execute(
                    """
                    INSERT INTO instruments(exchange, symbol, name, active, feed_code, price_source, asset_class)
                    VALUES ('NSE', ?, ?, 1, ?, 'exchange_api', ?)
                    """,
                    (symbol, symbol, symbol, infer_asset_class(symbol=symbol, name=symbol)),
                )

            conn.execute(
                """
                INSERT INTO dividends(symbol, entry_date, amount, reference_text, external_entry_id, source, created_at)
                VALUES (?, ?, ?, ?, ?, 'dividend_upload', ?)
                """,
                (
                    symbol_upper(symbol),
                    cand["entry_date"],
                    cand["amount"],
                    ref_txt or f"upload:{filename}",
                    ext_eid,
                    now_iso(),
                ),
            )
            existing_by_date[cand["entry_date"]].append(cand)
            if ext_eid:
                existing_eids.add(ext_eid)
                seen_eids.add(ext_eid)
            inserted += 1
        conn.commit()

    dedup_removed = dedupe_cross_source_dividends()
    recompute_holdings_and_signals(force_strategy=False)
    refresh_strategy_analytics(force=False)
    return {
        "inserted": inserted,
        "skipped_duplicates": skipped_duplicates,
        "skipped_id_duplicates": skipped_id_duplicates,
        "skipped_invalid": skipped_invalid,
        "cross_source_dedup_removed": dedup_removed,
        "entry_id_column_detected": bool(col_eid),
    }


def import_from_excel(xlsx_path):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    with db_connect() as conn:
        clear_core_tables(conn)

        portfolio = wb["Portfolio"] if "Portfolio" in wb.sheetnames else None
        if portfolio is None:
            raise RuntimeError("Portfolio sheet not found in workbook.")

        symbol_rows = []
        for row in range(4, 220):
            symbol = _clean_symbol(portfolio.cell(row, 3).value)  # C
            if not symbol:
                continue
            symbol_rows.append(symbol)

        seen = set()
        for symbol in symbol_rows:
            if symbol in seen:
                continue
            seen.add(symbol)
            conn.execute(
                """
                INSERT INTO instruments(exchange, symbol, name, active, feed_code, price_source, asset_class)
                VALUES ('NSE', ?, ?, 1, ?, 'exchange_api', ?)
                """,
                (symbol, symbol, symbol, infer_asset_class(symbol=symbol, name=symbol)),
            )

        for symbol in seen:
            if symbol not in wb.sheetnames:
                continue
            ws = wb[symbol]
            for row in range(16, 535):
                b_date = parse_excel_date(ws.cell(row, 1).value)  # A
                b_qty = parse_float(ws.cell(row, 2).value, 0.0)  # B
                b_price = parse_float(ws.cell(row, 3).value, 0.0)  # C
                if b_date and b_qty > 0 and b_price > 0:
                    amount = b_qty * b_price
                    conn.execute(
                        """
                        INSERT INTO trades(symbol, side, trade_date, quantity, price, amount, source, notes)
                        VALUES (?, 'BUY', ?, ?, ?, ?, 'excel', 'ledger_buy')
                        """,
                        (symbol, b_date.isoformat(), b_qty, b_price, amount),
                    )

                s_date = parse_excel_date(ws.cell(row, 5).value)  # E
                s_qty = parse_float(ws.cell(row, 6).value, 0.0)  # F
                s_price = parse_float(ws.cell(row, 8).value, 0.0)  # H
                if s_price <= 0:
                    s_amount = parse_float(ws.cell(row, 9).value, 0.0)  # I
                    if s_qty > 0:
                        s_price = s_amount / s_qty
                if s_date and s_qty > 0 and s_price > 0:
                    amount = s_qty * s_price
                    conn.execute(
                        """
                        INSERT INTO trades(symbol, side, trade_date, quantity, price, amount, source, notes)
                        VALUES (?, 'SELL', ?, ?, ?, ?, 'excel', 'ledger_sell')
                        """,
                        (symbol, s_date.isoformat(), s_qty, s_price, amount),
                    )

        # Cash ledger is intentionally not seeded from legacy workbook sheets.
        # Cash balance is maintained from dedicated payin/payout uploads only.

        conn.commit()

    refresh_latest_prices_from_exchange(max_runtime_sec=12)
    recompute_holdings_and_signals()


def import_tradebook_bytes(file_bytes, filename="tradebook.xlsx", collect_skipped=False, max_skipped_details=300):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Equity"] if "Equity" in wb.sheetnames else wb.worksheets[0]

    # Find header row dynamically to support broker export shifts and Kite raw exports.
    req_aliases = {
        "symbol": ("symbol", "trading symbol", "tradingsymbol"),
        "trade_date": ("trade date", "date", "trade_date"),
        "trade_type": ("trade type", "transaction type", "side", "type"),
        "quantity": ("quantity", "qty", "filled quantity", "traded quantity"),
        "price": ("price", "average price", "avg price", "trade price"),
    }
    opt_aliases = {
        "trade_id": ("trade id", "tradeid", "order id", "orderid", "exchange trade id", "exchange order id"),
        "exchange": ("exchange",),
        "segment": ("segment",),
    }

    header_row = None
    header_map = {}
    for r in range(1, min(ws.max_row, 60) + 1):
        probe = {}
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            key = normalize_header_key(v)
            if key and key not in probe:
                probe[key] = c

        def find_col(aliases):
            for a in aliases:
                idx = probe.get(a)
                if idx:
                    return idx
            return None

        req_cols = {
            k: find_col(v)
            for k, v in req_aliases.items()
        }
        if all(req_cols.values()):
            header_row = r
            header_map = {**req_cols}
            for k, aliases in opt_aliases.items():
                header_map[k] = find_col(aliases)
            break
    if header_row is None:
        raise RuntimeError(
            "Unsupported tradebook format. Expected at least Symbol, Trade Date, Trade Type, Quantity, Price (Kite raw format supported)."
        )

    col_symbol = header_map["symbol"]
    col_date = header_map["trade_date"]
    col_side = header_map["trade_type"]
    col_qty = header_map["quantity"]
    col_price = header_map["price"]
    col_trade_id = header_map.get("trade_id")
    col_exchange = header_map.get("exchange")
    col_segment = header_map.get("segment")

    inserted = 0
    skipped_duplicates = 0
    skipped_trade_id_duplicates = 0
    skipped_invalid = 0
    skipped_non_equity = 0
    new_symbols = set()
    skipped_items = []
    skipped_items_truncated = False

    def append_skipped(reason, row_no, symbol, trade_date, side, qty, price, ext_trade_id, message):
        nonlocal skipped_items_truncated
        if not collect_skipped:
            return
        if len(skipped_items) >= max_skipped_details:
            skipped_items_truncated = True
            return
        skipped_items.append(
            {
                "reason": reason,
                "row_number": int(row_no),
                "symbol": symbol_upper(symbol) if symbol else "",
                "trade_date": trade_date or "",
                "side": str(side or "").upper(),
                "quantity": round(parse_float(qty, 0.0), 6),
                "price": round(parse_float(price, 0.0), 6),
                "external_trade_id": normalize_external_trade_id(ext_trade_id),
                "message": message,
            }
        )

    with db_connect() as conn:
        existing_by_trip = defaultdict(list)
        existing_trade_ids = {
            normalize_external_trade_id(r["external_trade_id"])
            for r in conn.execute(
                """
                SELECT external_trade_id
                FROM trades
                WHERE external_trade_id IS NOT NULL AND TRIM(external_trade_id) <> ''
                """
            ).fetchall()
        }
        existing_trade_ids.discard(None)
        for r in conn.execute(
            """
            SELECT symbol, UPPER(side) AS side, trade_date, quantity, price, amount, source, external_trade_id
            FROM trades
            """
        ).fetchall():
            ex = {
                "symbol": r["symbol"].upper(),
                "side": r["side"].upper(),
                "trade_date": r["trade_date"],
                "quantity": float(r["quantity"]),
                "price": float(r["price"]),
                "amount": float(r["amount"]),
                "source": r["source"],
                "external_trade_id": normalize_external_trade_id(r["external_trade_id"]),
            }
            existing_by_trip[(ex["symbol"], ex["side"], ex["trade_date"])].append(ex)

        seen_trade_ids = set()
        for row in range(header_row + 1, ws.max_row + 1):
            symbol = _clean_symbol(ws.cell(row, col_symbol).value)
            if not symbol:
                continue
            symbol = symbol.upper()
            d = parse_excel_date(ws.cell(row, col_date).value)
            side_raw = str(ws.cell(row, col_side).value or "").strip().lower()
            qty = parse_float(ws.cell(row, col_qty).value, 0.0)
            price = parse_float(ws.cell(row, col_price).value, 0.0)
            if col_exchange:
                ex = str(ws.cell(row, col_exchange).value or "").strip().upper()
                if ex and ex not in ("NSE", "BSE"):
                    skipped_non_equity += 1
                    append_skipped(
                        "non_equity",
                        row,
                        symbol,
                        d.isoformat() if d else "",
                        side_raw,
                        qty,
                        price,
                        ws.cell(row, col_trade_id).value if col_trade_id else None,
                        f"Exchange {ex} is not NSE/BSE.",
                    )
                    continue
            if col_segment:
                seg = str(ws.cell(row, col_segment).value or "").strip().upper()
                if seg and seg not in ("EQ", "EQUITY"):
                    skipped_non_equity += 1
                    append_skipped(
                        "non_equity",
                        row,
                        symbol,
                        d.isoformat() if d else "",
                        side_raw,
                        qty,
                        price,
                        ws.cell(row, col_trade_id).value if col_trade_id else None,
                        f"Segment {seg} is not equity.",
                    )
                    continue
            if not d or qty <= 0 or price <= 0:
                skipped_invalid += 1
                append_skipped(
                    "invalid",
                    row,
                    symbol,
                    d.isoformat() if d else "",
                    side_raw,
                    qty,
                    price,
                    ws.cell(row, col_trade_id).value if col_trade_id else None,
                    "Missing/invalid date, quantity, or price.",
                )
                continue
            if side_raw not in ("buy", "sell"):
                skipped_invalid += 1
                append_skipped(
                    "invalid",
                    row,
                    symbol,
                    d.isoformat(),
                    side_raw,
                    qty,
                    price,
                    ws.cell(row, col_trade_id).value if col_trade_id else None,
                    "Side must be BUY or SELL.",
                )
                continue

            ext_trade_id = normalize_external_trade_id(ws.cell(row, col_trade_id).value) if col_trade_id else None
            side = side_raw.upper()
            trade_date = d.isoformat()
            amount = qty * price
            cand = {
                "symbol": symbol,
                "side": side,
                "trade_date": trade_date,
                "quantity": float(qty),
                "price": float(price),
                "amount": float(amount),
                "source": "tradebook_upload",
                "external_trade_id": ext_trade_id,
            }
            if ext_trade_id:
                if ext_trade_id in existing_trade_ids or ext_trade_id in seen_trade_ids:
                    skipped_duplicates += 1
                    skipped_trade_id_duplicates += 1
                    append_skipped(
                        "duplicate_trade_id",
                        row,
                        symbol,
                        trade_date,
                        side,
                        qty,
                        price,
                        ext_trade_id,
                        "Duplicate Trade ID detected.",
                    )
                    continue

            trip = (symbol, side, trade_date)
            if not ext_trade_id and any(_is_duplicate_trade(cand, ex) for ex in existing_by_trip.get(trip, [])):
                skipped_duplicates += 1
                append_skipped(
                    "duplicate_value_date",
                    row,
                    symbol,
                    trade_date,
                    side,
                    qty,
                    price,
                    ext_trade_id,
                    "Duplicate by symbol/side/date/qty/price.",
                )
                continue

            inst = conn.execute(
                "SELECT symbol FROM instruments WHERE UPPER(symbol) = ?",
                (symbol_upper(symbol),),
            ).fetchone()
            if inst is None:
                conn.execute(
                    """
                    INSERT INTO instruments(exchange, symbol, name, active, feed_code, price_source, asset_class)
                    VALUES ('NSE', ?, ?, 1, ?, 'exchange_api', ?)
                    """,
                    (symbol, symbol, symbol, infer_asset_class(symbol=symbol, name=symbol)),
                )
                new_symbols.add(symbol)

            conn.execute(
                """
                INSERT INTO trades(symbol, side, trade_date, quantity, price, amount, external_trade_id, source, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'tradebook_upload', ?)
                """,
                (symbol, side, trade_date, qty, price, amount, ext_trade_id, f"upload:{filename}"),
            )
            existing_by_trip[trip].append(cand)
            if ext_trade_id:
                existing_trade_ids.add(ext_trade_id)
                seen_trade_ids.add(ext_trade_id)
            inserted += 1

        conn.commit()

    removed = dedupe_cross_source_trades()
    recompute_holdings_and_signals()
    return {
        "inserted": inserted,
        "skipped_duplicates": skipped_duplicates,
        "skipped_trade_id_duplicates": skipped_trade_id_duplicates,
        "skipped_invalid": skipped_invalid,
        "skipped_non_equity": skipped_non_equity,
        "new_symbols_added": sorted(new_symbols),
        "cross_source_dedup_removed": removed,
        "trade_id_column_detected": bool(col_trade_id),
        "skipped_items": skipped_items,
        "skipped_items_truncated": skipped_items_truncated,
    }


def parse_history_date(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %b %y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def upsert_market_daily_prices(rows):
    if not rows:
        return 0
    normalized = []
    for r in rows:
        symbol = symbol_upper(r[0])
        price_date = str(r[1] or "").strip()
        close = parse_float(r[2], 0.0)
        source = str(r[3] or "unknown")
        fetched_at = str(r[4] or now_iso())
        if not symbol or not price_date or close <= 0:
            continue
        normalized.append((symbol, price_date, close, source, fetched_at))
    if not normalized:
        return 0
    with market_db_connect() as conn:
        conn.executemany(
            """
            INSERT INTO daily_prices(symbol, price_date, close, source, fetched_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(symbol, price_date) DO UPDATE SET
              close=excluded.close,
              source=excluded.source,
              fetched_at=excluded.fetched_at
            """,
            normalized,
        )
        conn.commit()
    return len(normalized)


def insert_market_daily_prices_if_missing(rows):
    if not rows:
        return 0
    normalized = []
    for r in rows:
        symbol = symbol_upper(r[0])
        price_date = str(r[1] or "").strip()
        close = parse_float(r[2], 0.0)
        source = str(r[3] or "unknown")
        fetched_at = str(r[4] or now_iso())
        if not symbol or not price_date or close <= 0:
            continue
        normalized.append((symbol, price_date, close, source, fetched_at))
    if not normalized:
        return 0
    with market_db_connect() as conn:
        conn.executemany(
            """
            INSERT OR IGNORE INTO daily_prices(symbol, price_date, close, source, fetched_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            normalized,
        )
        conn.commit()
        try:
            return int(conn.total_changes)
        except Exception:
            return 0


def delete_market_history_symbols(symbols):
    syms = sorted({symbol_upper(s) for s in (symbols or []) if symbol_upper(s)})
    if not syms:
        return 0
    placeholders = ",".join(["?"] * len(syms))
    with market_db_connect() as conn:
        n = conn.execute(
            f"DELETE FROM daily_prices WHERE UPPER(symbol) IN ({placeholders})",
            syms,
        ).rowcount
        conn.commit()
    return n


def history_symbol_universe(conn):
    inst_rows = conn.execute(
        "SELECT UPPER(symbol) AS symbol, exchange, feed_code FROM instruments"
    ).fetchall()
    inst_map = {
        symbol_upper(r["symbol"]): {
            "exchange": str(r["exchange"] or "NSE").upper(),
            "feed_code": str(r["feed_code"] or r["symbol"] or "").strip(),
        }
        for r in inst_rows
    }
    trade_rows = conn.execute(
        """
        SELECT UPPER(symbol) AS symbol, MIN(trade_date) AS first_trade_date
        FROM trades
        GROUP BY UPPER(symbol)
        ORDER BY UPPER(symbol)
        """
    ).fetchall()
    out = []
    seen = set()
    for r in trade_rows:
        sym = symbol_upper(r["symbol"])
        if not sym:
            continue
        seen.add(sym)
        meta = inst_map.get(sym, {})
        out.append(
            {
                "symbol": sym,
                "exchange": str(meta.get("exchange", "NSE")).upper(),
                "feed_code": str(meta.get("feed_code", sym) or sym),
                "first_trade_date": str(r["first_trade_date"] or ""),
            }
        )
    for sym, meta in inst_map.items():
        if sym in seen:
            continue
        out.append(
            {
                "symbol": sym,
                "exchange": str(meta.get("exchange", "NSE")).upper(),
                "feed_code": str(meta.get("feed_code", sym) or sym),
                "first_trade_date": "",
            }
        )
    out.sort(key=lambda x: x["symbol"])
    return out


def aggregate_price_ticks_daily(conn, symbol, from_s=None, to_s=None):
    where = ["UPPER(symbol) = ?"]
    params = [symbol_upper(symbol)]
    if from_s:
        where.append("SUBSTR(fetched_at,1,10) >= ?")
        params.append(str(from_s))
    if to_s:
        where.append("SUBSTR(fetched_at,1,10) <= ?")
        params.append(str(to_s))
    rows = conn.execute(
        f"""
        SELECT fetched_at, ltp
        FROM price_ticks
        WHERE {' AND '.join(where)} AND ltp > 0
        ORDER BY fetched_at
        """,
        params,
    ).fetchall()
    by_day = {}
    for r in rows:
        day_s = str(r["fetched_at"] or "")[:10]
        ltp = parse_float(r["ltp"], 0.0)
        if not day_s or ltp <= 0:
            continue
        by_day[day_s] = ltp
    return [(symbol_upper(symbol), d, px) for d, px in sorted(by_day.items())]


def upsert_today_market_history_from_latest(symbols=None):
    today_s = dt.date.today().isoformat()
    stamp = now_iso()
    gold_symbols = set()
    with db_connect() as conn:
        if symbols:
            syms = sorted({symbol_upper(s) for s in symbols if symbol_upper(s)})
            if not syms:
                return 0
            placeholders = ",".join(["?"] * len(syms))
            rows = conn.execute(
                f"SELECT symbol, ltp FROM latest_prices WHERE UPPER(symbol) IN ({placeholders})",
                syms,
            ).fetchall()
            inst_rows = conn.execute(
                f"""
                SELECT symbol, COALESCE(asset_class, 'EQUITY') AS asset_class
                FROM instruments
                WHERE UPPER(symbol) IN ({placeholders})
                """,
                syms,
            ).fetchall()
        else:
            rows = conn.execute("SELECT symbol, ltp FROM latest_prices").fetchall()
            inst_rows = conn.execute(
                "SELECT symbol, COALESCE(asset_class, 'EQUITY') AS asset_class FROM instruments"
            ).fetchall()
        for r in inst_rows:
            sym = symbol_upper(r["symbol"])
            if not sym:
                continue
            ac = normalize_asset_class(
                r["asset_class"],
                fallback=infer_asset_class(symbol=sym, name=sym),
            )
            if ac == ASSET_CLASS_GOLD:
                gold_symbols.add(sym)
    payload = []
    for r in rows:
        ltp = parse_float(r["ltp"], 0.0)
        if ltp <= 0:
            continue
        payload.append((symbol_upper(r["symbol"]), today_s, ltp, "latest_prices", stamp))
    if not payload:
        return 0
    gold_payload = [x for x in payload if symbol_upper(x[0]) in gold_symbols]
    non_gold_payload = [x for x in payload if symbol_upper(x[0]) not in gold_symbols]
    inserted_gold = insert_market_daily_prices_if_missing(gold_payload)
    upserted_non_gold = upsert_market_daily_prices(non_gold_payload)
    return int(inserted_gold + upserted_non_gold)


def _http_json(opener, url, headers, timeout=4):
    req = urllib.request.Request(url, headers=headers)
    with opener.open(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _http_json_post(opener, url, headers, payload, timeout=5):
    data = json.dumps(payload).encode("utf-8")
    req_headers = dict(headers or {})
    if "Content-Type" not in req_headers:
        req_headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=req_headers, method="POST")
    with opener.open(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _http_text(opener, url, headers, timeout=4):
    req = urllib.request.Request(url, headers=headers)
    with opener.open(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="ignore")


class MarketDataClient:
    def __init__(self):
        self.nse_opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(http.cookiejar.CookieJar())
        )
        self.bse_opener = urllib.request.build_opener()
        self.yahoo_opener = urllib.request.build_opener()
        self.google_opener = urllib.request.build_opener()
        self.nse_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://www.nseindia.com/",
        }
        self.bse_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://www.bseindia.com/",
            "Origin": "https://www.bseindia.com",
        }
        self.yahoo_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://finance.yahoo.com/",
        }
        self.google_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.google.com/",
        }
        self.screener_opener = urllib.request.build_opener()
        self.screener_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.screener.in/",
        }
        self.trendlyne_opener = urllib.request.build_opener()
        self.trendlyne_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://trendlyne.com/",
        }
        self.cnbc_opener = urllib.request.build_opener()
        self.cnbc_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.cnbctv18.com/",
        }
        self.stock_nse_india_opener = urllib.request.build_opener()
        self.stock_nse_india_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://www.nseindia.com/",
        }
        self.gold_opener = urllib.request.build_opener()
        self.gold_headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.thehindubusinessline.com/",
        }
        self.stock_nse_india_base = str(
            os.environ.get("STOCK_NSE_INDIA_API_BASE", STOCK_NSE_INDIA_API_BASE_DEFAULT)
        ).strip()
        self._nsetools_client = None
        self._nse_bootstrapped = False

    def _bootstrap_nse(self):
        if self._nse_bootstrapped:
            return
        req = urllib.request.Request("https://www.nseindia.com/", headers=self.nse_headers)
        with self.nse_opener.open(req, timeout=4):
            pass
        self._nse_bootstrapped = True

    def fetch_nse_quote(self, symbol):
        self._bootstrap_nse()
        q = urllib.parse.quote(symbol)
        url = f"https://www.nseindia.com/api/quote-equity?symbol={q}"
        data = _http_json(self.nse_opener, url, self.nse_headers)
        pinfo = data.get("priceInfo", {})
        ltp = parse_float(pinfo.get("lastPrice"), 0.0)
        change_abs = parse_float(pinfo.get("change"), 0.0)
        return ltp, change_abs

    def fetch_nse_history(self, symbol, from_date, to_date):
        self._bootstrap_nse()
        if isinstance(from_date, dt.date):
            from_s = from_date.strftime("%d-%m-%Y")
        else:
            from_s = (parse_history_date(from_date) or ist_now().date()).strftime("%d-%m-%Y")
        if isinstance(to_date, dt.date):
            to_s = to_date.strftime("%d-%m-%Y")
        else:
            to_s = (parse_history_date(to_date) or ist_now().date()).strftime("%d-%m-%Y")
        params = {
            "symbol": str(symbol).upper(),
            "series": '["EQ"]',
            "from": from_s,
            "to": to_s,
        }
        url = "https://www.nseindia.com/api/historical/cm/equity?" + urllib.parse.urlencode(params)
        data = _http_json(self.nse_opener, url, self.nse_headers)
        rows = data.get("data", []) or []
        out = []
        for r in rows:
            d = parse_history_date(
                r.get("CH_TIMESTAMP")
                or r.get("date")
                or r.get("Date")
                or r.get("priceDate")
            )
            px = parse_float(
                r.get("CH_CLOSING_PRICE")
                or r.get("CLOSE")
                or r.get("close")
                or r.get("CH_CLOSE_PRICE"),
                0.0,
            )
            if d and px > 0:
                out.append((d.isoformat(), px))
        out.sort(key=lambda x: x[0])
        return out

    def fetch_bse_quote(self, scrip_code):
        code = str(scrip_code).strip()
        if not code:
            return 0.0, 0.0
        url = f"https://api.bseindia.com/BseIndiaAPI/api/getScripHeaderData/w?scripcode={urllib.parse.quote(code)}"
        data = _http_json(self.bse_opener, url, self.bse_headers)
        header = data.get("Header", {})
        ltp = parse_float(header.get("LTP"), 0.0)
        change_abs = parse_float(header.get("Chg"), 0.0)
        return ltp, change_abs

    def fetch_quote(self, exchange, symbol, feed_code):
        ex = (exchange or "NSE").upper()
        if ex == "BSE":
            return self.fetch_bse_quote(feed_code or symbol)
        return self.fetch_nse_quote(symbol)

    def fetch_yahoo_quote(self, symbol, exchange="NSE"):
        sym = str(symbol or "").strip().upper()
        if not sym:
            return 0.0, 0.0
        suffixes = [".NS", ".BO"]
        if str(exchange or "").upper() == "BSE":
            suffixes = [".BO", ".NS"]
        for sfx in suffixes:
            quote_sym = f"{sym}{sfx}"
            url = f"https://query1.finance.yahoo.com/v7/finance/quote?symbols={urllib.parse.quote(quote_sym)}"
            try:
                data = _http_json(self.yahoo_opener, url, self.yahoo_headers)
            except Exception:
                continue
            rows = (((data or {}).get("quoteResponse") or {}).get("result") or [])
            if not rows:
                continue
            row = rows[0]
            ltp = parse_float(row.get("regularMarketPrice"), 0.0)
            change_abs = parse_float(row.get("regularMarketChange"), 0.0)
            if ltp > 0:
                return ltp, change_abs
        return 0.0, 0.0

    def fetch_screener_quote(self, symbol):
        sym = str(symbol).strip().upper()
        if not sym:
            return 0.0, 0.0
        urls = [
            f"https://www.screener.in/company/{urllib.parse.quote(sym)}/",
            f"https://www.screener.in/company/{urllib.parse.quote(sym)}/consolidated/",
        ]

        # Use Screener search API to discover the right slug first.
        try:
            search_url = f"https://www.screener.in/api/company/search/?q={urllib.parse.quote(sym)}"
            req = urllib.request.Request(
                search_url,
                headers={**self.screener_headers, "Accept": "application/json,text/plain,*/*"},
            )
            with self.screener_opener.open(req, timeout=4) as resp:
                arr = json.loads(resp.read().decode("utf-8", errors="ignore"))
            if isinstance(arr, list):
                for it in arr[:4]:
                    rel = str((it or {}).get("url") or "").strip()
                    if not rel:
                        continue
                    full = urllib.parse.urljoin("https://www.screener.in", rel)
                    if full not in urls:
                        urls.insert(0, full)
        except Exception:
            pass

        price_patterns = [
            r'Current\s+Price[\s\S]{0,260}?<span[^>]*class="number"[^>]*>\s*([0-9][0-9,]*\.?[0-9]*)\s*</span>',
            r'current_price["\']?\s*[:=]\s*([0-9][0-9,]*\.?[0-9]*)',
            r'"currentPrice"\s*:\s*"?(?:INR\s*)?([0-9][0-9,]*\.?[0-9]*)"?',
        ]
        for u in urls:
            try:
                html = _http_text(self.screener_opener, u, self.screener_headers, timeout=4)
            except Exception:
                continue
            for pat in price_patterns:
                m = re.search(pat, html, flags=re.IGNORECASE)
                if m:
                    ltp = parse_float(m.group(1), 0.0)
                    if ltp > 0:
                        return ltp, 0.0
        return 0.0, 0.0

    def fetch_google_quote(self, symbol, exchange="NSE"):
        sym = str(symbol or "").strip().upper()
        ex = str(exchange or "NSE").strip().upper()
        if not sym:
            return 0.0, 0.0

        url = f"https://www.google.com/finance/quote/{urllib.parse.quote(sym)}:{urllib.parse.quote(ex)}"
        try:
            html = _http_text(self.google_opener, url, self.google_headers, timeout=4)
        except Exception:
            return 0.0, 0.0

        # Guard: ensure page canonical matches requested symbol/exchange.
        cm = re.search(r'<link rel="canonical" href="([^"]+)"', html, flags=re.IGNORECASE)
        if cm:
            canon = str(cm.group(1) or "")
            if f"/FINANCE/QUOTE/{sym}:{ex}" not in canon.upper():
                return 0.0, 0.0

        # Strict extraction to avoid picking unrelated index values.
        pm = re.search(r'data-last-price\s*=\s*"([0-9][0-9,]*\.?[0-9]*)"', html, flags=re.IGNORECASE)
        if not pm:
            return 0.0, 0.0
        ltp = parse_float(pm.group(1), 0.0)
        if ltp <= 0:
            return 0.0, 0.0

        change_abs = 0.0
        cm1 = re.search(r'"priceChange"\s*:\s*\{\s*"raw"\s*:\s*([+\-]?[0-9][0-9,]*\.?[0-9]*)', html, flags=re.IGNORECASE)
        if cm1:
            change_abs = parse_float(cm1.group(1), 0.0)
        return ltp, change_abs

    def fetch_trendlyne_quote(self, symbol):
        sym = str(symbol or "").strip().upper()
        if not sym:
            return 0.0, 0.0
        urls = [
            f"https://trendlyne.com/equity/{urllib.parse.quote(sym)}/",
            f"https://trendlyne.com/equity/?q={urllib.parse.quote(sym)}",
        ]
        patterns = [
            r'"last_price"\s*:\s*([0-9][0-9,]*\.?[0-9]*)',
            r'"ltp"\s*:\s*([0-9][0-9,]*\.?[0-9]*)',
            r"LTP[^0-9]{1,20}([0-9][0-9,]*\.?[0-9]*)",
        ]
        for url in urls:
            try:
                html = _http_text(self.trendlyne_opener, url, self.trendlyne_headers, timeout=4)
            except Exception:
                continue
            for pat in patterns:
                m = re.search(pat, html, flags=re.IGNORECASE)
                if m:
                    return parse_float(m.group(1), 0.0), 0.0
        return 0.0, 0.0

    def fetch_cnbc_quote(self, symbol):
        sym = str(symbol or "").strip().upper()
        if not sym:
            return 0.0, 0.0
        urls = [
            f"https://www.cnbctv18.com/search/?q={urllib.parse.quote(sym)}",
            f"https://www.cnbctv18.com/market/stocks/{urllib.parse.quote(sym.lower())}-share-price.htm",
        ]
        patterns = [
            r'"lastPrice"\s*:\s*"?([0-9][0-9,]*\.?[0-9]*)"?',
            r'"ltp"\s*:\s*"?([0-9][0-9,]*\.?[0-9]*)"?',
            r"LTP[^0-9]{1,20}([0-9][0-9,]*\.?[0-9]*)",
        ]
        for url in urls:
            try:
                html = _http_text(self.cnbc_opener, url, self.cnbc_headers, timeout=4)
            except Exception:
                continue
            for pat in patterns:
                m = re.search(pat, html, flags=re.IGNORECASE)
                if m:
                    return parse_float(m.group(1), 0.0), 0.0
        return 0.0, 0.0

    def fetch_nsetools_quote(self, symbol):
        sym = str(symbol or "").strip().upper()
        if not sym:
            return 0.0, 0.0
        try:
            if self._nsetools_client is None:
                from nsetools import Nse

                self._nsetools_client = Nse()
            data = self._nsetools_client.get_quote(sym) or {}
        except Exception:
            return 0.0, 0.0
        if not isinstance(data, dict):
            return 0.0, 0.0
        ltp = parse_float(
            data.get("lastPrice")
            or data.get("closePrice")
            or data.get("ltp")
            or data.get("last"),
            0.0,
        )
        prev_close = parse_float(
            data.get("previousClose")
            or data.get("prevClose")
            or data.get("close"),
            0.0,
        )
        change_abs = parse_float(data.get("change"), 0.0)
        if ltp > 0 and abs(change_abs) <= 1e-9 and prev_close > 0:
            change_abs = ltp - prev_close
        if ltp > 0:
            return ltp, change_abs
        return 0.0, 0.0

    def fetch_stock_nse_india_quote(self, symbol):
        sym = str(symbol or "").strip().upper()
        if not sym:
            return 0.0, 0.0
        base = str(os.environ.get("STOCK_NSE_INDIA_API_BASE", self.stock_nse_india_base or "")).strip().rstrip("/")
        if not base:
            return 0.0, 0.0
        urls = [
            f"{base}/api/equity/{urllib.parse.quote(sym)}",
            f"{base}/equity/{urllib.parse.quote(sym)}",
            f"{base}/api/quote-equity?symbol={urllib.parse.quote(sym)}",
        ]
        for url in urls:
            try:
                raw = _http_json(self.stock_nse_india_opener, url, self.stock_nse_india_headers, timeout=4)
            except Exception:
                continue
            if isinstance(raw, dict):
                payload = raw.get("data", raw)
            else:
                payload = raw
            if not isinstance(payload, dict):
                continue
            pinfo = payload.get("priceInfo", {}) if isinstance(payload.get("priceInfo"), dict) else {}
            sec = payload.get("securityInfo", {}) if isinstance(payload.get("securityInfo"), dict) else {}
            ltp = parse_float(
                pinfo.get("lastPrice")
                or payload.get("lastPrice")
                or payload.get("ltp")
                or payload.get("closePrice")
                or payload.get("last"),
                0.0,
            )
            if ltp <= 0:
                continue
            change_abs = parse_float(
                pinfo.get("change")
                or payload.get("change")
                or payload.get("priceChange"),
                0.0,
            )
            prev_close = parse_float(
                pinfo.get("previousClose")
                or sec.get("previousClose")
                or payload.get("previousClose")
                or payload.get("prevClose"),
                0.0,
            )
            if abs(change_abs) <= 1e-9 and prev_close > 0:
                change_abs = ltp - prev_close
            return ltp, change_abs
        return 0.0, 0.0

    def _extract_gold_rate_24k_per_gram(self, text):
        raw = str(text or "")
        if not raw:
            return 0.0, 0.0

        ltp = 0.0
        change_abs = 0.0

        # Primary: Hindu BusinessLine text format, e.g. "24ct Gold Rs 7,497 / 1gram"
        p_24ct_gram = re.search(
            r"24\s*ct\s*gold[\s\S]{0,40}?([0-9][0-9,]*\.?[0-9]*)\s*/\s*1\s*gram",
            raw,
            flags=re.IGNORECASE,
        )
        if p_24ct_gram:
            ltp = parse_float(p_24ct_gram.group(1), 0.0)

        # Table format: "Today 1 Gram 24 Carat Gold Price in Kerala ... 1 gram<today><yesterday><change>"
        p_24_table = re.search(
            r"today\s*1\s*gram\s*24\s*carat\s*gold\s*price\s*in\s*kerala[\s\S]{0,700}?1\s*gram\s*[^0-9]{0,12}([0-9][0-9,]*\.?[0-9]*)\s*[^0-9]{0,12}([0-9][0-9,]*\.?[0-9]*)\s*[^0-9]{0,12}([+\-]?[0-9][0-9,]*\.?[0-9]*)",
            raw,
            flags=re.IGNORECASE,
        )
        if p_24_table:
            today_px = parse_float(p_24_table.group(1), 0.0)
            yday_px = parse_float(p_24_table.group(2), 0.0)
            table_chg = parse_float(p_24_table.group(3), 0.0)
            if today_px > 0:
                ltp = today_px
            if abs(table_chg) > 0:
                # Most pages print absolute difference; derive sign from today/yesterday when possible.
                if today_px > 0 and yday_px > 0:
                    change_abs = today_px - yday_px
                else:
                    change_abs = table_chg

        # Fallback extraction for 24K per-gram values from generic pages.
        if ltp <= 0:
            patterns = [
                r"24ct\s*gold\s*today[\s\S]{0,180}?[^0-9]{0,16}([0-9][0-9,]*\.?[0-9]*)",
                r"24K\s*Gold\s*/g\s*[^0-9]{0,24}([0-9][0-9,]*\.?[0-9]*)",
                r"price of gold in india today is\s*[^0-9]{0,24}([0-9][0-9,]*\.?[0-9]*)\s*per gram for 24 karat gold",
                r"today\s+24\s*carat\s+gold\s+rate\s+per\s+gram\s+in\s+india[\s\S]{0,220}\b1\s*[^0-9]{0,16}([0-9][0-9,]*\.?[0-9]*)",
                r"indian major cities gold rates today[\s\S]{0,220}\b24k today\b[^0-9]{0,16}([0-9][0-9,]*\.?[0-9]*)",
                r"24\s*carat\s*gold\s*price\s*in\s*kerala[\s\S]{0,220}?1\s*gram[^0-9]{0,16}([0-9][0-9,]*\.?[0-9]*)",
            ]
            for pat in patterns:
                m = re.search(pat, raw, flags=re.IGNORECASE)
                if not m:
                    continue
                ltp = parse_float(m.group(1), 0.0)
                if ltp > 0:
                    break

        if abs(change_abs) <= 1e-9:
            change_match = re.search(
                r"24K\s*Gold\s*/g[\s\S]{0,48}([+\-]\s*[^0-9]{0,12}[0-9][0-9,]*\.?[0-9]*)",
                raw,
                flags=re.IGNORECASE,
            )
            if change_match:
                token = str(change_match.group(1) or "")
                sign = -1.0 if "-" in token else 1.0
                numeric = re.sub(r"[^0-9.,]", "", token)
                mag = parse_float(numeric, 0.0)
                if mag > 0:
                    change_abs = sign * mag

        if ltp <= 0:
            return 0.0, 0.0
        return ltp, change_abs

    def fetch_google_gold_rate_quote(self):
        # Best-effort fallback via Google search snippets.
        queries = [
            "24 carat gold price in kerala today per gram",
            "today 1 gram 24 carat gold price in kerala",
            "gold rate today kerala 24ct 1 gram",
        ]
        for q in queries:
            url = "https://www.google.com/search?q=" + urllib.parse.quote(q)
            try:
                html = _http_text(self.google_opener, url, self.google_headers, timeout=5)
            except Exception:
                continue
            ltp, change_abs = self._extract_gold_rate_24k_per_gram(html)
            if ltp > 0:
                return ltp, change_abs
        return 0.0, 0.0

    def fetch_gold_rate_quote(self, symbol=None):
        # GOLD is modeled as physical grams; LTP is 24K India per-gram rate.
        urls = [
            "https://www.thehindubusinessline.com/gold-rate-today/Kerala/",
            "https://www.goodreturns.in/gold-rates/",
            "https://origin-www.goodreturns.in/gold-rates/",
        ]
        best = (0.0, 0.0)
        for url in urls:
            try:
                html = _http_text(self.gold_opener, url, self.gold_headers, timeout=5)
            except Exception:
                continue
            ltp, change_abs = self._extract_gold_rate_24k_per_gram(html)
            if ltp > 0:
                return ltp, change_abs
            if ltp > parse_float(best[0], 0.0):
                best = (ltp, change_abs)
        g_ltp, g_change = self.fetch_google_gold_rate_quote()
        if g_ltp > 0:
            return g_ltp, g_change
        return best

    def fetch_source_quote(self, source, exchange, symbol, feed_code):
        src = str(source or "").strip().lower()
        ex = str(exchange or "NSE").upper()
        sym = symbol_upper(symbol)
        if src == "nse_api":
            return self.fetch_nse_quote(sym)
        if src == "bse_api":
            return self.fetch_bse_quote(feed_code or sym)
        if src == "yahoo_finance":
            return self.fetch_yahoo_quote(sym, exchange=ex)
        if src == "google_scrape":
            return self.fetch_google_quote(sym, exchange=ex)
        if src == "screener_scrape":
            return self.fetch_screener_quote(sym)
        if src == "trendlyne_scrape":
            return self.fetch_trendlyne_quote(sym)
        if src == "cnbc_scrape":
            return self.fetch_cnbc_quote(sym)
        if src == "nsetools_api":
            return self.fetch_nsetools_quote(sym)
        if src == "stock_nse_india_api":
            return self.fetch_stock_nse_india_quote(sym)
        if src == "gold_rate_scrape":
            return self.fetch_gold_rate_quote(sym)

        # Dynamic adapter hook: source "foo_bar" -> fetch_foo_bar_quote / fetch_foo_quote.
        method_names = [f"fetch_{src}_quote"]
        if src.endswith("_scrape"):
            method_names.append(f"fetch_{src[:-7]}_quote")
        if src.endswith("_api"):
            method_names.append(f"fetch_{src[:-4]}_quote")
        if src.endswith("_finance"):
            method_names.append(f"fetch_{src[:-8]}_quote")
        for m in method_names:
            fn = getattr(self, m, None)
            if not callable(fn):
                continue
            try:
                return fn(sym)
            except TypeError:
                try:
                    return fn(sym, ex)
                except Exception:
                    continue
            except Exception:
                continue
        return 0.0, 0.0

    def collect_quotes(self, exchange, symbol, feed_code, source_order, asset_class=None):
        ex = str(exchange or "NSE").upper()
        sym = symbol_upper(symbol)
        normalized_sources = parse_source_list(",".join(source_order or []), DEFAULT_LIVE_QUOTE_SOURCES)
        ac = normalize_asset_class(asset_class, fallback=infer_asset_class(symbol=sym, name=sym))
        if ac == ASSET_CLASS_GOLD:
            normalized_sources = [s for s in normalized_sources if str(s or "").strip().lower() == "gold_rate_scrape"]
            if not normalized_sources:
                normalized_sources = ["gold_rate_scrape"]
        else:
            normalized_sources = [s for s in normalized_sources if str(s or "").strip().lower() != "gold_rate_scrape"]
        quotes = []
        attempts = []
        t0 = time.time()
        target_quotes = 2 if len(normalized_sources) >= 2 else 1
        for src in normalized_sources:
            elapsed = time.time() - t0
            if elapsed > 1.8:
                break
            ltp = 0.0
            change_abs = 0.0
            st = time.time()
            try:
                ltp, change_abs = self.fetch_source_quote(src, ex, sym, feed_code)
            except Exception:
                ltp = 0.0
                change_abs = 0.0
            elapsed_ms = round((time.time() - st) * 1000.0, 3)
            attempts.append(
                {
                    "source": src,
                    "success": bool(ltp > 0),
                    "ltp": float(ltp) if ltp > 0 else 0.0,
                    "latency_ms": elapsed_ms,
                }
            )
            if ltp > 0:
                quotes.append(
                    {
                        "source": src,
                        "ltp": float(ltp),
                        "change_abs": float(change_abs),
                        "latency_ms": elapsed_ms,
                    }
                )
                if len(quotes) >= target_quotes:
                    break
        return quotes, attempts

    def pick_quote_candidate(self, candidates, source_order, max_deviation_pct=LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT):
        if not candidates:
            return None, 0.0, []
        median_ltp = median_value([c["ltp"] for c in candidates])
        if median_ltp <= 0:
            chosen = max(candidates, key=lambda x: parse_float(x.get("ltp"), 0.0))
            return chosen, 0.0, list(candidates)

        tolerance = max(0.5, median_ltp * max(parse_float(max_deviation_pct, LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT), 1.0) / 100.0)
        filtered = [c for c in candidates if abs(parse_float(c.get("ltp"), 0.0) - median_ltp) <= tolerance]
        if not filtered:
            filtered = list(candidates)

        rank = {src: i for i, src in enumerate(parse_source_list(",".join(source_order or []), DEFAULT_LIVE_QUOTE_SOURCES))}
        filtered.sort(
            key=lambda c: (
                rank.get(str(c.get("source") or "").lower(), 999),
                abs(parse_float(c.get("ltp"), 0.0) - median_ltp),
            )
        )
        chosen = filtered[0]
        return chosen, median_ltp, filtered

    def fetch_multi_source_quote(
        self,
        exchange,
        symbol,
        feed_code,
        source_order=None,
        max_deviation_pct=LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT,
        asset_class=None,
    ):
        order = parse_source_list(",".join(source_order or []), DEFAULT_LIVE_QUOTE_SOURCES)
        candidates, attempts = self.collect_quotes(exchange, symbol, feed_code, order, asset_class=asset_class)
        chosen, median_ltp, filtered = self.pick_quote_candidate(candidates, order, max_deviation_pct=max_deviation_pct)
        return {
            "selected": chosen,
            "consensus_ltp": median_ltp,
            "candidates": candidates,
            "attempts": attempts,
            "filtered_candidates": filtered,
            "source_order": order,
        }

    def fetch_nse_all_indices(self):
        self._bootstrap_nse()
        url = "https://www.nseindia.com/api/allIndices"
        data = _http_json(self.nse_opener, url, self.nse_headers)
        rows = data.get("data", [])
        out = {}
        for r in rows:
            name = str(r.get("indexSymbol") or r.get("index") or "").strip()
            if not name:
                continue
            key = name.upper()
            out[key] = {
                "last": parse_float(r.get("last"), parse_float(r.get("lastPrice"), 0.0)),
                "change_abs": parse_float(r.get("variation"), parse_float(r.get("change"), 0.0)),
                "change_pct": parse_float(r.get("percentChange"), parse_float(r.get("pChange"), 0.0)),
            }
        return out

    def fetch_yahoo_indices(self):
        symbols = [
            "^NSEI",  # NIFTY 50
            "^NSEBANK",  # NIFTY BANK
            "^CNXIT",  # NIFTY IT
            "^CNXPHARMA",  # NIFTY PHARMA
            "NIFTY_MID_SELECT.NS",  # midcap proxy (best effort)
            "NIFTY_SMALLCAP_100.NS",  # smallcap proxy (best effort)
        ]
        url = "https://query1.finance.yahoo.com/v7/finance/quote?symbols=" + urllib.parse.quote(",".join(symbols))
        data = _http_json(self.yahoo_opener, url, self.yahoo_headers)
        rows = (((data or {}).get("quoteResponse") or {}).get("result") or [])
        out = {}
        key_map = {
            "^NSEI": "NIFTY 50",
            "^NSEBANK": "NIFTY BANK",
            "^CNXIT": "NIFTY IT",
            "^CNXPHARMA": "NIFTY PHARMA",
            "NIFTY_MID_SELECT.NS": "NIFTY MIDCAP 100",
            "NIFTY_SMALLCAP_100.NS": "NIFTY SMALLCAP 100",
        }
        for r in rows:
            raw_sym = str(r.get("symbol") or "").strip().upper()
            key = key_map.get(raw_sym)
            if not key:
                continue
            last = parse_float(r.get("regularMarketPrice"), 0.0)
            if last <= 0:
                continue
            out[key] = {
                "last": last,
                "change_abs": parse_float(r.get("regularMarketChange"), 0.0),
                "change_pct": parse_float(r.get("regularMarketChangePercent"), 0.0),
            }
        return out


def _macro_pick_index(indices, names):
    for n in names:
        key = str(n).upper()
        if key in indices:
            return indices[key]
    return None


def _build_local_proxy_indices():
    # Fallback macro approximation when external index feeds are blocked.
    buckets = {
        "NIFTY 50": ["RELIANCE", "HDFCBANK", "ICICIBANK", "SBIN", "LT", "ITC", "TCS", "INFY", "HINDUNILVR"],
        "NIFTY BANK": ["HDFCBANK", "ICICIBANK", "SBIN", "KOTAKBANK", "AXISBANK", "INDUSINDBK"],
        "NIFTY MIDCAP 100": ["TRENT", "IRFC", "BEL", "BSE", "PERSISTENT", "POLYCAB", "DIXON", "CUMMINSIND"],
        "NIFTY SMALLCAP 100": ["IRCON", "RVNL", "KAYNES", "KPITTECH", "IDFCFIRSTB", "IDEA"],
        "NIFTY IT": ["INFY", "TCS", "HCLTECH", "WIPRO", "TECHM", "LTIM"],
        "NIFTY PHARMA": ["SUNPHARMA", "DRREDDY", "CIPLA", "DIVISLAB", "AUROPHARMA", "LUPIN"],
    }
    all_syms = sorted({s for arr in buckets.values() for s in arr})
    if not all_syms:
        return {}
    placeholders = ",".join(["?"] * len(all_syms))
    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT symbol, ltp, change_abs
            FROM latest_prices
            WHERE UPPER(symbol) IN ({placeholders})
            """,
            all_syms,
        ).fetchall()
        tick_rows = conn.execute(
            f"""
            SELECT symbol, ltp, fetched_at
            FROM price_ticks
            WHERE UPPER(symbol) IN ({placeholders})
              AND fetched_at >= datetime('now','-3 day')
              AND ltp > 0
            ORDER BY symbol, fetched_at
            """,
            all_syms,
        ).fetchall()
    tick_by_symbol = defaultdict(list)
    for tr in tick_rows:
        tick_by_symbol[symbol_upper(tr["symbol"])].append(parse_float(tr["ltp"], 0.0))
    chg = {}
    for r in rows:
        sym = symbol_upper(r["symbol"])
        ltp = parse_float(r["ltp"], 0.0)
        change_abs = parse_float(r["change_abs"], 0.0)
        prev = ltp - change_abs
        if ltp > 0 and (abs(change_abs) <= 1e-9 or prev <= 0):
            ticks = [v for v in tick_by_symbol.get(sym, []) if v > 0]
            if len(ticks) >= 2 and ticks[0] > 0:
                change_abs = ticks[-1] - ticks[0]
                prev = ltp - change_abs
        if ltp > 0 and prev > 0:
            chg[sym] = {
                "last": ltp,
                "change_abs": change_abs,
                "change_pct": (change_abs / prev) * 100.0,
            }

    out = {}
    for name, syms in buckets.items():
        vals = [chg[s]["change_pct"] for s in syms if s in chg]
        if len(vals) < 2:
            continue
        avg_pct = sum(vals) / len(vals)
        out[name.upper()] = {
            "last": 0.0,
            "change_abs": 0.0,
            "change_pct": avg_pct,
        }
    return out


def build_macro_thoughts():
    client = MarketDataClient()
    source = "nse_indices"
    try:
        indices = client.fetch_nse_all_indices()
    except Exception as nse_ex:
        try:
            indices = client.fetch_yahoo_indices()
            source = "yahoo_indices"
        except Exception:
            indices = _build_local_proxy_indices()
            source = "local_proxy"
            if not indices:
                return {
                    "as_of": now_iso(),
                    "source": "unavailable",
                    "regime": "neutral",
                    "score": 0.0,
                    "confidence": 0.45,
                    "signals": {},
                    "thought": (
                        f"Macro feed unavailable (NSE blocked: {type(nse_ex).__name__}; no fallback feed). "
                        "Strategy is running with neutral macro bias."
                    ),
                }

    if not indices:
        if source != "local_proxy":
            indices = _build_local_proxy_indices()
            if indices:
                source = "local_proxy"
        if not indices:
            return {
                "as_of": now_iso(),
                "source": source,
                "regime": "neutral",
                "score": 0.0,
                "confidence": 0.45,
                "signals": {},
                "thought": "Macro feed returned no usable indices; strategy is running with neutral macro bias.",
            }

    nifty = _macro_pick_index(indices, ("NIFTY 50", "NIFTY 50 PR", "NIFTY TOTAL MARKET"))
    bank = _macro_pick_index(indices, ("NIFTY BANK", "NIFTY BANK PR"))
    mid = _macro_pick_index(indices, ("NIFTY MIDCAP 100", "NIFTY MIDCAP 150", "NIFTY MIDCAP SELECT"))
    small = _macro_pick_index(indices, ("NIFTY SMALLCAP 100", "NIFTY SMALLCAP 250"))
    it = _macro_pick_index(indices, ("NIFTY IT",))
    pharma = _macro_pick_index(indices, ("NIFTY PHARMA",))

    components = {
        "nifty50_change_pct": parse_float(nifty["change_pct"], 0.0) if nifty else 0.0,
        "banknifty_change_pct": parse_float(bank["change_pct"], 0.0) if bank else 0.0,
        "midcap_change_pct": parse_float(mid["change_pct"], 0.0) if mid else 0.0,
        "smallcap_change_pct": parse_float(small["change_pct"], 0.0) if small else 0.0,
        "it_change_pct": parse_float(it["change_pct"], 0.0) if it else 0.0,
        "pharma_change_pct": parse_float(pharma["change_pct"], 0.0) if pharma else 0.0,
    }
    available = [v for v in components.values() if abs(v) > 0]
    avg_move = (sum(available) / len(available)) if available else 0.0
    breadth_pos = sum(1 for v in available if v > 0)
    breadth_neg = sum(1 for v in available if v < 0)

    score = 0.0
    score += clamp(components["nifty50_change_pct"] / 1.2, -1.3, 1.3)
    score += clamp(components["banknifty_change_pct"] / 1.2, -1.0, 1.0)
    score += clamp((components["midcap_change_pct"] + components["smallcap_change_pct"]) / 2.0 / 1.5, -1.0, 1.0)
    score += clamp((components["it_change_pct"] + components["pharma_change_pct"]) / 2.0 / 1.8, -0.8, 0.8)
    score = round(score, 3)

    if score >= 1.4 and breadth_pos >= max(2, breadth_neg + 1):
        regime = "risk_on"
    elif score <= -1.4 and breadth_neg >= max(2, breadth_pos + 1):
        regime = "risk_off"
    else:
        regime = "neutral"

    confidence = clamp(0.45 + min(0.35, abs(score) * 0.11) + min(0.15, abs(avg_move) * 0.08), 0.45, 0.95)
    thought = (
        f"Macro {regime}: NIFTY50 {components['nifty50_change_pct']:.2f}%, "
        f"BANKNIFTY {components['banknifty_change_pct']:.2f}%, "
        f"MID/SMALL {components['midcap_change_pct']:.2f}%/{components['smallcap_change_pct']:.2f}%, "
        f"breadth +{breadth_pos}/-{breadth_neg} (source: {source})."
    )
    return {
        "as_of": now_iso(),
        "source": source,
        "regime": regime,
        "score": score,
        "confidence": round(confidence, 4),
        "signals": components,
        "thought": thought,
    }


def refresh_latest_prices_from_exchange(max_runtime_sec=12):
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT symbol, exchange, feed_code, COALESCE(asset_class, 'EQUITY') AS asset_class
            FROM instruments
            WHERE active = 1
            ORDER BY symbol
            """
        ).fetchall()
        latest_rows = conn.execute("SELECT symbol, ltp, updated_at FROM latest_prices").fetchall()
        holding_rows = conn.execute("SELECT symbol, qty, avg_cost FROM holdings").fetchall()
        quote_policy = get_live_quote_policy(conn)
        cursor_row = conn.execute(
            "SELECT value FROM app_config WHERE key = 'live_refresh_cursor'"
        ).fetchone()
        try:
            cursor = int(float(cursor_row["value"])) if cursor_row else 0
        except Exception:
            cursor = 0
        instruments = list(rows)
        if instruments:
            cursor = cursor % len(instruments)
            instruments = instruments[cursor:] + instruments[:cursor]
        else:
            cursor = 0
        nse_order = get_ranked_quote_sources(conn, quote_policy, exchange="NSE")
        bse_order = get_ranked_quote_sources(conn, quote_policy, exchange="BSE")
        today_ist_s = ist_now().date().isoformat()
        # Capture today's online tick series so fallback day-change can ignore early outlier ticks.
        tick_series_map = defaultdict(list)
        tick_rows = conn.execute(
            """
            SELECT symbol, ltp
            FROM price_ticks
            WHERE SUBSTR(fetched_at,1,10) = ?
              AND ltp > 0
            ORDER BY UPPER(symbol), fetched_at ASC
            """,
            (today_ist_s,),
        ).fetchall()
        for tr in tick_rows:
            su = symbol_upper(tr["symbol"])
            px = parse_float(tr["ltp"], 0.0)
            if su and px > 0:
                tick_series_map[su].append(px)
    latest_ltp_map = {symbol_upper(r["symbol"]): parse_float(r["ltp"], 0.0) for r in latest_rows}
    latest_updated_date_map = {
        symbol_upper(r["symbol"]): str(r["updated_at"] or "")[:10]
        for r in latest_rows
    }
    holding_map = {
        symbol_upper(r["symbol"]): {
            "qty": parse_float(r["qty"], 0.0),
            "avg_cost": parse_float(r["avg_cost"], 0.0),
        }
        for r in holding_rows
    }

    if not instruments:
        return

    gold_symbols = {
        symbol_upper(r["symbol"])
        for r in instruments
        if normalize_asset_class(
            r["asset_class"],
            fallback=infer_asset_class(symbol=r["symbol"], name=r["symbol"]),
        )
        == ASSET_CLASS_GOLD
    }

    client = MarketDataClient()
    updates = []
    sample_rows = []
    metric_events = []
    fetched_at = now_iso()
    t0 = time.time()
    processed_symbols = 0
    now_ist = ist_now()
    today_ist_s = now_ist.date().isoformat()
    prev_close_map = load_prev_close_map(None, [r["symbol"] for r in instruments], as_of_date=today_ist_s)
    allow_zero_qty_refresh = is_zero_qty_eod_window(now_ist)
    gold_quote_cache = None
    for inst in instruments:
        if max_runtime_sec is not None and (time.time() - t0) > max_runtime_sec:
            break
        processed_symbols += 1
        symbol = inst["symbol"]
        sym_u = symbol_upper(symbol)
        asset_class = normalize_asset_class(
            inst["asset_class"],
            fallback=infer_asset_class(symbol=inst["symbol"], name=inst["symbol"]),
        )
        hold_ref = holding_map.get(sym_u, {"qty": 0.0, "avg_cost": 0.0})
        qty = parse_float(hold_ref.get("qty"), 0.0)
        if qty <= 0:
            if not allow_zero_qty_refresh:
                continue
            if latest_updated_date_map.get(sym_u) == today_ist_s:
                continue

        if asset_class == ASSET_CLASS_GOLD:
            if gold_quote_cache is None:
                st = time.time()
                g_ltp = 0.0
                g_change = 0.0
                try:
                    g_ltp, g_change = client.fetch_gold_rate_quote(symbol)
                except Exception:
                    g_ltp, g_change = 0.0, 0.0
                lat_ms = round((time.time() - st) * 1000.0, 3)
                gold_quote_cache = {
                    "ltp": parse_float(g_ltp, 0.0),
                    "change_abs": parse_float(g_change, 0.0),
                    "latency_ms": lat_ms,
                    "source": "gold_rate_scrape",
                }
                metric_events.append(
                    {
                        "source": "gold_rate_scrape",
                        "success": bool(gold_quote_cache["ltp"] > 0),
                        "latency_ms": gold_quote_cache["latency_ms"],
                        "accuracy_error_pct": None,
                        "fetched_at": fetched_at,
                    }
                )
            ltp = parse_float((gold_quote_cache or {}).get("ltp"), 0.0)
            if ltp <= 0:
                continue
            # GOLD day-change should be derived from previously logged daily close.
            prev_close = parse_float(prev_close_map.get(sym_u), 0.0)
            if prev_close > 0:
                change_abs = ltp - prev_close
            else:
                change_abs = parse_float((gold_quote_cache or {}).get("change_abs"), 0.0)
            updates.append((symbol, ltp, change_abs, fetched_at, "gold_rate_scrape"))
            sample_rows.append(
                (
                    sym_u,
                    "gold_rate_scrape",
                    ltp,
                    change_abs,
                    parse_float((gold_quote_cache or {}).get("latency_ms"), 0.0),
                    0.0,
                    fetched_at,
                    1,
                    ltp,
                )
            )
            continue

        exchange = str(inst["exchange"] or "NSE").upper()
        source_order = nse_order if exchange == "NSE" else bse_order
        report = client.fetch_multi_source_quote(
            exchange=exchange,
            symbol=symbol,
            feed_code=inst["feed_code"],
            source_order=source_order,
            max_deviation_pct=quote_policy.get("max_deviation_pct", LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT),
            asset_class=asset_class,
        )
        selected = report.get("selected") or {}
        ltp = parse_float(selected.get("ltp"), 0.0)
        consensus_ltp = parse_float(report.get("consensus_ltp"), 0.0)
        selected_source = str(selected.get("source") or "")
        for a in report.get("attempts") or []:
            err = None
            if bool(a.get("success")) and consensus_ltp > 0:
                err = abs(parse_float(a.get("ltp"), 0.0) - consensus_ltp) / consensus_ltp * 100.0
            metric_events.append(
                {
                    "source": str(a.get("source") or "").lower(),
                    "success": bool(a.get("success")),
                    "latency_ms": parse_float(a.get("latency_ms"), 0.0),
                    "accuracy_error_pct": err,
                    "fetched_at": fetched_at,
                }
            )

        source = selected_source or "multi_source"
        prev_ltp = parse_float(latest_ltp_map.get(sym_u), 0.0)
        candidates = report.get("candidates") or []
        if ltp > 0 and not _quote_is_plausible(
            selected_source=source,
            ltp=ltp,
            prev_ltp=prev_ltp,
            qty=hold_ref.get("qty"),
            avg_cost=hold_ref.get("avg_cost"),
            candidates=candidates,
            max_dev_pct=quote_policy.get("max_deviation_pct", LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT),
            asset_class=asset_class,
            symbol=symbol,
        ):
            fallback = None
            for q in candidates:
                q_ltp = parse_float(q.get("ltp"), 0.0)
                q_source = str(q.get("source") or "")
                if q_ltp <= 0:
                    continue
                if _quote_is_plausible(
                    selected_source=q_source,
                    ltp=q_ltp,
                    prev_ltp=prev_ltp,
                    qty=hold_ref.get("qty"),
                    avg_cost=hold_ref.get("avg_cost"),
                    candidates=candidates,
                    max_dev_pct=quote_policy.get("max_deviation_pct", LIVE_QUOTE_MAX_DEVIATION_PCT_DEFAULT),
                    asset_class=asset_class,
                    symbol=symbol,
                ):
                    fallback = q
                    break
            if fallback:
                ltp = parse_float(fallback.get("ltp"), 0.0)
                selected_source = str(fallback.get("source") or selected_source or "")
                source = selected_source or source
            else:
                metric_events.append(
                    {
                        "source": source,
                        "success": False,
                        "latency_ms": 0.0,
                        "accuracy_error_pct": None,
                        "fetched_at": fetched_at,
                    }
                )
                ltp = 0.0
                selected_source = ""
        for q in candidates:
            err = None
            if consensus_ltp > 0:
                err = abs(parse_float(q.get("ltp"), 0.0) - consensus_ltp) / consensus_ltp * 100.0
            sample_rows.append(
                (
                    sym_u,
                    str(q.get("source") or ""),
                    parse_float(q.get("ltp"), 0.0),
                    parse_float(q.get("change_abs"), 0.0),
                    parse_float(q.get("latency_ms"), 0.0),
                    err,
                    fetched_at,
                    1 if selected_source and (str(q.get("source") or "") == selected_source) else 0,
                    consensus_ltp if consensus_ltp > 0 else None,
                )
            )
        if ltp <= 0:
            continue
        change_abs = resolve_preferred_equity_day_change_abs(
            exchange=exchange,
            selected_ltp=ltp,
            selected_change_abs=selected.get("change_abs"),
            candidates=candidates,
        )
        if abs(change_abs) <= 1e-9:
            prev_close = parse_float(prev_close_map.get(sym_u), 0.0)
            if prev_close > 0 and is_plausible_day_reference_price(ltp, prev_close):
                change_abs = ltp - prev_close
            else:
                first_tick = first_plausible_intraday_ltp(tick_series_map.get(sym_u), ltp)
                if first_tick > 0:
                    change_abs = ltp - first_tick
        updates.append((symbol, ltp, change_abs, fetched_at, selected_source or source))
        # light throttle to avoid API blocking
        time.sleep(0.12)

    if not updates and not metric_events:
        ensure_latest_prices_nonzero_from_last_trade()
        with db_connect() as conn:
            refresh_holdings_mark_to_market(conn)
            conn.commit()
        return

    with db_connect() as conn:
        if updates:
            conn.executemany(
                """
                INSERT INTO latest_prices(symbol, ltp, change_abs, updated_at) VALUES (?, ?, ?, ?)
                ON CONFLICT(symbol) DO UPDATE SET
                  ltp=excluded.ltp,
                  change_abs=excluded.change_abs,
                  updated_at=excluded.updated_at
                """,
                [(u[0], u[1], u[2], u[3]) for u in updates],
            )
            conn.executemany(
                "INSERT INTO price_ticks(symbol, ltp, change_abs, fetched_at, source) VALUES (?, ?, ?, ?, ?)",
                updates,
            )
        if sample_rows:
            conn.executemany(
                """
                INSERT INTO quote_samples(symbol, source, ltp, change_abs, latency_ms, accuracy_error_pct, fetched_at, selected, consensus_ltp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                sample_rows,
            )
        if metric_events:
            apply_quote_source_metrics(conn, metric_events)
        guard_replacements = sanitize_latest_price_outliers(conn, fetched_at)
        sanitize_latest_price_day_change_outliers(conn, fetched_at)
        if instruments:
            next_cursor = (cursor + processed_symbols) % len(instruments)
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('live_refresh_cursor', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (str(next_cursor), now_iso()),
            )
        conn.commit()
    if updates:
        today_s = str(fetched_at)[:10]
        hist_payload = [(u[0], today_s, u[1], f"live_{u[4]}", fetched_at) for u in updates if parse_float(u[1], 0.0) > 0]
        if guard_replacements:
            hist_payload.extend(
                [(g[0], today_s, g[1], f"live_{g[4]}", fetched_at) for g in guard_replacements if parse_float(g[1], 0.0) > 0]
            )
        gold_hist_payload = [r for r in hist_payload if symbol_upper(r[0]) in gold_symbols]
        non_gold_hist_payload = [r for r in hist_payload if symbol_upper(r[0]) not in gold_symbols]
        # GOLD: keep one daily first-hit snapshot (no overwrite) for stable day-change baseline.
        insert_market_daily_prices_if_missing(gold_hist_payload)
        # Non-GOLD: retain latest close snapshot per day.
        upsert_market_daily_prices(non_gold_hist_payload)
    ensure_latest_prices_nonzero_from_last_trade()
    with db_connect() as conn:
        refresh_holdings_mark_to_market(conn)
        conn.commit()


def sync_market_history(backfill_all=False, max_runtime_sec=180, max_symbols=None):
    t0 = time.time()
    runtime_cap = None
    if max_runtime_sec is not None:
        try:
            runtime_cap = max(1, int(max_runtime_sec))
        except Exception:
            runtime_cap = 180
    today = dt.date.today()
    stats = {
        "symbols_total": 0,
        "symbols_processed": 0,
        "history_rows_upserted": 0,
        "tick_rows_upserted": 0,
        "latest_rows_upserted": 0,
        "errors": [],
        "backfill_all": bool(backfill_all),
    }
    with db_connect() as conn:
        universe = history_symbol_universe(conn)
    if max_symbols is not None:
        try:
            cap = max(1, int(max_symbols))
            universe = universe[:cap]
        except Exception:
            pass
    stats["symbols_total"] = len(universe)
    if not universe:
        with db_connect() as conn:
            conn.execute(
                """
                INSERT INTO app_config(key, value, updated_at) VALUES ('history_last_sync_at', ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (now_iso(), now_iso()),
            )
            conn.commit()
        return stats

    client = MarketDataClient()
    symbol_batch = []
    for item in universe:
        if runtime_cap is not None and (time.time() - t0) > runtime_cap:
            break
        symbol = symbol_upper(item.get("symbol"))
        if not symbol:
            continue
        symbol_batch.append(symbol)
        first_trade = parse_history_date(item.get("first_trade_date")) or today
        if first_trade > today:
            first_trade = today

        with market_db_connect() as mconn:
            mx_row = mconn.execute(
                "SELECT MAX(price_date) AS mx FROM daily_prices WHERE UPPER(symbol) = ?",
                (symbol,),
            ).fetchone()
        max_existing = parse_history_date(mx_row["mx"]) if mx_row and mx_row["mx"] else None
        if backfill_all or not max_existing:
            fetch_from = first_trade
        else:
            fetch_from = max_existing + dt.timedelta(days=1)
        if fetch_from > today:
            fetch_from = today

        # Always seed from locally captured ticks (works for NSE/BSE and closed symbols if seen before).
        with db_connect() as conn:
            tick_daily = aggregate_price_ticks_daily(conn, symbol, fetch_from.isoformat(), today.isoformat())
        if tick_daily:
            tick_payload = [(symbol, d, px, "price_ticks_agg", now_iso()) for _, d, px in tick_daily]
            stats["tick_rows_upserted"] += upsert_market_daily_prices(tick_payload)

        if fetch_from <= today:
            cur = fetch_from
            while cur <= today:
                if runtime_cap is not None and (time.time() - t0) > runtime_cap:
                    break
                chunk_end = min(today, cur + dt.timedelta(days=365))
                try:
                    hist = client.fetch_nse_history(symbol, cur, chunk_end)
                    if hist:
                        rows = [(symbol, d, px, "nse_history", now_iso()) for d, px in hist if px > 0]
                        stats["history_rows_upserted"] += upsert_market_daily_prices(rows)
                except Exception as ex:
                    if len(stats["errors"]) < 24:
                        stats["errors"].append(f"{symbol}: {str(ex)}")
                cur = chunk_end + dt.timedelta(days=1)
                time.sleep(0.08)

        stats["symbols_processed"] += 1

    stats["latest_rows_upserted"] = upsert_today_market_history_from_latest(symbol_batch)
    with db_connect() as conn:
        stamp = now_iso()
        conn.execute(
            """
            INSERT INTO app_config(key, value, updated_at) VALUES ('history_last_sync_at', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
            """,
            (stamp, stamp),
        )
        conn.commit()
    return stats


def recompute_holdings_and_signals(force_strategy=True):
    with db_connect() as conn:
        conn.execute("DELETE FROM holdings")
        conn.execute("DELETE FROM lot_closures")
        conn.execute("DELETE FROM signals")
        params = get_active_params(conn)

        buy_l1 = params.get("buy_l1_discount", 0.03)
        buy_l2 = params.get("buy_l2_discount", 0.06)
        s1 = params.get("sell_s1_markup", 0.12)
        s2 = params.get("sell_s2_markup", 0.50)
        s3 = params.get("sell_s3_markup", 1.25)
        split_map = load_split_map(conn)
        dividend_map = dividend_amount_map(conn)

        symbols = [
            {
                "symbol": r["symbol"],
                "asset_class": normalize_asset_class(
                    r["asset_class"],
                    fallback=infer_asset_class(symbol=r["symbol"], name=r["name"]),
                ),
            }
            for r in conn.execute("SELECT symbol, COALESCE(asset_class, 'EQUITY') AS asset_class, name FROM instruments").fetchall()
        ]
        for row in symbols:
            symbol = row["symbol"]
            asset_class = row["asset_class"]
            symbol_q = symbol_upper(symbol)
            trades = conn.execute(
                """
                SELECT id, side, trade_date, quantity, price
                FROM trades
                WHERE UPPER(symbol) = ?
                ORDER BY trade_date, id
                """,
                (symbol_q,),
            ).fetchall()

            qty = 0.0
            avg_cost = 0.0
            realized = 0.0
            historical_buy_value = 0.0
            lots = deque()

            for t in trades:
                side = t["side"]
                d = t["trade_date"]
                q, p = adjusted_trade_values(
                    symbol, d, float(t["quantity"]), float(t["price"]), split_map
                )
                if side == "BUY":
                    old_qty = qty
                    qty += q
                    if qty > 0:
                        avg_cost = ((avg_cost * old_qty) + (p * q)) / qty
                    historical_buy_value += (q * p)
                    lots.append({"qty": q, "price": p, "date": d})
                else:
                    sell_qty = q if q > 0 else 0.0
                    matched = min(sell_qty, qty) if qty > 0 else 0.0
                    if matched > 0:
                        realized += (p - avg_cost) * matched
                    qty = max(0.0, qty - sell_qty)
                    if qty == 0:
                        avg_cost = 0.0

                    rem = sell_qty
                    while rem > 1e-9 and lots:
                        first = lots[0]
                        close_qty = min(first["qty"], rem)
                        lot_pnl = (p - first["price"]) * close_qty
                        conn.execute(
                            """
                            INSERT INTO lot_closures(symbol, close_date, qty_closed, buy_price, sell_price, realized_pnl)
                            VALUES (?, ?, ?, ?, ?, ?)
                            """,
                            (symbol, d, close_qty, first["price"], p, lot_pnl),
                        )
                        first["qty"] -= close_qty
                        rem -= close_qty
                        if first["qty"] <= 1e-9:
                            lots.popleft()

            ltp = get_effective_ltp_for_asset(conn, symbol, asset_class=asset_class, split_map=split_map)

            invested_raw = qty * avg_cost
            dividend_total = max(0.0, parse_float(dividend_map.get(symbol_q), 0.0))
            dividend_to_invested = min(max(invested_raw, 0.0), dividend_total)
            dividend_excess_realized = max(0.0, dividend_total - dividend_to_invested)
            invested = max(0.0, invested_raw - dividend_to_invested)
            market_value = qty * ltp
            unrealized = market_value - invested
            realized_total = realized + dividend_excess_realized
            base_capital = historical_buy_value if historical_buy_value > 0 else invested
            ret_pct = ((realized_total + unrealized) / base_capital * 100.0) if base_capital > 0 else 0.0

            conn.execute(
                """
                INSERT INTO holdings(symbol, qty, avg_cost, invested, market_value, realized_pnl, unrealized_pnl, total_return_pct, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (symbol, qty, avg_cost, invested, market_value, realized_total, unrealized, ret_pct, now_iso()),
            )

            if qty <= 0 or avg_cost <= 0:
                continue
            buy_signal = None
            sell_signal = None
            if ltp <= avg_cost * (1.0 - buy_l2):
                buy_signal = "BUY"
            elif ltp <= avg_cost * (1.0 - buy_l1):
                buy_signal = "B1"

            if ltp >= avg_cost * (1.0 + s3):
                sell_signal = "S3"
            elif ltp >= avg_cost * (1.0 + s2):
                sell_signal = "S2"
            elif ltp >= avg_cost * (1.0 + s1):
                sell_signal = "S1"

            score = ((ltp - avg_cost) / avg_cost) * 100.0
            if buy_signal or sell_signal:
                conn.execute(
                    """
                    INSERT INTO signals(symbol, signal_date, buy_signal, sell_signal, score, reason)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (symbol, now_iso(), buy_signal, sell_signal, score, "threshold_match"),
                )

        conn.commit()
    ensure_latest_prices_nonzero_from_last_trade()
    refresh_strategy_analytics(force=bool(force_strategy))


def date_range_from_basis(basis, from_s, to_s):
    today = dt.date.today()
    basis = (basis or "ytd").lower()
    if to_s:
        end = dt.date.fromisoformat(to_s)
    else:
        end = today
    if basis == "custom" and from_s:
        start = dt.date.fromisoformat(from_s)
    elif basis == "mtd":
        start = end.replace(day=1)
    elif basis == "qtd":
        q_month = ((end.month - 1) // 3) * 3 + 1
        start = dt.date(end.year, q_month, 1)
    elif basis == "1y":
        start = end - dt.timedelta(days=365)
    elif basis == "all":
        start = dt.date(2000, 1, 1)
    else:
        start = dt.date(end.year, 1, 1)
    return start, end


def build_snapshot_for_symbol(conn, symbol, up_to_date, split_map=None):
    if split_map is None:
        split_map = load_split_map(conn)
    symbol_q = symbol_upper(symbol)
    trades = conn.execute(
        """
        SELECT side, trade_date, quantity, price
        FROM trades
        WHERE UPPER(symbol) = ? AND trade_date <= ?
        ORDER BY trade_date, id
        """,
        (symbol_q, up_to_date.isoformat()),
    ).fetchall()
    qty = 0.0
    avg = 0.0
    realized = 0.0
    for t in trades:
        q, p = adjusted_trade_values(
            symbol, t["trade_date"], float(t["quantity"]), float(t["price"]), split_map
        )
        if t["side"] == "BUY":
            old = qty
            qty += q
            if qty > 0:
                avg = ((avg * old) + (p * q)) / qty
        else:
            m = min(q, qty)
            if m > 0:
                realized += (p - avg) * m
            qty = max(0.0, qty - q)
            if qty == 0:
                avg = 0.0
    ltp = get_effective_ltp(conn, symbol, split_map)
    invested = qty * avg
    unrealized = qty * ltp - invested
    return {
        "qty": qty,
        "avg_cost": avg,
        "invested": invested,
        "realized": realized,
        "unrealized": unrealized,
        "ltp": ltp,
    }


def open_lots_for_symbol(conn, symbol, split_map=None):
    if split_map is None:
        split_map = load_split_map(conn)
    symbol_q = symbol_upper(symbol)
    rows = conn.execute(
        """
        SELECT id, side, trade_date, quantity, price
        FROM trades
        WHERE UPPER(symbol) = ?
        ORDER BY trade_date, id
        """,
        (symbol_q,),
    ).fetchall()
    lots = deque()
    for r in rows:
        q, p = adjusted_trade_values(
            symbol, r["trade_date"], float(r["quantity"]), float(r["price"]), split_map
        )
        if q <= 0:
            continue
        if r["side"] == "BUY":
            lots.append(
                {
                    "trade_id": int(r["id"]),
                    "buy_date": r["trade_date"],
                    "buy_price": float(p),
                    "qty": float(q),
                }
            )
        else:
            rem = float(q)
            while rem > 1e-9 and lots:
                first = lots[0]
                matched = min(rem, first["qty"])
                first["qty"] -= matched
                rem -= matched
                if first["qty"] <= 1e-9:
                    lots.popleft()
    return [l for l in lots if l["qty"] > 1e-9]


def simulate_sell_for_symbol(conn, symbol, quantity, sell_price=None):
    split_map = load_split_map(conn)
    qty_req = float(quantity)
    if qty_req <= 0:
        raise ValueError("quantity_must_be_positive")

    ltp = get_effective_ltp(conn, symbol, split_map)
    px = parse_float(sell_price, 0.0)
    if px <= 0:
        px = float(ltp)
    if px <= 0:
        raise ValueError("sell_price_or_ltp_required")

    lots = open_lots_for_symbol(conn, symbol, split_map)
    available = sum(float(l["qty"]) for l in lots)
    remaining = qty_req
    lines = []
    total_cost = 0.0
    total_proceeds = 0.0
    total_profit = 0.0

    for lot in lots:
        if remaining <= 1e-9:
            break
        q = min(remaining, float(lot["qty"]))
        if q <= 0:
            continue
        buy_price = float(lot["buy_price"])
        cost = q * buy_price
        proceeds = q * px
        profit = proceeds - cost
        lines.append(
            {
                "buy_trade_id": lot["trade_id"],
                "buy_date": lot["buy_date"],
                "buy_price": round(buy_price, 4),
                "qty_sold": round(q, 4),
                "cost": round(cost, 2),
                "sell_price": round(px, 4),
                "proceeds": round(proceeds, 2),
                "profit": round(profit, 2),
            }
        )
        total_cost += cost
        total_proceeds += proceeds
        total_profit += profit
        remaining -= q

    matched = qty_req - max(remaining, 0.0)
    total_profit_pct = (total_profit / total_cost * 100.0) if total_cost > 0 else 0.0
    return {
        "symbol": symbol,
        "requested_qty": round(qty_req, 4),
        "available_qty": round(available, 4),
        "matched_qty": round(matched, 4),
        "unmatched_qty": round(max(remaining, 0.0), 4),
        "sell_price": round(px, 4),
        "ltp": round(float(ltp), 4),
        "total_cost": round(total_cost, 2),
        "total_proceeds": round(total_proceeds, 2),
        "total_profit": round(total_profit, 2),
        "total_profit_pct": round(total_profit_pct, 2),
        "lines": lines,
    }


def symbol_performance(conn, symbol, basis, from_s, to_s):
    start, end = date_range_from_basis(basis, from_s, to_s)
    split_map = load_split_map(conn)
    snap_s = build_snapshot_for_symbol(conn, symbol, start - dt.timedelta(days=1), split_map)
    snap_e = build_snapshot_for_symbol(conn, symbol, end, split_map)
    realized_delta = snap_e["realized"] - snap_s["realized"]
    unreal_delta = snap_e["unrealized"] - snap_s["unrealized"]
    pnl = realized_delta + unreal_delta
    base = snap_s["invested"] if snap_s["invested"] > 0 else max(snap_e["invested"], 1.0)
    return_pct = (pnl / base) * 100.0
    return {
        "symbol": symbol,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "pnl": round(pnl, 2),
        "return_pct": round(return_pct, 2),
        "realized_delta": round(realized_delta, 2),
        "unrealized_delta": round(unreal_delta, 2),
        "start_snapshot": snap_s,
        "end_snapshot": snap_e,
    }


def _parse_iso_date_basic(value):
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s[:10])
    except ValueError:
        return None


def _xnpv(rate, cashflows):
    if rate <= -0.999999:
        return float("inf")
    base_date = cashflows[0][0]
    base = 1.0 + rate
    total = 0.0
    for flow_date, amount in cashflows:
        years = max(0.0, (flow_date - base_date).days / 365.25)
        try:
            total += float(amount) / (base ** years)
        except (OverflowError, ZeroDivisionError):
            return float("inf") if amount > 0 else float("-inf")
    return total


def _solve_xirr(cashflows):
    flows = [(d, parse_float(a, 0.0)) for d, a in cashflows if d is not None and abs(parse_float(a, 0.0)) > 1e-9]
    flows.sort(key=lambda x: x[0])
    if len(flows) < 2:
        return None
    has_pos = any(a > 0 for _, a in flows)
    has_neg = any(a < 0 for _, a in flows)
    if not (has_pos and has_neg):
        return None

    lo = -0.9999
    hi = 1.0
    f_lo = _xnpv(lo, flows)
    f_hi = _xnpv(hi, flows)
    tries = 0
    while tries < 60 and (not math.isfinite(f_lo) or not math.isfinite(f_hi) or f_lo * f_hi > 0):
        hi = hi * 1.8 + 0.2
        if hi > 1000:
            break
        f_hi = _xnpv(hi, flows)
        tries += 1
    if (not math.isfinite(f_lo)) or (not math.isfinite(f_hi)) or (f_lo * f_hi > 0):
        return None

    for _ in range(120):
        mid = (lo + hi) / 2.0
        f_mid = _xnpv(mid, flows)
        if not math.isfinite(f_mid):
            lo = mid
            continue
        if abs(f_mid) <= 1e-7:
            return mid
        if f_lo * f_mid <= 0:
            hi = mid
            f_hi = f_mid
        else:
            lo = mid
            f_lo = f_mid
    return (lo + hi) / 2.0


def _portfolio_return_metrics(conn, market_value, cash_balance):
    rows = conn.execute(
        """
        SELECT entry_date, LOWER(entry_type) AS entry_type, amount
        FROM cash_ledger
        WHERE LOWER(entry_type) IN ('deposit','withdrawal')
        ORDER BY entry_date, id
        """
    ).fetchall()
    if not rows:
        return {"cagr_pct": 0.0, "xirr_pct": 0.0}

    investor_flows = []
    portfolio_flows = []
    for r in rows:
        d = _parse_iso_date_basic(r["entry_date"])
        if not d:
            continue
        et = str(r["entry_type"] or "").strip().lower()
        amt = parse_float(r["amount"], 0.0)
        if et == "deposit":
            investor_amt = -abs(amt)
            portfolio_amt = abs(amt)
        elif et == "withdrawal":
            investor_amt = abs(amt)
            portfolio_amt = -abs(amt)
        else:
            continue
        if abs(investor_amt) <= 1e-9:
            continue
        investor_flows.append((d, investor_amt))
        portfolio_flows.append((d, portfolio_amt))

    if not investor_flows:
        return {"cagr_pct": 0.0, "xirr_pct": 0.0}

    as_of = dt.date.today()
    ending_value = parse_float(market_value, 0.0) + parse_float(cash_balance, 0.0)

    xirr_rate = None
    if abs(ending_value) > 1e-9:
        xirr_rate = _solve_xirr(investor_flows + [(as_of, ending_value)])

    cagr_rate = None
    start_date = portfolio_flows[0][0] if portfolio_flows else None
    if start_date and ending_value > 0:
        total_days = max((as_of - start_date).days, 1)
        years = total_days / 365.25
        if years > 0:
            total_external = sum(v for _, v in portfolio_flows)
            weighted_capital = 0.0
            for flow_date, flow_amt in portfolio_flows:
                weight = clamp((as_of - flow_date).days / total_days, 0.0, 1.0)
                weighted_capital += flow_amt * weight
            if abs(weighted_capital) > 1e-9:
                period_return = (ending_value - total_external) / weighted_capital
                growth_factor = 1.0 + period_return
                if growth_factor > 0:
                    cagr_rate = growth_factor ** (1.0 / years) - 1.0
            if cagr_rate is None and total_external > 0:
                growth_factor = ending_value / total_external
                if growth_factor > 0:
                    cagr_rate = growth_factor ** (1.0 / years) - 1.0

    return {
        "cagr_pct": round((cagr_rate or 0.0) * 100.0, 2),
        "xirr_pct": round((xirr_rate or 0.0) * 100.0, 2),
    }


def portfolio_summary(conn):
    row = conn.execute(
        """
        SELECT
          COALESCE(SUM(invested),0) AS market_deployment,
          COALESCE(SUM(market_value),0) AS market_value,
          COALESCE(SUM(realized_pnl),0) AS realized_pnl,
          COALESCE(SUM(unrealized_pnl),0) AS unrealized_pnl
        FROM holdings h
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(h.symbol)
        WHERE UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
        """
    ).fetchone()
    cash_row = conn.execute(
        """
        SELECT
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'deposit' THEN amount ELSE 0 END),0) AS deposits_total,
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'withdrawal' THEN -amount ELSE 0 END),0) AS withdrawals_total,
          COALESCE(SUM(amount),0) AS cash_balance
        FROM cash_ledger
        """
    ).fetchone()
    market_deployment = float(row["market_deployment"])
    market = float(row["market_value"])
    realized = float(row["realized_pnl"])
    unrealized = float(row["unrealized_pnl"])
    deposits_total = parse_float(cash_row["deposits_total"], 0.0)
    withdrawals_total = parse_float(cash_row["withdrawals_total"], 0.0)
    hand_invested = deposits_total - withdrawals_total
    total_pnl = realized + unrealized
    ret_pct = (total_pnl / hand_invested * 100.0) if hand_invested > 0 else 0.0
    deployment_ret_pct = (total_pnl / market_deployment * 100.0) if market_deployment > 0 else 0.0
    day_rows = conn.execute(
        """
        SELECT
          h.symbol,
          COALESCE(h.qty,0) AS qty,
          COALESCE(lp.ltp,0) AS ltp,
          COALESCE(lp.change_abs,0) AS change_abs
        FROM holdings h
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(h.symbol)
        LEFT JOIN latest_prices lp ON lp.symbol = h.symbol
        WHERE UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
        """
    ).fetchall()
    today_pnl = 0.0
    prev_day_value = 0.0
    prev_close_map = load_prev_close_map(conn, [r["symbol"] for r in day_rows])
    split_map = load_split_map(conn)
    for r in day_rows:
        symbol = symbol_upper(r["symbol"])
        qty = parse_float(r["qty"], 0.0)
        ltp = parse_float(r["ltp"], 0.0)
        if ltp <= 0:
            ltp = get_effective_ltp(conn, r["symbol"], split_map)
        change_abs = resolve_effective_change_abs(conn, symbol, ltp, r["change_abs"], prev_close_map=prev_close_map)
        today_pnl += qty * change_abs
        prev_close = ltp - change_abs
        if prev_close <= 0:
            prev_close = parse_float(prev_close_map.get(symbol), 0.0)
        if prev_close > 0:
            prev_day_value += qty * prev_close
    today_change_pct = (today_pnl / prev_day_value * 100.0) if prev_day_value > 0 else 0.0
    return_metrics = _portfolio_return_metrics(conn, market, parse_float(cash_row["cash_balance"], 0.0))
    return {
        "invested": round(hand_invested, 2),
        "hand_invested": round(hand_invested, 2),
        "market_deployment": round(market_deployment, 2),
        "market_value": round(market, 2),
        "realized_pnl": round(realized, 2),
        "unrealized_pnl": round(unrealized, 2),
        "total_pnl": round(total_pnl, 2),
        "total_return_pct": round(ret_pct, 2),
        "deployment_return_pct": round(deployment_ret_pct, 2),
        "today_pnl": round(today_pnl, 2),
        "today_change_pct": round(today_change_pct, 2),
        "cagr_pct": round(parse_float(return_metrics.get("cagr_pct"), 0.0), 2),
        "xirr_pct": round(parse_float(return_metrics.get("xirr_pct"), 0.0), 2),
        "cash_balance": round(parse_float(cash_row["cash_balance"], 0.0), 2),
        "as_of": now_iso(),
    }


def cashflow_summary(conn):
    row = conn.execute(
        """
        SELECT
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'deposit' THEN amount ELSE 0 END),0) AS deposits_total,
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'withdrawal' THEN -amount ELSE 0 END),0) AS withdrawals_total,
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'investment' THEN -amount ELSE 0 END),0) AS investment_spend_total,
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'charge' THEN -amount ELSE 0 END),0) AS charges_total,
          COALESCE(SUM(CASE WHEN LOWER(entry_type) = 'trade_credit' THEN amount ELSE 0 END),0) AS trade_credit_total,
          COALESCE(SUM(amount),0) AS cash_balance,
          COUNT(*) AS entries,
          MIN(entry_date) AS from_date,
          MAX(entry_date) AS to_date
        FROM cash_ledger
        """
    ).fetchone()
    deposits_total = parse_float(row["deposits_total"], 0.0)
    withdrawals_total = parse_float(row["withdrawals_total"], 0.0)
    investment_spend_total = parse_float(row["investment_spend_total"], 0.0)
    charges_total = parse_float(row["charges_total"], 0.0)
    net_hand_investment_total = deposits_total - withdrawals_total
    net_hand_after_charges_total = net_hand_investment_total - charges_total
    return {
        "entries": int(row["entries"]),
        "deposits_total": round(deposits_total, 2),
        "withdrawals_total": round(withdrawals_total, 2),
        "net_hand_investment_total": round(net_hand_investment_total, 2),
        "net_hand_after_charges_total": round(net_hand_after_charges_total, 2),
        "investment_spend_total": round(investment_spend_total, 2),
        "charges_total": round(charges_total, 2),
        "trade_credit_total": round(parse_float(row["trade_credit_total"], 0.0), 2),
        "invested_plus_charges_total": round(investment_spend_total + charges_total, 2),
        "cash_balance": round(parse_float(row["cash_balance"], 0.0), 2),
        "from_date": row["from_date"],
        "to_date": row["to_date"],
    }


def list_cashflows(conn, from_s=None, to_s=None, entry_type=None, text_q=None):
    where = ["1=1"]
    params = []
    if from_s:
        where.append("entry_date >= ?")
        params.append(from_s)
    if to_s:
        where.append("entry_date <= ?")
        params.append(to_s)
    if entry_type and entry_type.lower() in ("deposit", "withdrawal", "investment", "trade_credit", "charge"):
        where.append("LOWER(entry_type) = ?")
        params.append(entry_type.lower())
    if text_q:
        where.append("UPPER(COALESCE(reference_text,'')) LIKE ?")
        params.append(f"%{str(text_q).upper()}%")
    sql = (
        "SELECT id, entry_date, entry_type, amount, reference_text, external_entry_id, source "
        "FROM cash_ledger WHERE "
        + " AND ".join(where)
        + " ORDER BY entry_date DESC, id DESC"
    )
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    return rows


def dividend_summary(conn):
    row = conn.execute(
        """
        SELECT
          COUNT(*) AS entries,
          COALESCE(SUM(amount),0) AS total_dividend,
          MIN(entry_date) AS from_date,
          MAX(entry_date) AS to_date
        FROM dividends
        """
    ).fetchone()
    by_symbol = conn.execute(
        """
        SELECT symbol, COALESCE(SUM(amount),0) AS total_dividend
        FROM dividends
        GROUP BY symbol
        ORDER BY total_dividend DESC, symbol
        LIMIT 5
        """
    ).fetchall()
    by_year = conn.execute(
        """
        SELECT
          STRFTIME('%Y', entry_date) AS year,
          COUNT(*) AS entries,
          COALESCE(SUM(amount), 0) AS total_dividend
        FROM dividends
        WHERE entry_date IS NOT NULL AND TRIM(entry_date) <> ''
        GROUP BY year
        ORDER BY year DESC
        """
    ).fetchall()
    return {
        "entries": int(row["entries"] or 0),
        "total_dividend": round(parse_float(row["total_dividend"], 0.0), 2),
        "from_date": row["from_date"],
        "to_date": row["to_date"],
        "top_symbols": [
            {"symbol": r["symbol"], "total_dividend": round(parse_float(r["total_dividend"], 0.0), 2)}
            for r in by_symbol
        ],
        "by_year": [
            {
                "year": r["year"],
                "entries": int(r["entries"] or 0),
                "total_dividend": round(parse_float(r["total_dividend"], 0.0), 2),
            }
            for r in by_year
        ],
    }


def list_dividends(conn, from_s=None, to_s=None, symbol=None, text_q=None):
    where = ["1=1"]
    params = []
    if from_s:
        where.append("entry_date >= ?")
        params.append(from_s)
    if to_s:
        where.append("entry_date <= ?")
        params.append(to_s)
    if symbol:
        where.append("UPPER(symbol) = ?")
        params.append(symbol_upper(symbol))
    if text_q:
        where.append(
            "(UPPER(COALESCE(reference_text,'')) LIKE ? OR UPPER(COALESCE(external_entry_id,'')) LIKE ?)"
        )
        like_q = f"%{str(text_q).upper()}%"
        params.extend([like_q, like_q])
    sql = (
        "SELECT id, symbol, entry_date, amount, reference_text, external_entry_id, source, created_at "
        "FROM dividends WHERE "
        + " AND ".join(where)
        + " ORDER BY entry_date DESC, id DESC"
    )
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


def dividend_amount_map(conn, from_s=None, to_s=None):
    where = ["1=1"]
    params = []
    if from_s:
        where.append("entry_date >= ?")
        params.append(from_s)
    if to_s:
        where.append("entry_date <= ?")
        params.append(to_s)
    rows = conn.execute(
        """
        SELECT UPPER(symbol) AS symbol, COALESCE(SUM(amount),0) AS total
        FROM dividends
        WHERE """
        + " AND ".join(where)
        + " GROUP BY UPPER(symbol)"
        ,
        params,
    ).fetchall()
    return {symbol_upper(r["symbol"]): parse_float(r["total"], 0.0) for r in rows}


def apply_dividend_adjustment(item, dividend_total=0.0):
    item["dividend_amount"] = round(parse_float(dividend_total, 0.0), 2)
    return item


def portfolio_performance(conn, basis, from_s, to_s):
    symbols = [
        r["symbol"]
        for r in conn.execute(
            "SELECT symbol FROM instruments WHERE UPPER(COALESCE(asset_class, 'EQUITY')) <> 'GOLD'"
        ).fetchall()
    ]
    details = []
    total_pnl = 0.0
    total_base = 0.0
    start, end = date_range_from_basis(basis, from_s, to_s)
    for symbol in symbols:
        p = symbol_performance(conn, symbol, basis, from_s, to_s)
        details.append(p)
        total_pnl += p["pnl"]
        snap_s = p["start_snapshot"]["invested"]
        snap_e = p["end_snapshot"]["invested"]
        total_base += snap_s if snap_s > 0 else max(snap_e, 1.0)
    ret = (total_pnl / total_base * 100.0) if total_base > 0 else 0.0
    return {
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "pnl": round(total_pnl, 2),
        "return_pct": round(ret, 2),
        "details": details,
    }


def portfolio_timeseries(conn, from_s=None, to_s=None):
    today = dt.date.today()
    end = dt.date.fromisoformat(to_s) if to_s else today
    if from_s:
        start = dt.date.fromisoformat(from_s)
    else:
        min_row = conn.execute(
            """
            SELECT MIN(t.trade_date) AS d
            FROM trades t
            LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(t.symbol)
            WHERE UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
            """
        ).fetchone()
        min_d = str(min_row["d"] or "").strip() if min_row else ""
        if min_d:
            try:
                start = dt.date.fromisoformat(min_d)
            except ValueError:
                start = end - dt.timedelta(days=365)
        else:
            start = end - dt.timedelta(days=365)
    rows = conn.execute(
        """
        SELECT t.symbol, t.side, t.trade_date, t.quantity, t.price
        FROM trades t
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(t.symbol)
        WHERE t.trade_date >= ? AND t.trade_date <= ?
          AND UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
        ORDER BY t.trade_date, t.id
        """,
        (start.isoformat(), end.isoformat()),
    ).fetchall()
    by_day = defaultdict(list)
    for r in rows:
        by_day[r["trade_date"]].append(r)

    pre_rows = conn.execute(
        """
        SELECT t.symbol, t.side, t.trade_date, t.quantity, t.price
        FROM trades t
        LEFT JOIN instruments i ON UPPER(i.symbol) = UPPER(t.symbol)
        WHERE t.trade_date < ?
          AND UPPER(COALESCE(i.asset_class, 'EQUITY')) <> 'GOLD'
        ORDER BY t.trade_date, t.id
        """,
        (start.isoformat(),),
    ).fetchall()
    split_map = load_split_map(conn)
    cash_rows = conn.execute(
        """
        SELECT entry_date, entry_type, amount
        FROM cash_ledger
        WHERE entry_date >= ? AND entry_date <= ?
          AND LOWER(entry_type) IN ('deposit','withdrawal')
        ORDER BY entry_date, id
        """,
        (start.isoformat(), end.isoformat()),
    ).fetchall()
    cash_by_day = defaultdict(list)
    for r in cash_rows:
        cash_by_day[str(r["entry_date"])].append(r)

    pre_cash_rows = conn.execute(
        """
        SELECT entry_type, amount
        FROM cash_ledger
        WHERE entry_date < ?
          AND LOWER(entry_type) IN ('deposit','withdrawal')
        ORDER BY entry_date, id
        """,
        (start.isoformat(),),
    ).fetchall()

    def hand_delta(entry_type, amount):
        et = str(entry_type or "").strip().lower()
        amt = parse_float(amount, 0.0)
        if et == "deposit":
            return abs(amt)
        if et == "withdrawal":
            return -abs(amt)
        return 0.0

    state = defaultdict(lambda: {"qty": 0.0, "avg": 0.0, "realized": 0.0})
    hand_invested = 0.0
    for r in pre_cash_rows:
        hand_invested += hand_delta(r["entry_type"], r["amount"])
    for r in pre_rows:
        symbol = symbol_upper(r["symbol"])
        s = state[symbol]
        q, p = adjusted_trade_values(
            symbol, r["trade_date"], float(r["quantity"]), float(r["price"]), split_map
        )
        if r["side"] == "BUY":
            old = s["qty"]
            s["qty"] += q
            if s["qty"] > 0:
                s["avg"] = ((s["avg"] * old) + (p * q)) / s["qty"]
        else:
            m = min(q, s["qty"])
            if m > 0:
                s["realized"] += (p - s["avg"]) * m
            s["qty"] = max(0.0, s["qty"] - q)
            if s["qty"] == 0:
                s["avg"] = 0.0

    latest_ltp_map = {
        symbol_upper(r["symbol"]): parse_float(r["ltp"], 0.0)
        for r in conn.execute("SELECT symbol, ltp FROM latest_prices").fetchall()
    }
    symbols_scope = sorted(
        {
            symbol_upper(r["symbol"])
            for r in list(rows) + list(pre_rows)
            if symbol_upper(r["symbol"])
        }
    )
    daily_ltp_map = defaultdict(dict)  # symbol -> {YYYY-MM-DD: daily close}
    pre_start_ltp = {}  # symbol -> last known close before start date
    start_s = start.isoformat()
    if symbols_scope:
        placeholders = ",".join(["?"] * len(symbols_scope))
        with market_db_connect() as mconn:
            hist_rows = mconn.execute(
                f"""
                SELECT symbol, price_date, close
                FROM daily_prices
                WHERE UPPER(symbol) IN ({placeholders})
                  AND price_date <= ?
                  AND close > 0
                ORDER BY symbol, price_date
                """,
                symbols_scope + [end.isoformat()],
            ).fetchall()
        for r in hist_rows:
            symbol = symbol_upper(r["symbol"])
            day_s = str(r["price_date"] or "")
            ltp = parse_float(r["close"], 0.0)
            if not symbol or not day_s or ltp <= 0:
                continue
            if day_s < start_s:
                pre_start_ltp[symbol] = ltp
            else:
                daily_ltp_map[symbol][day_s] = ltp
    rolling_ltp = dict(pre_start_ltp)

    points = []
    cur = start
    while cur <= end:
        d = cur.isoformat()
        for r in cash_by_day.get(d, []):
            hand_invested += hand_delta(r["entry_type"], r["amount"])
        for r in by_day.get(d, []):
            symbol = symbol_upper(r["symbol"])
            s = state[symbol]
            q, p = adjusted_trade_values(
                symbol, r["trade_date"], float(r["quantity"]), float(r["price"]), split_map
            )
            if r["side"] == "BUY":
                old = s["qty"]
                s["qty"] += q
                if s["qty"] > 0:
                    s["avg"] = ((s["avg"] * old) + (p * q)) / s["qty"]
            else:
                m = min(q, s["qty"])
                if m > 0:
                    s["realized"] += (p - s["avg"]) * m
                s["qty"] = max(0.0, s["qty"] - q)
                if s["qty"] == 0:
                    s["avg"] = 0.0

        invested = 0.0
        market = 0.0
        realized = 0.0
        is_end_day = (d == end.isoformat())
        for symbol, s in state.items():
            day_ltp = daily_ltp_map.get(symbol, {}).get(d)
            if day_ltp and day_ltp > 0:
                rolling_ltp[symbol] = day_ltp
            ltp = parse_float(rolling_ltp.get(symbol), 0.0)
            if is_end_day:
                end_ltp = parse_float(latest_ltp_map.get(symbol), 0.0)
                if end_ltp > 0:
                    ltp = end_ltp
            if ltp <= 0:
                ltp = parse_float(s["avg"], 0.0)
            if ltp > 0:
                rolling_ltp[symbol] = ltp
            invested += s["qty"] * s["avg"]
            market += s["qty"] * ltp
            realized += s["realized"]
        points.append(
            {
                "date": d,
                "investment": round(hand_invested, 2),
                "invested": round(invested, 2),
                "hand_invested": round(hand_invested, 2),
                "market_value": round(market, 2),
                "realized_pnl": round(realized, 2),
                "total_value": round(market + realized, 2),
            }
        )
        cur += dt.timedelta(days=1)
    return points


def json_response(handler, payload, status=HTTPStatus.OK):
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    def _is_client_disconnect_error(exc):
        if isinstance(exc, (BrokenPipeError, ConnectionResetError, ConnectionAbortedError)):
            return True
        if isinstance(exc, OSError):
            if getattr(exc, "errno", None) in (errno.EPIPE, errno.ECONNRESET, errno.ECONNABORTED):
                return True
            if getattr(exc, "winerror", None) in (10053, 10054):
                return True
        return False

    try:
        handler.send_response(status)
        handler.send_header("Content-Type", "application/json; charset=utf-8")
        handler.send_header("Content-Length", str(len(raw)))
        handler.send_header("Cache-Control", "no-store")
        handler.send_header("Access-Control-Allow-Origin", "*")
        handler.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")
        handler.send_header("Access-Control-Allow-Headers", "Content-Type")
        handler.end_headers()
        handler.wfile.write(raw)
    except Exception as exc:
        if _is_client_disconnect_error(exc):
            return
        raise


class AppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(WEB_DIR), **kwargs)

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/v1/"):
            try:
                self._activate_request_tenant(parsed)
                self.handle_api_get(parsed)
            except ValueError as exc:
                json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            except Exception as exc:
                json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
            finally:
                clear_request_tenant_key()
            return

        if parsed.path in ("/", ""):
            return super().do_GET()
        target = WEB_DIR / parsed.path.lstrip("/")
        if target.exists() and target.is_file():
            return super().do_GET()
        self.path = "/index.html"
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/v1/"):
            json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)
            return
        try:
            self._activate_request_tenant(parsed)
            self.handle_api_post(parsed)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
        finally:
            clear_request_tenant_key()

    def do_PUT(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/v1/"):
            json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)
            return
        try:
            self._activate_request_tenant(parsed)
            self.handle_api_put(parsed)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
        finally:
            clear_request_tenant_key()

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/v1/"):
            json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)
            return
        try:
            self._activate_request_tenant(parsed)
            self.handle_api_delete(parsed)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
        finally:
            clear_request_tenant_key()

    def _read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        if not raw:
            return {}
        return json.loads(raw.decode("utf-8"))

    def _activate_request_tenant(self, parsed):
        qs = parse_qs(parsed.query or "")
        tenant_raw = qs.get("tenant", [None])[0]
        if tenant_raw is None:
            tenant_raw = self.headers.get("X-Tenant")
        if tenant_raw is None or not str(tenant_raw).strip():
            return set_request_tenant_key(get_active_tenant_key())
        key = sanitize_tenant_key(tenant_raw)
        if not key:
            raise ValueError("invalid_tenant")
        valid = {t["key"] for t in load_tenant_meta().get("tenants", [])}
        if key not in valid:
            raise ValueError("tenant_not_found")
        return set_request_tenant_key(key)

    def handle_api_get(self, parsed):
        path = parsed.path
        qs = parse_qs(parsed.query)

        if path == "/api/v1/tenants":
            json_response(
                self,
                {
                    "max_tenants": TENANT_MAX_COUNT,
                    "active_tenant": get_active_tenant_key(),
                    "request_tenant": get_current_tenant_key(),
                    "items": list_tenants(),
                },
            )
            return

        if path == "/api/v1/tenants/active":
            current = get_current_tenant_key()
            p = tenant_paths(current)
            json_response(
                self,
                {
                    "max_tenants": TENANT_MAX_COUNT,
                    "active_tenant": get_active_tenant_key(),
                    "request_tenant": current,
                    "paths": {
                        "data_dir": str(p["data_dir"]),
                        "db_path": str(p["db_path"]),
                        "market_db_path": str(p["market_db_path"]),
                        "upload_dir": str(p["upload_dir"]),
                        "backup_dir": str(p["backup_dir"]),
                    },
                },
            )
            return

        with db_connect() as conn:
            if path == "/api/v1/health":
                c1 = conn.execute("SELECT COUNT(*) AS c FROM trades").fetchone()["c"]
                c2 = conn.execute("SELECT COUNT(*) AS c FROM holdings").fetchone()["c"]
                json_response(
                    self,
                    {
                        "ok": True,
                        "tenant": get_current_tenant_key(),
                        "trades": c1,
                        "holdings": c2,
                    },
                )
                return

            if path == "/api/v1/scrips":
                split_map = load_split_map(conn)
                peak_split_map = load_peak_split_map(conn)
                div_total_map = dividend_amount_map(conn)
                rows = conn.execute(
                    """
                    SELECT
                      i.symbol,
                      i.exchange,
                      COALESCE(i.asset_class, 'EQUITY') AS asset_class,
                      i.feed_code,
                      COALESCE(h.qty,0) AS qty,
                      COALESCE(h.avg_cost,0) AS avg_cost,
                      COALESCE(h.invested,0) AS invested,
                      COALESCE(h.market_value,0) AS market_value,
                      COALESCE(h.realized_pnl,0) AS realized_pnl,
                      COALESCE(h.unrealized_pnl,0) AS unrealized_pnl,
                      COALESCE(h.total_return_pct,0) AS total_return_pct,
                      COALESCE((
                        SELECT sr.action
                        FROM strategy_recommendations sr
                        WHERE UPPER(sr.symbol) = UPPER(i.symbol)
                          AND sr.run_date = (SELECT MAX(run_date) FROM strategy_runs)
                        ORDER BY sr.priority DESC, sr.id DESC
                        LIMIT 1
                      ), '') AS strategy_action,
                      COALESCE((SELECT COUNT(*) FROM trades t WHERE UPPER(t.symbol)=UPPER(i.symbol)),0) AS trade_count,
                      COALESCE(lp.ltp,0) AS ltp,
                      COALESCE(lp.change_abs,0) AS change_abs,
                      lp.updated_at AS price_updated_at,
                      COALESCE(g.min_value,0) AS min_value,
                      g.max_value AS max_value
                    FROM instruments i
                    LEFT JOIN holdings h ON h.symbol = i.symbol
                    LEFT JOIN latest_prices lp ON lp.symbol = i.symbol
                    LEFT JOIN scrip_position_guards g ON UPPER(g.symbol) = UPPER(i.symbol)
                    WHERE i.active = 1
                    ORDER BY market_value DESC, i.symbol
                    """
                ).fetchall()
                signals = {
                    r["symbol"]: r
                    for r in conn.execute(
                        """
                        SELECT symbol, buy_signal, sell_signal, score
                        FROM signals
                        ORDER BY signal_date DESC
                        """
                    ).fetchall()
                }
                prev_close_map = load_prev_close_map(conn, [r["symbol"] for r in rows])
                out = []
                for r in rows:
                    item = dict(r)
                    if parse_float(item.get("ltp"), 0.0) <= 0:
                        item["ltp"] = round(
                            get_effective_ltp_for_asset(
                                conn,
                                r["symbol"],
                                asset_class=item.get("asset_class"),
                                split_map=split_map,
                            ),
                            4,
                        )
                    su = symbol_upper(item.get("symbol"))
                    apply_dividend_adjustment(item, dividend_total=div_total_map.get(su, 0.0))
                    sig = signals.get(r["symbol"])
                    item["buy_signal"] = sig["buy_signal"] if sig else None
                    item["sell_signal"] = sig["sell_signal"] if sig else None
                    item["signal_score"] = round(sig["score"], 2) if sig else None
                    peak = peak_traded_metrics(conn, r["symbol"], peak_split_map)
                    item["peak_traded_price"] = round(peak["peak_traded_price"], 4)
                    item["pct_from_peak_traded"] = round(peak["pct_from_peak_traded"], 2)
                    item["peak_buy_price"] = round(peak["peak_buy_price"], 4)
                    item["pct_from_peak_buy"] = round(peak["pct_from_peak_buy"], 2)
                    enrich_holding_metrics(item, conn=conn, prev_close_map=prev_close_map)
                    out.append(item)
                json_response(self, {"items": out})
                return

            if path == "/api/v1/scrips/position-guards":
                symbols_q = str(qs.get("symbols", [""])[0] or "").strip()
                symbols = [symbol_upper(x) for x in symbols_q.split(",") if symbol_upper(x)] if symbols_q else None
                items = list_scrip_position_guards(conn, symbols=symbols)
                json_response(self, {"items": items, "count": len(items)})
                return

            if path == "/api/v1/rebalance/lot/active":
                active_lot = get_active_rebalance_lot(conn)
                json_response(self, _rebalance_lot_payload(conn, active_lot))
                return

            if path == "/api/v1/rebalance/closed-history":
                limit_raw = qs.get("limit", ["250"])[0]
                include_raw = qs.get("include_completed", ["0"])[0]
                try:
                    include_completed = parse_bool(include_raw, default=False)
                except ValueError:
                    json_response(self, {"error": "include_completed_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
                out = list_rebalance_closed_history(conn, limit=limit_raw, include_buyback_completed=include_completed)
                json_response(self, out)
                return

            if path == "/api/v1/rebalance/suggestions":
                active_lot = get_active_rebalance_lot(conn)
                if active_lot:
                    json_response(self, _rebalance_lot_payload(conn, active_lot))
                    return
                side = str(qs.get("side", ["SELL"])[0] or "SELL").strip().upper()
                percent_raw = qs.get("percent", ["10"])[0]
                try:
                    out = build_rebalance_suggestions(conn, side=side, percent=percent_raw)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                json_response(self, out)
                return

            if path == "/api/v1/daily-target/plan":
                seed_raw = qs.get("seed_capital", ["10000"])[0]
                target_raw = qs.get("target_profit_pct", ["1"])[0]
                top_n_raw = qs.get("top_n", ["5"])[0]
                recalc_raw = qs.get("recalibrate", ["1"])[0]
                try:
                    recalibrate = parse_bool(recalc_raw, default=True)
                except ValueError:
                    json_response(self, {"error": "recalibrate_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
                try:
                    out = get_or_create_daily_target_plan(
                        conn,
                        seed_capital=seed_raw,
                        target_profit_pct=target_raw,
                        top_n=top_n_raw,
                        recalibrate=recalibrate,
                    )
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                json_response(self, out)
                return

            if path == "/api/v1/daily-target/history":
                limit_raw = qs.get("limit", ["120"])[0]
                date_from_raw = qs.get("date_from", [""])[0]
                date_to_raw = qs.get("date_to", [""])[0]
                state_raw = qs.get("state", ["all"])[0]
                try:
                    out = list_daily_target_plan_history(
                        conn,
                        limit=limit_raw,
                        date_from=date_from_raw,
                        date_to=date_to_raw,
                        state_filter=state_raw,
                    )
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                json_response(self, out)
                return

            if path.startswith("/api/v1/scrips/") and path.endswith("/trades"):
                symbol_req = path.split("/")[4]
                symbol = resolve_symbol(conn, symbol_req)
                if not symbol:
                    json_response(self, {"error": "symbol_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                from_s = qs.get("from", [None])[0]
                to_s = qs.get("to", [None])[0]
                where = ["UPPER(symbol) = ?"]
                params = [symbol_upper(symbol)]
                if from_s:
                    where.append("trade_date >= ?")
                    params.append(from_s)
                if to_s:
                    where.append("trade_date <= ?")
                    params.append(to_s)
                sql = (
                    "SELECT id, symbol, side, trade_date, quantity, price, amount, external_trade_id, notes "
                    "FROM trades WHERE " + " AND ".join(where) + " ORDER BY trade_date DESC, id DESC"
                )
                ac_row = conn.execute(
                    "SELECT COALESCE(asset_class, 'EQUITY') AS asset_class FROM instruments WHERE UPPER(symbol)=? LIMIT 1",
                    (symbol_upper(symbol),),
                ).fetchone()
                current_ltp = round(
                    get_effective_ltp_for_asset(
                        conn,
                        symbol,
                        asset_class=(ac_row["asset_class"] if ac_row else ASSET_CLASS_EQUITY),
                    ),
                    4,
                )
                split_map = load_split_map(conn)
                symbol_q = symbol_upper(symbol)
                timeline = conn.execute(
                    """
                    SELECT id, side, trade_date, quantity, price
                    FROM trades
                    WHERE UPPER(symbol) = ?
                    ORDER BY trade_date, id
                    """,
                    (symbol_q,),
                ).fetchall()
                buy_lots = deque()
                per_trade = {}
                for t in timeline:
                    trade_id = int(parse_float(t["id"], 0.0))
                    side = str(t["side"] or "").strip().upper()
                    q_raw = parse_float(t["quantity"], 0.0)
                    p_raw = parse_float(t["price"], 0.0)
                    q, p = adjusted_trade_values(symbol, t["trade_date"], q_raw, p_raw, split_map)
                    metric = {
                        "current_ltp": current_ltp,
                        "current_pnl": 0.0,
                        "pnl_basis": "unknown",
                        "matched_qty": 0.0,
                        "unmatched_qty": 0.0,
                        "matched_avg_buy_price": None,
                    }
                    if side == "BUY":
                        mtm = 0.0
                        if current_ltp > 0 and q > 0 and p > 0:
                            mtm = (current_ltp - p) * q
                        metric["current_pnl"] = mtm
                        metric["pnl_basis"] = "mark_to_market"
                        if q > 0 and p > 0:
                            buy_lots.append({"qty": q, "buy_price": p})
                    elif side == "SELL":
                        sell_qty = max(0.0, q)
                        rem = sell_qty
                        matched_qty = 0.0
                        matched_cost = 0.0
                        while rem > 1e-9 and buy_lots:
                            first = buy_lots[0]
                            close_qty = min(rem, parse_float(first.get("qty"), 0.0))
                            if close_qty <= 0:
                                buy_lots.popleft()
                                continue
                            first["qty"] = parse_float(first.get("qty"), 0.0) - close_qty
                            rem -= close_qty
                            matched_qty += close_qty
                            matched_cost += close_qty * parse_float(first.get("buy_price"), 0.0)
                            if parse_float(first.get("qty"), 0.0) <= 1e-9:
                                buy_lots.popleft()
                        avg_buy = (matched_cost / matched_qty) if matched_qty > 0 else None
                        realized = ((p * matched_qty) - matched_cost) if matched_qty > 0 else 0.0
                        metric["current_pnl"] = realized
                        metric["pnl_basis"] = "realized_fifo"
                        metric["matched_qty"] = matched_qty
                        metric["unmatched_qty"] = max(0.0, sell_qty - matched_qty)
                        metric["matched_avg_buy_price"] = avg_buy
                    per_trade[trade_id] = metric
                rows = []
                total_current_pnl = 0.0
                total_buy_mtm_pnl = 0.0
                total_sell_realized_pnl = 0.0
                for r in conn.execute(sql, params).fetchall():
                    item = dict(r)
                    tid = int(parse_float(item.get("id"), 0.0))
                    side = str(item.get("side", "")).strip().upper()
                    metric = per_trade.get(
                        tid,
                        {
                            "current_ltp": current_ltp,
                            "current_pnl": 0.0,
                            "pnl_basis": "unknown",
                            "matched_qty": 0.0,
                            "unmatched_qty": 0.0,
                            "matched_avg_buy_price": None,
                        },
                    )
                    current_pnl = parse_float(metric.get("current_pnl"), 0.0)
                    item["current_ltp"] = round(parse_float(metric.get("current_ltp"), current_ltp), 4)
                    item["current_pnl"] = round(current_pnl, 2)
                    item["pnl_basis"] = str(metric.get("pnl_basis", "unknown"))
                    item["matched_qty"] = round(parse_float(metric.get("matched_qty"), 0.0), 4)
                    item["unmatched_qty"] = round(parse_float(metric.get("unmatched_qty"), 0.0), 4)
                    avg_buy = metric.get("matched_avg_buy_price")
                    item["matched_avg_buy_price"] = round(parse_float(avg_buy, 0.0), 4) if avg_buy is not None else None
                    total_current_pnl += current_pnl
                    if side == "BUY":
                        total_buy_mtm_pnl += current_pnl
                    elif side == "SELL":
                        total_sell_realized_pnl += current_pnl
                    rows.append(item)
                json_response(
                    self,
                    {
                        "symbol": symbol,
                        "current_ltp": current_ltp,
                        "total_current_pnl": round(total_current_pnl, 2),
                        "total_buy_mtm_pnl": round(total_buy_mtm_pnl, 2),
                        "total_sell_realized_pnl": round(total_sell_realized_pnl, 2),
                        "items": rows,
                    },
                )
                return

            if path.startswith("/api/v1/scrips/") and path.endswith("/performance"):
                symbol_req = path.split("/")[4]
                symbol = resolve_symbol(conn, symbol_req)
                if not symbol:
                    json_response(self, {"error": "symbol_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                basis = qs.get("basis", ["ytd"])[0]
                from_s = qs.get("from", [None])[0]
                to_s = qs.get("to", [None])[0]
                perf = symbol_performance(conn, symbol, basis, from_s, to_s)
                json_response(self, perf)
                return

            if path.startswith("/api/v1/scrips/") and path.count("/") == 4:
                symbol_req = path.split("/")[4]
                symbol = resolve_symbol(conn, symbol_req)
                if not symbol:
                    json_response(self, {"error": "symbol_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                row = conn.execute(
                    """
                    SELECT
                      i.symbol,
                      i.exchange,
                      COALESCE(i.asset_class, 'EQUITY') AS asset_class,
                      i.feed_code,
                      COALESCE(h.qty,0) AS qty,
                      COALESCE(h.avg_cost,0) AS avg_cost,
                      COALESCE(h.invested,0) AS invested,
                      COALESCE(h.market_value,0) AS market_value,
                      COALESCE(h.realized_pnl,0) AS realized_pnl,
                      COALESCE(h.unrealized_pnl,0) AS unrealized_pnl,
                      COALESCE(h.total_return_pct,0) AS total_return_pct,
                      COALESCE((
                        SELECT sr.action
                        FROM strategy_recommendations sr
                        WHERE UPPER(sr.symbol) = UPPER(i.symbol)
                          AND sr.run_date = (SELECT MAX(run_date) FROM strategy_runs)
                        ORDER BY sr.priority DESC, sr.id DESC
                        LIMIT 1
                      ), '') AS strategy_action,
                      h.updated_at,
                      COALESCE(lp.ltp,0) AS ltp,
                      COALESCE(lp.change_abs,0) AS change_abs,
                      COALESCE(g.min_value,0) AS min_value,
                      g.max_value AS max_value
                    FROM instruments i
                    LEFT JOIN holdings h ON h.symbol = i.symbol
                    LEFT JOIN latest_prices lp ON lp.symbol = i.symbol
                    LEFT JOIN scrip_position_guards g ON UPPER(g.symbol) = UPPER(i.symbol)
                    WHERE UPPER(i.symbol) = ?
                    """,
                    (symbol_upper(symbol),),
                ).fetchone()
                if not row:
                    json_response(self, {"error": "symbol_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                signal = conn.execute(
                    "SELECT buy_signal, sell_signal, score, signal_date FROM signals WHERE UPPER(symbol) = ? ORDER BY signal_date DESC LIMIT 1",
                    (symbol_upper(symbol),),
                ).fetchone()
                payload = dict(row)
                if parse_float(payload.get("ltp"), 0.0) <= 0:
                    payload["ltp"] = round(
                        get_effective_ltp_for_asset(
                            conn,
                            payload["symbol"],
                            asset_class=payload.get("asset_class"),
                        ),
                        4,
                    )
                su = symbol_upper(payload.get("symbol"))
                div_total_map = dividend_amount_map(conn)
                apply_dividend_adjustment(payload, dividend_total=div_total_map.get(su, 0.0))
                payload["signal"] = dict(signal) if signal else None
                peak = peak_traded_metrics(conn, payload["symbol"])
                payload["peak_traded_price"] = round(peak["peak_traded_price"], 4)
                payload["pct_from_peak_traded"] = round(peak["pct_from_peak_traded"], 2)
                payload["peak_buy_price"] = round(peak["peak_buy_price"], 4)
                payload["pct_from_peak_buy"] = round(peak["pct_from_peak_buy"], 2)
                enrich_holding_metrics(payload, conn=conn)
                json_response(self, payload)
                return

            if path == "/api/v1/analytics/peak-diff":
                symbols = [r["symbol"] for r in conn.execute("SELECT symbol FROM instruments ORDER BY symbol").fetchall()]
                split_map = load_split_map(conn)
                peak_split_map = load_peak_split_map(conn)
                pending = pending_peak_split_candidates(conn)
                items = []
                for symbol in symbols:
                    ltp = get_effective_ltp(conn, symbol, split_map)
                    peak = peak_traded_metrics(conn, symbol, peak_split_map)
                    items.append(
                        {
                            "symbol": symbol,
                            "ltp": ltp,
                            "peak_traded_price": round(peak["peak_traded_price"], 4),
                            "pct_from_peak_traded": round(peak["pct_from_peak_traded"], 2),
                            "peak_buy_price": round(peak["peak_buy_price"], 4),
                            "pct_from_peak_buy": round(peak["pct_from_peak_buy"], 2),
                        }
                    )
                items.sort(
                    key=lambda x: parse_float(
                        x.get("pct_from_peak_traded"),
                        parse_float(x.get("pct_from_peak_buy"), 0.0),
                    )
                )
                json_response(self, {"items": items, "pending_split_candidates": pending})
                return

            if path == "/api/v1/portfolio/summary":
                json_response(self, portfolio_summary(conn))
                return

            if path == "/api/v1/portfolio/performance":
                basis = qs.get("basis", ["ytd"])[0]
                from_s = qs.get("from", [None])[0]
                to_s = qs.get("to", [None])[0]
                json_response(self, portfolio_performance(conn, basis, from_s, to_s))
                return

            if path == "/api/v1/portfolio/timeseries":
                from_s = qs.get("from", [None])[0]
                to_s = qs.get("to", [None])[0]
                points = portfolio_timeseries(conn, from_s, to_s)
                json_response(self, {"points": points})
                return

            if path == "/api/v1/signals":
                rows = conn.execute(
                    "SELECT symbol, signal_date, buy_signal, sell_signal, score, reason FROM signals ORDER BY score DESC"
                ).fetchall()
                json_response(self, {"items": [dict(r) for r in rows]})
                return

            if path == "/api/v1/strategy/sets":
                sets = [dict(r) for r in conn.execute("SELECT * FROM strategy_sets ORDER BY id").fetchall()]
                for s in sets:
                    params = conn.execute(
                        "SELECT key, value FROM strategy_parameters WHERE set_id = ? ORDER BY key", (s["id"],)
                    ).fetchall()
                    s["parameters"] = [{"key": p["key"], "value": p["value"]} for p in params]
                json_response(self, {"items": sets})
                return

            if path == "/api/v1/strategy/insights":
                insights = load_latest_strategy_insights(conn)
                today = dt.date.today().isoformat()
                needs_refresh = insights is None or insights.get("run_date") != today
                if not needs_refresh and insights is not None:
                    try:
                        created = dt.datetime.fromisoformat(str(insights.get("generated_at")))
                        age_sec = (dt.datetime.now() - created).total_seconds()
                        needs_refresh = age_sec >= get_strategy_refresh_interval(conn)
                        if not needs_refresh:
                            lp_ts = conn.execute("SELECT MAX(updated_at) AS ts FROM latest_prices").fetchone()["ts"]
                            if lp_ts:
                                try:
                                    lp_dt = dt.datetime.fromisoformat(str(lp_ts))
                                    if lp_dt > (created + dt.timedelta(seconds=2)):
                                        needs_refresh = True
                                except Exception:
                                    needs_refresh = True
                    except Exception:
                        needs_refresh = True
                if needs_refresh:
                    fresh = build_strategy_insights(conn, run_date=today)
                    persist_strategy_insights(conn, fresh)
                    conn.commit()
                    insights = load_latest_strategy_insights(conn)
                json_response(self, {"item": insights})
                return

            if path == "/api/v1/strategy/audits":
                limit_s = qs.get("limit", ["25"])[0]
                try:
                    limit_n = max(1, min(200, int(limit_s)))
                except Exception:
                    limit_n = 25
                json_response(self, list_strategy_audit_runs(conn, limit=limit_n))
                return

            if path == "/api/v1/strategy/audits/latest":
                out = list_strategy_audit_runs(conn, limit=1)
                json_response(self, out.get("latest") or {"latest": None})
                return

            if path == "/api/v1/intel/summary":
                limit_s = qs.get("limit", ["40"])[0]
                try:
                    limit_n = max(5, min(200, int(limit_s)))
                except Exception:
                    limit_n = 40
                json_response(self, intelligence_summary(conn, limit=limit_n))
                return

            if path == "/api/v1/harvest/plan":
                target_loss_raw = qs.get("target_loss", ["0"])[0]
                run_llm_raw = qs.get("run_llm", ["0"])[0]
                try:
                    target_loss = max(0.0, parse_float(target_loss_raw, 0.0))
                except Exception:
                    json_response(self, {"error": "target_loss_must_be_numeric"}, HTTPStatus.BAD_REQUEST)
                    return
                try:
                    run_llm = parse_bool(run_llm_raw, default=False)
                except ValueError:
                    json_response(self, {"error": "run_llm_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
                json_response(self, build_tax_harvest_plan(conn, target_loss=target_loss, run_llm=run_llm))
                return

            if path == "/api/v1/loss-lots":
                json_response(self, build_loss_lot_analysis(conn))
                return

            if path == "/api/v1/intel/charts":
                limit_s = qs.get("limit", ["80"])[0]
                symbol_q = qs.get("symbol", [""])[0]
                try:
                    limit_n = max(1, min(500, int(limit_s)))
                except Exception:
                    limit_n = 80
                items = list_chart_snapshots(conn, limit=limit_n, symbol=symbol_q)
                json_response(
                    self,
                    {
                        "config": get_chart_agent_config(conn),
                        "items": items,
                        "count": len(items),
                    },
                )
                return

            if path == "/api/v1/intel/autopilot":
                cfg = get_intel_autopilot_config(conn)
                seed_lines = [x.strip() for x in str(cfg.get("query_seed") or "").splitlines() if x.strip()]
                symbol_q = _collect_autopilot_symbol_queries(conn, symbols_limit=int(cfg.get("symbols_limit", 20)))
                preview_queries = []
                seen = set()
                for q in seed_lines + symbol_q:
                    k = q.lower()
                    if k in seen:
                        continue
                    seen.add(k)
                    preview_queries.append(q)
                    if len(preview_queries) >= 60:
                        break
                json_response(
                    self,
                    {
                        "config": cfg,
                        "preview_queries": preview_queries,
                        "preview_query_count": len(preview_queries),
                    },
                )
                return

            if path == "/api/v1/config/live":
                cfg = get_live_config(conn)
                cfg["source"] = "multi_source_agent"
                json_response(self, cfg)
                return

            if path == "/api/v1/agents/status":
                json_response(self, {"items": build_agents_status(conn)})
                return

            if path == "/api/v1/attention":
                json_response(self, build_attention_console_payload(conn))
                return

            if path == "/api/v1/agents/backtest/history":
                limit_s = qs.get("limit", ["30"])[0]
                try:
                    limit_n = max(1, min(200, int(limit_s)))
                except Exception:
                    limit_n = 30
                json_response(self, {"items": list_backtest_runs(conn, limit=limit_n)})
                return

            if path == "/api/v1/agents/software-performance":
                limit_s = qs.get("limit", ["40"])[0]
                try:
                    limit_n = max(1, min(300, int(limit_s)))
                except Exception:
                    limit_n = 40
                snaps = list_software_perf_snapshots(conn, limit=limit_n)
                acts = list_software_perf_actions(conn, limit=max(20, min(500, limit_n * 2)))
                json_response(
                    self,
                    {
                        "config": get_software_perf_agent_config(conn),
                        "latest": snaps[0] if snaps else None,
                        "snapshots": snaps,
                        "actions": acts,
                    },
                )
                return

            if path == "/api/v1/llm/config":
                json_response(self, get_llm_runtime_config(conn, include_secret=False))
                return

            if path == "/api/v1/agents/risk-analysis":
                limit_s = qs.get("limit", ["40"])[0]
                try:
                    limit_n = max(1, min(300, int(limit_s)))
                except Exception:
                    limit_n = 40
                snaps = list_risk_analysis_snapshots(conn, limit=limit_n)
                json_response(
                    self,
                    {
                        "config": get_risk_agent_config(conn),
                        "latest": snaps[0] if snaps else None,
                        "snapshots": snaps,
                    },
                )
                return

            if path == "/api/v1/prices/status":
                row = conn.execute(
                    "SELECT MAX(updated_at) AS updated_at, COUNT(*) AS c FROM latest_prices"
                ).fetchone()
                json_response(
                    self,
                    {
                        "updated_at": row["updated_at"],
                        "scrips_with_price": row["c"],
                    },
                )
                return

            if path == "/api/v1/prices/sources":
                symbol_q = symbol_upper(qs.get("symbol", [""])[0])
                limit_s = qs.get("limit", ["200"])[0]
                try:
                    limit_n = max(1, min(2000, int(limit_s)))
                except Exception:
                    limit_n = 200
                params = []
                where = ["1=1"]
                if symbol_q:
                    where.append("UPPER(symbol) = ?")
                    params.append(symbol_q)
                rows = conn.execute(
                    """
                    SELECT id, symbol, source, ltp, change_abs, latency_ms, accuracy_error_pct, fetched_at, selected, consensus_ltp
                    FROM quote_samples
                    WHERE """
                    + " AND ".join(where)
                    + " ORDER BY id DESC LIMIT ?",
                    params + [limit_n],
                ).fetchall()
                json_response(self, {"items": [dict(r) for r in rows]})
                return

            if path == "/api/v1/prices/source-ranking":
                policy = get_live_quote_policy(conn)
                ranking = quote_source_ranking(conn)
                nse_order = get_ranked_quote_sources(conn, policy, exchange="NSE")
                bse_order = get_ranked_quote_sources(conn, policy, exchange="BSE")
                json_response(
                    self,
                    {
                        "policy": policy,
                        "nse_order": nse_order,
                        "bse_order": bse_order,
                        "items": ranking,
                    },
                )
                return

            if path == "/api/v1/prices/history/status":
                with market_db_connect() as mconn:
                    h = mconn.execute(
                        """
                        SELECT
                          COUNT(*) AS rows_total,
                          COUNT(DISTINCT symbol) AS symbols_total,
                          MIN(price_date) AS from_date,
                          MAX(price_date) AS to_date
                        FROM daily_prices
                        """
                    ).fetchone()
                cfg = get_history_sync_config(conn)
                json_response(
                    self,
                    {
                        "rows_total": int(h["rows_total"] or 0),
                        "symbols_total": int(h["symbols_total"] or 0),
                        "from_date": h["from_date"],
                        "to_date": h["to_date"],
                        "sync_enabled": cfg["enabled"],
                        "sync_interval_seconds": cfg["interval_seconds"],
                        "last_sync_at": cfg.get("last_sync_at"),
                        "db_path": str(tenant_paths(get_current_tenant_key())["market_db_path"]),
                    },
                )
                return

            if path == "/api/v1/cashflows":
                from_s = qs.get("from", [None])[0]
                to_s = qs.get("to", [None])[0]
                entry_type = qs.get("entry_type", [None])[0]
                text_q = qs.get("q", [None])[0]
                items = list_cashflows(conn, from_s=from_s, to_s=to_s, entry_type=entry_type, text_q=text_q)
                summary = cashflow_summary(conn)
                json_response(self, {"items": items, "summary": summary})
                return

            if path == "/api/v1/dividends":
                from_s = qs.get("from", [None])[0]
                to_s = qs.get("to", [None])[0]
                symbol = qs.get("symbol", [None])[0]
                text_q = qs.get("q", [None])[0]
                items = list_dividends(conn, from_s=from_s, to_s=to_s, symbol=symbol, text_q=text_q)
                summary = dividend_summary(conn)
                json_response(self, {"items": items, "summary": summary})
                return

            if path == "/api/v1/assistant/approvals":
                status = qs.get("status", [None])[0]
                limit = qs.get("limit", [50])[0]
                try:
                    limit_n = int(limit)
                except Exception:
                    limit_n = 50
                items = list_agent_approvals(conn, status=status, limit=limit_n)
                pending_count = int(
                    conn.execute(
                        "SELECT COUNT(*) AS c FROM agent_approvals WHERE LOWER(status) = 'pending'"
                    ).fetchone()["c"]
                )
                json_response(self, {"items": items, "pending_count": pending_count})
                return

            if path == "/api/v1/assistant/verification":
                approval_limit_s = qs.get("approval_limit", ["120"])[0]
                action_limit_s = qs.get("action_limit", ["180"])[0]
                try:
                    approval_limit_n = max(1, min(500, int(approval_limit_s)))
                except Exception:
                    approval_limit_n = 120
                try:
                    action_limit_n = max(10, min(500, int(action_limit_s)))
                except Exception:
                    action_limit_n = 180
                json_response(
                    self,
                    build_approval_verification_payload(
                        conn,
                        approval_limit=approval_limit_n,
                        action_limit=action_limit_n,
                    ),
                )
                return

            if path == "/api/v1/corporate-actions/splits":
                symbol = qs.get("symbol", [None])[0]
                if symbol:
                    rows = conn.execute(
                        """
                        SELECT id, symbol, effective_date, factor, note, created_at
                        FROM corporate_actions
                        WHERE action_type='SPLIT' AND UPPER(symbol) = ?
                        ORDER BY effective_date DESC
                        """,
                        (symbol_upper(symbol),),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        """
                        SELECT id, symbol, effective_date, factor, note, created_at
                        FROM corporate_actions
                        WHERE action_type='SPLIT'
                        ORDER BY symbol, effective_date DESC
                        """
                    ).fetchall()
                json_response(self, {"items": [dict(r) for r in rows]})
                return

        json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)

    def handle_api_post(self, parsed):
        if parsed.path == "/api/v1/tenants":
            body = self._read_json()
            key = body.get("key")
            name = body.get("name")
            try:
                activate = parse_bool(body.get("activate"), default=False)
            except ValueError:
                json_response(self, {"error": "activate_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                tenant = create_tenant(key=key, name=name, activate=activate)
            except ValueError as ex:
                json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                return
            json_response(
                self,
                {
                    "ok": True,
                    "tenant": tenant,
                    "max_tenants": TENANT_MAX_COUNT,
                    "active_tenant": get_active_tenant_key(),
                    "items": list_tenants(),
                },
            )
            return

        if parsed.path == "/api/v1/sync/excel":
            body = self._read_json()
            confirm_replace = parse_bool(body.get("confirm_replace"), default=False)
            if not confirm_replace:
                json_response(
                    self,
                    {"error": "confirm_replace_required", "message": "Set confirm_replace=true to replace DB using workbook import."},
                    HTTPStatus.BAD_REQUEST,
                )
                return
            xlsx_path = body.get("xlsx_path")
            if not xlsx_path:
                json_response(self, {"error": "xlsx_path_required"}, HTTPStatus.BAD_REQUEST)
                return
            import_from_excel(xlsx_path)
            with db_connect() as conn:
                summary = portfolio_summary(conn)
            json_response(self, {"ok": True, "mode": "replace_from_excel", "summary": summary})
            return

        if parsed.path == "/api/v1/assistant/chat":
            body = self._read_json()
            message = body.get("message", "")
            with db_connect() as conn:
                out = assistant_chat_response(conn, message)
            if out.get("executed") and out.get("intent") == "delete_by_notes":
                recompute_holdings_and_signals()
            json_response(self, out)
            return

        if parsed.path == "/api/v1/llm/test":
            with db_connect() as conn:
                try:
                    out = test_llm_runtime(conn)
                except Exception as ex:
                    json_response(
                        self,
                        {
                            "ok": False,
                            "status": "error",
                            "message": str(ex),
                            "config": get_llm_runtime_config(conn, include_secret=False),
                        },
                        HTTPStatus.BAD_REQUEST,
                    )
                    return
            json_response(self, out)
            return

        if parsed.path == "/api/v1/agents/backtest/run":
            body = self._read_json()
            from_date = body.get("from_date")
            to_date = body.get("to_date")
            horizon_days = body.get("horizon_days", 20)
            apply_tuning = parse_bool(body.get("apply_tuning"), default=False)
            fix_data_pipes = parse_bool(body.get("fix_data_pipes"), default=False)
            min_samples = body.get("min_samples")
            if min_samples is not None:
                try:
                    min_samples = int(float(min_samples))
                except Exception:
                    json_response(self, {"error": "min_samples_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
            try:
                result = run_agent_backtest(
                    from_date=from_date,
                    to_date=to_date,
                    horizon_days=int(float(horizon_days)),
                    apply_tuning=bool(apply_tuning),
                    fix_data_pipes=bool(fix_data_pipes),
                    min_samples=min_samples,
                )
            except Exception as ex:
                json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                return
            json_response(self, result)
            return

        if parsed.path == "/api/v1/intel/docs":
            body = self._read_json()
            doc_type = body.get("doc_type")
            source = body.get("source")
            source_ref = body.get("source_ref")
            doc_date = body.get("doc_date")
            title = body.get("title")
            content = body.get("content")
            run_strategy = parse_bool(body.get("run_strategy"), default=True)
            if not str(content or "").strip():
                json_response(self, {"error": "content_required"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                try:
                    result = analyze_and_store_intelligence_document(
                        conn=conn,
                        doc_type=doc_type,
                        source=source,
                        source_ref=source_ref,
                        doc_date=doc_date,
                        title=title,
                        content=content,
                    )
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                summary = intelligence_summary(conn, limit=20)
            strategy_item = None
            if run_strategy:
                strategy_item = refresh_strategy_analytics(force=True)
            json_response(
                self,
                {
                    "ok": True,
                    "result": result,
                    "summary": summary,
                    "strategy_refreshed": bool(run_strategy),
                    "strategy_item": strategy_item,
                },
            )
            return

        if parsed.path == "/api/v1/intel/financials":
            body = self._read_json()
            run_strategy = parse_bool(body.get("run_strategy"), default=True)
            with db_connect() as conn:
                try:
                    row = upsert_company_financial_row(conn, body)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                fin_sig = financial_signal_for_symbol(conn, row["symbol"])
                summary = intelligence_summary(conn, limit=20)
            strategy_item = None
            if run_strategy:
                strategy_item = refresh_strategy_analytics(force=True)
            json_response(
                self,
                {
                    "ok": True,
                    "row": row,
                    "financial_signal": fin_sig,
                    "summary": summary,
                    "strategy_refreshed": bool(run_strategy),
                    "strategy_item": strategy_item,
                },
            )
            return

        if parsed.path.startswith("/api/v1/assistant/approvals/") and parsed.path.endswith("/decision"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 7:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                approval_id = int(parts[5])
            except ValueError:
                json_response(self, {"error": "invalid_approval_id"}, HTTPStatus.BAD_REQUEST)
                return
            body = self._read_json()
            decision = str(body.get("decision", "")).strip().lower()
            note = str(body.get("note", "")).strip()
            with db_connect() as conn:
                try:
                    out = resolve_agent_approval(conn, approval_id, decision, note=note)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
            if out.get("executed") and out.get("intent") == "delete_by_notes":
                recompute_holdings_and_signals()
            json_response(self, out)
            return

        if parsed.path == "/api/v1/trades/override":
            body = self._read_json()
            symbol = symbol_upper(body.get("symbol"))
            side = str(body.get("side", "")).strip().upper()
            trade_date = str(body.get("trade_date", "")).strip()
            qty = parse_float(body.get("quantity"), 0.0)
            price = parse_float(body.get("price"), 0.0)
            notes = str(body.get("notes", "manual override")).strip()
            ext_trade_id = normalize_external_trade_id(body.get("external_trade_id"))
            if not symbol:
                json_response(self, {"error": "symbol_required"}, HTTPStatus.BAD_REQUEST)
                return
            if side not in ("BUY", "SELL"):
                json_response(self, {"error": "side_must_be_buy_or_sell"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                d = dt.date.fromisoformat(trade_date)
                trade_date = d.isoformat()
            except ValueError:
                json_response(self, {"error": "trade_date_must_be_iso_yyyy_mm_dd"}, HTTPStatus.BAD_REQUEST)
                return
            if qty <= 0 or price <= 0:
                json_response(self, {"error": "quantity_and_price_must_be_positive"}, HTTPStatus.BAD_REQUEST)
                return
            amount = qty * price
            with db_connect() as conn:
                trade_symbol = resolve_symbol(conn, symbol) or symbol
                inst = conn.execute(
                    "SELECT symbol FROM instruments WHERE UPPER(symbol) = ?",
                    (symbol_upper(trade_symbol),),
                ).fetchone()
                if inst is None:
                    conn.execute(
                        """
                        INSERT INTO instruments(exchange, symbol, name, active, feed_code, price_source, asset_class)
                        VALUES ('NSE', ?, ?, 1, ?, 'exchange_api', ?)
                        """,
                        (
                            trade_symbol,
                            trade_symbol,
                            trade_symbol,
                            infer_asset_class(symbol=trade_symbol, name=trade_symbol),
                        ),
                    )
                cur = conn.execute(
                    """
                    INSERT INTO trades(symbol, side, trade_date, quantity, price, amount, external_trade_id, source, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'manual_override', ?)
                    """,
                    (trade_symbol, side, trade_date, qty, price, amount, ext_trade_id, notes),
                )
                trade_id = cur.lastrowid
                reconcile_daily_target_trade_links(conn)
                conn.commit()
            recompute_holdings_and_signals()
            json_response(
                self,
                {
                    "ok": True,
                    "trade_id": trade_id,
                    "symbol": symbol,
                    "side": side,
                    "trade_date": trade_date,
                    "source": "manual_override",
                },
            )
            return

        if parsed.path == "/api/v1/strategy/refresh":
            item = refresh_strategy_analytics(force=True)
            json_response(self, {"ok": True, "item": item})
            return

        if parsed.path == "/api/v1/strategy/audits/run":
            body = self._read_json()
            try:
                refresh_first = parse_bool(body.get("refresh_strategy", False), default=False)
            except ValueError:
                json_response(self, {"error": "refresh_strategy_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                try:
                    out = run_strategy_audit(conn, refresh_strategy=refresh_first)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
            json_response(self, out)
            return

        if parsed.path.startswith("/api/v1/scrips/") and parsed.path.endswith("/sell-simulate"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) < 6:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            symbol_req = parts[4]
            body = self._read_json()
            qty = parse_float(body.get("quantity"), 0.0)
            sell_price = body.get("sell_price")
            if qty <= 0:
                json_response(self, {"error": "quantity_must_be_positive"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                symbol = resolve_symbol(conn, symbol_req)
                if not symbol:
                    json_response(self, {"error": "symbol_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                try:
                    result = simulate_sell_for_symbol(conn, symbol, qty, sell_price)
                except ValueError as exc:
                    json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
                    return
            json_response(self, result)
            return

        if parsed.path.startswith("/api/v1/scrips/") and parsed.path.endswith("/trades"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) < 6:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            symbol = symbol_upper(parts[4])
            body = self._read_json()
            side = str(body.get("side", "")).strip().upper()
            trade_date = str(body.get("trade_date", "")).strip()
            qty = parse_float(body.get("quantity"), 0.0)
            price = parse_float(body.get("price"), 0.0)
            notes = str(body.get("notes", "")).strip()
            if side not in ("BUY", "SELL"):
                json_response(self, {"error": "side_must_be_buy_or_sell"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                d = dt.date.fromisoformat(trade_date)
                trade_date = d.isoformat()
            except ValueError:
                json_response(self, {"error": "trade_date_must_be_iso_yyyy_mm_dd"}, HTTPStatus.BAD_REQUEST)
                return
            if qty <= 0 or price <= 0:
                json_response(self, {"error": "quantity_and_price_must_be_positive"}, HTTPStatus.BAD_REQUEST)
                return
            amount = qty * price

            with db_connect() as conn:
                trade_symbol = resolve_symbol(conn, symbol) or symbol
                inst = conn.execute(
                    "SELECT symbol FROM instruments WHERE UPPER(symbol) = ?",
                    (symbol_upper(trade_symbol),),
                ).fetchone()
                if inst is None:
                    conn.execute(
                        """
                        INSERT INTO instruments(exchange, symbol, name, active, feed_code, price_source, asset_class)
                        VALUES ('NSE', ?, ?, 1, ?, 'exchange_api', ?)
                        """,
                        (
                            trade_symbol,
                            trade_symbol,
                            trade_symbol,
                            infer_asset_class(symbol=trade_symbol, name=trade_symbol),
                        ),
                    )
                existing_rows = conn.execute(
                    """
                    SELECT symbol, UPPER(side) AS side, trade_date, quantity, price, amount, source, external_trade_id
                    FROM trades
                    WHERE UPPER(symbol) = ? AND UPPER(side) = ? AND trade_date = ?
                    """,
                    (symbol_upper(trade_symbol), side, trade_date),
                ).fetchall()
                candidate = {
                    "symbol": symbol_upper(trade_symbol),
                    "side": side,
                    "trade_date": trade_date,
                    "quantity": float(qty),
                    "price": float(price),
                    "amount": float(amount),
                    "source": "manual_entry",
                }
                for r in existing_rows:
                    ex = {
                        "symbol": symbol_upper(r["symbol"]),
                        "side": str(r["side"]).upper(),
                        "trade_date": r["trade_date"],
                        "quantity": float(r["quantity"]),
                        "price": float(r["price"]),
                        "amount": float(r["amount"]),
                        "source": r["source"],
                        "external_trade_id": normalize_external_trade_id(r["external_trade_id"]),
                    }
                    # Manual entries use strict duplicate detection to avoid blocking legitimate corrections.
                    if _is_exact_duplicate_trade(candidate, ex):
                        json_response(
                            self,
                            {
                                "error": "duplicate_trade_detected",
                                "message": "Exact duplicate exists for same symbol/side/date/qty/amount.",
                            },
                            HTTPStatus.CONFLICT,
                        )
                        return

                cur = conn.execute(
                    """
                    INSERT INTO trades(symbol, side, trade_date, quantity, price, amount, source, notes)
                    VALUES (?, ?, ?, ?, ?, ?, 'manual_entry', ?)
                    """,
                    (trade_symbol, side, trade_date, qty, price, amount, notes),
                )
                trade_id = cur.lastrowid
                reconcile_daily_target_trade_links(conn)
                conn.commit()
            recompute_holdings_and_signals()
            json_response(self, {"ok": True, "trade_id": trade_id, "symbol": trade_symbol})
            return

        if parsed.path == "/api/v1/rebalance/lot/lock":
            body = self._read_json()
            side = str(body.get("side", "SELL") or "SELL").strip().upper()
            percent = body.get("percent", 10)
            with db_connect() as conn:
                try:
                    out = lock_rebalance_lot(conn, side=side, percent=percent)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
            json_response(self, out)
            return

        if parsed.path == "/api/v1/rebalance/lot/reset":
            with db_connect() as conn:
                try:
                    out = reset_active_rebalance_lot(conn)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
            json_response(self, out)
            return

        if parsed.path == "/api/v1/daily-target/reset":
            with db_connect() as conn:
                try:
                    out = reset_daily_target_plan(conn)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
            json_response(self, out)
            return

        if parsed.path == "/api/v1/upload/tradebook":
            body = self._read_json()
            b64 = body.get("content_base64")
            filename = str(body.get("filename", "tradebook.xlsx"))
            include_skipped = parse_bool(body.get("include_skipped"), default=True)
            if not b64:
                json_response(self, {"error": "content_base64_required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                raw = base64.b64decode(b64.encode("utf-8"))
            except Exception:
                json_response(self, {"error": "invalid_base64_payload"}, HTTPStatus.BAD_REQUEST)
                return
            safe_name = filename.replace("\\", "_").replace("/", "_")
            upload_dir = tenant_paths(get_current_tenant_key())["upload_dir"]
            upload_dir.mkdir(parents=True, exist_ok=True)
            save_path = upload_dir / f"{now_iso().replace(':','-')}_{safe_name}"
            save_path.write_bytes(raw)
            stats = import_tradebook_bytes(raw, filename=safe_name, collect_skipped=include_skipped)
            with db_connect() as conn:
                reconcile_daily_target_trade_links(conn)
                conn.commit()
                summary = portfolio_summary(conn)
            json_response(self, {"ok": True, "stats": stats, "summary": summary})
            return

        if parsed.path == "/api/v1/upload/cashflow":
            body = self._read_json()
            b64 = body.get("content_base64")
            filename = str(body.get("filename", "cashflow.xlsx"))
            replace_all = parse_bool(body.get("replace_all"), default=False)
            if not b64:
                json_response(self, {"error": "content_base64_required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                raw = base64.b64decode(b64.encode("utf-8"))
            except Exception:
                json_response(self, {"error": "invalid_base64_payload"}, HTTPStatus.BAD_REQUEST)
                return
            safe_name = filename.replace("\\", "_").replace("/", "_")
            upload_dir = tenant_paths(get_current_tenant_key())["upload_dir"]
            upload_dir.mkdir(parents=True, exist_ok=True)
            save_path = upload_dir / f"{now_iso().replace(':','-')}_{safe_name}"
            save_path.write_bytes(raw)
            stats = import_cashflow_bytes(raw, filename=safe_name, replace_existing=replace_all)
            with db_connect() as conn:
                summary = portfolio_summary(conn)
                cash = cashflow_summary(conn)
            json_response(
                self,
                {
                    "ok": True,
                    "replaced_existing": replace_all,
                    "stats": stats,
                    "summary": summary,
                    "cashflow_summary": cash,
                },
            )
            return

        if parsed.path == "/api/v1/upload/dividends":
            body = self._read_json()
            b64 = body.get("content_base64")
            filename = str(body.get("filename", "dividends.xlsx"))
            replace_all = parse_bool(body.get("replace_all"), default=False)
            if not b64:
                json_response(self, {"error": "content_base64_required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                raw = base64.b64decode(b64.encode("utf-8"))
            except Exception:
                json_response(self, {"error": "invalid_base64_payload"}, HTTPStatus.BAD_REQUEST)
                return
            safe_name = filename.replace("\\", "_").replace("/", "_")
            upload_dir = tenant_paths(get_current_tenant_key())["upload_dir"]
            upload_dir.mkdir(parents=True, exist_ok=True)
            save_path = upload_dir / f"{now_iso().replace(':','-')}_{safe_name}"
            save_path.write_bytes(raw)
            stats = import_dividend_bytes(raw, filename=safe_name, replace_existing=replace_all)
            with db_connect() as conn:
                summary = portfolio_summary(conn)
                dsum = dividend_summary(conn)
            json_response(
                self,
                {
                    "ok": True,
                    "replaced_existing": replace_all,
                    "stats": stats,
                    "summary": summary,
                    "dividend_summary": dsum,
                },
            )
            return

        if parsed.path == "/api/v1/prices/refresh":
            refresh_latest_prices_from_exchange(max_runtime_sec=20)
            recompute_holdings_and_signals()
            with db_connect() as conn:
                updated = conn.execute("SELECT MAX(updated_at) AS ts FROM latest_prices").fetchone()["ts"]
                count = conn.execute("SELECT COUNT(*) AS c FROM latest_prices").fetchone()["c"]
            json_response(self, {"ok": True, "updated_at": updated, "scrips_with_price": count})
            return

        if parsed.path == "/api/v1/prices/history/backfill":
            body = self._read_json()
            full = parse_bool(body.get("full"), default=True)
            try:
                max_runtime_sec = int(body.get("max_runtime_sec", 240))
            except Exception:
                json_response(self, {"error": "max_runtime_sec_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                return
            max_symbols = body.get("max_symbols")
            if max_symbols is not None:
                try:
                    max_symbols = int(max_symbols)
                except Exception:
                    json_response(self, {"error": "max_symbols_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
            stats = sync_market_history(
                backfill_all=full,
                max_runtime_sec=max_runtime_sec,
                max_symbols=max_symbols,
            )
            with market_db_connect() as mconn:
                h = mconn.execute(
                    """
                    SELECT
                      COUNT(*) AS rows_total,
                      COUNT(DISTINCT symbol) AS symbols_total,
                      MIN(price_date) AS from_date,
                      MAX(price_date) AS to_date
                    FROM daily_prices
                    """
                ).fetchone()
            json_response(
                self,
                {
                    "ok": True,
                    "stats": stats,
                    "history": {
                        "rows_total": int(h["rows_total"] or 0),
                        "symbols_total": int(h["symbols_total"] or 0),
                        "from_date": h["from_date"],
                        "to_date": h["to_date"],
                    },
                },
            )
            return

        if parsed.path == "/api/v1/scrips/bulk-delete":
            body = self._read_json()
            symbols_raw = body.get("symbols", [])
            if not isinstance(symbols_raw, list):
                json_response(self, {"error": "symbols_must_be_array"}, HTTPStatus.BAD_REQUEST)
                return
            symbols = []
            seen = set()
            for s in symbols_raw:
                su = symbol_upper(s)
                if not su or su in seen:
                    continue
                seen.add(su)
                symbols.append(su)
            if not symbols:
                json_response(self, {"error": "symbols_required"}, HTTPStatus.BAD_REQUEST)
                return

            placeholders = ",".join(["?"] * len(symbols))
            with db_connect() as conn:
                found = {
                    r["sym"]
                    for r in conn.execute(
                        f"SELECT DISTINCT UPPER(symbol) AS sym FROM instruments WHERE UPPER(symbol) IN ({placeholders})",
                        symbols,
                    ).fetchall()
                }
                found.update(
                    {
                        r["sym"]
                        for r in conn.execute(
                            f"SELECT DISTINCT UPPER(symbol) AS sym FROM trades WHERE UPPER(symbol) IN ({placeholders})",
                            symbols,
                        ).fetchall()
                    }
                )
                not_found = [s for s in symbols if s not in found]
                d_trades = conn.execute(
                    f"DELETE FROM trades WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_prices = conn.execute(
                    f"DELETE FROM latest_prices WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_holdings = conn.execute(
                    f"DELETE FROM holdings WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_lots = conn.execute(
                    f"DELETE FROM lot_closures WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_signals = conn.execute(
                    f"DELETE FROM signals WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_actions = conn.execute(
                    f"DELETE FROM corporate_actions WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_ticks = conn.execute(
                    f"DELETE FROM price_ticks WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_dividends = conn.execute(
                    f"DELETE FROM dividends WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_guards = conn.execute(
                    f"DELETE FROM scrip_position_guards WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                d_instruments = conn.execute(
                    f"DELETE FROM instruments WHERE UPPER(symbol) IN ({placeholders})",
                    symbols,
                ).rowcount
                conn.commit()
            d_history = delete_market_history_symbols(symbols)
            recompute_holdings_and_signals()
            json_response(
                self,
                {
                    "ok": True,
                    "requested_symbols": symbols,
                    "deleted_symbols": sorted(found),
                    "not_found_symbols": not_found,
                    "deleted_rows": {
                        "instruments": d_instruments,
                        "trades": d_trades,
                        "latest_prices": d_prices,
                        "holdings": d_holdings,
                        "lot_closures": d_lots,
                        "signals": d_signals,
                        "corporate_actions": d_actions,
                        "price_ticks": d_ticks,
                        "dividends": d_dividends,
                        "position_guards": d_guards,
                        "market_history": d_history,
                    },
                },
            )
            return

        if parsed.path == "/api/v1/scrips":
            body = self._read_json()
            symbol = symbol_upper(body.get("symbol", ""))
            exchange = str(body.get("exchange", "NSE")).strip().upper()
            feed_code = str(body.get("feed_code", symbol)).strip()
            asset_class_raw = body.get("asset_class", None)
            ltp = parse_float(body.get("ltp"), 0.0)
            if not symbol:
                json_response(self, {"error": "symbol_required"}, HTTPStatus.BAD_REQUEST)
                return
            if exchange not in ("NSE", "BSE"):
                json_response(self, {"error": "exchange_must_be_nse_or_bse"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                asset_class = (
                    normalize_asset_class(asset_class_raw, fallback=infer_asset_class(symbol=symbol, name=symbol))
                    if asset_class_raw is not None
                    else infer_asset_class(symbol=symbol, name=symbol)
                )
            except ValueError as ex:
                json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                exists = conn.execute(
                    "SELECT symbol FROM instruments WHERE UPPER(symbol) = ?",
                    (symbol_upper(symbol),),
                ).fetchone()
                if exists:
                    json_response(self, {"error": "symbol_exists", "symbol": exists["symbol"]}, HTTPStatus.CONFLICT)
                    return
                conn.execute(
                    """
                    INSERT INTO instruments(exchange, symbol, name, active, feed_code, price_source, asset_class)
                    VALUES (?, ?, ?, 1, ?, 'exchange_api', ?)
                    """,
                    (exchange, symbol, symbol, feed_code, asset_class),
                )
                if ltp > 0:
                    conn.execute(
                        "INSERT INTO latest_prices(symbol, ltp, change_abs, updated_at) VALUES (?, ?, 0, ?)",
                        (symbol, ltp, now_iso()),
                    )
                conn.commit()
                exists_after = conn.execute(
                    "SELECT symbol FROM instruments WHERE UPPER(symbol) = ?", (symbol_upper(symbol),)
                ).fetchone()
            refresh_latest_prices_from_exchange()
            recompute_holdings_and_signals()
            with db_connect() as conn:
                instrument_count = conn.execute("SELECT COUNT(*) AS c FROM instruments").fetchone()["c"]
            json_response(
                self,
                {
                    "ok": True,
                    "added_symbol": symbol,
                    "asset_class": asset_class,
                    "present_after_insert": bool(exists_after),
                    "instrument_count": instrument_count,
                },
            )
            return

        if parsed.path == "/api/v1/corporate-actions/splits":
            body = self._read_json()
            symbol = symbol_upper(body.get("symbol", ""))
            effective_date = str(body.get("effective_date", "")).strip()
            factor = float(body.get("factor", 1))
            note = str(body.get("note", "")).strip()
            if not symbol:
                json_response(self, {"error": "symbol_required"}, HTTPStatus.BAD_REQUEST)
                return
            if factor <= 0:
                json_response(self, {"error": "factor_must_be_positive"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                dt.date.fromisoformat(effective_date)
            except ValueError:
                json_response(self, {"error": "effective_date_must_be_iso_yyyy_mm_dd"}, HTTPStatus.BAD_REQUEST)
                return

            with db_connect() as conn:
                conn.execute(
                    """
                    INSERT INTO corporate_actions(symbol, action_type, effective_date, factor, note, created_at)
                    VALUES (?, 'SPLIT', ?, ?, ?, ?)
                    """,
                    (symbol, effective_date, factor, note, now_iso()),
                )
                conn.commit()
            recompute_holdings_and_signals()
            json_response(self, {"ok": True})
            return
        json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)

    def handle_api_put(self, parsed):
        if parsed.path == "/api/v1/tenants/active":
            body = self._read_json()
            tenant_key = body.get("tenant") or body.get("key")
            if not str(tenant_key or "").strip():
                json_response(self, {"error": "tenant_required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                key = set_active_tenant_key(tenant_key, persist=True)
                with tenant_context(key):
                    init_db()
                set_request_tenant_key(key)
            except ValueError as ex:
                json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                return
            json_response(
                self,
                {
                    "ok": True,
                    "active_tenant": key,
                    "max_tenants": TENANT_MAX_COUNT,
                    "items": list_tenants(),
                },
            )
            return

        if parsed.path == "/api/v1/llm/config":
            body = self._read_json()
            api_key_raw = body.get("api_key", "__UNCHANGED__")
            model_raw = body.get("model", "__UNCHANGED__")
            api_url_raw = body.get("api_url", "__UNCHANGED__")
            if model_raw != "__UNCHANGED__" and not str(model_raw or "").strip():
                json_response(self, {"error": "model_required"}, HTTPStatus.BAD_REQUEST)
                return
            if api_url_raw != "__UNCHANGED__":
                parsed_url = urllib.parse.urlparse(str(api_url_raw or "").strip() or LLM_DEFAULT_API_URL)
                if parsed_url.scheme not in ("http", "https") or not parsed_url.netloc:
                    json_response(self, {"error": "api_url_invalid"}, HTTPStatus.BAD_REQUEST)
                    return
            set_llm_runtime_config(
                api_key=api_key_raw,
                model=model_raw,
                api_url=api_url_raw,
            )
            with db_connect() as conn:
                cfg = get_llm_runtime_config(conn, include_secret=False)
                _update_llm_runtime_status(
                    conn,
                    "ready" if cfg.get("configured") else "not_configured",
                    "" if cfg.get("configured") else "LLM API key not configured.",
                    checked_at="",
                )
                conn.commit()
                json_response(self, {"ok": True, "config": get_llm_runtime_config(conn, include_secret=False)})
            return

        if parsed.path == "/api/v1/scrips/position-guards":
            body = self._read_json()
            items = body.get("items", [])
            if not isinstance(items, list):
                json_response(self, {"error": "items_must_be_array"}, HTTPStatus.BAD_REQUEST)
                return
            updated = []
            with db_connect() as conn:
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    symbol = item.get("symbol")
                    min_value = item.get("min_value", 0.0)
                    max_value = item.get("max_value", None)
                    try:
                        row = set_scrip_position_guard(conn, symbol, min_value=min_value, max_value=max_value)
                    except ValueError as ex:
                        json_response(self, {"error": str(ex), "symbol": symbol_upper(symbol)}, HTTPStatus.BAD_REQUEST)
                        return
                    updated.append(row)
                conn.commit()
            json_response(self, {"ok": True, "items": updated, "count": len(updated)})
            return

        if parsed.path.startswith("/api/v1/rebalance/closed-history/items/") and parsed.path.endswith("/buyback"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 8:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                item_id = int(parts[6])
            except Exception:
                json_response(self, {"error": "invalid_lot_item_id"}, HTTPStatus.BAD_REQUEST)
                return
            body = self._read_json()
            try:
                buyback_completed = parse_bool(body.get("buyback_completed"), default=True)
            except ValueError:
                json_response(self, {"error": "buyback_completed_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                return
            buyback_price = body.get("buyback_price", None)
            buyback_at = body.get("buyback_at", None)
            note = str(body.get("note", "") or "")
            include_completed_raw = body.get("include_completed", False)
            try:
                include_completed = parse_bool(include_completed_raw, default=False)
            except ValueError:
                json_response(self, {"error": "include_completed_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                try:
                    out = set_rebalance_buyback_status(
                        conn,
                        item_id=item_id,
                        buyback_completed=buyback_completed,
                        buyback_price=buyback_price,
                        buyback_at=buyback_at,
                        note=note,
                    )
                    out["history"] = list_rebalance_closed_history(
                        conn, limit=300, include_buyback_completed=include_completed
                    )
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
            json_response(self, out)
            return

        if parsed.path.startswith("/api/v1/rebalance/lot/items/"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 7:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                item_id = int(parts[6])
            except Exception:
                json_response(self, {"error": "invalid_lot_item_id"}, HTTPStatus.BAD_REQUEST)
                return
            body = self._read_json()
            state_raw = body.get("state")
            planned_qty_raw = body.get("planned_qty", None)
            planned_qty_provided = planned_qty_raw is not None
            completed = None
            if state_raw is None:
                try:
                    completed = parse_bool(body.get("completed"), default=None)
                except ValueError:
                    json_response(self, {"error": "completed_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
                if completed is None and (not planned_qty_provided):
                    json_response(self, {"error": "state_or_completed_or_planned_qty_required"}, HTTPStatus.BAD_REQUEST)
                    return
            note = str(body.get("note", "") or "")
            executed_price = body.get("executed_price", None)
            executed_at = body.get("executed_at", None)
            with db_connect() as conn:
                try:
                    out = None
                    if planned_qty_provided:
                        out = set_rebalance_lot_item_planned_qty(conn, item_id=item_id, planned_qty=planned_qty_raw)
                    if state_raw is None:
                        if completed is not None:
                            out = set_rebalance_lot_item_completed(conn, item_id=item_id, completed=completed, note=note)
                    else:
                        out = set_rebalance_lot_item_status(
                            conn,
                            item_id=item_id,
                            state=state_raw,
                            note=note,
                            executed_price=executed_price,
                            executed_at=executed_at,
                        )
                    if out is None:
                        json_response(self, {"error": "no_valid_update_fields"}, HTTPStatus.BAD_REQUEST)
                        return
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
            json_response(self, out)
            return

        if parsed.path.startswith("/api/v1/daily-target/pairs/"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 6:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                pair_id = int(parts[5])
            except Exception:
                json_response(self, {"error": "invalid_daily_target_pair_id"}, HTTPStatus.BAD_REQUEST)
                return
            body = self._read_json()
            state_raw = body.get("state")
            if state_raw is None:
                json_response(self, {"error": "state_required"}, HTTPStatus.BAD_REQUEST)
                return
            note = str(body.get("note", "") or "")
            executed_sell_price = body.get("executed_sell_price", None)
            executed_sell_at = body.get("executed_sell_at", None)
            executed_buy_price = body.get("executed_buy_price", None)
            executed_buy_at = body.get("executed_buy_at", None)
            with db_connect() as conn:
                try:
                    out = update_daily_target_pair(
                        conn,
                        pair_id=pair_id,
                        state=state_raw,
                        note=note,
                        executed_sell_price=executed_sell_price,
                        executed_sell_at=executed_sell_at,
                        executed_buy_price=executed_buy_price,
                        executed_buy_at=executed_buy_at,
                    )
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
            json_response(self, out)
            return

        if parsed.path.startswith("/api/v1/agents/") and parsed.path.endswith("/control"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 6:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            agent = str(parts[4] or "").strip().lower()
            body = self._read_json()
            enabled = body.get("enabled")
            run_now_raw = body.get("run_now", False)
            interval_seconds = body.get("interval_seconds")
            min_samples = body.get("min_samples")
            max_docs = body.get("max_docs")
            symbols_limit = body.get("symbols_limit")
            sources = body.get("sources")
            query_seed = body.get("query_seed")
            auto_tune = body.get("auto_tune")
            write_changes = body.get("write_changes")
            core_objective = body.get("core_objective")
            retention_days = body.get("retention_days")
            lookback_days = body.get("lookback_days")
            winsorize_pct = body.get("winsorize_pct")
            if enabled is not None:
                try:
                    enabled = parse_bool(enabled)
                except ValueError:
                    json_response(self, {"error": "enabled_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
            if auto_tune is not None:
                try:
                    auto_tune = parse_bool(auto_tune)
                except ValueError:
                    json_response(self, {"error": "auto_tune_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
            if write_changes is not None:
                try:
                    write_changes = parse_bool(write_changes)
                except ValueError:
                    json_response(self, {"error": "write_changes_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
            try:
                run_now = parse_bool(run_now_raw, default=False)
            except ValueError:
                json_response(self, {"error": "run_now_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                return
            if interval_seconds is not None:
                try:
                    interval_seconds = int(float(interval_seconds))
                except Exception:
                    json_response(self, {"error": "interval_seconds_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
                if interval_seconds <= 0:
                    json_response(self, {"error": "interval_seconds_must_be_positive"}, HTTPStatus.BAD_REQUEST)
                    return
            if min_samples is not None:
                try:
                    min_samples = int(float(min_samples))
                except Exception:
                    json_response(self, {"error": "min_samples_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
            if max_docs is not None:
                try:
                    max_docs = int(float(max_docs))
                except Exception:
                    json_response(self, {"error": "max_docs_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
            if symbols_limit is not None:
                try:
                    symbols_limit = int(float(symbols_limit))
                except Exception:
                    json_response(self, {"error": "symbols_limit_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
            if lookback_days is not None:
                try:
                    lookback_days = int(float(lookback_days))
                except Exception:
                    json_response(self, {"error": "lookback_days_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
            if winsorize_pct is not None:
                try:
                    winsorize_pct = float(winsorize_pct)
                except Exception:
                    json_response(self, {"error": "winsorize_pct_must_be_number"}, HTTPStatus.BAD_REQUEST)
                    return
            if retention_days is not None:
                try:
                    retention_days = int(float(retention_days))
                except Exception:
                    json_response(self, {"error": "retention_days_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                    return
                if retention_days <= 0:
                    json_response(self, {"error": "retention_days_must_be_positive"}, HTTPStatus.BAD_REQUEST)
                    return

            if (
                enabled is None
                and not run_now
                and interval_seconds is None
                and min_samples is None
                and max_docs is None
                and symbols_limit is None
                and sources is None
                and query_seed is None
                and auto_tune is None
                and write_changes is None
                and core_objective is None
                and retention_days is None
                and lookback_days is None
                and winsorize_pct is None
            ):
                json_response(self, {"error": "no_action_requested"}, HTTPStatus.BAD_REQUEST)
                return

            run_result = {}
            if agent == "market":
                set_live_config(enabled=enabled, interval_seconds=interval_seconds)
                if run_now:
                    refresh_latest_prices_from_exchange(max_runtime_sec=20)
                    recompute_holdings_and_signals(force_strategy=False)
                    with db_connect() as conn:
                        row = conn.execute(
                            "SELECT MAX(updated_at) AS updated_at, COUNT(*) AS c FROM latest_prices"
                        ).fetchone()
                    run_result = {
                        "updated_at": row["updated_at"],
                        "scrips_with_price": int(row["c"] or 0),
                    }
            elif agent == "history":
                set_history_sync_config(enabled=enabled, interval_seconds=interval_seconds)
                if run_now:
                    run_result = sync_market_history(backfill_all=False, max_runtime_sec=180)
            elif agent == "strategy":
                set_strategy_config(enabled=enabled, interval_seconds=interval_seconds)
                if run_now:
                    run_result = refresh_strategy_analytics(force=True)
            elif agent == "backup":
                if interval_seconds is not None:
                    json_response(
                        self,
                        {"error": "backup_interval_fixed", "message": "Backup interval is fixed at 2 days."},
                        HTTPStatus.BAD_REQUEST,
                    )
                    return
                set_backup_config(enabled=enabled)
                if run_now:
                    run_result = {"backup_file": backup_database()}
            elif agent == "self_learning":
                interval_days = None
                if interval_seconds is not None:
                    interval_days = max(1, int(math.ceil(interval_seconds / 86400.0)))
                set_self_learning_config(enabled=enabled, interval_days=interval_days, min_samples=min_samples)
                if run_now:
                    run_result = run_agent_backtest(
                        from_date=(dt.date.today() - dt.timedelta(days=420)).isoformat(),
                        to_date=dt.date.today().isoformat(),
                        horizon_days=20,
                        apply_tuning=True,
                        fix_data_pipes=True,
                        min_samples=min_samples,
                    )
                    set_self_learning_config(last_run_at=now_iso())
            elif agent == "intel_autopilot":
                set_intel_autopilot_config(
                    enabled=enabled,
                    interval_seconds=interval_seconds,
                    max_docs=max_docs,
                    symbols_limit=symbols_limit,
                    sources=sources,
                    query_seed=query_seed,
                )
                if run_now:
                    run_result = run_intelligence_autopilot_once(max_runtime_sec=60, force=True)
                    set_intel_autopilot_config(last_run_at=now_iso())
            elif agent == "chart_intel":
                set_chart_agent_config(
                    enabled=enabled,
                    interval_seconds=interval_seconds,
                    sources=sources,
                )
                if run_now:
                    run_result = run_chart_intel_agent_once(max_runtime_sec=60, force=True)
                    if int(parse_float((run_result or {}).get("updated"), 0.0)) > 0:
                        try:
                            refresh_strategy_analytics(force=True)
                        except Exception:
                            pass
            elif agent == "software_performance":
                set_software_perf_agent_config(
                    enabled=enabled,
                    interval_seconds=interval_seconds,
                    auto_tune=auto_tune,
                    write_changes=write_changes,
                    core_objective=core_objective,
                    retention_days=retention_days,
                )
                if run_now:
                    run_result = run_software_perf_agent_once(
                        max_runtime_sec=max(20, min(180, int(parse_float(interval_seconds, 60)))),
                        force=True,
                    )
            elif agent == "risk_analysis":
                set_risk_agent_config(
                    enabled=enabled,
                    interval_seconds=interval_seconds,
                    lookback_days=lookback_days,
                    winsorize_pct=winsorize_pct,
                )
                if run_now:
                    run_result = run_risk_analysis_agent_once(
                        max_runtime_sec=max(20, min(180, int(parse_float(interval_seconds, 45)))),
                        force=True,
                    )
            elif agent == "tax_monitor":
                with db_connect() as conn:
                    set_tax_monitor_config(
                        conn,
                        enabled=enabled,
                        interval_seconds=interval_seconds,
                    )
                    conn.commit()
                if run_now:
                    run_result = run_tax_rate_monitor_once(force=True, timeout=max(6, min(20, int(parse_float(interval_seconds, 8)))))
            else:
                json_response(self, {"error": "invalid_agent"}, HTTPStatus.BAD_REQUEST)
                return

            with db_connect() as conn:
                items = build_agents_status(conn)
            status_item = next((x for x in items if x.get("agent") == agent), None)
            json_response(
                self,
                {
                    "ok": True,
                    "agent": agent,
                    "status": status_item,
                    "run_result": run_result,
                },
            )
            return

        if parsed.path == "/api/v1/config/live":
            body = self._read_json()
            enabled = body.get("enabled")
            interval_seconds = body.get("interval_seconds")
            quote_sources = body.get("quote_sources")
            quote_max_deviation_pct = body.get("quote_max_deviation_pct")
            quote_top_k = body.get("quote_top_k")
            quote_explore_ratio = body.get("quote_explore_ratio")
            if enabled is not None:
                try:
                    enabled = parse_bool(enabled)
                except ValueError:
                    json_response(self, {"error": "enabled_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                    return
            if interval_seconds is not None:
                interval_seconds = int(interval_seconds)
            set_live_config(
                enabled=enabled,
                interval_seconds=interval_seconds,
                quote_sources=quote_sources,
                quote_max_deviation_pct=quote_max_deviation_pct,
                quote_top_k=quote_top_k,
                quote_explore_ratio=quote_explore_ratio,
            )
            with db_connect() as conn:
                json_response(self, get_live_config(conn))
            return

        if parsed.path == "/api/v1/prices/source-ranking":
            body = self._read_json()
            source = str(body.get("source", "")).strip().lower()
            enabled = body.get("enabled")
            notes = body.get("notes")
            if not source:
                json_response(self, {"error": "source_required"}, HTTPStatus.BAD_REQUEST)
                return
            if enabled is None:
                json_response(self, {"error": "enabled_required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                enabled_bool = parse_bool(enabled)
            except ValueError:
                json_response(self, {"error": "enabled_must_be_boolean"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                try:
                    set_quote_source_enabled(conn, source, enabled_bool, notes=notes)
                except ValueError as ex:
                    json_response(self, {"error": str(ex)}, HTTPStatus.BAD_REQUEST)
                    return
                conn.commit()
                policy = get_live_quote_policy(conn)
                json_response(
                    self,
                    {
                        "ok": True,
                        "policy": policy,
                        "nse_order": get_ranked_quote_sources(conn, policy, exchange="NSE"),
                        "bse_order": get_ranked_quote_sources(conn, policy, exchange="BSE"),
                        "items": quote_source_ranking(conn),
                    },
                )
            return

        if parsed.path.startswith("/api/v1/analytics/peak-splits/") and parsed.path.endswith("/review"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 7:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                split_id = int(parts[5])
            except ValueError:
                json_response(self, {"error": "invalid_split_id"}, HTTPStatus.BAD_REQUEST)
                return
            body = self._read_json()
            decision = str(body.get("decision", "")).strip().lower()
            if decision not in ("apply", "ignore"):
                json_response(self, {"error": "decision_must_be_apply_or_ignore"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                ensure_peak_split_reviews_table(conn)
                split_row = conn.execute(
                    "SELECT id FROM corporate_actions WHERE id = ? AND action_type='SPLIT'",
                    (split_id,),
                ).fetchone()
                if not split_row:
                    json_response(self, {"error": "split_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                conn.execute(
                    """
                    INSERT INTO peak_split_reviews(corporate_action_id, decision, decided_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(corporate_action_id) DO UPDATE SET
                      decision=excluded.decision,
                      decided_at=excluded.decided_at
                    """,
                    (split_id, decision, now_iso()),
                )
                conn.commit()
                pending_left = len(pending_peak_split_candidates(conn))
            json_response(self, {"ok": True, "split_id": split_id, "decision": decision, "pending_left": pending_left})
            return

        if parsed.path == "/api/v1/strategy/sets/active":
            body = self._read_json()
            try:
                set_id = int(body.get("id"))
            except Exception:
                json_response(self, {"error": "id_must_be_integer"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                exists = conn.execute("SELECT id FROM strategy_sets WHERE id = ?", (set_id,)).fetchone()
                if not exists:
                    json_response(self, {"error": "set_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                conn.execute("UPDATE strategy_sets SET is_active = 0")
                conn.execute("UPDATE strategy_sets SET is_active = 1 WHERE id = ?", (set_id,))
                conn.commit()
            recompute_holdings_and_signals()
            json_response(self, {"ok": True})
            return

        if parsed.path.startswith("/api/v1/strategy/sets/") and parsed.path.endswith("/parameters"):
            parts = parsed.path.split("/")
            try:
                set_id = int(parts[5])
            except Exception:
                json_response(self, {"error": "invalid_set_id"}, HTTPStatus.BAD_REQUEST)
                return
            body = self._read_json()
            params = body.get("parameters", [])
            if not isinstance(params, list):
                json_response(self, {"error": "parameters_must_be_array"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                exists = conn.execute("SELECT id FROM strategy_sets WHERE id = ?", (set_id,)).fetchone()
                if not exists:
                    json_response(self, {"error": "set_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                for p in params:
                    if not isinstance(p, dict):
                        json_response(self, {"error": "parameter_item_must_be_object"}, HTTPStatus.BAD_REQUEST)
                        return
                    key = str(p.get("key", "")).strip()
                    if not key:
                        json_response(self, {"error": "parameter_key_required"}, HTTPStatus.BAD_REQUEST)
                        return
                    try:
                        value = float(p.get("value"))
                    except Exception:
                        json_response(self, {"error": "parameter_value_must_be_number"}, HTTPStatus.BAD_REQUEST)
                        return
                    existing = conn.execute(
                        "SELECT id FROM strategy_parameters WHERE set_id = ? AND key = ?",
                        (set_id, key),
                    ).fetchone()
                    if existing:
                        conn.execute(
                            "UPDATE strategy_parameters SET value = ? WHERE id = ?",
                            (value, existing["id"]),
                        )
                    else:
                        conn.execute(
                            "INSERT INTO strategy_parameters(set_id, key, value) VALUES (?, ?, ?)",
                            (set_id, key, value),
                        )
                conn.commit()
            recompute_holdings_and_signals()
            json_response(self, {"ok": True})
            return

        json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)

    def handle_api_delete(self, parsed):
        if parsed.path.startswith("/api/v1/trades/"):
            try:
                trade_id = int(parsed.path.split("/")[-1])
            except ValueError:
                json_response(self, {"error": "invalid_trade_id"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                row = conn.execute("SELECT id, symbol FROM trades WHERE id = ?", (trade_id,)).fetchone()
                if row is None:
                    json_response(self, {"error": "trade_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                conn.execute("DELETE FROM trades WHERE id = ?", (trade_id,))
                conn.commit()
            recompute_holdings_and_signals()
            json_response(self, {"ok": True, "trade_id": trade_id, "symbol": row["symbol"]})
            return

        if parsed.path.startswith("/api/v1/scrips/"):
            parts = parsed.path.rstrip("/").split("/")
            if len(parts) != 5:
                json_response(self, {"error": "invalid_path"}, HTTPStatus.BAD_REQUEST)
                return
            requested_symbol = parts[4].strip()
            symbol_q = symbol_upper(requested_symbol)
            if not symbol_q:
                json_response(self, {"error": "symbol_required"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                exists = resolve_symbol(conn, symbol_q)
                d_trades = conn.execute("DELETE FROM trades WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_prices = conn.execute("DELETE FROM latest_prices WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_holdings = conn.execute("DELETE FROM holdings WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_lots = conn.execute("DELETE FROM lot_closures WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_signals = conn.execute("DELETE FROM signals WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_actions = conn.execute("DELETE FROM corporate_actions WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_ticks = conn.execute("DELETE FROM price_ticks WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_dividends = conn.execute("DELETE FROM dividends WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_guards = conn.execute("DELETE FROM scrip_position_guards WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                d_instruments = conn.execute("DELETE FROM instruments WHERE UPPER(symbol) = ?", (symbol_q,)).rowcount
                total = d_trades + d_prices + d_holdings + d_lots + d_signals + d_actions + d_ticks + d_dividends + d_guards + d_instruments
                if total == 0 and not exists:
                    json_response(self, {"error": "symbol_not_found"}, HTTPStatus.NOT_FOUND)
                    return
                conn.commit()
            d_history = delete_market_history_symbols([symbol_q])
            recompute_holdings_and_signals()
            json_response(
                self,
                {
                    "ok": True,
                    "requested_symbol": requested_symbol,
                    "deleted_symbol": exists or symbol_q,
                    "deleted_rows": {
                        "instruments": d_instruments,
                        "trades": d_trades,
                        "latest_prices": d_prices,
                        "holdings": d_holdings,
                        "lot_closures": d_lots,
                        "signals": d_signals,
                        "corporate_actions": d_actions,
                        "price_ticks": d_ticks,
                        "dividends": d_dividends,
                        "position_guards": d_guards,
                        "market_history": d_history,
                    },
                },
            )
            return

        if parsed.path.startswith("/api/v1/corporate-actions/splits/"):
            try:
                split_id = int(parsed.path.split("/")[-1])
            except Exception:
                json_response(self, {"error": "invalid_split_id"}, HTTPStatus.BAD_REQUEST)
                return
            with db_connect() as conn:
                ensure_peak_split_reviews_table(conn)
                conn.execute("DELETE FROM peak_split_reviews WHERE corporate_action_id = ?", (split_id,))
                deleted = conn.execute("DELETE FROM corporate_actions WHERE id = ? AND action_type='SPLIT'", (split_id,)).rowcount
                conn.commit()
            if deleted <= 0:
                json_response(self, {"error": "split_not_found"}, HTTPStatus.NOT_FOUND)
                return
            recompute_holdings_and_signals()
            json_response(self, {"ok": True})
            return
        json_response(self, {"error": "not_found"}, HTTPStatus.NOT_FOUND)

    def log_message(self, format, *args):
        return



def _resolve_repo_relative_path(raw_path):
    raw = str(raw_path or "").strip()
    if not raw:
        return None
    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = (ROOT / p).resolve()
    else:
        p = p.resolve()
    return p


def resolve_xlsx_path(cli_path=None, env_var="PORTFOLIO_XLSX_PATH"):
    explicit = _resolve_repo_relative_path(cli_path)
    if explicit is not None:
        return explicit
    env_raw = str(os.environ.get(env_var, "") or "").strip()
    env_path = _resolve_repo_relative_path(env_raw)
    if env_path is not None:
        return env_path
    for candidate in DEFAULT_XLSX_CANDIDATE_PATHS:
        try:
            p = Path(candidate).resolve()
        except Exception:
            continue
        if p.exists() and p.is_file():
            return p
    return None


def validate(xlsx_path):
    init_db()
    import_from_excel(xlsx_path)
    with db_connect() as conn:
        counts = {
            "instruments": conn.execute("SELECT COUNT(*) AS c FROM instruments").fetchone()["c"],
            "trades": conn.execute("SELECT COUNT(*) AS c FROM trades").fetchone()["c"],
            "holdings": conn.execute("SELECT COUNT(*) AS c FROM holdings").fetchone()["c"],
            "signals": conn.execute("SELECT COUNT(*) AS c FROM signals").fetchone()["c"],
        }
        if counts["trades"] == 0:
            raise RuntimeError("Validation failed: no trades imported.")
        if counts["holdings"] == 0:
            raise RuntimeError("Validation failed: no holdings computed.")
        summary = portfolio_summary(conn)
        print("Validation OK")
        print(json.dumps({"counts": counts, "summary": summary}, indent=2))


def live_price_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = 10
        try:
            with db_connect() as conn:
                cfg = get_live_config(conn)
            sleep_for = cfg["interval_seconds"]
            if cfg["enabled"]:
                refresh_latest_prices_from_exchange(max_runtime_sec=max(10, min(60, sleep_for * 2)))
                recompute_holdings_and_signals(force_strategy=False)
        except Exception:
            # keep worker alive; next cycle retries
            pass
        stop_event.wait(timeout=max(LIVE_REFRESH_MIN_SEC, sleep_for))


def db_backup_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = DB_BACKUP_INTERVAL_SEC
        try:
            with db_connect() as conn:
                cfg = get_backup_config(conn)
            sleep_for = cfg["interval_seconds"]
            if cfg["enabled"]:
                backup_database()
        except Exception:
            pass
        stop_event.wait(timeout=max(60 * 60, int(sleep_for)))


def repo_data_sync_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = REPO_SYNC_INTERVAL_DEFAULT_SEC
        try:
            with tenant_context(DEFAULT_TENANT_KEY):
                with db_connect() as conn:
                    cfg = get_repo_sync_config(conn)
                sleep_for = int(cfg.get("interval_seconds", REPO_SYNC_INTERVAL_DEFAULT_SEC))
                if cfg.get("enabled"):
                    err = ""
                    if cfg.get("auto_push"):
                        try:
                            sync_repo_data_snapshots_to_git()
                        except Exception as ex:
                            err = str(ex)
                    else:
                        try:
                            export_repo_data_snapshots()
                        except Exception as ex:
                            err = str(ex)
                    with db_connect() as conn:
                        run_at = now_iso()
                        conn.execute(
                            """
                            INSERT INTO app_config(key, value, updated_at) VALUES ('repo_sync_last_run_at', ?, ?)
                            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                            """,
                            (run_at, run_at),
                        )
                        conn.execute(
                            """
                            INSERT INTO app_config(key, value, updated_at) VALUES ('repo_sync_last_error', ?, ?)
                            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                            """,
                            (err, run_at),
                        )
                        conn.commit()
        except Exception:
            pass
        stop_event.wait(timeout=max(REPO_SYNC_MIN_INTERVAL_SEC, int(sleep_for)))


def strategy_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = STRATEGY_REFRESH_MIN_SEC
        try:
            with db_connect() as conn:
                cfg = get_strategy_config(conn)
            sleep_for = cfg["interval_seconds"]
            if cfg["enabled"]:
                refresh_strategy_analytics(force=False)
                maybe_run_self_learning_once()
        except Exception:
            pass
        stop_event.wait(timeout=max(300, int(sleep_for)))


def history_worker(stop_event):
    first_cycle = True
    while not stop_event.is_set():
        sleep_for = HISTORY_REFRESH_DEFAULT_SEC
        try:
            with db_connect() as conn:
                cfg = get_history_sync_config(conn)
            sleep_for = cfg["interval_seconds"]
            if cfg["enabled"]:
                sync_market_history(
                    backfill_all=first_cycle,
                    max_runtime_sec=max(90, min(900, sleep_for)),
                )
            first_cycle = False
        except Exception:
            pass
        stop_event.wait(timeout=max(HISTORY_REFRESH_MIN_SEC, int(sleep_for)))


def intelligence_autopilot_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = 12 * 60 * 60
        try:
            with db_connect() as conn:
                cfg = get_intel_autopilot_config(conn)
            sleep_for = int(cfg.get("interval_seconds", sleep_for))
            if cfg.get("enabled"):
                maybe_run_intel_autopilot_once()
        except Exception:
            pass
        stop_event.wait(timeout=max(60 * 15, int(sleep_for)))


def chart_intel_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = CHART_AGENT_INTERVAL_DEFAULT_SEC
        try:
            with db_connect() as conn:
                cfg = get_chart_agent_config(conn)
            sleep_for = int(cfg.get("interval_seconds", CHART_AGENT_INTERVAL_DEFAULT_SEC))
            if cfg.get("enabled"):
                maybe_run_chart_intel_agent_once()
        except Exception:
            pass
        stop_event.wait(timeout=max(CHART_AGENT_MIN_INTERVAL_SEC, int(sleep_for)))


def software_performance_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC
        try:
            with db_connect() as conn:
                cfg = get_software_perf_agent_config(conn)
            sleep_for = int(cfg.get("interval_seconds", SOFTWARE_PERF_AGENT_INTERVAL_DEFAULT_SEC))
            if cfg.get("enabled"):
                maybe_run_software_perf_agent_once()
        except Exception:
            pass
        stop_event.wait(timeout=max(SOFTWARE_PERF_AGENT_MIN_INTERVAL_SEC, int(sleep_for)))


def tax_monitor_worker(stop_event):
    while not stop_event.is_set():
        sleep_for = TAX_MONITOR_INTERVAL_DEFAULT_SEC
        try:
            with db_connect() as conn:
                cfg = get_tax_monitor_config(conn)
            sleep_for = int(cfg.get("interval_seconds", TAX_MONITOR_INTERVAL_DEFAULT_SEC))
            if cfg.get("enabled"):
                run_tax_rate_monitor_once(force=False, timeout=max(6, min(20, int(sleep_for // 3600) + 8)))
                with db_connect() as conn:
                    refresh_attention_alerts(conn)
                    conn.commit()
        except Exception:
            pass
        stop_event.wait(timeout=max(TAX_MONITOR_MIN_INTERVAL_SEC, int(sleep_for)))


def serve(port):
    WEB_DIR.mkdir(exist_ok=True)
    ensure_tenant_bootstrap()
    init_db()
    repair_results = repair_all_tenants_market_data_once()
    export_repo_data_snapshots()
    stop_event = threading.Event()
    worker = threading.Thread(target=live_price_worker, args=(stop_event,), daemon=True)
    backup_worker = threading.Thread(target=db_backup_worker, args=(stop_event,), daemon=True)
    repo_sync_worker_thread = threading.Thread(target=repo_data_sync_worker, args=(stop_event,), daemon=True)
    strat_worker = threading.Thread(target=strategy_worker, args=(stop_event,), daemon=True)
    hist_worker = threading.Thread(target=history_worker, args=(stop_event,), daemon=True)
    intel_worker = threading.Thread(target=intelligence_autopilot_worker, args=(stop_event,), daemon=True)
    chart_worker = threading.Thread(target=chart_intel_worker, args=(stop_event,), daemon=True)
    software_worker = threading.Thread(target=software_performance_worker, args=(stop_event,), daemon=True)
    risk_worker = threading.Thread(target=risk_analysis_worker, args=(stop_event,), daemon=True)
    tax_worker = threading.Thread(target=tax_monitor_worker, args=(stop_event,), daemon=True)
    worker.start()
    backup_worker.start()
    repo_sync_worker_thread.start()
    strat_worker.start()
    hist_worker.start()
    intel_worker.start()
    chart_worker.start()
    software_worker.start()
    risk_worker.start()
    tax_worker.start()
    server = ThreadingHTTPServer(("127.0.0.1", port), AppHandler)
    active_key = get_active_tenant_key()
    p = tenant_paths(active_key)
    print(f"Portfolio Agent running at http://127.0.0.1:{port}")
    print(f"Active tenant: {active_key}")
    print(f"Using database: {p['db_path']}")
    print("Startup mode: DB authoritative (no automatic Excel import).")
    print(f"DB backup interval: every {DB_BACKUP_INTERVAL_SEC // 86400} days in {p['backup_dir']}")
    if repair_results:
        print("Startup market-data repair:")
        for r in repair_results:
            print(
                "  "
                f"{r['tenant']}: "
                f"ticks_purged={r['purged_price_ticks']}, "
                f"samples_purged={r['purged_quote_samples']}, "
                f"ltp_repairs={r['ltp_repairs']}, "
                f"day_change_repairs={r['day_change_repairs']}"
            )
    try:
        server.serve_forever()
    finally:
        stop_event.set()


def main():
    parser = argparse.ArgumentParser(description="Portfolio Agent")
    parser.add_argument("--port", type=int, default=8080)
    parser.add_argument("--validate", action="store_true")
    parser.add_argument("--xlsx-path", type=str, default=None)
    parser.add_argument(
        "--seed-xlsx",
        type=str,
        default=None,
        help="One-time workbook import that replaces portfolio tables before server starts.",
    )
    args = parser.parse_args()

    xlsx_path = resolve_xlsx_path(args.xlsx_path)
    if args.validate:
        if not xlsx_path:
            searched = "\n".join([f"- {p}" for p in DEFAULT_XLSX_CANDIDATE_PATHS])
            raise RuntimeError(
                "No workbook path found for --validate. Provide --xlsx-path <file> or set PORTFOLIO_XLSX_PATH. "
                + "Searched default relative paths:\n"
                + searched
            )
        validate(str(xlsx_path))
    else:
        if args.seed_xlsx:
            seed_xlsx = _resolve_repo_relative_path(args.seed_xlsx)
            if seed_xlsx is None or not seed_xlsx.exists() or not seed_xlsx.is_file():
                raise RuntimeError(f"seed_xlsx_not_found: {args.seed_xlsx}")
            import_from_excel(str(seed_xlsx))
        serve(args.port)


# Module wiring: keep `app.py` compatibility while delegating software-performance
# agent logic into `portfolio_agent/software_performance.py`.
get_software_perf_agent_config = _mod_get_software_perf_agent_config
set_software_perf_agent_config = _mod_set_software_perf_agent_config
list_software_perf_snapshots = _mod_list_software_perf_snapshots
list_software_perf_actions = _mod_list_software_perf_actions
run_software_perf_agent_once = _mod_run_software_perf_agent_once
maybe_run_software_perf_agent_once = _mod_maybe_run_software_perf_agent_once
software_performance_worker = _mod_software_performance_worker
get_risk_agent_config = _mod_get_risk_agent_config
set_risk_agent_config = _mod_set_risk_agent_config
list_risk_analysis_snapshots = _mod_list_risk_analysis_snapshots
run_risk_analysis_agent_once = _mod_run_risk_analysis_agent_once
maybe_run_risk_analysis_agent_once = _mod_maybe_run_risk_analysis_agent_once
risk_analysis_worker = _mod_risk_analysis_worker


if __name__ == "__main__":
    main()

